#!/usr/bin/env python3
"""
Reset Alert Prices Script

Clears all price data (2-min, 10-min, EOD) and resets status to "Pending"
so that the fixed V2 scripts can repopulate with correct historical data.

Usage:
    python3 reset_alert_prices.py [--all] [--date YYYY-MM-DD]

Examples:
    python3 reset_alert_prices.py --all                    # Reset all alerts
    python3 reset_alert_prices.py --date 2025-11-10        # Reset specific date
    python3 reset_alert_prices.py --date 2025-11-10 --date 2025-11-11  # Multiple dates
"""

import sys
import logging
import argparse
from typing import List, Set
import config
from alert_excel_logger import AlertExcelLogger

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/alert_excel_updates.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class AlertPriceResetter:
    """Resets price data in alert tracking Excel."""

    def __init__(self):
        """Initialize Excel logger."""
        self.excel_logger = AlertExcelLogger(config.ALERT_EXCEL_PATH)
        logger.info(f"Excel logger initialized: {config.ALERT_EXCEL_PATH}")

    def reset_all_prices(self) -> int:
        """
        Clear all price data and reset status for ALL alerts.

        Returns:
            Number of alerts reset
        """
        logger.info("=" * 60)
        logger.info("Resetting ALL alert prices...")
        logger.info("=" * 60)

        reset_count = 0

        try:
            # Process all sheets
            for sheet_name in set(self.excel_logger.SHEET_NAMES.values()):
                ws = self.excel_logger.workbook[sheet_name]

                # Skip if no data
                if ws.max_row < 2:
                    continue

                sheet_reset_count = 0

                # Determine column numbers based on sheet type (RSI columns added)
                if sheet_name == "ATR_Breakout_alerts":
                    price_2min_col = 25  # Column Y (ATR sheet with RSI)
                    price_10min_col = 26  # Column Z
                    price_eod_col = 27   # Column AA
                    status_col = 28      # Column AB
                else:
                    price_2min_col = 22  # Column V (Standard sheets with RSI)
                    price_10min_col = 23  # Column W
                    price_eod_col = 24   # Column X
                    status_col = 25      # Column Y

                # Iterate through all data rows (skip header)
                for row_num in range(2, ws.max_row + 1):
                    symbol = ws.cell(row=row_num, column=3).value  # Column C
                    date = ws.cell(row=row_num, column=1).value    # Column A

                    # Clear price columns
                    ws.cell(row=row_num, column=price_2min_col, value="")  # Price 2min
                    ws.cell(row=row_num, column=price_10min_col, value="")  # Price 10min
                    ws.cell(row=row_num, column=price_eod_col, value="")   # Price EOD

                    # Clear any cell fill colors for these columns
                    from openpyxl.styles import PatternFill
                    no_fill = PatternFill(fill_type=None)
                    ws.cell(row=row_num, column=price_2min_col).fill = no_fill
                    ws.cell(row=row_num, column=price_10min_col).fill = no_fill
                    ws.cell(row=row_num, column=price_eod_col).fill = no_fill

                    # Reset status to "Pending"
                    ws.cell(row=row_num, column=status_col, value="Pending")  # Status

                    sheet_reset_count += 1

                reset_count += sheet_reset_count
                logger.info(f"  {sheet_name}: Reset {sheet_reset_count} alerts")

            # Save workbook
            if reset_count > 0:
                self.excel_logger._save_workbook()
                logger.info(f"✓ Successfully reset {reset_count} alerts")

            return reset_count

        except Exception as e:
            logger.error(f"Error resetting prices: {e}", exc_info=True)
            return reset_count

    def reset_prices_by_date(self, target_dates: List[str]) -> int:
        """
        Clear price data and reset status for specific dates.

        Args:
            target_dates: List of date strings in YYYY-MM-DD format

        Returns:
            Number of alerts reset
        """
        logger.info("=" * 60)
        logger.info(f"Resetting alert prices for dates: {', '.join(target_dates)}")
        logger.info("=" * 60)

        reset_count = 0

        try:
            # Process all sheets
            for sheet_name in set(self.excel_logger.SHEET_NAMES.values()):
                ws = self.excel_logger.workbook[sheet_name]

                # Skip if no data
                if ws.max_row < 2:
                    continue

                sheet_reset_count = 0

                # Determine column numbers based on sheet type (RSI columns added)
                if sheet_name == "ATR_Breakout_alerts":
                    price_2min_col = 25  # Column Y (ATR sheet with RSI)
                    price_10min_col = 26  # Column Z
                    price_eod_col = 27   # Column AA
                    status_col = 28      # Column AB
                else:
                    price_2min_col = 22  # Column V (Standard sheets with RSI)
                    price_10min_col = 23  # Column W
                    price_eod_col = 24   # Column X
                    status_col = 25      # Column Y

                # Iterate through all data rows (skip header)
                for row_num in range(2, ws.max_row + 1):
                    date = ws.cell(row=row_num, column=1).value  # Column A
                    symbol = ws.cell(row=row_num, column=3).value  # Column C

                    # Check if this row's date matches any target date
                    if date in target_dates:
                        # Clear price columns
                        ws.cell(row=row_num, column=price_2min_col, value="")  # Price 2min
                        ws.cell(row=row_num, column=price_10min_col, value="")  # Price 10min
                        ws.cell(row=row_num, column=price_eod_col, value="")   # Price EOD

                        # Clear any cell fill colors
                        from openpyxl.styles import PatternFill
                        no_fill = PatternFill(fill_type=None)
                        ws.cell(row=row_num, column=price_2min_col).fill = no_fill
                        ws.cell(row=row_num, column=price_10min_col).fill = no_fill
                        ws.cell(row=row_num, column=price_eod_col).fill = no_fill

                        # Reset status to "Pending"
                        ws.cell(row=row_num, column=status_col, value="Pending")  # Status

                        sheet_reset_count += 1
                        logger.info(f"  Reset: {symbol} on {date}")

                reset_count += sheet_reset_count

                if sheet_reset_count > 0:
                    logger.info(f"  {sheet_name}: Reset {sheet_reset_count} alerts")

            # Save workbook
            if reset_count > 0:
                self.excel_logger._save_workbook()
                logger.info(f"✓ Successfully reset {reset_count} alerts")

            return reset_count

        except Exception as e:
            logger.error(f"Error resetting prices: {e}", exc_info=True)
            return reset_count

    def show_summary(self):
        """Show summary of current alert status before reset."""
        logger.info("=" * 60)
        logger.info("CURRENT ALERT STATUS SUMMARY")
        logger.info("=" * 60)

        try:
            total_alerts = 0
            complete_count = 0
            partial_count = 0
            pending_count = 0

            for sheet_name in set(self.excel_logger.SHEET_NAMES.values()):
                ws = self.excel_logger.workbook[sheet_name]

                if ws.max_row < 2:
                    continue

                sheet_total = ws.max_row - 1  # Exclude header
                total_alerts += sheet_total

                # Determine status column based on sheet type (RSI columns added)
                if sheet_name == "ATR_Breakout_alerts":
                    status_col = 28  # Column AB (ATR sheet with RSI)
                else:
                    status_col = 25  # Column Y (Standard sheets with RSI)

                # Count status
                for row_num in range(2, ws.max_row + 1):
                    status = ws.cell(row=row_num, column=status_col).value  # Status column

                    if status == "Complete":
                        complete_count += 1
                    elif status == "Partial":
                        partial_count += 1
                    else:
                        pending_count += 1

                logger.info(f"  {sheet_name}: {sheet_total} alerts")

            logger.info("")
            logger.info(f"Total Alerts: {total_alerts}")
            logger.info(f"  - Complete: {complete_count}")
            logger.info(f"  - Partial: {partial_count}")
            logger.info(f"  - Pending: {pending_count}")
            logger.info("=" * 60)

        except Exception as e:
            logger.error(f"Error showing summary: {e}")

    def close(self):
        """Close resources."""
        if self.excel_logger:
            self.excel_logger.close()


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Reset alert prices to allow repopulation with correct historical data"
    )
    parser.add_argument(
        '--all',
        action='store_true',
        help='Reset ALL alerts in the entire workbook'
    )
    parser.add_argument(
        '--date',
        action='append',
        dest='dates',
        help='Reset alerts for specific date(s) in YYYY-MM-DD format (can specify multiple times)'
    )
    parser.add_argument(
        '--summary',
        action='store_true',
        help='Show summary of current status without resetting'
    )

    args = parser.parse_args()

    try:
        resetter = AlertPriceResetter()

        # Show summary if requested
        if args.summary:
            resetter.show_summary()
            resetter.close()
            return 0

        # Validate arguments
        if not args.all and not args.dates:
            logger.error("ERROR: Must specify either --all or --date")
            parser.print_help()
            return 1

        # Show current status before reset
        resetter.show_summary()

        # Confirm reset
        print("\n⚠️  WARNING: This will CLEAR all price data (2-min, 10-min, EOD)")
        print("⚠️  and reset status to 'Pending' for the selected alerts.")
        print()

        if args.all:
            print("🔴 You are about to reset ALL alerts in the workbook!")
        else:
            print(f"🟡 You are about to reset alerts for dates: {', '.join(args.dates)}")

        print()
        response = input("Are you sure you want to continue? (yes/no): ")

        if response.lower() not in ['yes', 'y']:
            logger.info("Reset cancelled by user")
            return 0

        # Perform reset
        if args.all:
            reset_count = resetter.reset_all_prices()
        else:
            reset_count = resetter.reset_prices_by_date(args.dates)

        logger.info("=" * 60)
        logger.info(f"RESET COMPLETE: {reset_count} alerts reset")
        logger.info("=" * 60)
        logger.info("")
        logger.info("Next steps:")
        logger.info("1. Refresh Kite token: ./generate_kite_token.py")
        logger.info("2. Update 2-min prices: ./venv/bin/python3 update_alert_prices_v2.py --2min")
        logger.info("3. Update 10-min prices: ./venv/bin/python3 update_alert_prices_v2.py --10min")
        logger.info("4. Update EOD prices: ./venv/bin/python3 update_eod_prices_v2.py --date YYYY-MM-DD")
        logger.info("")

        resetter.close()
        return 0

    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
        return 1


if __name__ == "__main__":
    sys.exit(main())
