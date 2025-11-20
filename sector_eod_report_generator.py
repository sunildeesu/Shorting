"""
Sector EOD Report Generator - Generate Excel reports for end-of-day sector analysis
Creates comprehensive sector fund flow and performance reports
"""

import os
import logging
from datetime import datetime
from typing import Dict, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class SectorEODReportGenerator:
    """Generates Excel reports for sector analysis at end of day"""

    def __init__(self, report_dir: str = "data/sector_eod_reports"):
        """
        Initialize report generator

        Args:
            report_dir: Base directory for saving reports
        """
        self.report_dir = report_dir
        os.makedirs(report_dir, exist_ok=True)

    def generate_report(self, sector_analysis: Dict, report_date: datetime = None) -> str:
        """
        Generate Excel report for sector analysis

        Args:
            sector_analysis: Sector analysis dict with sectors and metrics
            report_date: Date for the report (defaults to today)

        Returns:
            Path to generated Excel file
        """
        if report_date is None:
            report_date = datetime.now()

        try:
            # Create year/month directory structure
            year = report_date.strftime("%Y")
            month = report_date.strftime("%m")
            report_dir_dated = os.path.join(self.report_dir, year, month)
            os.makedirs(report_dir_dated, exist_ok=True)

            # Generate filename
            filename = f"sector_analysis_{report_date.strftime('%Y%m%d')}.xlsx"
            filepath = os.path.join(report_dir_dated, filename)

            # Create workbook
            wb = openpyxl.Workbook()

            # Generate sheets
            self._create_summary_sheet(wb, sector_analysis, report_date)
            self._create_detailed_metrics_sheet(wb, sector_analysis)
            self._create_fund_flow_sheet(wb, sector_analysis)

            # Remove default sheet if it exists
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            # Save workbook
            wb.save(filepath)
            logger.info(f"Sector EOD report generated: {filepath}")

            return filepath

        except Exception as e:
            logger.error(f"Error generating sector EOD report: {e}", exc_info=True)
            return None

    def _create_summary_sheet(self, wb: openpyxl.Workbook, sector_analysis: Dict, report_date: datetime):
        """Create summary sheet with sector rankings"""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws['A1'] = "SECTOR PERFORMANCE SUMMARY"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Date: {report_date.strftime('%d %b %Y')}"
        ws['A2'].font = Font(size=12)

        # Headers
        headers = [
            "Rank", "Sector", "10-Min Change %", "30-Min Change %",
            "Momentum Score", "Volume Ratio", "Market Cap (Cr)",
            "Stocks Up", "Stocks Down", "Total Stocks", "Breadth %", "Status"
        ]

        header_row = 4
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # Sort sectors by 10-min performance
        sectors = sector_analysis.get('sectors', {})
        sorted_sectors = sorted(
            sectors.items(),
            key=lambda x: x[1].get('price_change_10min', 0),
            reverse=True
        )

        # Data rows
        row = header_row + 1
        for rank, (sector, data) in enumerate(sorted_sectors, 1):
            sector_name = sector.replace('_', ' ').title()
            price_change_10min = data.get('price_change_10min', 0)
            price_change_30min = data.get('price_change_30min', 0)
            momentum = data.get('momentum_score_10min', 0)
            volume_ratio = data.get('volume_ratio', 1.0)
            market_cap = data.get('total_market_cap_cr', 0)
            stocks_up = data.get('stocks_up_10min', 0)
            stocks_down = data.get('stocks_down_10min', 0)
            total_stocks = data.get('total_stocks', 0)
            breadth_pct = (stocks_up / total_stocks * 100) if total_stocks > 0 else 0

            # Determine status
            if price_change_10min > 0.5:
                status = "Strong Inflow"
                status_color = "00B050"
            elif price_change_10min > 0:
                status = "Inflow"
                status_color = "92D050"
            elif price_change_10min > -0.5:
                status = "Outflow"
                status_color = "FFC000"
            else:
                status = "Strong Outflow"
                status_color = "FF0000"

            # Write data
            ws.cell(row=row, column=1, value=rank)
            ws.cell(row=row, column=2, value=sector_name)
            ws.cell(row=row, column=3, value=price_change_10min)
            ws.cell(row=row, column=4, value=price_change_30min)
            ws.cell(row=row, column=5, value=momentum)
            ws.cell(row=row, column=6, value=volume_ratio)
            ws.cell(row=row, column=7, value=market_cap)
            ws.cell(row=row, column=8, value=stocks_up)
            ws.cell(row=row, column=9, value=stocks_down)
            ws.cell(row=row, column=10, value=total_stocks)
            ws.cell(row=row, column=11, value=breadth_pct)
            ws.cell(row=row, column=12, value=status)

            # Format numbers
            ws.cell(row=row, column=3).number_format = '0.00'
            ws.cell(row=row, column=4).number_format = '0.00'
            ws.cell(row=row, column=5).number_format = '0.00'
            ws.cell(row=row, column=6).number_format = '0.00'
            ws.cell(row=row, column=7).number_format = '#,##0'
            ws.cell(row=row, column=11).number_format = '0.0'

            # Color status cell
            ws.cell(row=row, column=12).fill = PatternFill(
                start_color=status_color, end_color=status_color, fill_type="solid"
            )
            ws.cell(row=row, column=12).font = Font(bold=True, color="FFFFFF")

            # Color price change cells
            if price_change_10min > 0:
                ws.cell(row=row, column=3).font = Font(color="00B050", bold=True)
            elif price_change_10min < 0:
                ws.cell(row=row, column=3).font = Font(color="FF0000", bold=True)

            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 10
        ws.column_dimensions['L'].width = 15

    def _create_detailed_metrics_sheet(self, wb: openpyxl.Workbook, sector_analysis: Dict):
        """Create detailed metrics sheet with all timeframes"""
        ws = wb.create_sheet("Detailed Metrics")

        # Title
        ws['A1'] = "DETAILED SECTOR METRICS"
        ws['A1'].font = Font(size=14, bold=True)

        # Headers
        headers = [
            "Sector",
            "5-Min Change %", "10-Min Change %", "30-Min Change %",
            "5-Min Momentum", "10-Min Momentum", "30-Min Momentum",
            "Volume (Current)", "Volume (Avg)", "Volume Ratio",
            "Market Cap (Cr)", "Total Stocks", "Participation %"
        ]

        header_row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Sort sectors alphabetically
        sectors = sector_analysis.get('sectors', {})
        sorted_sectors = sorted(sectors.items(), key=lambda x: x[0])

        # Data rows
        row = header_row + 1
        for sector, data in sorted_sectors:
            sector_name = sector.replace('_', ' ').title()

            ws.cell(row=row, column=1, value=sector_name)
            ws.cell(row=row, column=2, value=data.get('price_change_5min', 0))
            ws.cell(row=row, column=3, value=data.get('price_change_10min', 0))
            ws.cell(row=row, column=4, value=data.get('price_change_30min', 0))
            ws.cell(row=row, column=5, value=data.get('momentum_score_5min', 0))
            ws.cell(row=row, column=6, value=data.get('momentum_score_10min', 0))
            ws.cell(row=row, column=7, value=data.get('momentum_score_30min', 0))
            ws.cell(row=row, column=8, value=data.get('total_volume', 0))
            ws.cell(row=row, column=9, value=data.get('total_volume', 0) / data.get('volume_ratio', 1.0) if data.get('volume_ratio', 1.0) != 0 else 0)
            ws.cell(row=row, column=10, value=data.get('volume_ratio', 1.0))
            ws.cell(row=row, column=11, value=data.get('total_market_cap_cr', 0))
            ws.cell(row=row, column=12, value=data.get('total_stocks', 0))
            ws.cell(row=row, column=13, value=data.get('participation_pct', 0))

            # Format numbers
            for col in [2, 3, 4, 5, 6, 7]:
                ws.cell(row=row, column=col).number_format = '0.00'
            for col in [8, 9, 11]:
                ws.cell(row=row, column=col).number_format = '#,##0'
            ws.cell(row=row, column=10).number_format = '0.00'
            ws.cell(row=row, column=13).number_format = '0.0'

            row += 1

        # Adjust column widths
        for col in range(1, 14):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_fund_flow_sheet(self, wb: openpyxl.Workbook, sector_analysis: Dict):
        """Create fund flow analysis sheet"""
        ws = wb.create_sheet("Fund Flow")

        # Title
        ws['A1'] = "SECTOR FUND FLOW ANALYSIS"
        ws['A1'].font = Font(size=14, bold=True)

        # Calculate overall market stats
        sectors = sector_analysis.get('sectors', {})
        total_market_cap = sum(s.get('total_market_cap_cr', 0) for s in sectors.values())
        total_stocks = sum(s.get('total_stocks', 0) for s in sectors.values())
        total_up = sum(s.get('stocks_up_10min', 0) for s in sectors.values())
        total_down = sum(s.get('stocks_down_10min', 0) for s in sectors.values())

        # Market summary
        ws['A3'] = "Market Summary"
        ws['A3'].font = Font(bold=True, size=12)
        ws['A4'] = f"Total Market Cap: ₹{total_market_cap:,.0f} Cr"
        ws['A5'] = f"Total Stocks: {total_stocks}"
        ws['A6'] = f"Stocks Up: {total_up} ({total_up/total_stocks*100:.1f}%)" if total_stocks > 0 else "Stocks Up: 0"
        ws['A7'] = f"Stocks Down: {total_down} ({total_down/total_stocks*100:.1f}%)" if total_stocks > 0 else "Stocks Down: 0"

        # Fund Flow table
        ws['A9'] = "FUND FLOW BY SECTOR"
        ws['A9'].font = Font(bold=True, size=12)

        headers = [
            "Sector", "Market Cap (Cr)", "% of Total", "10-Min Change %",
            "Implied Flow (Cr)", "Volume Ratio", "Momentum", "Flow Status"
        ]

        header_row = 10
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Sort by market cap
        sorted_sectors = sorted(
            sectors.items(),
            key=lambda x: x[1].get('total_market_cap_cr', 0),
            reverse=True
        )

        # Data rows
        row = header_row + 1
        for sector, data in sorted_sectors:
            sector_name = sector.replace('_', ' ').title()
            market_cap = data.get('total_market_cap_cr', 0)
            pct_of_total = (market_cap / total_market_cap * 100) if total_market_cap > 0 else 0
            price_change = data.get('price_change_10min', 0)
            implied_flow = market_cap * (price_change / 100)  # Approximate fund flow
            volume_ratio = data.get('volume_ratio', 1.0)
            momentum = data.get('momentum_score_10min', 0)

            # Determine flow status
            if price_change > 0.5 and volume_ratio > 1.2:
                flow_status = "Strong Inflow"
                status_color = "00B050"
            elif price_change > 0:
                flow_status = "Moderate Inflow"
                status_color = "92D050"
            elif price_change > -0.5:
                flow_status = "Moderate Outflow"
                status_color = "FFC000"
            else:
                flow_status = "Strong Outflow"
                status_color = "FF0000"

            ws.cell(row=row, column=1, value=sector_name)
            ws.cell(row=row, column=2, value=market_cap)
            ws.cell(row=row, column=3, value=pct_of_total)
            ws.cell(row=row, column=4, value=price_change)
            ws.cell(row=row, column=5, value=implied_flow)
            ws.cell(row=row, column=6, value=volume_ratio)
            ws.cell(row=row, column=7, value=momentum)
            ws.cell(row=row, column=8, value=flow_status)

            # Format numbers
            ws.cell(row=row, column=2).number_format = '#,##0'
            ws.cell(row=row, column=3).number_format = '0.0'
            ws.cell(row=row, column=4).number_format = '0.00'
            ws.cell(row=row, column=5).number_format = '#,##0'
            ws.cell(row=row, column=6).number_format = '0.00'
            ws.cell(row=row, column=7).number_format = '0.00'

            # Color status cell
            ws.cell(row=row, column=8).fill = PatternFill(
                start_color=status_color, end_color=status_color, fill_type="solid"
            )
            ws.cell(row=row, column=8).font = Font(bold=True, color="FFFFFF")

            # Color implied flow
            if implied_flow > 0:
                ws.cell(row=row, column=5).font = Font(color="00B050", bold=True)
            elif implied_flow < 0:
                ws.cell(row=row, column=5).font = Font(color="FF0000", bold=True)

            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 18


# Global singleton instance
_sector_eod_report_generator: Optional[SectorEODReportGenerator] = None


def get_sector_eod_report_generator() -> SectorEODReportGenerator:
    """
    Get singleton instance of SectorEODReportGenerator

    Returns:
        SectorEODReportGenerator instance
    """
    global _sector_eod_report_generator
    if _sector_eod_report_generator is None:
        _sector_eod_report_generator = SectorEODReportGenerator()
    return _sector_eod_report_generator
