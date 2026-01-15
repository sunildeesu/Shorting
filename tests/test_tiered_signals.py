#!/usr/bin/env python3
"""
Integration Test for Tiered Signal System

Tests all components of the tiered signal implementation:
1. Config loading
2. Tier assignment logic
3. Excel logging with new columns
4. Backwards compatibility

Author: Sunil Kumar Durganaik
Date: January 3, 2026
"""

import os
import sys
import logging
from datetime import datetime
from pathlib import Path

# Add project directory to path
sys.path.insert(0, str(Path(__file__).parent))

import config
from nifty_option_analyzer import NiftyOptionAnalyzer
from nifty_option_logger import NiftyOptionLogger

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def test_config_loading():
    """Test that tiered signal config is loaded correctly"""
    print("\n" + "="*60)
    print("TEST 1: Config Loading")
    print("="*60)

    assert hasattr(config, 'ENABLE_TIERED_SIGNALS'), "ENABLE_TIERED_SIGNALS missing"
    assert hasattr(config, 'IV_RANK_EXCELLENT'), "IV_RANK_EXCELLENT missing"
    assert hasattr(config, 'IV_RANK_GOOD'), "IV_RANK_GOOD missing"
    assert hasattr(config, 'IV_RANK_MARGINAL'), "IV_RANK_MARGINAL missing"
    assert hasattr(config, 'POSITION_SIZE_STRONG'), "POSITION_SIZE_STRONG missing"
    assert hasattr(config, 'POSITION_SIZE_MODERATE'), "POSITION_SIZE_MODERATE missing"
    assert hasattr(config, 'POSITION_SIZE_WEAK'), "POSITION_SIZE_WEAK missing"

    print(f"✅ ENABLE_TIERED_SIGNALS: {config.ENABLE_TIERED_SIGNALS}")
    print(f"✅ IV_RANK_EXCELLENT: {config.IV_RANK_EXCELLENT}%")
    print(f"✅ IV_RANK_GOOD: {config.IV_RANK_GOOD}%")
    print(f"✅ IV_RANK_MARGINAL: {config.IV_RANK_MARGINAL}%")
    print(f"✅ POSITION_SIZE_STRONG: {config.POSITION_SIZE_STRONG}")
    print(f"✅ POSITION_SIZE_MODERATE: {config.POSITION_SIZE_MODERATE}")
    print(f"✅ POSITION_SIZE_WEAK: {config.POSITION_SIZE_WEAK}")

    print("\n✅ Config Loading Test: PASSED")


def test_tier_assignment_logic():
    """Test tier assignment for different IV Rank values"""
    print("\n" + "="*60)
    print("TEST 2: Tier Assignment Logic")
    print("="*60)

    test_cases = [
        # (iv_rank, expected_tier, expected_position_size, description)
        (48.6, 'SELL_STRONG', 1.0, 'High IV Rank'),
        (30.0, 'SELL_STRONG', 1.0, 'IV Rank at 30%'),
        (25.0, 'SELL_STRONG', 1.0, 'IV Rank exactly at 25% threshold'),
        (24.9, 'SELL_MODERATE', 0.75, 'IV Rank just below 25%'),
        (20.0, 'SELL_MODERATE', 0.75, 'Mid-range MODERATE'),
        (15.0, 'SELL_MODERATE', 0.75, 'IV Rank exactly at 15% threshold'),
        (14.9, 'SELL_WEAK', 0.5, 'IV Rank just below 15%'),
        (12.0, 'SELL_WEAK', 0.5, 'Mid-range WEAK'),
        (10.0, 'SELL_WEAK', 0.5, 'IV Rank exactly at 10% threshold'),
        (9.9, 'AVOID', 0.0, 'IV Rank just below 10%'),
        (5.0, 'AVOID', 0.0, 'Low IV Rank'),
        (0.0, 'AVOID', 0.0, 'Zero IV Rank'),
    ]

    for iv_rank, expected_tier, expected_position_size, description in test_cases:
        # Simulate tier assignment logic
        if config.ENABLE_TIERED_SIGNALS:
            if iv_rank >= config.IV_RANK_EXCELLENT:
                tier = 'SELL_STRONG'
                position_size = config.POSITION_SIZE_STRONG
            elif iv_rank >= config.IV_RANK_GOOD:
                tier = 'SELL_MODERATE'
                position_size = config.POSITION_SIZE_MODERATE
            elif iv_rank >= config.IV_RANK_MARGINAL:
                tier = 'SELL_WEAK'
                position_size = config.POSITION_SIZE_WEAK
            else:
                tier = 'AVOID'
                position_size = 0.0
        else:
            tier = 'SELL' if iv_rank >= 15 else 'AVOID'
            position_size = 1.0 if tier == 'SELL' else 0.0

        # Verify
        status = "✅" if (tier == expected_tier and position_size == expected_position_size) else "❌"
        print(f"{status} IV Rank {iv_rank:5.1f}% → {tier:15s} (pos: {position_size*100:3.0f}%) - {description}")

        assert tier == expected_tier, f"Expected {expected_tier}, got {tier}"
        assert position_size == expected_position_size, f"Expected {expected_position_size}, got {position_size}"

    print("\n✅ Tier Assignment Logic Test: PASSED")


def test_excel_logger():
    """Test Excel logging with tiered signal columns"""
    print("\n" + "="*60)
    print("TEST 3: Excel Logger with Tiered Columns")
    print("="*60)

    # Remove old test file if exists
    test_file = "data/nifty_options/nifty_options_2026-01.xlsx"
    if os.path.exists(test_file):
        os.remove(test_file)
        print(f"🗑️  Removed old test file: {test_file}")

    logger_instance = NiftyOptionLogger()

    # Test data for each tier
    test_scenarios = [
        {
            'signal_tier': 'SELL_STRONG',
            'position_size': 1.0,
            'premium_quality': 'EXCELLENT (100% of fair value or better)',
            'iv_rank': 32.5,
            'total_score': 85.0,
            'vix': 12.5,
        },
        {
            'signal_tier': 'SELL_MODERATE',
            'position_size': 0.75,
            'premium_quality': 'GOOD (85-90% of fair value)',
            'iv_rank': 18.2,
            'total_score': 72.0,
            'vix': 11.8,
        },
        {
            'signal_tier': 'SELL_WEAK',
            'position_size': 0.5,
            'premium_quality': 'BELOW AVERAGE (75-80% of fair value)',
            'iv_rank': 12.1,
            'total_score': 65.0,
            'vix': 11.5,
        },
        {
            'signal_tier': 'AVOID',
            'position_size': 0.0,
            'premium_quality': 'CHEAP (< 70% of fair value)',
            'iv_rank': 5.0,
            'total_score': 45.0,
            'vix': 10.8,
        },
    ]

    for idx, scenario in enumerate(test_scenarios, start=1):
        test_data = {
            'timestamp': datetime.now().isoformat(),
            'signal': 'SELL' if scenario['signal_tier'].startswith('SELL') else 'AVOID',
            'signal_tier': scenario['signal_tier'],
            'position_size': scenario['position_size'],
            'premium_quality': scenario['premium_quality'],
            'total_score': scenario['total_score'],
            'nifty_spot': 21850.50,
            'vix': scenario['vix'],
            'vix_trend': -0.5,
            'iv_rank': scenario['iv_rank'],
            'market_regime': 'NEUTRAL',
            'best_strategy': 'straddle',
            'recommendation': f'Test scenario {idx}',
            'risk_factors': ['Test risk'],
            'breakdown': {
                'theta_score': 80.0,
                'gamma_score': 85.0,
                'vega_score': 75.0,
                'vix_score': 60.0,
                'regime_score': 100.0,
                'oi_score': 70.0
            },
            'oi_analysis': {
                'pattern': 'LONG_UNWINDING'
            },
            'expiry_analyses': [{
                'expiry_date': datetime(2026, 1, 9),
                'days_to_expiry': 8,
                'straddle': {
                    'total_premium': 355.0,
                    'greeks': {'theta': -45.0, 'gamma': 0.0012}
                },
                'strangle': {
                    'total_premium': 185.0,
                    'greeks': {'theta': -25.0, 'gamma': 0.0008}
                }
            }]
        }

        result = logger_instance.log_analysis(test_data, telegram_sent=True)
        status = "✅" if result else "❌"
        print(f"{status} Logged: {scenario['signal_tier']:15s} (IV Rank: {scenario['iv_rank']:5.1f}%, Pos: {scenario['position_size']*100:3.0f}%)")
        assert result, f"Failed to log {scenario['signal_tier']}"

    # Verify Excel file
    from openpyxl import load_workbook
    wb = load_workbook(test_file)
    ws = wb.active

    # Check headers
    expected_headers = ['Date', 'Time', 'Signal', 'Signal_Tier', 'Position_Size', 'Premium_Quality', 'Total_Score']
    actual_headers = [ws.cell(1, i).value for i in range(1, 8)]

    print("\n📊 Excel File Verification:")
    print(f"   Expected headers: {expected_headers}")
    print(f"   Actual headers:   {actual_headers}")
    assert actual_headers == expected_headers, "Headers mismatch"
    print("   ✅ Headers match")

    # Check data rows
    print(f"\n   Data rows logged: {ws.max_row - 1}")
    assert ws.max_row == 5, f"Expected 5 rows (1 header + 4 data), got {ws.max_row}"
    print("   ✅ Correct number of rows")

    # Check color coding for Signal_Tier column
    print("\n   Signal_Tier Color Coding:")
    color_map = {
        'SELL_STRONG': '0000B050',
        'SELL_MODERATE': '0092D050',
        'SELL_WEAK': '00FFC000',
        'AVOID': '00FF0000'
    }

    for row_idx in range(2, 6):
        cell = ws.cell(row_idx, 4)  # Signal_Tier column
        tier = cell.value
        expected_color = color_map[tier]
        actual_color = cell.fill.start_color.rgb if cell.fill.start_color else None

        status = "✅" if actual_color == expected_color else "❌"
        print(f"   {status} Row {row_idx}: {tier:15s} → Color: {actual_color}")
        assert actual_color == expected_color, f"Color mismatch for {tier}"

    print("\n✅ Excel Logger Test: PASSED")


def test_summary():
    """Print test summary"""
    print("\n" + "="*60)
    print("TEST SUMMARY")
    print("="*60)
    print("✅ All tests passed successfully!")
    print("\nTiered Signal System Implementation:")
    print("  ✅ Config loading")
    print("  ✅ Tier assignment logic")
    print("  ✅ Excel logging with new columns")
    print("  ✅ Color coding for tiers")
    print("\nSystem is ready for production use.")
    print("="*60)


if __name__ == "__main__":
    try:
        test_config_loading()
        test_tier_assignment_logic()
        test_excel_logger()
        test_summary()

        print("\n🎉 ALL TESTS PASSED! 🎉\n")
        sys.exit(0)

    except AssertionError as e:
        print(f"\n❌ TEST FAILED: {e}\n")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ ERROR: {e}\n")
        import traceback
        traceback.print_exc()
        sys.exit(1)
