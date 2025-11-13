"""
Alert Excel Logger

Logs all trading alerts (5min, 10min, 30min, volume_spike) to a cumulative Excel file
with separate sheets per alert type. Supports later price updates for tracking.
"""

import os
import logging
from datetime import datetime
from typing import Dict, Optional, List
from pathlib import Path
import fcntl
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class AlertExcelLogger:
    """
    Manages cumulative Excel workbook for alert tracking with price history.

    Structure:
    - Single file: data/alerts/alert_tracking.xlsx
    - 4 sheets: 5min_alerts, 10min_alerts, 30min_alerts, Volume_Spike_alerts
    - Each sheet tracks: alert details, prices at 2min/10min/EOD
    """

    # Sheet names for different alert types
    SHEET_NAMES = {
        "5min": "5min_alerts",
        "5min_rise": "5min_alerts",
        "10min": "10min_alerts",
        "10min_rise": "10min_alerts",
        "30min": "30min_alerts",
        "30min_rise": "30min_alerts",
        "volume_spike": "Volume_Spike_alerts",
        "volume_spike_rise": "Volume_Spike_alerts",
        "atr_breakout": "ATR_Breakout_alerts"
    }

    # Column headers for standard drop/rise alerts
    HEADERS = [
        "Date", "Time", "Symbol", "Direction",
        "Alert Price", "Previous Price", "Change %", "Change (Rs)",
        "Volume", "Avg Volume", "Volume Multiplier",
        "Market Cap (Cr)", "Telegram Sent",
        "RSI(9)", "RSI(14)", "RSI(21)",
        "RSI 9vs14", "RSI 9vs21", "RSI 14vs21",
        "RSI Recent Cross", "RSI Summary",
        "Price 2min", "Price 10min", "Price EOD",
        "Status", "Row ID"
    ]

    # Column headers for ATR breakout alerts
    ATR_HEADERS = [
        "Date", "Time", "Symbol",
        "Open", "Entry Level", "Current Price", "Breakout Distance",
        "ATR(20)", "ATR(30)", "Volatility Filter",
        "Stop Loss", "Risk Amount", "Risk %",
        "Volume", "Market Cap (Cr)", "Telegram Sent",
        "RSI(9)", "RSI(14)", "RSI(21)",
        "RSI 9vs14", "RSI 9vs21", "RSI 14vs21",
        "RSI Recent Cross", "RSI Summary",
        "Price 2min", "Price 10min", "Price EOD",
        "Status", "Row ID", "Day of Week"
    ]

    def __init__(self, excel_path: str):
        """
        Initialize AlertExcelLogger.

        Args:
            excel_path: Path to Excel file (will be created if not exists)
        """
        self.excel_path = excel_path
        self.workbook = None
        self.file_handle = None

        # Ensure directory exists
        Path(excel_path).parent.mkdir(parents=True, exist_ok=True)

        # Load or create workbook
        self._load_or_create_workbook()

    def _load_or_create_workbook(self):
        """Load existing workbook or create new one with sheets and headers."""
        if os.path.exists(self.excel_path):
            try:
                self.workbook = openpyxl.load_workbook(self.excel_path)
                logger.info(f"Loaded existing Excel workbook: {self.excel_path}")

                # Ensure all required sheets exist
                for sheet_name in set(self.SHEET_NAMES.values()):
                    if sheet_name not in self.workbook.sheetnames:
                        self._create_sheet(sheet_name)
                        logger.info(f"Added missing sheet: {sheet_name}")
            except Exception as e:
                logger.error(f"Error loading workbook: {e}. Creating new workbook.")
                self._create_new_workbook()
        else:
            self._create_new_workbook()

    def _create_new_workbook(self):
        """Create new workbook with all required sheets and headers."""
        self.workbook = openpyxl.Workbook()

        # Remove default sheet
        if "Sheet" in self.workbook.sheetnames:
            del self.workbook["Sheet"]

        # Create all required sheets
        for sheet_name in set(self.SHEET_NAMES.values()):
            self._create_sheet(sheet_name)

        logger.info(f"Created new Excel workbook: {self.excel_path}")

    def _create_sheet(self, sheet_name: str):
        """
        Create a new sheet with formatted headers.

        Args:
            sheet_name: Name of sheet to create
        """
        ws = self.workbook.create_sheet(sheet_name)

        # Determine which headers to use based on sheet type
        if sheet_name == "ATR_Breakout_alerts":
            headers = self.ATR_HEADERS
            column_widths = {
                'A': 12,  # Date
                'B': 10,  # Time
                'C': 12,  # Symbol
                'D': 12,  # Open
                'E': 12,  # Entry Level
                'F': 12,  # Current Price
                'G': 14,  # Breakout Distance
                'H': 10,  # ATR(20)
                'I': 10,  # ATR(30)
                'J': 12,  # Volatility Filter
                'K': 12,  # Stop Loss
                'L': 12,  # Risk Amount
                'M': 10,  # Risk %
                'N': 13,  # Volume
                'O': 13,  # Market Cap
                'P': 12,  # Telegram
                'Q': 10,  # RSI(9)
                'R': 10,  # RSI(14)
                'S': 10,  # RSI(21)
                'T': 13,  # RSI 9vs14
                'U': 13,  # RSI 9vs21
                'V': 13,  # RSI 14vs21
                'W': 18,  # RSI Recent Cross
                'X': 15,  # RSI Summary
                'Y': 12,  # Price 2min
                'Z': 12,  # Price 10min
                'AA': 12,  # Price EOD
                'AB': 10,  # Status
                'AC': 25,  # Row ID
                'AD': 12   # Day of Week
            }
        else:
            headers = self.HEADERS
            column_widths = {
                'A': 12,  # Date
                'B': 10,  # Time
                'C': 12,  # Symbol
                'D': 10,  # Direction
                'E': 12,  # Alert Price
                'F': 12,  # Previous Price
                'G': 10,  # Change %
                'H': 11,  # Change Rs
                'I': 13,  # Volume
                'J': 13,  # Avg Volume
                'K': 12,  # Volume Mult
                'L': 13,  # Market Cap
                'M': 12,  # Telegram
                'N': 10,  # RSI(9)
                'O': 10,  # RSI(14)
                'P': 10,  # RSI(21)
                'Q': 13,  # RSI 9vs14
                'R': 13,  # RSI 9vs21
                'S': 13,  # RSI 14vs21
                'T': 18,  # RSI Recent Cross
                'U': 15,  # RSI Summary
                'V': 12,  # Price 2min
                'W': 12,  # Price 10min
                'X': 12,  # Price EOD
                'Y': 10,  # Status
                'Z': 25   # Row ID
            }

        # Write headers
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)

            # Header formatting
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # Set column widths
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

    def log_alert(
        self,
        symbol: str,
        alert_type: str,
        drop_percent: float,
        current_price: float,
        previous_price: float,
        volume_data: Optional[Dict] = None,
        market_cap_cr: Optional[float] = None,
        telegram_sent: bool = False,
        timestamp: Optional[datetime] = None,
        rsi_analysis: Optional[Dict] = None
    ) -> bool:
        """
        Log an alert to the appropriate sheet.

        Args:
            symbol: Stock symbol (e.g., "RELIANCE")
            alert_type: Type of alert (5min, 10min, 30min, volume_spike, etc.)
            drop_percent: Percentage change (positive for rise, negative for drop)
            current_price: Current stock price
            previous_price: Previous stock price
            volume_data: Optional volume information dict
            market_cap_cr: Optional market cap in crores
            telegram_sent: Whether Telegram alert was sent
            timestamp: Optional timestamp (defaults to now)
            rsi_analysis: Optional RSI analysis dict with RSI values and crossovers

        Returns:
            True if logged successfully, False otherwise
        """
        try:
            # Determine sheet name
            sheet_name = self.SHEET_NAMES.get(alert_type)
            if not sheet_name:
                logger.error(f"Unknown alert type: {alert_type}")
                return False

            # Get worksheet
            ws = self.workbook[sheet_name]

            # Prepare timestamp
            if timestamp is None:
                timestamp = datetime.now()

            date_str = timestamp.strftime("%Y-%m-%d")
            time_str = timestamp.strftime("%H:%M:%S")

            # Determine direction based on alert type
            # - Rise alerts have "_rise" suffix (5min_rise, 10min_rise, etc.)
            # - Drop alerts have no suffix (5min, 10min, volume_spike, etc.)
            # The drop_percent value (positive or negative) indicates if prediction was correct,
            # but doesn't change the alert direction
            direction = "Rise" if "_rise" in alert_type else "Drop"

            # Calculate absolute change
            change_rs = previous_price - current_price

            # Extract volume data
            volume = volume_data.get("current_volume", "") if volume_data else ""
            avg_volume = volume_data.get("avg_volume", "") if volume_data else ""
            volume_multiplier = ""
            if volume_data and avg_volume and avg_volume > 0:
                volume_multiplier = f"{volume / avg_volume:.2f}x"

            # Generate unique row ID
            row_id = f"{symbol}_{alert_type}_{timestamp.strftime('%Y%m%d_%H%M%S')}"

            # Find next empty row
            next_row = ws.max_row + 1

            # Extract RSI data
            rsi_9 = rsi_14 = rsi_21 = ""
            rsi_9vs14 = rsi_9vs21 = rsi_14vs21 = ""
            rsi_recent_cross = rsi_summary = ""

            if rsi_analysis:
                rsi_9 = round(rsi_analysis.get('rsi_9'), 2) if rsi_analysis.get('rsi_9') else ""
                rsi_14 = round(rsi_analysis.get('rsi_14'), 2) if rsi_analysis.get('rsi_14') else ""
                rsi_21 = round(rsi_analysis.get('rsi_21'), 2) if rsi_analysis.get('rsi_21') else ""

                # Format crossover status
                crossovers = rsi_analysis.get('crossovers', {})
                if '9_14' in crossovers:
                    c = crossovers['9_14']
                    if c.get('status') and c.get('strength') is not None:
                        arrow = "↑" if c['status'] == 'above' else "↓"
                        sign = "+" if c['strength'] >= 0 else ""
                        rsi_9vs14 = f"9{arrow}14 ({sign}{c['strength']})"

                if '9_21' in crossovers:
                    c = crossovers['9_21']
                    if c.get('status') and c.get('strength') is not None:
                        arrow = "↑" if c['status'] == 'above' else "↓"
                        sign = "+" if c['strength'] >= 0 else ""
                        rsi_9vs21 = f"9{arrow}21 ({sign}{c['strength']})"

                if '14_21' in crossovers:
                    c = crossovers['14_21']
                    if c.get('status') and c.get('strength') is not None:
                        arrow = "↑" if c['status'] == 'above' else "↓"
                        sign = "+" if c['strength'] >= 0 else ""
                        rsi_14vs21 = f"14{arrow}21 ({sign}{c['strength']})"

                # Find most recent crossover across all pairs
                recent_crosses = []
                for pair, c in crossovers.items():
                    recent = c.get('recent_cross', {})
                    if recent.get('occurred'):
                        bars_ago = recent.get('bars_ago', 0)
                        direction_text = recent.get('direction', '').capitalize()
                        emoji = "🟢" if direction_text == 'Bullish' else "🔴"
                        recent_crosses.append(f"{emoji} {direction_text} {bars_ago}b ago")

                rsi_recent_cross = "; ".join(recent_crosses) if recent_crosses else "None"
                rsi_summary = rsi_analysis.get('summary', "")

            # Write data
            row_data = [
                date_str,                           # Date
                time_str,                           # Time
                symbol,                             # Symbol
                direction,                          # Direction
                round(current_price, 2),           # Alert Price
                round(previous_price, 2),          # Previous Price
                round(drop_percent, 2),            # Change %
                round(change_rs, 2),               # Change (Rs)
                volume,                             # Volume
                avg_volume,                         # Avg Volume
                volume_multiplier,                  # Volume Multiplier
                round(market_cap_cr, 2) if market_cap_cr else "",  # Market Cap
                "Yes" if telegram_sent else "No",   # Telegram Sent
                rsi_9,                              # RSI(9)
                rsi_14,                             # RSI(14)
                rsi_21,                             # RSI(21)
                rsi_9vs14,                          # RSI 9vs14
                rsi_9vs21,                          # RSI 9vs21
                rsi_14vs21,                         # RSI 14vs21
                rsi_recent_cross,                   # RSI Recent Cross
                rsi_summary,                        # RSI Summary
                "",                                 # Price 2min (to be filled)
                "",                                 # Price 10min (to be filled)
                "",                                 # Price EOD (to be filled)
                "Pending",                          # Status
                row_id                              # Row ID
            ]

            for col_num, value in enumerate(row_data, start=1):
                cell = ws.cell(row=next_row, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # Apply borders
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            # Save workbook with file locking
            self._save_workbook()

            logger.info(f"Logged alert: {symbol} {alert_type} at {timestamp}")
            return True

        except Exception as e:
            logger.error(f"Error logging alert: {e}", exc_info=True)
            return False

    def get_pending_updates(self, min_age_minutes: int = 0) -> Dict[str, List[Dict]]:
        """
        Get all alerts that need price updates (Status != 'Complete').

        Args:
            min_age_minutes: Minimum age of alert in minutes to include

        Returns:
            Dict with sheet names as keys and list of pending alert data as values
        """
        pending_updates = {}

        try:
            for sheet_name in set(self.SHEET_NAMES.values()):
                ws = self.workbook[sheet_name]
                sheet_pending = []

                # Iterate through rows (skip header)
                for row_num in range(2, ws.max_row + 1):
                    # Determine column numbers based on sheet type
                    if sheet_name == "ATR_Breakout_alerts":
                        status_col = 28  # AB
                        row_id_col = 29  # AC
                    else:
                        status_col = 25  # Y
                        row_id_col = 26  # Z

                    status = ws.cell(row=row_num, column=status_col).value

                    if status != "Complete":
                        # Extract alert data
                        date_str = ws.cell(row=row_num, column=1).value
                        time_str = ws.cell(row=row_num, column=2).value
                        symbol = ws.cell(row=row_num, column=3).value
                        row_id = ws.cell(row=row_num, column=row_id_col).value

                        # Check age if required
                        if min_age_minutes > 0:
                            try:
                                alert_time = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
                                age_minutes = (datetime.now() - alert_time).total_seconds() / 60
                                if age_minutes < min_age_minutes:
                                    continue
                            except:
                                pass  # Include if we can't parse time

                        sheet_pending.append({
                            "row_num": row_num,
                            "symbol": symbol,
                            "date": date_str,
                            "time": time_str,
                            "row_id": row_id,
                            "status": status
                        })

                if sheet_pending:
                    pending_updates[sheet_name] = sheet_pending

            return pending_updates

        except Exception as e:
            logger.error(f"Error getting pending updates: {e}", exc_info=True)
            return {}

    def _get_color_for_price_change(
        self,
        direction: str,
        percent_change: float
    ) -> Optional[PatternFill]:
        """
        Calculate cell color based on price change relative to 2min price.

        Args:
            direction: "Drop" or "Rise"
            percent_change: Percentage change from 2min price (positive = price increased)

        Returns:
            PatternFill object with appropriate color, or None if no color
        """
        # For DROP alerts:
        # - Negative change (price dropped) = Green (good)
        # - Positive change (price rose) = Red (bad)
        # For RISE alerts: opposite

        if direction == "Drop":
            is_good = percent_change < 0  # Price dropped further
        else:  # Rise
            is_good = percent_change > 0  # Price rose further

        # Calculate color intensity based on magnitude
        abs_change = abs(percent_change)

        # Color scale: 0-0.5% (light), 0.5-1.5% (medium), 1.5%+ (dark)
        if abs_change < 0.5:
            intensity = "light"
        elif abs_change < 1.5:
            intensity = "medium"
        else:
            intensity = "dark"

        # Green shades for good movement
        green_colors = {
            "light": "C6EFCE",    # Light green
            "medium": "92D050",   # Medium green
            "dark": "00B050"      # Dark green
        }

        # Red shades for bad movement
        red_colors = {
            "light": "FFC7CE",    # Light red
            "medium": "FF6B6B",   # Medium red
            "dark": "C00000"      # Dark red
        }

        if is_good:
            color = green_colors[intensity]
        else:
            color = red_colors[intensity]

        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def update_prices(
        self,
        updates: List[Dict[str, any]],
        price_column: str,  # "2min", "10min", or "EOD"
        auto_complete_eod: bool = False
    ) -> int:
        """
        Update price columns for specific alerts.

        Args:
            updates: List of dicts with keys: row_id, sheet_name, price
            price_column: Which price column to update ("2min", "10min", "EOD")
            auto_complete_eod: If True and price_column is "EOD", mark status as Complete

        Returns:
            Number of rows updated
        """
        # Column mappings for standard alerts (with RSI columns)
        standard_column_map = {
            "2min": 22,   # Column V
            "10min": 23,  # Column W
            "EOD": 24     # Column X
        }

        # Column mappings for ATR alerts (with RSI columns)
        atr_column_map = {
            "2min": 25,   # Column Y
            "10min": 26,  # Column Z
            "EOD": 27     # Column AA
        }

        if price_column not in standard_column_map:
            logger.error(f"Invalid price column: {price_column}")
            return 0

        updated_count = 0

        try:
            for update in updates:
                sheet_name = update.get("sheet_name")
                row_id = update.get("row_id")
                price = update.get("price")

                if not all([sheet_name, row_id, price is not None]):
                    continue

                # Determine column numbers based on sheet type
                is_atr_sheet = (sheet_name == "ATR_Breakout_alerts")
                if is_atr_sheet:
                    col_num = atr_column_map[price_column]
                    price_2min_col = 25  # Y
                    price_10min_col = 26  # Z
                    price_eod_col = 27  # AA
                    status_col = 28  # AB
                    row_id_col = 29  # AC
                else:
                    col_num = standard_column_map[price_column]
                    price_2min_col = 22  # V
                    price_10min_col = 23  # W
                    price_eod_col = 24  # X
                    status_col = 25  # Y
                    row_id_col = 26  # Z

                # Find the row with matching row_id
                ws = self.workbook[sheet_name]
                for row_num in range(2, ws.max_row + 1):
                    cell_row_id = ws.cell(row=row_num, column=row_id_col).value

                    if cell_row_id == row_id:
                        # Update price
                        price_cell = ws.cell(row=row_num, column=col_num, value=round(price, 2))

                        # Apply color coding for 10min and EOD prices (not 2min as it's the reference)
                        if price_column in ["10min", "EOD"]:
                            price_2min = ws.cell(row=row_num, column=price_2min_col).value
                            direction = ws.cell(row=row_num, column=4).value  # Column D

                            if price_2min and price_2min > 0 and direction:
                                # Calculate percentage change from 2min price
                                percent_change = ((price - price_2min) / price_2min) * 100

                                # Apply color
                                color_fill = self._get_color_for_price_change(direction, percent_change)
                                if color_fill:
                                    price_cell.fill = color_fill

                        # Update status
                        price_2min = ws.cell(row=row_num, column=price_2min_col).value
                        price_10min = ws.cell(row=row_num, column=price_10min_col).value
                        price_eod = ws.cell(row=row_num, column=price_eod_col).value

                        if price_eod:
                            new_status = "Complete"
                        elif price_10min:
                            new_status = "Partial"
                        elif price_2min:
                            new_status = "Partial"
                        else:
                            new_status = "Pending"

                        # Override for EOD auto-complete
                        if auto_complete_eod and price_column == "EOD":
                            new_status = "Complete"

                        ws.cell(row=row_num, column=status_col, value=new_status)

                        updated_count += 1
                        break

            if updated_count > 0:
                self._save_workbook()
                logger.info(f"Updated {updated_count} rows with {price_column} prices")

            return updated_count

        except Exception as e:
            logger.error(f"Error updating prices: {e}", exc_info=True)
            return updated_count

    def _save_workbook(self):
        """Save workbook with file locking for concurrent access protection."""
        try:
            # Open file handle for locking
            with open(self.excel_path, 'wb') as f:
                # Acquire exclusive lock
                fcntl.flock(f.fileno(), fcntl.LOCK_EX)
                try:
                    self.workbook.save(f)
                finally:
                    # Release lock
                    fcntl.flock(f.fileno(), fcntl.LOCK_UN)
        except Exception as e:
            logger.error(f"Error saving workbook: {e}", exc_info=True)
            raise

    def sort_all_sheets_by_date(self) -> int:
        """
        Sort all sheets by Date and Time columns (oldest to newest).

        Returns:
            Number of sheets sorted
        """
        sorted_count = 0

        try:
            for sheet_name in set(self.SHEET_NAMES.values()):
                ws = self.workbook[sheet_name]

                # Check if there's data to sort (need at least 2 rows including header)
                if ws.max_row < 2:
                    continue

                # Read all data rows values only (skip header)
                data_rows = []
                for row_num in range(2, ws.max_row + 1):
                    row_values = []
                    for col_num in range(1, len(self.HEADERS) + 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        row_values.append(cell.value)
                    data_rows.append(row_values)

                # Sort by date (column 0) and time (column 1)
                def sort_key(row):
                    date_val = row[0] if row[0] else "0000-00-00"
                    time_val = row[1] if row[1] else "00:00:00"
                    return f"{date_val} {time_val}"

                data_rows.sort(key=sort_key)

                # Write sorted data back with basic formatting
                for row_idx, row_values in enumerate(data_rows, start=2):
                    for col_idx, value in enumerate(row_values, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)

                        # Apply basic formatting
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )

                sorted_count += 1
                logger.info(f"Sorted {sheet_name}: {len(data_rows)} alerts")

            if sorted_count > 0:
                self._save_workbook()
                logger.info(f"Sorted {sorted_count} sheets by date/time")

            return sorted_count

        except Exception as e:
            logger.error(f"Error sorting sheets: {e}", exc_info=True)
            return sorted_count

    def fix_all_directions(self) -> int:
        """
        Fix direction column for all existing alerts based on correct logic.

        Returns:
            Number of rows fixed
        """
        fixed_count = 0

        try:
            for sheet_name in set(self.SHEET_NAMES.values()):
                ws = self.workbook[sheet_name]

                # Determine column numbers based on sheet type
                if sheet_name == "ATR_Breakout_alerts":
                    row_id_col = 29  # AC
                else:
                    row_id_col = 26  # Z

                # Iterate through rows (skip header)
                for row_num in range(2, ws.max_row + 1):
                    row_id = ws.cell(row=row_num, column=row_id_col).value
                    change_percent = ws.cell(row=row_num, column=7).value  # Column G (Change %)

                    if not row_id or change_percent is None:
                        continue

                    # Extract alert_type from row_id (format: SYMBOL_ALERTTYPE_TIMESTAMP)
                    parts = row_id.split('_')
                    if len(parts) < 3:
                        continue

                    # Alert type is the second part (e.g., "5min", "volume_spike", "5min_rise")
                    # It could be parts[1] or parts[1]_parts[2] if it contains "rise"
                    alert_type = parts[1]
                    if len(parts) > 2 and parts[2] == "rise":
                        alert_type = f"{parts[1]}_rise"

                    # Calculate correct direction based on alert_type only
                    # Rise alerts have "_rise" suffix, drop alerts don't
                    correct_direction = "Rise" if "_rise" in alert_type else "Drop"

                    # Update direction if incorrect
                    current_direction = ws.cell(row=row_num, column=4).value  # Column D
                    if current_direction != correct_direction:
                        ws.cell(row=row_num, column=4, value=correct_direction)
                        fixed_count += 1

            if fixed_count > 0:
                self._save_workbook()
                logger.info(f"Fixed direction for {fixed_count} alerts")

            return fixed_count

        except Exception as e:
            logger.error(f"Error fixing directions: {e}", exc_info=True)
            return fixed_count

    def apply_color_formatting_to_all(self) -> int:
        """
        Apply color formatting to all existing price data in the workbook.

        This iterates through all sheets and applies gradient color coding
        to 10min and EOD price columns based on 2min reference price.

        Returns:
            Number of cells colored
        """
        colored_count = 0

        try:
            for sheet_name in set(self.SHEET_NAMES.values()):
                ws = self.workbook[sheet_name]

                # Determine column numbers based on sheet type
                if sheet_name == "ATR_Breakout_alerts":
                    price_2min_col = 25  # Y
                    price_10min_col = 26  # Z
                    price_eod_col = 27  # AA
                else:
                    price_2min_col = 22  # V
                    price_10min_col = 23  # W
                    price_eod_col = 24  # X

                # Iterate through rows (skip header)
                for row_num in range(2, ws.max_row + 1):
                    direction = ws.cell(row=row_num, column=4).value  # Column D
                    price_2min = ws.cell(row=row_num, column=price_2min_col).value
                    price_10min = ws.cell(row=row_num, column=price_10min_col).value
                    price_eod = ws.cell(row=row_num, column=price_eod_col).value

                    # Skip if no reference price
                    if not price_2min or price_2min <= 0 or not direction:
                        continue

                    # Color 10min price
                    if price_10min and price_10min > 0:
                        percent_change = ((price_10min - price_2min) / price_2min) * 100
                        color_fill = self._get_color_for_price_change(direction, percent_change)
                        if color_fill:
                            ws.cell(row=row_num, column=price_10min_col).fill = color_fill
                            colored_count += 1

                    # Color EOD price
                    if price_eod and price_eod > 0:
                        percent_change = ((price_eod - price_2min) / price_2min) * 100
                        color_fill = self._get_color_for_price_change(direction, percent_change)
                        if color_fill:
                            ws.cell(row=row_num, column=price_eod_col).fill = color_fill
                            colored_count += 1

            if colored_count > 0:
                self._save_workbook()
                logger.info(f"Applied color formatting to {colored_count} cells")

            return colored_count

        except Exception as e:
            logger.error(f"Error applying color formatting: {e}", exc_info=True)
            return colored_count

    def log_atr_breakout(
        self,
        symbol: str,
        today_open: float,
        entry_level: float,
        current_price: float,
        breakout_distance: float,
        atr_20: float,
        atr_30: float,
        volatility_filter_passed: bool,
        stop_loss: float,
        risk_amount: float,
        risk_percent: float,
        volume: int,
        market_cap_cr: Optional[float] = None,
        telegram_sent: bool = False,
        timestamp: Optional[datetime] = None,
        rsi_analysis: Optional[Dict] = None
    ) -> bool:
        """
        Log an ATR breakout alert to the ATR_Breakout_alerts sheet.

        Args:
            symbol: Stock symbol (e.g., "RELIANCE")
            today_open: Today's opening price
            entry_level: Breakout entry level (Open + 2.5×ATR)
            current_price: Current stock price
            breakout_distance: Distance above entry level
            atr_20: ATR(20) value
            atr_30: ATR(30) value
            volatility_filter_passed: Whether ATR(20) < ATR(30)
            stop_loss: Stop loss level
            risk_amount: Risk in rupees
            risk_percent: Risk as percentage
            volume: Current volume
            market_cap_cr: Market cap in crores
            telegram_sent: Whether Telegram alert was sent
            timestamp: Alert timestamp (defaults to now)
            rsi_analysis: Optional RSI analysis dict with RSI values and crossovers

        Returns:
            True if logged successfully, False otherwise
        """
        try:
            sheet_name = "ATR_Breakout_alerts"
            ws = self.workbook[sheet_name]

            # Prepare timestamp
            if timestamp is None:
                timestamp = datetime.now()

            date_str = timestamp.strftime("%Y-%m-%d")
            time_str = timestamp.strftime("%H:%M:%S")
            day_of_week = timestamp.strftime("%A")

            # Generate unique row ID
            row_id = f"{symbol}_{timestamp.strftime('%Y%m%d_%H%M%S')}"

            # Find next empty row
            next_row = ws.max_row + 1

            # Extract RSI data (same logic as log_alert)
            rsi_9 = rsi_14 = rsi_21 = ""
            rsi_9vs14 = rsi_9vs21 = rsi_14vs21 = ""
            rsi_recent_cross = rsi_summary = ""

            if rsi_analysis:
                rsi_9 = round(rsi_analysis.get('rsi_9'), 2) if rsi_analysis.get('rsi_9') else ""
                rsi_14 = round(rsi_analysis.get('rsi_14'), 2) if rsi_analysis.get('rsi_14') else ""
                rsi_21 = round(rsi_analysis.get('rsi_21'), 2) if rsi_analysis.get('rsi_21') else ""

                # Format crossover status
                crossovers = rsi_analysis.get('crossovers', {})
                if '9_14' in crossovers:
                    c = crossovers['9_14']
                    if c.get('status') and c.get('strength') is not None:
                        arrow = "↑" if c['status'] == 'above' else "↓"
                        sign = "+" if c['strength'] >= 0 else ""
                        rsi_9vs14 = f"9{arrow}14 ({sign}{c['strength']})"

                if '9_21' in crossovers:
                    c = crossovers['9_21']
                    if c.get('status') and c.get('strength') is not None:
                        arrow = "↑" if c['status'] == 'above' else "↓"
                        sign = "+" if c['strength'] >= 0 else ""
                        rsi_9vs21 = f"9{arrow}21 ({sign}{c['strength']})"

                if '14_21' in crossovers:
                    c = crossovers['14_21']
                    if c.get('status') and c.get('strength') is not None:
                        arrow = "↑" if c['status'] == 'above' else "↓"
                        sign = "+" if c['strength'] >= 0 else ""
                        rsi_14vs21 = f"14{arrow}21 ({sign}{c['strength']})"

                # Find most recent crossover across all pairs
                recent_crosses = []
                for pair, c in crossovers.items():
                    recent = c.get('recent_cross', {})
                    if recent.get('occurred'):
                        bars_ago = recent.get('bars_ago', 0)
                        direction_text = recent.get('direction', '').capitalize()
                        emoji = "🟢" if direction_text == 'Bullish' else "🔴"
                        recent_crosses.append(f"{emoji} {direction_text} {bars_ago}b ago")

                rsi_recent_cross = "; ".join(recent_crosses) if recent_crosses else "None"
                rsi_summary = rsi_analysis.get('summary', "")

            # Prepare row data (matching ATR_HEADERS order)
            row_data = [
                date_str,  # A: Date
                time_str,  # B: Time
                symbol,  # C: Symbol
                today_open,  # D: Open
                entry_level,  # E: Entry Level
                current_price,  # F: Current Price
                breakout_distance,  # G: Breakout Distance
                atr_20,  # H: ATR(20)
                atr_30,  # I: ATR(30)
                "PASSED" if volatility_filter_passed else "FAILED",  # J: Volatility Filter
                stop_loss,  # K: Stop Loss
                risk_amount,  # L: Risk Amount
                risk_percent,  # M: Risk %
                volume,  # N: Volume
                market_cap_cr if market_cap_cr else "N/A",  # O: Market Cap
                "Yes" if telegram_sent else "No",  # P: Telegram Sent
                rsi_9,  # Q: RSI(9)
                rsi_14,  # R: RSI(14)
                rsi_21,  # S: RSI(21)
                rsi_9vs14,  # T: RSI 9vs14
                rsi_9vs21,  # U: RSI 9vs21
                rsi_14vs21,  # V: RSI 14vs21
                rsi_recent_cross,  # W: RSI Recent Cross
                rsi_summary,  # X: RSI Summary
                "",  # Y: Price 2min (to be filled later)
                "",  # Z: Price 10min (to be filled later)
                "",  # AA: Price EOD (to be filled later)
                "Pending",  # AB: Status
                row_id,  # AC: Row ID
                day_of_week  # AD: Day of Week
            ]

            # Write row data
            for col_num, value in enumerate(row_data, start=1):
                cell = ws.cell(row=next_row, column=col_num, value=value)

                # Apply number formatting
                if col_num in [4, 5, 6, 7, 8, 9, 11, 12, 25, 26, 27]:  # Price/ATR columns (Y, Z, AA)
                    cell.number_format = '0.00'
                elif col_num == 13:  # Risk % (M)
                    cell.number_format = '0.00%'
                elif col_num in [14]:  # Volume (N)
                    cell.number_format = '#,##0'
                elif col_num == 15:  # Market Cap (O)
                    if market_cap_cr:
                        cell.number_format = '#,##0'
                elif col_num in [17, 18, 19]:  # RSI values (Q, R, S)
                    if value != "":
                        cell.number_format = '0.00'

                # Center alignment for most columns
                if col_num not in [29]:  # Except Row ID (AC)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Save workbook
            self._save_workbook()

            logger.info(f"ATR breakout logged for {symbol} at ₹{current_price:.2f} "
                       f"(Entry: ₹{entry_level:.2f}, SL: ₹{stop_loss:.2f})")
            return True

        except Exception as e:
            logger.error(f"Failed to log ATR breakout for {symbol}: {e}", exc_info=True)
            return False

    def close(self):
        """Close workbook and release resources."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None

        if self.file_handle:
            self.file_handle.close()
            self.file_handle = None
