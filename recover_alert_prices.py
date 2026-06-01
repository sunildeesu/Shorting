#!/usr/bin/env python3
"""
Recover corrupted Price 2min, Price 10min, and Price EOD columns in alert_tracking.xlsx.

Corrupted rows have RSI crossover strings in the Status column (not "Complete").
Uses Kite 15-minute candles for Price 2min / Price 10min (approximate),
and Kite day candles for Price EOD.

Run: venv/bin/python3 recover_alert_prices.py
Requires valid KITE_ACCESS_TOKEN in environment.
"""

import os
import json
import time
import shutil
import logging
from datetime import datetime, timedelta, date, time as dt_time

import openpyxl
from kiteconnect import KiteConnect

import config

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

ALERT_FILE = "data/alerts/alert_tracking.xlsx"
BACKUP_FILE = "data/alerts/alert_tracking_backup.xlsx"
TOKEN_FILE = "data/instrument_tokens.json"
EOD_CACHE_FILE = "data/eod_cache/historical_cache.json"
CANDLE_CACHE_FILE = "data/recover_candle_cache.json"

# Sheets to process and their Price 2min/10min/EOD column indices (1-based)
SHEETS = {
    "10min_alerts":       {"price2": 14, "price10": 15, "priceeod": 16, "status": 17},
    "5min_alerts":        {"price2": 14, "price10": 15, "priceeod": 16, "status": 17},
    "30min_alerts":       {"price2": 14, "price10": 15, "priceeod": 16, "status": 17},
    "Volume_Spike_alerts": {"price2": 14, "price10": 15, "priceeod": 16, "status": 17},
}


def load_candle_cache() -> dict:
    if os.path.exists(CANDLE_CACHE_FILE):
        with open(CANDLE_CACHE_FILE) as f:
            return json.load(f)
    return {}


def save_candle_cache(cache: dict):
    with open(CANDLE_CACHE_FILE, "w") as f:
        json.dump(cache, f)


def load_eod_cache() -> dict:
    """Returns dict: symbol -> {date_str -> close_price}"""
    result = {}
    if not os.path.exists(EOD_CACHE_FILE):
        return result
    with open(EOD_CACHE_FILE) as f:
        raw = json.load(f)
    for sym, v in raw.items():
        result[sym] = {}
        for d in v.get("data", []):
            date_key = d["date"][:10]
            result[sym][date_key] = d["close"]
    return result


def get_15min_candles(kite, token: int, target_date: date, candle_cache: dict) -> list:
    """Fetch 15-minute candles for target_date, using cache."""
    cache_key = f"15min_{token}_{target_date}"
    if cache_key in candle_cache:
        return candle_cache[cache_key]

    try:
        from_dt = datetime.combine(target_date, dt_time(9, 0))
        to_dt = datetime.combine(target_date, dt_time(15, 35))
        raw = kite.historical_data(token, from_dt, to_dt, "15minute")
        candles = [
            {
                "date": c["date"].isoformat() if hasattr(c["date"], "isoformat") else str(c["date"]),
                "close": c["close"],
            }
            for c in raw
        ]
        candle_cache[cache_key] = candles
        time.sleep(0.35)
    except Exception as e:
        logger.warning(f"15min fetch failed for token={token} date={target_date}: {e}")
        candle_cache[cache_key] = []
    return candle_cache[cache_key]


def get_day_close(kite, token: int, target_date: date, candle_cache: dict) -> float | None:
    """Fetch day close for target_date, using cache."""
    cache_key = f"day_{token}_{target_date}"
    if cache_key in candle_cache:
        return candle_cache[cache_key]

    try:
        from_dt = datetime.combine(target_date, dt_time(0, 0))
        to_dt = datetime.combine(target_date, dt_time(23, 59))
        raw = kite.historical_data(token, from_dt, to_dt, "day")
        close = raw[-1]["close"] if raw else None
        candle_cache[cache_key] = close
        time.sleep(0.35)
    except Exception as e:
        logger.warning(f"Day fetch failed for token={token} date={target_date}: {e}")
        candle_cache[cache_key] = None
    return candle_cache[cache_key]


def price_at_offset(candles: list, target_date: date, alert_time: dt_time, offset_min: int) -> float | None:
    """
    Return the close of the 15-min candle that covers alert_time + offset_min.
    Each candle at timestamp T covers [T, T+15min).
    We return the close of the last candle whose start <= target_time.
    """
    target_dt = datetime.combine(target_date, alert_time) + timedelta(minutes=offset_min)
    target_time = target_dt.time()

    result = None
    for c in candles:
        try:
            candle_dt = datetime.fromisoformat(c["date"]) if isinstance(c["date"], str) else c["date"]
            candle_time = candle_dt.time()
        except Exception:
            continue
        if candle_time <= target_time:
            result = c["close"]
        else:
            break
    return result


def is_corrupted(row: tuple, status_col: int) -> bool:
    """A row is corrupted if Status is not 'Complete'."""
    if not row[0]:  # no date
        return False
    status = str(row[status_col - 1]) if row[status_col - 1] else ""
    return status != "Complete"


def run():
    if not config.KITE_API_KEY or not config.KITE_ACCESS_TOKEN:
        print("ERROR: KITE_API_KEY and KITE_ACCESS_TOKEN required.")
        return

    kite = KiteConnect(api_key=config.KITE_API_KEY)
    kite.set_access_token(config.KITE_ACCESS_TOKEN)
    logger.info("Kite connected")

    with open(TOKEN_FILE) as f:
        tokens = json.load(f)

    eod_cache = load_eod_cache()
    candle_cache = load_candle_cache()

    # Backup
    if not os.path.exists(BACKUP_FILE):
        shutil.copy2(ALERT_FILE, BACKUP_FILE)
        logger.info(f"Backup created: {BACKUP_FILE}")
    else:
        logger.info(f"Backup already exists: {BACKUP_FILE}")

    wb = openpyxl.load_workbook(ALERT_FILE)

    total_recovered = 0
    total_skipped = 0
    cache_saves = 0

    for sheet_name, cols in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found, skipping")
            continue

        ws = wb[sheet_name]
        p2_col = cols["price2"]
        p10_col = cols["price10"]
        peod_col = cols["priceeod"]
        status_col = cols["status"]

        sheet_recovered = 0
        sheet_skipped = 0

        logger.info(f"\n=== Processing {sheet_name} ===")

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:
                continue
            if not is_corrupted(row, status_col):
                continue

            symbol = str(row[2]) if row[2] else ""
            date_val = row[0]
            time_val = row[1]

            # Parse date
            try:
                if isinstance(date_val, str):
                    alert_date = date.fromisoformat(date_val[:10])
                elif isinstance(date_val, datetime):
                    alert_date = date_val.date()
                else:
                    alert_date = date_val
            except Exception:
                sheet_skipped += 1
                continue

            # Parse time
            try:
                if isinstance(time_val, str):
                    parts = time_val.replace(":", " ").split()
                    alert_time = dt_time(int(parts[0]), int(parts[1]))
                elif isinstance(time_val, dt_time):
                    alert_time = time_val
                elif isinstance(time_val, datetime):
                    alert_time = time_val.time()
                else:
                    alert_time = dt_time(9, 25)
            except Exception:
                alert_time = dt_time(9, 25)

            # Look up token
            token = tokens.get(symbol)
            if token is None:
                logger.debug(f"  No token for {symbol}, skipping row {row_idx}")
                sheet_skipped += 1
                continue

            date_str = str(alert_date)

            # ── Price 2min ──
            candles = get_15min_candles(kite, token, alert_date, candle_cache)
            price_2min = price_at_offset(candles, alert_date, alert_time, 2)
            price_10min = price_at_offset(candles, alert_date, alert_time, 10)

            # ── Price EOD ──
            price_eod = eod_cache.get(symbol, {}).get(date_str)
            if price_eod is None:
                price_eod = get_day_close(kite, token, alert_date, candle_cache)

            if price_2min is None and price_10min is None and price_eod is None:
                logger.debug(f"  No data for {symbol} on {alert_date}, skipping row {row_idx}")
                sheet_skipped += 1
                continue

            # Write to cells
            if price_2min is not None:
                ws.cell(row=row_idx, column=p2_col).value = round(price_2min, 2)
            if price_10min is not None:
                ws.cell(row=row_idx, column=p10_col).value = round(price_10min, 2)
            if price_eod is not None:
                ws.cell(row=row_idx, column=peod_col).value = round(price_eod, 2)

            sheet_recovered += 1
            total_recovered += 1

            if sheet_recovered % 50 == 0:
                logger.info(f"  {sheet_name}: {sheet_recovered} recovered so far...")
                save_candle_cache(candle_cache)
                cache_saves += 1

        logger.info(f"  {sheet_name}: recovered={sheet_recovered}, skipped={sheet_skipped}")
        total_skipped += sheet_skipped

    save_candle_cache(candle_cache)
    wb.save(ALERT_FILE)

    print(f"\n{'='*60}")
    print(f"RECOVERY COMPLETE")
    print(f"  Recovered rows : {total_recovered}")
    print(f"  Skipped rows   : {total_skipped}")
    print(f"  Cache saves    : {cache_saves + 1}")
    print(f"  File saved     : {ALERT_FILE}")
    print(f"  Backup at      : {BACKUP_FILE}")
    print(f"{'='*60}")


if __name__ == "__main__":
    run()
