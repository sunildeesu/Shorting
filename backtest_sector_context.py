#!/usr/bin/env python3
"""
Backtest: Does sector context improve short signal quality?

Hypothesis: Stock dropping while sector is flat/green (stock-specific) →
better short than stock dropping in an already-weak sector (broad move).

DATA SOURCES:
1. alert_pnl_tracker.xlsx — 60 clean drop alerts with actual P&L (2026-02-25 to 2026-05-13)
2. sector_snapshots/     — sector day% for some of those dates
3. alert_tracking.xlsx   — 370 drop alerts (2025-10 to 2026-01), most price outcome data corrupted

METHOD A (Peer count — sector co-movement proxy):
  For each alert, count how many OTHER sector stocks also alerted that day.
  0 peers  → ISOLATED  (stock-specific)
  1 peer   → PARTIAL
  2+ peers → BROAD     (sector-wide)
  Uses clean P&L data from pnl_tracker.

METHOD B (Sector snapshot — direct sector day% when available):
  Uses price_change_day from sector_snapshots where available (14 dates).
  Small sample but more direct measurement.

TO DO PROPERLY (requires Kite token):
  Run `backtest_sector_kite.py` — fetches intraday index candles to get
  the sector's exact % change at alert time.
"""

import json
import os
import openpyxl
from collections import defaultdict
from datetime import datetime, timedelta


PNL_FILE = "data/alerts/alert_pnl_tracker.xlsx"
TRACKING_FILE = "data/alerts/alert_tracking.xlsx"
SNAPSHOT_DIR = "data/sector_snapshots"
SECTOR_MAP_FILE = "data/stock_sectors.json"


def load_pnl_alerts():
    """Load clean drop alerts from alert_pnl_tracker.xlsx (2026-02-25 onwards)."""
    wb = openpyxl.load_workbook(PNL_FILE)
    ws = wb.active
    alerts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        direction = str(row[4]).lower() if row[4] else ""
        if "drop" not in direction: continue
        date_str = str(row[0])
        time_str = str(row[1]) if row[1] else "09:25"
        symbol = row[2]
        alert_price = row[6]
        pnl_15m = row[10]   # P&L % 15min
        pnl_30m = row[11]   # P&L % 30min

        if not alert_price or pnl_15m is None:
            continue

        try:
            alert_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
        except Exception:
            try:
                alert_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
            except Exception:
                continue

        alerts.append({
            "date": date_str,
            "datetime": alert_dt,
            "symbol": symbol,
            "alert_price": float(alert_price),
            "pnl_15m": float(pnl_15m),
            "pnl_30m": float(pnl_30m) if pnl_30m is not None else None,
            "type": str(row[3]) if row[3] else "5min",
        })
    return alerts


def load_tracking_alerts():
    """Load drop alerts from alert_tracking.xlsx. Only use rows with valid prices (±10%)."""
    wb = openpyxl.load_workbook(TRACKING_FILE)
    ws = wb.active
    alerts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or row[3] != "Drop": continue
        date_str = str(row[0])
        symbol = row[2]
        ap = row[4]
        p2 = row[13]
        p10 = row[14]
        peod = row[15]
        if not ap: continue

        ap = float(ap)
        # Filter: require 2min price within ±10% of alert price
        if p2 is None: continue
        p2 = float(p2)
        if abs((p2 - ap) / ap) > 0.10: continue

        try:
            time_str = str(row[1]).split(".")[0]
            alert_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
        except Exception:
            continue

        alerts.append({
            "date": date_str,
            "datetime": alert_dt,
            "symbol": symbol,
            "alert_price": ap,
            "price_2min": p2,
            "price_10min": float(p10) if p10 is not None and abs((float(p10) - ap) / ap) < 0.10 else None,
            "price_eod": float(peod) if peod is not None and abs((float(peod) - ap) / ap) < 0.15 else None,
        })
    return alerts


def load_sector_snapshots():
    """Load sector snapshots. Returns dict: date -> (time, sectors_dict)."""
    snapshots = {}
    for fname in os.listdir(SNAPSHOT_DIR):
        if not fname.endswith(".json"): continue
        date_key = fname.replace(".json", "")
        with open(os.path.join(SNAPSHOT_DIR, fname)) as f:
            data = json.load(f)
        for t in ["09:30", "12:30", "15:15"]:
            if t in data and data[t].get("sectors"):
                snapshots[date_key] = (t, data[t]["sectors"])
                break
    return snapshots


def load_sector_map():
    with open(SECTOR_MAP_FILE) as f:
        return json.load(f)


def count_peer_alerts(alerts, sector_map, window_hours=2):
    """
    For each alert, count how many OTHER sector stocks also dropped that day
    within window_hours. Returns list of (alert, peer_count) tuples.
    """
    # Build index: date -> list of (datetime, symbol, sector)
    day_index = defaultdict(list)
    for a in alerts:
        sector = sector_map.get(a["symbol"])
        if sector:
            day_index[a["date"]].append((a["datetime"], a["symbol"], sector))

    result = []
    window = timedelta(hours=window_hours)
    for a in alerts:
        sector = sector_map.get(a["symbol"])
        if not sector:
            result.append((a, None))
            continue
        peers = sum(
            1 for dt, sym, sec in day_index[a["date"]]
            if sec == sector and sym != a["symbol"] and abs(dt - a["datetime"]) <= window
        )
        result.append((a, peers))
    return result


def categorize_peers(peer_count):
    if peer_count is None: return "UNKNOWN"
    if peer_count == 0: return "ISOLATED"
    if peer_count == 1: return "PARTIAL"
    return "BROAD"


def calc_stats(values):
    vals = [v for v in values if v is not None]
    if not vals:
        return {"n": 0, "win_rate": None, "avg": None, "n_wins": 0, "n_losses": 0}
    wins = [v for v in vals if v < 0]  # negative P&L% = profitable short
    return {
        "n": len(vals),
        "win_rate": len(wins) / len(vals) * 100,
        "avg": sum(vals) / len(vals),
        "avg_win": sum(wins) / len(wins) if wins else None,
        "avg_loss": sum(v for v in vals if v >= 0) / (len(vals) - len(wins)) if len(vals) > len(wins) else None,
        "n_wins": len(wins),
        "n_losses": len(vals) - len(wins),
    }


def print_stats_row(label, s15, s30=None):
    n = s15["n"]
    if n == 0:
        print(f"  {label:<14}  N=0")
        return
    wr = f"{s15['win_rate']:.0f}%" if s15["win_rate"] is not None else "N/A"
    avg = f"{s15['avg']:+.2f}%" if s15["avg"] is not None else "N/A"
    aw = f"{s15['avg_win']:+.2f}%" if s15["avg_win"] is not None else "N/A"
    al = f"{s15['avg_loss']:+.2f}%" if s15["avg_loss"] is not None else "N/A"

    s30_str = ""
    if s30 and s30["n"] > 0:
        wr30 = f"{s30['win_rate']:.0f}%"
        avg30 = f"{s30['avg']:+.2f}%"
        s30_str = f"  | 30m WR: {wr30}  avg: {avg30}"

    print(f"  {label:<14}  N={n:>3}  WinRate: {wr}  AvgRet: {avg}  (W:{aw} / L:{al}){s30_str}")


def run():
    sector_map = load_sector_map()
    snapshots = load_sector_snapshots()

    # ─── DATASET A: alert_pnl_tracker.xlsx (clean data, 2026-02+) ──────────────
    pnl_alerts = load_pnl_alerts()
    print(f"\n{'='*72}")
    print("BACKTEST: Sector Context vs Short Signal Quality")
    print(f"{'='*72}")
    print(f"\nDATASET A — alert_pnl_tracker.xlsx (clean P&L data)")
    print(f"  Drop alerts with P&L 15m: {len(pnl_alerts)}")
    if pnl_alerts:
        dates = [a["date"] for a in pnl_alerts]
        print(f"  Date range: {min(dates)} to {max(dates)}")

    # METHOD A1: Peer count using pnl_tracker alerts
    print(f"\n{'─'*72}")
    print("METHOD A: Peer Drop Count  (stock-specific vs sector-wide move)")
    print("  ISOLATED = 0 sector peers also dropped that day  ← hypothesis: best short")
    print("  PARTIAL  = 1 sector peer")
    print("  BROAD    = 2+ sector peers  ← hypothesis: weakest short signal")
    print(f"{'─'*72}")

    pnl_with_peers = count_peer_alerts(pnl_alerts, sector_map)
    buckets_a = defaultdict(lambda: {"15m": [], "30m": []})
    for a, peers in pnl_with_peers:
        cat = categorize_peers(peers)
        buckets_a[cat]["15m"].append(a["pnl_15m"])
        buckets_a[cat]["30m"].append(a["pnl_30m"])

    for cat in ["ISOLATED", "PARTIAL", "BROAD"]:
        s15 = calc_stats(buckets_a[cat]["15m"])
        s30 = calc_stats(buckets_a[cat]["30m"])
        print_stats_row(cat, s15, s30)

    # METHOD A2: Sector snapshot day% (limited dates)
    print(f"\n{'─'*72}")
    print("METHOD B: Sector Day % Change from Snapshot (where available)")
    print(f"  (price_change_day from sector_snapshots — only {sum(1 for a in pnl_alerts if a['date'] in snapshots and snapshots[a['date']][1])} overlap dates)")
    print(f"{'─'*72}")

    buckets_b = defaultdict(lambda: {"15m": [], "30m": [], "sector_pcts": []})
    snap_matched = 0
    for a in pnl_alerts:
        day_snap = snapshots.get(a["date"])
        if not day_snap: continue
        snap_time, sectors = day_snap
        sector = sector_map.get(a["symbol"])
        if not sector: continue
        sector_data = sectors.get(sector)
        if not sector_data: continue

        sector_day = sector_data.get("price_change_day", 0)
        # Fall back to 30min if day% not available
        if sector_day == 0:
            sector_day = sector_data.get("price_change_30min", 0)
        if sector_day == 0: continue

        if sector_day < -0.3:
            cat = "FALLING"
        elif sector_day > 0.3:
            cat = "RISING"
        else:
            cat = "FLAT"

        buckets_b[cat]["15m"].append(a["pnl_15m"])
        buckets_b[cat]["30m"].append(a["pnl_30m"])
        buckets_b[cat]["sector_pcts"].append(sector_day)
        snap_matched += 1

    print(f"  Alerts matched to sector snapshot: {snap_matched}")
    for cat in ["RISING", "FLAT", "FALLING"]:
        s15 = calc_stats(buckets_b[cat]["15m"])
        s30 = calc_stats(buckets_b[cat]["30m"])
        sp = buckets_b[cat]["sector_pcts"]
        avg_s = f"{sum(sp)/len(sp):+.2f}%" if sp else "N/A"
        prefix = f"{cat} (avg sector: {avg_s})"
        print_stats_row(prefix, s15, s30)

    # ─── DATASET B: alert_tracking.xlsx (valid rows only, 2025-10 to 2026-01) ──
    tracking_alerts = load_tracking_alerts()
    print(f"\n\n{'─'*72}")
    print(f"DATASET B — alert_tracking.xlsx (only rows with valid price data)")
    print(f"  Drop alerts with valid Price 2min (within ±10%): {len(tracking_alerts)}")
    if tracking_alerts:
        dates_t = [a["date"] for a in tracking_alerts]
        print(f"  Date range: {min(dates_t)} to {max(dates_t)}")

    if len(tracking_alerts) >= 10:
        tracking_with_peers = count_peer_alerts(tracking_alerts, sector_map)
        buckets_c = defaultdict(lambda: {"2min": [], "10min": [], "eod": []})
        for a, peers in tracking_with_peers:
            cat = categorize_peers(peers)
            ap = a["alert_price"]
            if a["price_2min"] is not None:
                buckets_c[cat]["2min"].append((a["price_2min"] - ap) / ap * 100)
            if a["price_10min"] is not None:
                buckets_c[cat]["10min"].append((a["price_10min"] - ap) / ap * 100)
            if a["price_eod"] is not None:
                buckets_c[cat]["eod"].append((a["price_eod"] - ap) / ap * 100)

        print(f"\n  Peer count analysis (2min/10min/EOD returns):")
        for cat in ["ISOLATED", "PARTIAL", "BROAD"]:
            s2 = calc_stats(buckets_c[cat]["2min"])
            s10 = calc_stats(buckets_c[cat]["10min"])
            seod = calc_stats(buckets_c[cat]["eod"])
            if s2["n"] == 0: continue
            wr2 = f"{s2['win_rate']:.0f}%" if s2["win_rate"] is not None else "N/A"
            wr10 = f"{s10['win_rate']:.0f}%" if s10["win_rate"] is not None else "N/A"
            wreod = f"{seod['win_rate']:.0f}%" if seod["win_rate"] is not None else "N/A"
            print(f"  {cat:<14}  N={s2['n']:>3}  2min WR: {wr2}  avg: {s2['avg']:+.2f}%"
                  f"  | 10min WR: {wr10}  avg: {s10['avg']:+.2f}% if {s10['n']}"
                  f"  | EOD WR: {wreod}")

    # ─── VERDICT ────────────────────────────────────────────────────────────────
    print(f"\n\n{'='*72}")
    print("VERDICT")
    print(f"{'='*72}")

    iso = calc_stats(buckets_a["ISOLATED"]["15m"])
    broad = calc_stats(buckets_a["BROAD"]["15m"])

    print(f"\nMethod A (Peer count, Dataset A — {len(pnl_alerts)} alerts):")
    if iso["n"] >= 5 and broad["n"] >= 5:
        diff = (iso["win_rate"] or 0) - (broad["win_rate"] or 0)
        avg_diff = (iso["avg"] or 0) - (broad["avg"] or 0)
        print(f"  ISOLATED: N={iso['n']}  WinRate={iso['win_rate']:.0f}%  Avg15m={iso['avg']:+.2f}%")
        print(f"  BROAD   : N={broad['n']}  WinRate={broad['win_rate']:.0f}%  Avg15m={broad['avg']:+.2f}%")
        if diff > 5:
            print(f"  ✅ CONFIRMED (+{diff:.0f}pp): Stock-specific drops DO produce better shorts")
        elif diff > 0:
            print(f"  ⚠️  MARGINAL (+{diff:.0f}pp): Slight edge for stock-specific drops")
        elif diff < -5:
            print(f"  ❌ REVERSED ({diff:.0f}pp): Sector-wide drops worked better as shorts")
        else:
            print(f"  ➖ NEUTRAL ({diff:+.0f}pp): No meaningful difference in this dataset")
        print(f"     Avg 15m return diff: {avg_diff:+.2f}pp")
    else:
        print(f"  ⚠️  Small samples: ISOLATED={iso['n']}, BROAD={broad['n']} — not conclusive")

    r_snap = calc_stats(buckets_b["RISING"]["15m"])
    f_snap = calc_stats(buckets_b["FALLING"]["15m"])
    print(f"\nMethod B (Sector day%, Dataset A — {snap_matched} matched alerts):")
    if r_snap["n"] >= 3 and f_snap["n"] >= 3:
        diff2 = (r_snap["win_rate"] or 0) - (f_snap["win_rate"] or 0)
        print(f"  Stock drop vs RISING sector: N={r_snap['n']}  WR={r_snap['win_rate']:.0f}%  Avg={r_snap['avg']:+.2f}%")
        print(f"  Stock drop vs FALLING sector: N={f_snap['n']}  WR={f_snap['win_rate']:.0f}%  Avg={f_snap['avg']:+.2f}%")
        if diff2 > 5:
            print(f"  ✅ CONFIRMED (+{diff2:.0f}pp): Drops against rising sector are better shorts")
        elif diff2 > 0:
            print(f"  ⚠️  MARGINAL (+{diff2:.0f}pp): Slight edge")
        elif diff2 < -5:
            print(f"  ❌ REVERSED: Sector-wide weakness actually produced better shorts")
        else:
            print(f"  ➖ NEUTRAL ({diff2:+.0f}pp): No clear edge")
    else:
        print(f"  ⚠️  Too few samples: RISING={r_snap['n']}, FALLING={f_snap['n']}")
        print(f"     Need more sector snapshot overlap. Run backtest_sector_kite.py for full analysis.")

    print(f"\n{'─'*72}")
    print("NOTE: For a proper backtest, use backtest_sector_kite.py which fetches")
    print("intraday Kite candles to get the EXACT sector state at alert time.")
    print(f"{'='*72}\n")


if __name__ == "__main__":
    run()
