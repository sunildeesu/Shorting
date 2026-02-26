#!/usr/bin/env python3
"""
NIFTY Option Analysis Logger

Logs daily NIFTY option selling analysis to Excel files organized by month.
Creates monthly reports in data/nifty_options/ directory.

Author: Sunil Kumar Durganaik
"""

import os
import logging
from datetime import datetime
from typing import Dict, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)


class NiftyOptionLogger:
    """Logs NIFTY option analysis results to monthly Excel files"""

    # Column headers for the Excel report
    HEADERS = [
        "Date", "Time", "Signal",
        # NEW COLUMNS (Tiered Signal System - Jan 3, 2026)
        "Signal_Tier", "Position_Size", "Premium_Quality",
        # END NEW COLUMNS
        "Total_Score",
        "NIFTY_Spot", "VIX", "VIX_Trend", "VIX_Score", "IV_Rank",
        "Market_Regime", "Regime_Score",
        "OI_Pattern", "OI_Score",
        "Theta_Score", "Gamma_Score", "Vega_Score",
        "Best_Strategy", "Expiry_1", "Days_To_Expiry_1",
        "Straddle_Premium", "Straddle_Theta", "Straddle_Gamma",
        "Strangle_Premium", "Strangle_Theta", "Strangle_Gamma",
        "Recommendation", "Risk_Factors", "Telegram_Sent"
    ]

    def __init__(self, base_dir: str = "data/nifty_options"):
        """
        Initialize NIFTY option logger

        Args:
            base_dir: Base directory for storing Excel files
        """
        self.base_dir = base_dir
        os.makedirs(base_dir, exist_ok=True)

    def _get_monthly_file_path(self, date: datetime) -> str:
        """Get Excel file path for the given month"""
        year_month = date.strftime("%Y-%m")
        filename = f"nifty_options_{year_month}.xlsx"
        return os.path.join(self.base_dir, filename)

    def _initialize_workbook(self, filepath: str) -> Workbook:
        """
        Initialize or load workbook with headers

        Args:
            filepath: Path to Excel file

        Returns:
            Workbook object
        """
        if os.path.exists(filepath):
            try:
                return load_workbook(filepath)
            except Exception as e:
                logger.warning(f"Could not load existing workbook: {e}. Creating new one.")

        # Create new workbook with headers
        wb = Workbook()
        ws = wb.active
        ws.title = "NIFTY_Options_Analysis"

        # Write headers
        for col_num, header in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)

            # Style headers
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # Set column widths (updated for tiered signals - 3 new columns after C)
        column_widths = {
            'A': 12,  # Date
            'B': 10,  # Time
            'C': 8,   # Signal
            'D': 14,  # Signal_Tier (NEW)
            'E': 12,  # Position_Size (NEW)
            'F': 22,  # Premium_Quality (NEW)
            'G': 12,  # Total_Score (was D)
            'H': 12,  # NIFTY_Spot (was E)
            'I': 8,   # VIX (was F)
            'J': 10,  # VIX_Trend (was G)
            'K': 10,  # VIX_Score (was H)
            'L': 10,  # IV_Rank (was I)
            'M': 14,  # Market_Regime (was J)
            'N': 12,  # Regime_Score (was K)
            'O': 16,  # OI_Pattern (was L)
            'P': 10,  # OI_Score (was M)
            'Q': 12,  # Theta_Score (was N)
            'R': 12,  # Gamma_Score (was O)
            'S': 12,  # Vega_Score (was P)
            'T': 14,  # Best_Strategy (was Q)
            'U': 12,  # Expiry_1 (was R)
            'V': 16,  # Days_To_Expiry_1 (was S)
            'W': 16,  # Straddle_Premium (was T)
            'X': 14,  # Straddle_Theta (was U)
            'Y': 14,  # Straddle_Gamma (was V)
            'Z': 16,  # Strangle_Premium (was W)
            'AA': 14,  # Strangle_Theta (was X)
            'AB': 14,  # Strangle_Gamma (was Y)
            'AC': 40,  # Recommendation (was Z)
            'AD': 40,  # Risk_Factors (was AA)
            'AE': 12   # Telegram_Sent (was AB)
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        return wb

    def log_analysis(
        self,
        analysis_data: Dict,
        telegram_sent: bool = False,
        timestamp: Optional[datetime] = None
    ) -> bool:
        """
        Log NIFTY option analysis to Excel

        Args:
            analysis_data: Analysis result dict from NiftyOptionAnalyzer
            telegram_sent: Whether Telegram alert was sent
            timestamp: Analysis timestamp (default: now)

        Returns:
            True if logged successfully
        """
        try:
            if timestamp is None:
                timestamp = datetime.now()

            # Get monthly file path
            filepath = self._get_monthly_file_path(timestamp)

            # Load or create workbook
            wb = self._initialize_workbook(filepath)
            ws = wb.active

            # Extract data
            signal = analysis_data.get('signal', 'HOLD')
            # NEW: Tiered signal fields
            signal_tier = analysis_data.get('signal_tier', signal)
            position_size = analysis_data.get('position_size', 1.0)
            premium_quality = analysis_data.get('premium_quality', 'TRADEABLE')
            # END NEW
            total_score = analysis_data.get('total_score', 0)
            nifty_spot = analysis_data.get('nifty_spot', 0)
            vix = analysis_data.get('vix', 0)
            vix_trend = analysis_data.get('vix_trend', 0)
            iv_rank = analysis_data.get('iv_rank', 50.0)
            market_regime = analysis_data.get('market_regime', 'UNKNOWN')
            best_strategy = analysis_data.get('best_strategy', 'straddle')
            recommendation = analysis_data.get('recommendation', '')
            risk_factors = analysis_data.get('risk_factors', [])
            breakdown = analysis_data.get('breakdown', {})
            oi_analysis = analysis_data.get('oi_analysis', {})
            expiry_analyses = analysis_data.get('expiry_analyses', [])

            # Get first expiry data
            first_expiry = expiry_analyses[0] if expiry_analyses else {}
            expiry_date = first_expiry.get('expiry_date')
            days_to_expiry = first_expiry.get('days_to_expiry', 0)
            straddle_data = first_expiry.get('straddle', {})
            strangle_data = first_expiry.get('strangle', {})

            # Prepare row data
            date_str = timestamp.strftime("%Y-%m-%d")
            time_str = timestamp.strftime("%H:%M:%S")
            expiry_str = expiry_date.strftime("%Y-%m-%d") if expiry_date else ""
            risk_factors_str = "; ".join(risk_factors)
            # Format position size as percentage string
            position_size_str = f"{int(position_size * 100)}%"
            # Extract premium quality label (remove explanation in parentheses)
            premium_quality_label = premium_quality.split(' (')[0]

            row_data = [
                date_str,
                time_str,
                signal,
                # NEW: Tiered signal columns
                signal_tier,
                position_size_str,
                premium_quality_label,
                # END NEW
                round(total_score, 1),
                round(nifty_spot, 2),
                round(vix, 2),
                round(vix_trend, 2),
                round(breakdown.get('vix_score', 0), 1),
                round(iv_rank, 1),
                market_regime,
                round(breakdown.get('regime_score', 0), 1),
                oi_analysis.get('pattern', 'UNKNOWN'),
                round(breakdown.get('oi_score', 0), 1),
                round(breakdown.get('theta_score', 0), 1),
                round(breakdown.get('gamma_score', 0), 1),
                round(breakdown.get('vega_score', 0), 1),
                best_strategy.upper(),
                expiry_str,
                days_to_expiry,
                round(straddle_data.get('total_premium', 0), 2),
                round(abs(straddle_data.get('greeks', {}).get('theta', 0)), 2),
                round(straddle_data.get('greeks', {}).get('gamma', 0), 6),
                round(strangle_data.get('total_premium', 0), 2),
                round(abs(strangle_data.get('greeks', {}).get('theta', 0)), 2),
                round(strangle_data.get('greeks', {}).get('gamma', 0), 6),
                recommendation,
                risk_factors_str,
                "Yes" if telegram_sent else "No"
            ]

            # Write row
            next_row = ws.max_row + 1
            for col_num, value in enumerate(row_data, start=1):
                cell = ws.cell(row=next_row, column=col_num, value=value)

                # Apply number formatting (updated for new columns - shifted by 3)
                if col_num in [7, 11, 12, 14, 16, 17, 18, 19]:  # Score columns (was 4,8,9,11,13,14,15,16)
                    cell.number_format = '0.0'
                elif col_num == 8:  # NIFTY Spot (was 5)
                    cell.number_format = '0.00'
                elif col_num in [9, 10]:  # VIX, VIX_Trend (was 6, 7)
                    cell.number_format = '0.00'
                elif col_num in [23, 24, 26, 27]:  # Premium, Theta (was 20, 21, 23, 24)
                    cell.number_format = '0.00'
                elif col_num in [25, 28]:  # Gamma (was 22, 25)
                    cell.number_format = '0.000000'

                # Center alignment
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                # Apply borders
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # Color code signal column
                if col_num == 3:  # Signal column
                    if signal == 'SELL':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(bold=True, color="006100")
                    elif signal == 'HOLD':
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = Font(bold=True, color="9C6500")
                    else:  # AVOID
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(bold=True, color="9C0006")

                # NEW: Color code Signal_Tier column (tiered signal system)
                if col_num == 4:  # Signal_Tier column
                    if signal_tier == 'SELL_STRONG':
                        # Dark green - excellent premium quality
                        cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                    elif signal_tier == 'SELL_MODERATE':
                        # Light green - good premium quality
                        cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                        cell.font = Font(bold=True, color="375623")
                    elif signal_tier == 'SELL_WEAK':
                        # Orange - marginal premium quality
                        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                        cell.font = Font(bold=True, color="974806")
                    elif signal_tier == 'AVOID':
                        # Red - poor premium quality
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                    # END NEW

            # Save workbook
            wb.save(filepath)
            from google_drive_sync import sync_to_drive
            sync_to_drive(filepath, "NiftyOptions")
            logger.info(f"NIFTY option analysis logged: {signal} (Score: {total_score:.1f}) to {filepath}")

            return True

        except Exception as e:
            logger.error(f"Error logging NIFTY option analysis: {e}", exc_info=True)
            return False

    def get_recent_analyses(self, days: int = 7) -> list:
        """
        Get recent analyses from Excel files

        Args:
            days: Number of days to look back

        Returns:
            List of analysis records
        """
        try:
            # Get files for last N days
            current_date = datetime.now()
            records = []

            # Check last 2 months to cover the day range
            for month_offset in range(2):
                target_date = datetime(current_date.year, current_date.month, 1)
                if month_offset > 0:
                    # Go back one month
                    if target_date.month == 1:
                        target_date = datetime(target_date.year - 1, 12, 1)
                    else:
                        target_date = datetime(target_date.year, target_date.month - 1, 1)

                filepath = self._get_monthly_file_path(target_date)

                if not os.path.exists(filepath):
                    continue

                wb = load_workbook(filepath, read_only=True)
                ws = wb.active

                # Read all rows (skip header)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:  # If date exists
                        try:
                            record_date = datetime.strptime(row[0], "%Y-%m-%d")
                            if (current_date - record_date).days <= days:
                                records.append({
                                    'date': row[0],
                                    'time': row[1],
                                    'signal': row[2],
                                    'score': row[3],
                                    'nifty_spot': row[4],
                                    'vix': row[5]
                                })
                        except:
                            continue

                wb.close()

            return sorted(records, key=lambda x: x['date'], reverse=True)

        except Exception as e:
            logger.error(f"Error reading recent analyses: {e}")
            return []


if __name__ == "__main__":
    # Test the logger
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )

    option_logger = NiftyOptionLogger()

    # Test data
    test_data = {
        'timestamp': datetime.now().isoformat(),
        'signal': 'SELL',
        # NEW: Tiered signal test data
        'signal_tier': 'SELL_STRONG',
        'position_size': 1.0,
        'premium_quality': 'EXCELLENT (100% of fair value or better)',
        # END NEW
        'total_score': 75.5,
        'nifty_spot': 21850.50,
        'vix': 14.2,
        'vix_trend': -0.5,
        'iv_rank': 32.5,
        'market_regime': 'NEUTRAL',
        'best_strategy': 'straddle',
        'recommendation': 'Good conditions for option selling',
        'risk_factors': ['VIX slightly elevated', 'Monitor for breakout'],
        'breakdown': {
            'theta_score': 80.0,
            'gamma_score': 85.0,
            'vega_score': 75.0,
            'vix_score': 60.0,
            'regime_score': 100.0,
            'oi_score': 70.0
        },
        'oi_analysis': {
            'pattern': 'LONG_UNWINDING'
        },
        'expiry_analyses': [{
            'expiry_date': datetime(2026, 1, 9),
            'days_to_expiry': 8,
            'straddle': {
                'total_premium': 355.0,
                'greeks': {'theta': -45.0, 'gamma': 0.0012}
            },
            'strangle': {
                'total_premium': 185.0,
                'greeks': {'theta': -25.0, 'gamma': 0.0008}
            }
        }]
    }

    # Log test data
    result = option_logger.log_analysis(test_data, telegram_sent=True)
    print(f"Test logging: {'Success' if result else 'Failed'}")

    # Get recent analyses
    recent = option_logger.get_recent_analyses(days=7)
    print(f"\nRecent analyses: {len(recent)} records found")
    for record in recent[:3]:
        print(f"  {record['date']} {record['time']}: {record['signal']} ({record['score']:.1f})")
