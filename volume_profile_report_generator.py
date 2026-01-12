#!/usr/bin/env python3
"""
Volume Profile Report Generator
Generates Excel reports for volume profile analysis with conditional formatting.
"""

import os
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import config

logger = logging.getLogger(__name__)


class VolumeProfileReportGenerator:
    """Generates Excel reports for volume profile analysis"""

    def __init__(self, base_dir: str = None):
        """
        Initialize report generator.

        Args:
            base_dir: Base directory for reports (default: from config)
        """
        self.base_dir = base_dir or config.VOLUME_PROFILE_REPORT_DIR

    def generate_report(self,
                       profile_results: List[Dict],
                       analysis_time: datetime,
                       execution_window: str) -> str:
        """
        Generate Excel report with volume profile findings.

        Report structure: data/volume_profile_reports/YYYY/MM/
        Files: volume_profile_eod_YYYY-MM-DD.xlsx

        Args:
            profile_results: List of volume profile results
            analysis_time: Time of analysis
            execution_window: "3:25PM" (end of day)

        Returns:
            Path to generated Excel file
        """
        # Create folder structure (YYYY/MM)
        year = analysis_time.strftime('%Y')
        month = analysis_time.strftime('%m')
        report_dir = os.path.join(self.base_dir, year, month)
        Path(report_dir).mkdir(parents=True, exist_ok=True)

        # Generate file name (simple: end-of-day)
        date_str = analysis_time.strftime('%Y-%m-%d')
        file_name = f'volume_profile_eod_{date_str}.xlsx'
        file_path = os.path.join(report_dir, file_name)

        logger.info(f"Generating report: {file_path}")

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Volume Profiles"

        # Write header
        self._write_header(ws, analysis_time, execution_window)

        # Write data
        self._write_data(ws, profile_results)

        # Apply formatting
        self._apply_formatting(ws, profile_results)

        # Auto-size columns
        self._auto_size_columns(ws)

        # Save workbook
        wb.save(file_path)
        logger.info(f"Report saved: {file_path}")

        return file_path

    def _write_header(self, ws, analysis_time: datetime, execution_window: str):
        """Write column headers"""
        headers = [
            "Stock Symbol",
            "Profile Shape",
            "POC Price",
            "POC Position %",
            "Day High",
            "Day Low",
            "Day Range",
            "Value Area High",
            "Value Area Low",
            "Total Volume",
            "Confidence Score",
            "Interpretation"
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _write_data(self, ws, profile_results: List[Dict]):
        """Write volume profile data"""
        # Sort by confidence (descending)
        sorted_results = sorted(profile_results, key=lambda x: x.get('confidence', 0), reverse=True)

        row = 2
        for result in sorted_results:
            # Interpretation (Continuation/Strength Framework)
            shape = result.get('profile_shape', 'UNKNOWN')
            if shape == 'P-SHAPE':
                interpretation = 'Bullish (Strength at Highs)'
            elif shape == 'B-SHAPE':
                interpretation = 'Bearish (Weakness at Lows)'
            elif shape == 'BALANCED':
                interpretation = 'Neutral'
            else:
                interpretation = 'N/A'

            # POC position as percentage
            poc_position_pct = result.get('poc_position', 0) * 100

            # Write row
            ws.cell(row=row, column=1).value = result.get('symbol', '')
            ws.cell(row=row, column=2).value = shape
            ws.cell(row=row, column=3).value = result.get('poc_price', 0)
            ws.cell(row=row, column=4).value = poc_position_pct
            ws.cell(row=row, column=5).value = result.get('day_high', 0)
            ws.cell(row=row, column=6).value = result.get('day_low', 0)
            ws.cell(row=row, column=7).value = result.get('day_range', 0)
            ws.cell(row=row, column=8).value = result.get('value_area_high', 0)
            ws.cell(row=row, column=9).value = result.get('value_area_low', 0)
            ws.cell(row=row, column=10).value = result.get('total_volume', 0)
            ws.cell(row=row, column=11).value = result.get('confidence', 0)
            ws.cell(row=row, column=12).value = interpretation

            row += 1

    def _apply_formatting(self, ws, profile_results: List[Dict]):
        """Apply conditional formatting"""
        # Colors
        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')  # P-shape (bearish)
        green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')  # B-shape (bullish)
        bold_font = Font(bold=True)

        # Start from row 2 (after headers)
        for row in range(2, len(profile_results) + 2):
            shape = ws.cell(row=row, column=2).value
            confidence = ws.cell(row=row, column=11).value or 0

            # Apply background color based on shape
            if shape == 'P-SHAPE':
                for col in range(1, 13):
                    ws.cell(row=row, column=col).fill = red_fill
            elif shape == 'B-SHAPE':
                for col in range(1, 13):
                    ws.cell(row=row, column=col).fill = green_fill

            # Bold text for high confidence (>= 7.5)
            if confidence >= config.VOLUME_PROFILE_MIN_CONFIDENCE:
                for col in range(1, 13):
                    ws.cell(row=row, column=col).font = bold_font

            # Center alignment for numeric columns
            for col in [2, 4, 11]:  # Shape, POC Position %, Confidence
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')

            # Number formatting
            ws.cell(row=row, column=3).number_format = '0.00'  # POC Price
            ws.cell(row=row, column=4).number_format = '0.0'   # POC Position %
            ws.cell(row=row, column=5).number_format = '0.00'  # Day High
            ws.cell(row=row, column=6).number_format = '0.00'  # Day Low
            ws.cell(row=row, column=7).number_format = '0.00'  # Day Range
            ws.cell(row=row, column=8).number_format = '0.00'  # Value Area High
            ws.cell(row=row, column=9).number_format = '0.00'  # Value Area Low
            ws.cell(row=row, column=10).number_format = '#,##0'  # Total Volume
            ws.cell(row=row, column=11).number_format = '0.0'  # Confidence

    def _auto_size_columns(self, ws):
        """Auto-size columns based on content"""
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter

            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 30)  # Cap at 30
            ws.column_dimensions[column].width = adjusted_width


if __name__ == "__main__":
    # Test report generation
    logging.basicConfig(level=logging.INFO)

    # Sample data
    sample_results = [
        {
            'symbol': 'RELIANCE',
            'profile_shape': 'P-SHAPE',
            'poc_price': 2450.50,
            'poc_position': 0.895,
            'day_high': 2470.0,
            'day_low': 2400.0,
            'day_range': 70.0,
            'value_area_high': 2465.0,
            'value_area_low': 2430.0,
            'total_volume': 8500000,
            'confidence': 8.5
        },
        {
            'symbol': 'INFY',
            'profile_shape': 'B-SHAPE',
            'poc_price': 1520.25,
            'poc_position': 0.123,
            'day_high': 1550.0,
            'day_low': 1510.0,
            'day_range': 40.0,
            'value_area_high': 1535.0,
            'value_area_low': 1515.0,
            'total_volume': 6200000,
            'confidence': 9.2
        },
        {
            'symbol': 'TCS',
            'profile_shape': 'BALANCED',
            'poc_price': 3890.0,
            'poc_position': 0.450,
            'day_high': 3920.0,
            'day_low': 3860.0,
            'day_range': 60.0,
            'value_area_high': 3905.0,
            'value_area_low': 3875.0,
            'total_volume': 4100000,
            'confidence': 5.0
        }
    ]

    generator = VolumeProfileReportGenerator()
    report_path = generator.generate_report(
        profile_results=sample_results,
        analysis_time=datetime.now(),
        execution_window="3:00PM"
    )

    print(f"\n✅ Test report generated: {report_path}\n")
