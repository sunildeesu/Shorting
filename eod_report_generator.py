#!/usr/bin/env python3
"""
EOD Report Generator - Creates Excel reports with analysis findings
Organizes reports by month/year in folder structure
"""

import os
from datetime import datetime
from typing import List, Dict
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class EODReportGenerator:
    """Generates Excel reports for EOD stock analysis"""

    def __init__(self, base_dir: str = "data/eod_reports"):
        """
        Initialize report generator

        Args:
            base_dir: Base directory for reports (default: data/eod_reports)
        """
        self.base_dir = base_dir

    def get_report_path(self, analysis_date: datetime = None) -> str:
        """
        Get report file path with YYYY/MM folder structure

        Args:
            analysis_date: Date of analysis (default: today)

        Returns:
            Full path to report file
        """
        if analysis_date is None:
            analysis_date = datetime.now()

        year = analysis_date.strftime('%Y')
        month = analysis_date.strftime('%m')
        filename = f"eod_analysis_{analysis_date.strftime('%Y-%m-%d')}.xlsx"

        # Create directory structure: data/eod_reports/YYYY/MM/
        report_dir = os.path.join(self.base_dir, year, month)
        os.makedirs(report_dir, exist_ok=True)

        return os.path.join(report_dir, filename)

    def generate_report(
        self,
        volume_results: List[Dict],
        pattern_results: List[Dict],
        quote_data: Dict[str, Dict],
        historical_data_map: Dict[str, List[Dict]],
        analysis_date: datetime = None
    ) -> str:
        """
        Generate Excel report with all findings

        Args:
            volume_results: List of volume analysis results
            pattern_results: List of pattern detection results
            quote_data: Quote data for price information
            historical_data_map: Historical data for EOD closing prices
            analysis_date: Date of analysis (default: today)

        Returns:
            Path to generated report file
        """
        if analysis_date is None:
            analysis_date = datetime.now()

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EOD Analysis"

        # Merge results (volume + patterns)
        merged_data = self._merge_results(volume_results, pattern_results, quote_data, historical_data_map)

        # Filter to only include stocks with findings
        findings = [
            stock for stock in merged_data
            if stock['has_volume_spike'] or stock['has_patterns']
        ]

        # Write header
        self._write_header(ws, analysis_date)

        # Write data
        self._write_data(ws, findings)

        # Format worksheet
        self._format_worksheet(ws)

        # Save report
        report_path = self.get_report_path(analysis_date)
        wb.save(report_path)

        logger.info(f"Report generated: {report_path} ({len(findings)} stocks with findings)")

        return report_path

    def _merge_results(
        self,
        volume_results: List[Dict],
        pattern_results: List[Dict],
        quote_data: Dict[str, Dict],
        historical_data_map: Dict[str, List[Dict]]
    ) -> List[Dict]:
        """Merge volume and pattern results into single dataset"""
        # Create lookup dictionaries
        volume_map = {r['symbol']: r for r in volume_results}
        pattern_map = {r['symbol']: r for r in pattern_results}

        # Get all unique symbols
        all_symbols = set(volume_map.keys()) | set(pattern_map.keys())

        merged = []
        for symbol in all_symbols:
            volume_data = volume_map.get(symbol, {})
            pattern_data = pattern_map.get(symbol, {})

            # Get EOD closing price from historical data (last day's close)
            historical_data = historical_data_map.get(symbol, [])
            current_price = 0
            open_price = 0

            if historical_data and len(historical_data) > 0:
                # Use the last day's closing price for EOD report
                current_price = historical_data[-1].get('close', 0)
                open_price = historical_data[-1].get('open', 0)

            # Calculate price change percentage
            price_change_pct = 0
            if open_price > 0:
                price_change_pct = ((current_price - open_price) / open_price) * 100

            # Extract buy, target, stop loss prices, and confidence score from pattern details
            buy_price = None
            target_price = None
            stop_loss = None
            confidence_score = None
            volume_ratio = None
            pattern_details = pattern_data.get('pattern_details', {})

            # Get buy/target/stop_loss/confidence from any detected pattern (prioritize first pattern)
            for pattern_name, details in pattern_details.items():
                if details and 'buy_price' in details:
                    buy_price = details['buy_price']
                    target_price = details['target_price']
                    stop_loss = details.get('stop_loss')
                    confidence_score = details.get('confidence_score')
                    volume_ratio = details.get('volume_ratio', 1.0)
                    break  # Use first pattern's prices

            merged.append({
                'symbol': symbol.replace('.NS', ''),
                'has_volume_spike': volume_data.get('has_spike_15min', False) or volume_data.get('has_spike_30min', False),
                'volume_spike_15min': volume_data.get('has_spike_15min', False),
                'volume_spike_30min': volume_data.get('has_spike_30min', False),
                'volume_15min': volume_data.get('volume_15min', 0),
                'volume_30min': volume_data.get('volume_30min', 0),
                'spike_ratio_15min': volume_data.get('spike_ratio_15min', 0),
                'spike_ratio_30min': volume_data.get('spike_ratio_30min', 0),
                'has_patterns': pattern_data.get('has_patterns', False),
                'patterns': ', '.join(pattern_data.get('patterns_found', [])),
                'current_price': current_price,
                'price_change_pct': price_change_pct,
                'buy_price': buy_price,
                'target_price': target_price,
                'stop_loss': stop_loss,
                'confidence_score': confidence_score,
                'volume_ratio': volume_ratio
            })

        return merged

    def _write_header(self, ws, analysis_date: datetime):
        """Write report header"""
        # Title row
        ws.merge_cells('A1:Q1')
        title_cell = ws['A1']
        title_cell.value = f"End-of-Day Stock Analysis - {analysis_date.strftime('%d %B %Y')}"
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Column headers (row 3)
        headers = [
            'Stock',
            '15-Min Spike',
            '15-Min Volume',
            '15-Min Ratio',
            '30-Min Spike',
            '30-Min Volume',
            '30-Min Ratio',
            'Chart Patterns',
            'Current Price',
            'Price Change %',
            'Buy/Entry Price',
            'Target Price',
            'Stop Loss',
            'Confidence',
            'Volume',
            'Signal',
            'Notes'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[3].height = 25

    def _write_data(self, ws, findings: List[Dict]):
        """Write data rows"""
        row_num = 4

        for stock in findings:
            ws.cell(row=row_num, column=1, value=stock['symbol'])
            ws.cell(row=row_num, column=2, value='YES' if stock['volume_spike_15min'] else 'NO')
            ws.cell(row=row_num, column=3, value=f"{stock['volume_15min']:,}")
            ws.cell(row=row_num, column=4, value=f"{stock['spike_ratio_15min']:.2f}x" if stock['volume_spike_15min'] else '-')
            ws.cell(row=row_num, column=5, value='YES' if stock['volume_spike_30min'] else 'NO')
            ws.cell(row=row_num, column=6, value=f"{stock['volume_30min']:,}")
            ws.cell(row=row_num, column=7, value=f"{stock['spike_ratio_30min']:.2f}x" if stock['volume_spike_30min'] else '-')
            ws.cell(row=row_num, column=8, value=stock['patterns'])
            ws.cell(row=row_num, column=9, value=f"₹{stock['current_price']:.2f}" if stock['current_price'] > 0 else '-')
            ws.cell(row=row_num, column=10, value=f"{stock['price_change_pct']:+.2f}%" if stock['current_price'] > 0 else '-')

            # Buy/Entry Price and Target Price
            if stock['buy_price']:
                ws.cell(row=row_num, column=11, value=f"₹{stock['buy_price']:.2f}")
            else:
                ws.cell(row=row_num, column=11, value='-')

            if stock['target_price']:
                ws.cell(row=row_num, column=12, value=f"₹{stock['target_price']:.2f}")
            else:
                ws.cell(row=row_num, column=12, value='-')

            # Stop Loss (Column 13)
            if stock['stop_loss']:
                ws.cell(row=row_num, column=13, value=f"₹{stock['stop_loss']:.2f}")
            else:
                ws.cell(row=row_num, column=13, value='-')

            # Confidence Score (Column 14)
            if stock['confidence_score']:
                ws.cell(row=row_num, column=14, value=f"{stock['confidence_score']:.1f}/10")
            else:
                ws.cell(row=row_num, column=14, value='-')

            # Volume Ratio (Column 15)
            if stock['volume_ratio']:
                ws.cell(row=row_num, column=15, value=f"{stock['volume_ratio']:.1f}x")
            else:
                ws.cell(row=row_num, column=15, value='-')

            # Determine signal (Column 16)
            signal = self._determine_signal(stock)
            ws.cell(row=row_num, column=16, value=signal)

            # Add notes (Column 17)
            notes = self._generate_notes(stock)
            ws.cell(row=row_num, column=17, value=notes)

            row_num += 1

    def _format_worksheet(self, ws):
        """Apply formatting to worksheet"""
        # Set column widths
        column_widths = {
            'A': 15,  # Stock
            'B': 12,  # 15-Min Spike
            'C': 15,  # 15-Min Volume
            'D': 12,  # 15-Min Ratio
            'E': 12,  # 30-Min Spike
            'F': 15,  # 30-Min Volume
            'G': 12,  # 30-Min Ratio
            'H': 30,  # Chart Patterns
            'I': 15,  # Current Price
            'J': 15,  # Price Change %
            'K': 15,  # Buy/Entry Price
            'L': 15,  # Target Price
            'M': 15,  # Stop Loss
            'N': 12,  # Confidence
            'O': 12,  # Volume
            'P': 12,  # Signal
            'Q': 40   # Notes
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Apply borders to all cells with data
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=17):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Center align specific columns (spike columns, confidence, volume)
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=2, max_col=7):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Center align confidence and volume columns
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=14, max_col=15):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Apply conditional formatting for confidence score column (N = 14)
        for row_num in range(4, ws.max_row + 1):
            confidence_cell = ws.cell(row=row_num, column=14)
            confidence_value = confidence_cell.value

            # Extract numeric value from "X.X/10" format
            if confidence_value and confidence_value != '-':
                try:
                    score = float(confidence_value.split('/')[0])
                    if score >= 8.0:
                        # High confidence: Green
                        confidence_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        confidence_cell.font = Font(color="006100", bold=True)
                    elif score >= 7.0:
                        # Medium confidence: Yellow
                        confidence_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        confidence_cell.font = Font(color="9C5700", bold=True)
                    else:
                        # Low confidence: Red (shouldn't appear due to min_confidence filter)
                        confidence_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        confidence_cell.font = Font(color="9C0006", bold=True)
                except:
                    pass

        # Apply conditional formatting for signal column (P = 16)
        for row_num in range(4, ws.max_row + 1):
            signal_cell = ws.cell(row=row_num, column=16)
            signal_value = signal_cell.value

            if signal_value == 'Bullish':
                signal_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                signal_cell.font = Font(color="006100", bold=True)
            elif signal_value == 'Bearish':
                signal_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                signal_cell.font = Font(color="9C0006", bold=True)
            elif signal_value == 'Mixed':
                signal_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                signal_cell.font = Font(color="9C5700", bold=True)

    def _determine_signal(self, stock: Dict) -> str:
        """Determine trading signal based on findings"""
        patterns = stock['patterns'].upper()

        # Check for both bullish and bearish patterns
        # Phase 1 patterns + Phase 2 high-probability patterns (65-80% win rates)
        has_bullish = ('RESISTANCE_BREAKOUT' in patterns or
                       'DOUBLE_BOTTOM' in patterns or
                       'CUP_HANDLE' in patterns or
                       'INVERSE_HEAD_SHOULDERS' in patterns or
                       'BULL_FLAG' in patterns or
                       'ASCENDING_TRIANGLE' in patterns or
                       'FALLING_WEDGE' in patterns)
        has_bearish = 'SUPPORT_BREAKOUT' in patterns or 'DOUBLE_TOP' in patterns

        # If both exist, return Mixed signal
        if has_bullish and has_bearish:
            return 'Mixed'
        elif has_bullish:
            return 'Bullish'
        elif has_bearish:
            return 'Bearish'
        elif stock['has_volume_spike']:
            return 'Watch'
        else:
            return 'Neutral'

    def _generate_notes(self, stock: Dict) -> str:
        """Generate notes about the stock"""
        notes = []

        # Check for mixed signals
        patterns = stock['patterns'].upper()
        # Phase 1 + Phase 2 patterns
        has_bullish = ('RESISTANCE_BREAKOUT' in patterns or
                       'DOUBLE_BOTTOM' in patterns or
                       'CUP_HANDLE' in patterns or
                       'INVERSE_HEAD_SHOULDERS' in patterns or
                       'BULL_FLAG' in patterns or
                       'ASCENDING_TRIANGLE' in patterns or
                       'FALLING_WEDGE' in patterns)
        has_bearish = 'SUPPORT_BREAKOUT' in patterns or 'DOUBLE_TOP' in patterns

        if has_bullish and has_bearish:
            notes.append("⚠️ MIXED SIGNALS - Both bullish and bearish patterns detected")

        if stock['volume_spike_15min'] and stock['volume_spike_30min']:
            notes.append(f"Strong EOD volume spike ({stock['spike_ratio_30min']:.1f}x avg)")
        elif stock['volume_spike_15min']:
            notes.append(f"15-min volume spike ({stock['spike_ratio_15min']:.1f}x)")
        elif stock['volume_spike_30min']:
            notes.append(f"30-min volume spike ({stock['spike_ratio_30min']:.1f}x)")

        if stock['patterns']:
            notes.append(f"Patterns: {stock['patterns']}")

        # Add confidence note
        if stock['confidence_score'] and stock['confidence_score'] >= 8.0:
            notes.append("High confidence setup")
        elif stock['confidence_score'] and stock['confidence_score'] >= 7.0:
            notes.append("Medium confidence")

        if stock['price_change_pct'] > 2:
            notes.append(f"Strong gain today (+{stock['price_change_pct']:.1f}%)")
        elif stock['price_change_pct'] < -2:
            notes.append(f"Strong drop today ({stock['price_change_pct']:.1f}%)")

        return '; '.join(notes) if notes else 'Monitor for follow-through'
