#!/usr/bin/env python3
"""
Verify that the current price in EOD report is the EOD closing price
"""

import openpyxl
from datetime import datetime
from kiteconnect import KiteConnect
import config

def verify_eod_current_price(report_path: str):
    """
    Verify that current price in report matches EOD closing price from historical data
    """
    print(f"Verifying EOD current price in: {report_path}\n")

    # Initialize Kite
    kite = KiteConnect(api_key=config.KITE_API_KEY)
    kite.set_access_token(config.KITE_ACCESS_TOKEN)

    # Load report
    wb = openpyxl.load_workbook(report_path)
    ws = wb.active

    # Get instruments
    instruments = kite.instruments("NSE")
    instrument_map = {inst['tradingsymbol']: inst['instrument_token']
                     for inst in instruments if inst['segment'] == 'NSE'}

    print(f"{'='*100}")
    print(f"{'Stock':<15} {'Report Price':<15} {'EOD Close':<15} {'EOD Open':<15} {'Match?':<10} {'Comment':<30}")
    print(f"{'='*100}")

    verification_results = []

    for row_num in range(4, min(ws.max_row + 1, 10)):  # Check first 6 stocks
        stock = ws.cell(row=row_num, column=1).value
        if not stock:
            break

        # Get current price from report (Column 9)
        report_price_cell = ws.cell(row=row_num, column=9).value
        if not report_price_cell or report_price_cell == '-':
            continue

        # Extract numeric value
        report_price = float(str(report_price_cell).replace('₹', '').replace(',', ''))

        # Get instrument token
        instrument_token = instrument_map.get(stock)
        if not instrument_token:
            print(f"{stock:<15} {report_price:<15.2f} {'N/A':<15} {'N/A':<15} {'?':<10} No instrument found")
            continue

        # Fetch historical data for Nov 4, 2025
        try:
            from_date = datetime(2025, 11, 4)
            to_date = datetime(2025, 11, 4, 23, 59, 59)

            historical_data = kite.historical_data(
                instrument_token=instrument_token,
                from_date=from_date,
                to_date=to_date,
                interval="day"
            )

            if not historical_data:
                print(f"{stock:<15} {report_price:<15.2f} {'No data':<15} {'No data':<15} {'?':<10} No historical data")
                continue

            eod_close = historical_data[-1]['close']
            eod_open = historical_data[-1]['open']

            # Check if report price matches EOD close
            price_diff = abs(report_price - eod_close)
            matches_close = price_diff < 0.5  # Allow 0.50 rupee tolerance

            # Check if it incorrectly matches open
            matches_open = abs(report_price - eod_open) < 0.5

            status = "✅ CORRECT" if matches_close else "❌ WRONG"
            comment = ""

            if matches_close:
                comment = "Using EOD close ✅"
            elif matches_open:
                comment = "⚠️ Using EOD open (BUG!)"
                status = "❌ BUG"
            else:
                comment = f"Mismatch (diff: ₹{price_diff:.2f})"

            print(f"{stock:<15} ₹{report_price:<14.2f} ₹{eod_close:<14.2f} ₹{eod_open:<14.2f} {status:<10} {comment:<30}")

            verification_results.append({
                'stock': stock,
                'matches_close': matches_close,
                'matches_open': matches_open and not matches_close
            })

        except Exception as e:
            print(f"{stock:<15} {report_price:<15.2f} {'Error':<15} {'Error':<15} {'?':<10} {str(e)[:30]}")

    print(f"{'='*100}")

    # Summary
    if verification_results:
        correct_count = sum(1 for r in verification_results if r['matches_close'])
        using_open_count = sum(1 for r in verification_results if r['matches_open'])

        print(f"\nVerification Summary:")
        print(f"  Total stocks checked: {len(verification_results)}")
        print(f"  Using EOD close (correct): {correct_count}")
        print(f"  Using EOD open (bug): {using_open_count}")

        if correct_count == len(verification_results):
            print(f"\n✅ SUCCESS! All stocks are using EOD closing price correctly!")
            return True
        elif using_open_count > 0:
            print(f"\n❌ BUG FOUND! Some stocks are still using EOD opening price instead of closing price!")
            return False
        else:
            print(f"\n⚠️ WARNING! Price mismatches detected but not using open price.")
            return False
    else:
        print("\n⚠️ No stocks verified")
        return False

if __name__ == "__main__":
    report_path = "data/eod_reports/2025/11/eod_analysis_2025-11-04.xlsx"
    verify_eod_current_price(report_path)
