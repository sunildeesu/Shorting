#!/usr/bin/env python3
"""
5-Minute Alert Precursor Analysis

Analyzes historical 5-min alerts to find predictive patterns
that appear 1-10 minutes before alerts trigger.

Usage:
    python analyze_5min_precursors.py --days 3
    python analyze_5min_precursors.py --from-date 2026-02-06

Author: Claude Opus 4.5
Date: 2026-02-09
"""

import argparse
import logging
import os
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import config

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class PreAlertAnalyzer:
    """
    Analyzes pre-alert patterns to find predictive signals.

    Workflow:
    1. Load 5-min alerts from Excel (last N days)
    2. For each alert, query central_quotes.db for preceding 10 minutes
    3. Compute per-minute indicators
    4. Identify patterns that precede alerts
    5. Generate statistical report
    """

    def __init__(self, excel_path: str = None, db_path: str = None):
        """
        Initialize analyzer.

        Args:
            excel_path: Path to alert_tracking.xlsx
            db_path: Path to central_quotes.db
        """
        self.excel_path = excel_path or config.ALERT_EXCEL_PATH
        self.db_path = db_path or "data/central_quotes.db"
        self.conn = None

    def _get_db_connection(self) -> sqlite3.Connection:
        """Get database connection (reader mode)."""
        if self.conn is None:
            self.conn = sqlite3.connect(self.db_path, timeout=30)
            self.conn.execute("PRAGMA query_only = ON")
        return self.conn

    def load_5min_alerts(self, days: int = 3, from_date: str = None) -> List[Dict]:
        """
        Load 5-min alerts from Excel for analysis.

        Args:
            days: Number of days to analyze
            from_date: Optional start date (YYYY-MM-DD)

        Returns:
            List of alert dicts
        """
        if not os.path.exists(self.excel_path):
            logger.error(f"Alert Excel file not found: {self.excel_path}")
            return []

        workbook = openpyxl.load_workbook(self.excel_path, read_only=True)

        # Check for 5min_alerts sheet
        if "5min_alerts" not in workbook.sheetnames:
            logger.error("Sheet '5min_alerts' not found in workbook")
            return []

        sheet = workbook["5min_alerts"]

        # Calculate cutoff date
        if from_date:
            cutoff_date = datetime.strptime(from_date, "%Y-%m-%d")
        else:
            cutoff_date = datetime.now() - timedelta(days=days)

        logger.info(f"Loading alerts since {cutoff_date.strftime('%Y-%m-%d')}")

        alerts = []
        row_count = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:  # Skip empty rows
                continue

            row_count += 1

            try:
                # Parse date and time (columns A, B)
                date_val = row[0]
                time_val = row[1]

                # Handle different date formats
                if isinstance(date_val, datetime):
                    date_str = date_val.strftime("%Y-%m-%d")
                else:
                    date_str = str(date_val)

                if isinstance(time_val, datetime):
                    time_str = time_val.strftime("%H:%M:%S")
                else:
                    time_str = str(time_val)

                # Parse full datetime
                try:
                    alert_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    # Try without seconds
                    alert_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")

                # Skip if before cutoff
                if alert_dt < cutoff_date:
                    continue

                # Extract columns based on actual structure:
                # Date(0), Time(1), Symbol(2), Direction(3), Alert Price(4), Previous Price(5),
                # Change %(6), Change Rs(7), Volume(8), Avg Volume(9), Volume Multiplier(10)

                # Parse volume multiplier (may have 'x' suffix like "1.67x")
                vol_mult_raw = row[10]
                if vol_mult_raw:
                    vol_mult_str = str(vol_mult_raw).replace('x', '').strip()
                    try:
                        vol_mult = float(vol_mult_str)
                    except ValueError:
                        vol_mult = 0
                else:
                    vol_mult = 0

                alert = {
                    'datetime': alert_dt,
                    'date': date_str,
                    'time': time_str,
                    'symbol': str(row[2]) if row[2] else '',
                    'direction': str(row[3]) if row[3] else 'Drop',
                    'alert_price': float(row[4]) if row[4] else 0,
                    'previous_price': float(row[5]) if row[5] else 0,
                    'change_pct': float(row[6]) if row[6] else 0,
                    'volume': int(row[8]) if row[8] else 0,
                    'avg_volume': float(row[9]) if row[9] else 0,
                    'volume_multiplier': vol_mult,
                }

                if alert['symbol']:  # Only add valid alerts
                    alerts.append(alert)

            except Exception as e:
                logger.warning(f"Error parsing row {row_count}: {e}")
                continue

        workbook.close()
        logger.info(f"Loaded {len(alerts)} alerts from {row_count} rows")
        return alerts

    def get_pre_alert_data(self, symbol: str, alert_time: datetime, minutes_back: int = 10) -> List[Dict]:
        """
        Query central_quotes.db for data before alert.

        Args:
            symbol: Stock symbol
            alert_time: Alert timestamp
            minutes_back: Minutes of history to fetch

        Returns:
            List of minute-by-minute data points
        """
        conn = self._get_db_connection()
        cursor = conn.cursor()

        # Calculate time range
        start_time = alert_time - timedelta(minutes=minutes_back)
        start_str = start_time.strftime('%Y-%m-%d %H:%M:00')
        end_str = alert_time.strftime('%Y-%m-%d %H:%M:00')

        cursor.execute("""
            SELECT timestamp, price, volume
            FROM stock_quotes
            WHERE symbol = ?
            AND timestamp >= ?
            AND timestamp < ?
            ORDER BY timestamp ASC
        """, (symbol, start_str, end_str))

        data = []
        for row in cursor.fetchall():
            data.append({
                'timestamp': row[0],
                'price': row[1],
                'volume': row[2] or 0
            })

        return data

    def analyze_volume_buildup(self, pre_data: List[Dict], alert_volume_mult: float) -> Dict:
        """
        Analyze volume buildup pattern before alert.

        Returns per-minute volume ratios and when 1.0x was first crossed.
        """
        if len(pre_data) < 2:
            return {'has_data': False}

        # Use first data point's volume as baseline
        baseline_volume = pre_data[0]['volume']
        if baseline_volume <= 0:
            # Find first non-zero volume
            for d in pre_data:
                if d['volume'] > 0:
                    baseline_volume = d['volume']
                    break

        if baseline_volume <= 0:
            return {'has_data': False}

        volume_ratios = []
        first_1x_minute = None

        for i, d in enumerate(pre_data):
            minutes_before = len(pre_data) - i  # T-10, T-9, ...
            ratio = d['volume'] / baseline_volume if baseline_volume > 0 else 0
            volume_ratios.append({
                'minutes_before': minutes_before,
                'ratio': ratio,
                'raw_volume': d['volume']
            })

            if first_1x_minute is None and ratio >= 1.0:
                first_1x_minute = minutes_before

        # Calculate acceleration (is volume growing faster?)
        acceleration = 0
        if len(volume_ratios) >= 4:
            early_avg = sum(r['ratio'] for r in volume_ratios[:3]) / 3
            late_avg = sum(r['ratio'] for r in volume_ratios[-3:]) / 3
            acceleration = late_avg - early_avg

        return {
            'has_data': True,
            'volume_ratios': volume_ratios,
            'first_1x_minute': first_1x_minute,
            'acceleration': acceleration,
            'baseline_volume': baseline_volume
        }

    def analyze_price_momentum(self, pre_data: List[Dict], direction: str) -> Dict:
        """
        Analyze price momentum pattern before alert.

        Returns per-minute price changes and when thresholds were crossed.
        """
        if len(pre_data) < 2:
            return {'has_data': False}

        baseline_price = pre_data[0]['price']
        if baseline_price <= 0:
            return {'has_data': False}

        price_changes = []
        first_025_minute = None
        first_050_minute = None
        first_075_minute = None

        is_drop = direction.lower() in ['drop', 'down', 'fall']

        for i, d in enumerate(pre_data):
            minutes_before = len(pre_data) - i

            if is_drop:
                # For drops, calculate decline percentage
                change_pct = ((baseline_price - d['price']) / baseline_price) * 100
            else:
                # For rises, calculate gain percentage
                change_pct = ((d['price'] - baseline_price) / baseline_price) * 100

            price_changes.append({
                'minutes_before': minutes_before,
                'change_pct': change_pct,
                'price': d['price']
            })

            # Track threshold crossings
            if first_025_minute is None and change_pct >= 0.25:
                first_025_minute = minutes_before
            if first_050_minute is None and change_pct >= 0.50:
                first_050_minute = minutes_before
            if first_075_minute is None and change_pct >= 0.75:
                first_075_minute = minutes_before

        # Calculate velocity (average % change per minute in last 5 minutes)
        if len(price_changes) >= 5:
            late_changes = [p['change_pct'] for p in price_changes[-5:]]
            velocity = max(late_changes) / 5 if late_changes else 0
        else:
            velocity = 0

        return {
            'has_data': True,
            'price_changes': price_changes,
            'first_025_minute': first_025_minute,
            'first_050_minute': first_050_minute,
            'first_075_minute': first_075_minute,
            'velocity': velocity,
            'baseline_price': baseline_price
        }

    def check_1min_precursor(self, symbol: str, alert_time: datetime) -> Dict:
        """
        Check if a 1-min alert preceded this 5-min alert.
        """
        if not os.path.exists(self.excel_path):
            return {'found': False}

        workbook = openpyxl.load_workbook(self.excel_path, read_only=True)

        if "1min_alerts" not in workbook.sheetnames:
            workbook.close()
            return {'found': False}

        sheet = workbook["1min_alerts"]

        # Look for 1-min alerts 1-5 minutes before the 5-min alert
        window_start = alert_time - timedelta(minutes=5)
        window_end = alert_time - timedelta(minutes=1)

        found_precursor = None

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            try:
                date_val = row[0]
                time_val = row[1]
                row_symbol = str(row[2]) if row[2] else ''

                if row_symbol != symbol:
                    continue

                # Parse datetime
                if isinstance(date_val, datetime):
                    date_str = date_val.strftime("%Y-%m-%d")
                else:
                    date_str = str(date_val)

                if isinstance(time_val, datetime):
                    time_str = time_val.strftime("%H:%M:%S")
                else:
                    time_str = str(time_val)

                try:
                    row_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    row_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")

                if window_start <= row_dt <= window_end:
                    delta_minutes = (alert_time - row_dt).total_seconds() / 60
                    found_precursor = {
                        'time': row_dt,
                        'delta_minutes': delta_minutes,
                        'direction': str(row[3]) if row[3] else 'Unknown'
                    }
                    break

            except Exception:
                continue

        workbook.close()

        if found_precursor:
            return {'found': True, **found_precursor}
        return {'found': False}

    def analyze_single_alert(self, alert: Dict) -> Optional[Dict]:
        """
        Complete analysis for one 5-min alert.
        """
        symbol = alert['symbol']
        alert_time = alert['datetime']
        direction = alert['direction']

        # Get pre-alert data (10 minutes before)
        pre_data = self.get_pre_alert_data(symbol, alert_time, minutes_back=10)

        if len(pre_data) < 5:
            logger.warning(f"Insufficient data for {symbol} at {alert_time} (only {len(pre_data)} points)")
            return None

        # Run all analyses
        volume_analysis = self.analyze_volume_buildup(pre_data, alert.get('volume_multiplier', 1.25))
        price_analysis = self.analyze_price_momentum(pre_data, direction)
        precursor_check = self.check_1min_precursor(symbol, alert_time)

        return {
            'alert': alert,
            'pre_data_points': len(pre_data),
            'volume': volume_analysis,
            'price': price_analysis,
            'precursor': precursor_check
        }

    def run_analysis(self, days: int = 3, from_date: str = None) -> Dict:
        """
        Run full analysis on historical alerts.

        Returns comprehensive statistics.
        """
        # Load alerts
        alerts = self.load_5min_alerts(days=days, from_date=from_date)

        if not alerts:
            logger.error("No alerts found to analyze")
            return {'error': 'No alerts found'}

        logger.info(f"Analyzing {len(alerts)} alerts...")

        # Analyze each alert
        results = []
        for i, alert in enumerate(alerts):
            if i > 0 and i % 10 == 0:
                logger.info(f"Progress: {i}/{len(alerts)} alerts analyzed")

            result = self.analyze_single_alert(alert)
            if result:
                results.append(result)

        logger.info(f"Successfully analyzed {len(results)} alerts")

        # Compile statistics
        stats = self._compile_statistics(results)

        return {
            'alerts_analyzed': len(results),
            'total_alerts': len(alerts),
            'results': results,
            'statistics': stats
        }

    def _compile_statistics(self, results: List[Dict]) -> Dict:
        """
        Compile aggregate statistics from results.
        """
        if not results:
            return {}

        # Per-minute aggregates
        minute_stats = defaultdict(lambda: {
            'volume_ratios': [],
            'price_changes': [],
            'count': 0
        })

        # Threshold crossing counts
        volume_1x_crossings = defaultdict(int)  # minute -> count
        price_025_crossings = defaultdict(int)
        price_050_crossings = defaultdict(int)

        precursor_count = 0
        precursor_led_to_alert = 0

        time_of_day_counts = defaultdict(int)

        for r in results:
            # Volume analysis
            if r['volume'].get('has_data'):
                for vr in r['volume'].get('volume_ratios', []):
                    minute_stats[vr['minutes_before']]['volume_ratios'].append(vr['ratio'])
                    minute_stats[vr['minutes_before']]['count'] += 1

                if r['volume'].get('first_1x_minute'):
                    volume_1x_crossings[r['volume']['first_1x_minute']] += 1

            # Price analysis
            if r['price'].get('has_data'):
                for pc in r['price'].get('price_changes', []):
                    minute_stats[pc['minutes_before']]['price_changes'].append(pc['change_pct'])

                if r['price'].get('first_025_minute'):
                    price_025_crossings[r['price']['first_025_minute']] += 1
                if r['price'].get('first_050_minute'):
                    price_050_crossings[r['price']['first_050_minute']] += 1

            # Precursor check
            if r['precursor'].get('found'):
                precursor_count += 1

            # Time of day
            hour = r['alert']['datetime'].hour
            time_of_day_counts[hour] += 1

        # Calculate averages per minute
        per_minute_summary = {}
        for minute in sorted(minute_stats.keys(), reverse=True):
            data = minute_stats[minute]
            avg_vol = sum(data['volume_ratios']) / len(data['volume_ratios']) if data['volume_ratios'] else 0
            avg_price = sum(data['price_changes']) / len(data['price_changes']) if data['price_changes'] else 0

            # Calculate what % of alerts showed signal at this minute
            vol_signal_pct = (sum(1 for v in data['volume_ratios'] if v >= 1.0) / len(data['volume_ratios']) * 100) if data['volume_ratios'] else 0
            price_signal_pct = (sum(1 for p in data['price_changes'] if p >= 0.5) / len(data['price_changes']) * 100) if data['price_changes'] else 0

            per_minute_summary[minute] = {
                'avg_volume_ratio': round(avg_vol, 2),
                'avg_price_change': round(avg_price, 2),
                'pct_volume_above_1x': round(vol_signal_pct, 1),
                'pct_price_above_05': round(price_signal_pct, 1),
                'sample_size': len(data['volume_ratios'])
            }

        return {
            'per_minute': per_minute_summary,
            'volume_1x_first_seen_at': dict(volume_1x_crossings),
            'price_050_first_seen_at': dict(price_050_crossings),
            'precursor_count': precursor_count,
            'precursor_pct': round(precursor_count / len(results) * 100, 1) if results else 0,
            'time_of_day': dict(time_of_day_counts)
        }

    def generate_report(self, analysis_result: Dict, output_path: str = None):
        """
        Generate Excel report with findings.
        """
        if 'error' in analysis_result:
            logger.error(f"Cannot generate report: {analysis_result['error']}")
            return

        if output_path is None:
            # Create reports directory
            reports_dir = Path("data/reports")
            reports_dir.mkdir(parents=True, exist_ok=True)
            output_path = str(reports_dir / f"5min_precursor_analysis_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx")

        workbook = openpyxl.Workbook()

        # Sheet 1: Minute-by-Minute Summary
        ws1 = workbook.active
        ws1.title = "Minute Summary"

        headers1 = ["Minutes Before", "Avg Volume Ratio", "Avg Price Change %",
                   "% With Vol > 1.0x", "% With Price > 0.5%", "Sample Size"]
        ws1.append(headers1)

        stats = analysis_result.get('statistics', {})
        per_minute = stats.get('per_minute', {})

        for minute in sorted(per_minute.keys(), reverse=True):
            data = per_minute[minute]
            ws1.append([
                f"T-{minute}",
                f"{data['avg_volume_ratio']:.2f}x",
                f"{data['avg_price_change']:.2f}%",
                f"{data['pct_volume_above_1x']:.1f}%",
                f"{data['pct_price_above_05']:.1f}%",
                data['sample_size']
            ])

        # Sheet 2: Alert Details
        ws2 = workbook.create_sheet("Alert Details")

        headers2 = ["Date", "Time", "Symbol", "Direction", "Change %", "Vol Mult",
                   "First Vol>1x (min)", "First Price>0.5% (min)", "Had 1min Precursor"]
        ws2.append(headers2)

        for r in analysis_result.get('results', []):
            alert = r['alert']
            vol = r.get('volume', {})
            price = r.get('price', {})
            precursor = r.get('precursor', {})

            ws2.append([
                alert['date'],
                alert['time'],
                alert['symbol'],
                alert['direction'],
                f"{alert['change_pct']:.2f}%",
                f"{alert['volume_multiplier']:.2f}x",
                vol.get('first_1x_minute', '-'),
                price.get('first_050_minute', '-'),
                "Yes" if precursor.get('found') else "No"
            ])

        # Sheet 3: Key Findings
        ws3 = workbook.create_sheet("Key Findings")

        ws3.append(["Metric", "Value"])
        ws3.append(["Total Alerts Analyzed", analysis_result.get('alerts_analyzed', 0)])
        ws3.append([""])
        ws3.append(["1-Min Precursor Rate", f"{stats.get('precursor_pct', 0):.1f}%"])
        ws3.append([""])

        # Find earliest reliable signal
        ws3.append(["EARLIEST RELIABLE SIGNALS:"])
        for minute in sorted(per_minute.keys(), reverse=True):
            data = per_minute[minute]
            if data['pct_volume_above_1x'] >= 40:
                ws3.append([f"Volume > 1.0x first seen at T-{minute}", f"{data['pct_volume_above_1x']:.1f}% of alerts"])
                break

        for minute in sorted(per_minute.keys(), reverse=True):
            data = per_minute[minute]
            if data['pct_price_above_05'] >= 40:
                ws3.append([f"Price > 0.5% first seen at T-{minute}", f"{data['pct_price_above_05']:.1f}% of alerts"])
                break

        # Sheet 4: Time of Day Analysis
        ws4 = workbook.create_sheet("Time of Day")
        ws4.append(["Hour", "Alert Count"])

        for hour in sorted(stats.get('time_of_day', {}).keys()):
            ws4.append([f"{hour}:00", stats['time_of_day'][hour]])

        # Save
        workbook.save(output_path)
        logger.info(f"Report saved to {output_path}")

        return output_path

    def print_summary(self, analysis_result: Dict):
        """
        Print summary to console.
        """
        if 'error' in analysis_result:
            print(f"Error: {analysis_result['error']}")
            return

        stats = analysis_result.get('statistics', {})
        per_minute = stats.get('per_minute', {})

        print("\n" + "=" * 60)
        print("5-MINUTE ALERT PRECURSOR ANALYSIS (10-Min Window)")
        print("=" * 60)
        print(f"Alerts Analyzed: {analysis_result.get('alerts_analyzed', 0)}")
        print()

        print("MINUTE-BY-MINUTE PROGRESSION:")
        print("-" * 60)
        print(f"{'Minute':<10} {'Avg Vol':<12} {'Avg Price':<12} {'% Vol>1x':<12} {'% Price>0.5%':<12}")
        print("-" * 60)

        for minute in sorted(per_minute.keys(), reverse=True):
            data = per_minute[minute]
            marker = ""
            if data['pct_volume_above_1x'] >= 40 and not hasattr(self, '_vol_marked'):
                marker = " <-- volume signal"
                self._vol_marked = True
            if data['pct_price_above_05'] >= 40 and not hasattr(self, '_price_marked'):
                marker = " <-- price signal"
                self._price_marked = True

            print(f"T-{minute:<8} {data['avg_volume_ratio']:.2f}x{'':<7} "
                  f"{data['avg_price_change']:.2f}%{'':<7} "
                  f"{data['pct_volume_above_1x']:.1f}%{'':<7} "
                  f"{data['pct_price_above_05']:.1f}%{marker}")

        print("-" * 60)
        print()

        print("KEY FINDINGS:")
        print(f"  - 1-Min Alert Precursor Rate: {stats.get('precursor_pct', 0):.1f}%")
        print()

        print("ACTIONABLE THRESHOLDS:")
        for minute in sorted(per_minute.keys(), reverse=True):
            data = per_minute[minute]
            if data['pct_volume_above_1x'] >= 40:
                print(f"  - Volume > 1.0x at T-{minute}: {data['pct_volume_above_1x']:.1f}% hit rate, {minute} min lead time")
                break

        for minute in sorted(per_minute.keys(), reverse=True):
            data = per_minute[minute]
            if data['pct_price_above_05'] >= 40:
                print(f"  - Price > 0.5% at T-{minute}: {data['pct_price_above_05']:.1f}% hit rate, {minute} min lead time")
                break

        if stats.get('precursor_pct', 0) >= 20:
            print(f"  - 1-min alert fired: {stats.get('precursor_pct', 0):.1f}% hit rate, 2-3 min lead time")

        print()

    def close(self):
        """Close database connection."""
        if self.conn:
            self.conn.close()
            self.conn = None


def main():
    parser = argparse.ArgumentParser(description="Analyze 5-min alert precursor patterns")
    parser.add_argument("--days", type=int, default=3, help="Number of days to analyze (default: 3)")
    parser.add_argument("--from-date", type=str, help="Start date (YYYY-MM-DD)")
    parser.add_argument("--no-report", action="store_true", help="Skip Excel report generation")

    args = parser.parse_args()

    analyzer = PreAlertAnalyzer()

    try:
        # Run analysis
        result = analyzer.run_analysis(days=args.days, from_date=args.from_date)

        # Print summary
        analyzer.print_summary(result)

        # Generate report
        if not args.no_report and 'error' not in result:
            report_path = analyzer.generate_report(result)
            print(f"\nReport saved to: {report_path}")

    finally:
        analyzer.close()


if __name__ == "__main__":
    main()
