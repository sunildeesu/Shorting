#!/Users/sunildeesu/myProjects/ShortIndicator/venv/bin/python3
"""
ATR Breakout Strategy Backtesting
Simulates the ATR breakout strategy over historical data and generates detailed Excel report

Strategy Being Tested:
1. Entry: Price crosses above Open + (2.5 × ATR(20))
2. Filter: ATR(20) < ATR(30) (volatility contracting)
3. Stop Loss: Entry - (0.5 × ATR(20))
4. Exit: Friday close OR Stop loss hit

Author: Sunil Kumar Durganaik
"""

import json
import os
import sys
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple
import pandas as pd
import pandas_ta as ta
from kiteconnect import KiteConnect
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import config

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/atr_backtest.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class ATRBacktester:
    """Backtests the ATR breakout strategy"""

    def __init__(self, start_date: datetime, end_date: datetime):
        """Initialize the backtester"""
        self.start_date = start_date
        self.end_date = end_date

        # Initialize Kite Connect
        self.kite = KiteConnect(api_key=config.KITE_API_KEY)
        self.kite.set_access_token(config.KITE_ACCESS_TOKEN)

        # Load stocks and instrument tokens
        self.stocks = self._load_stock_list()
        self.instrument_tokens = self._load_instrument_tokens()

        # Results storage
        self.all_trades = []
        self.daily_signals = []

        logger.info(f"Backtester initialized: {start_date.date()} to {end_date.date()}")
        logger.info(f"Testing {len(self.stocks)} stocks")

    def _load_stock_list(self) -> List[str]:
        """Load F&O stock list"""
        try:
            with open(config.STOCK_LIST_FILE, 'r') as f:
                data = json.load(f)
                return data['stocks']
        except Exception as e:
            logger.error(f"Failed to load stock list: {e}")
            return []

    def _load_instrument_tokens(self) -> Dict[str, int]:
        """Load instrument tokens"""
        tokens_file = "data/instrument_tokens.json"
        try:
            if os.path.exists(tokens_file):
                with open(tokens_file, 'r') as f:
                    return json.load(f)
            else:
                logger.warning("Instrument tokens not found, fetching...")
                return self._fetch_instrument_tokens()
        except Exception as e:
            logger.error(f"Failed to load tokens: {e}")
            return {}

    def _fetch_instrument_tokens(self) -> Dict[str, int]:
        """Fetch instrument tokens from Kite"""
        try:
            instruments = self.kite.instruments("NSE")
            token_map = {}
            for inst in instruments:
                if inst['tradingsymbol'] in self.stocks:
                    token_map[inst['tradingsymbol']] = inst['instrument_token']

            os.makedirs("data", exist_ok=True)
            with open("data/instrument_tokens.json", 'w') as f:
                json.dump(token_map, f, indent=2)

            return token_map
        except Exception as e:
            logger.error(f"Failed to fetch tokens: {e}")
            return {}

    def fetch_historical_data(
        self,
        symbol: str,
        from_date: datetime,
        to_date: datetime
    ) -> Optional[pd.DataFrame]:
        """Fetch historical data for backtesting"""
        try:
            if symbol not in self.instrument_tokens:
                logger.warning(f"{symbol}: No instrument token")
                return None

            token = self.instrument_tokens[symbol]

            # Fetch with extra buffer for ATR calculation
            buffer_days = 60
            adjusted_from = from_date - timedelta(days=buffer_days)

            data = self.kite.historical_data(
                instrument_token=token,
                from_date=adjusted_from.date(),
                to_date=to_date.date(),
                interval="day"
            )

            if not data:
                return None

            df = pd.DataFrame(data)
            df.columns = df.columns.str.lower()
            df['date'] = pd.to_datetime(df['date']).dt.tz_localize(None)

            return df

        except Exception as e:
            logger.error(f"{symbol}: Failed to fetch data: {e}")
            return None

    def calculate_atr(self, df: pd.DataFrame, period: int) -> pd.Series:
        """Calculate ATR using pandas-ta"""
        try:
            atr_series = ta.atr(
                high=df['high'],
                low=df['low'],
                close=df['close'],
                length=period
            )
            return atr_series
        except Exception as e:
            logger.error(f"ATR calculation failed: {e}")
            return pd.Series([None] * len(df))

    def simulate_stock(self, symbol: str) -> List[Dict]:
        """Simulate ATR strategy for a single stock"""
        trades = []

        try:
            # Fetch historical data
            df = self.fetch_historical_data(symbol, self.start_date, self.end_date)

            if df is None or len(df) < config.ATR_PERIOD_LONG:
                return trades

            # Calculate ATR(20) and ATR(30)
            df['atr_20'] = self.calculate_atr(df, config.ATR_PERIOD_SHORT)
            df['atr_30'] = self.calculate_atr(df, config.ATR_PERIOD_LONG)

            # NEW: Calculate 20-day MA for price filter
            if config.ATR_PRICE_FILTER:
                df['ma_20'] = df['close'].rolling(window=config.ATR_PRICE_MA_PERIOD).mean()

            # NEW: Calculate 20-day average volume for volume filter
            if config.ATR_VOLUME_FILTER:
                df['avg_volume'] = df['volume'].rolling(window=20).mean()

            # Filter to backtest period
            df = df[df['date'] >= self.start_date].copy()

            # Track open position
            open_position = None

            # Iterate through each day
            for idx, row in df.iterrows():
                date = row['date']
                open_price = row['open']
                high_price = row['high']
                low_price = row['low']
                close_price = row['close']
                volume = row['volume']
                atr_20 = row['atr_20']
                atr_30 = row['atr_30']

                # Skip if ATR not available
                if pd.isna(atr_20) or pd.isna(atr_30):
                    continue

                # Calculate entry level and stop loss
                entry_level = open_price + (config.ATR_ENTRY_MULTIPLIER * atr_20)
                stop_loss = entry_level - (config.ATR_STOP_MULTIPLIER * atr_20)

                # Check volatility filter
                volatility_filter_passed = atr_20 < atr_30 if config.ATR_FILTER_CONTRACTION else True

                # NEW: Check price trend filter
                price_filter_passed = True
                if config.ATR_PRICE_FILTER and 'ma_20' in df.columns:
                    ma_20 = row.get('ma_20')
                    if pd.notna(ma_20):
                        price_filter_passed = close_price > ma_20

                # NEW: Check volume confirmation filter
                volume_filter_passed = True
                if config.ATR_VOLUME_FILTER and 'avg_volume' in df.columns:
                    avg_volume = row.get('avg_volume')
                    if pd.notna(avg_volume):
                        volume_filter_passed = volume >= (avg_volume * config.ATR_VOLUME_MULTIPLIER)

                # If no position, check for entry
                if open_position is None:
                    # Check if price broke out during the day AND all filters pass
                    if (high_price >= entry_level and
                        volatility_filter_passed and
                        price_filter_passed and
                        volume_filter_passed):
                        # Enter position
                        open_position = {
                            'symbol': symbol,
                            'entry_date': date,
                            'entry_price': entry_level,  # Assume entry at entry_level
                            'stop_loss': stop_loss,
                            'atr_20': atr_20,
                            'atr_30': atr_30,
                            'entry_open': open_price,
                            'volatility_filter': 'PASSED',
                            'volume': volume
                        }

                # If position open, check for exit
                elif open_position is not None:
                    exit_triggered = False
                    exit_price = None
                    exit_reason = None

                    # Check stop loss (intraday low)
                    if low_price <= open_position['stop_loss']:
                        exit_triggered = True
                        exit_price = open_position['stop_loss']
                        exit_reason = 'STOP_LOSS'

                    # Check Friday exit (if enabled)
                    elif date.weekday() == 4 and config.ATR_FRIDAY_EXIT:
                        exit_triggered = True
                        exit_price = close_price
                        exit_reason = 'FRIDAY_EXIT'

                    # Exit position if triggered
                    if exit_triggered:
                        entry_price = open_position['entry_price']
                        pnl = exit_price - entry_price
                        pnl_percent = (pnl / entry_price) * 100

                        trade = {
                            'symbol': symbol,
                            'entry_date': open_position['entry_date'].strftime('%Y-%m-%d'),
                            'entry_day': open_position['entry_date'].strftime('%A'),
                            'exit_date': date.strftime('%Y-%m-%d'),
                            'exit_day': date.strftime('%A'),
                            'holding_days': (date - open_position['entry_date']).days,
                            'entry_price': entry_price,
                            'exit_price': exit_price,
                            'stop_loss': open_position['stop_loss'],
                            'pnl': pnl,
                            'pnl_percent': pnl_percent,
                            'result': 'WIN' if pnl > 0 else 'LOSS',
                            'exit_reason': exit_reason,
                            'atr_20': open_position['atr_20'],
                            'atr_30': open_position['atr_30'],
                            'volatility_filter': open_position['volatility_filter'],
                            'volume': open_position['volume']
                        }

                        trades.append(trade)
                        open_position = None

            # Close any open position at end of backtest period
            if open_position is not None:
                last_row = df.iloc[-1]
                entry_price = open_position['entry_price']
                exit_price = last_row['close']
                pnl = exit_price - entry_price
                pnl_percent = (pnl / entry_price) * 100

                trade = {
                    'symbol': symbol,
                    'entry_date': open_position['entry_date'].strftime('%Y-%m-%d'),
                    'entry_day': open_position['entry_date'].strftime('%A'),
                    'exit_date': last_row['date'].strftime('%Y-%m-%d'),
                    'exit_day': last_row['date'].strftime('%A'),
                    'holding_days': (last_row['date'] - open_position['entry_date']).days,
                    'entry_price': entry_price,
                    'exit_price': exit_price,
                    'stop_loss': open_position['stop_loss'],
                    'pnl': pnl,
                    'pnl_percent': pnl_percent,
                    'result': 'WIN' if pnl > 0 else 'LOSS',
                    'exit_reason': 'BACKTEST_END',
                    'atr_20': open_position['atr_20'],
                    'atr_30': open_position['atr_30'],
                    'volatility_filter': open_position['volatility_filter'],
                    'volume': open_position['volume']
                }

                trades.append(trade)

            return trades

        except Exception as e:
            logger.error(f"{symbol}: Simulation failed: {e}")
            return trades

    def run_backtest(self, max_stocks: Optional[int] = None):
        """Run backtest on all stocks"""
        logger.info("=" * 80)
        logger.info("STARTING ATR STRATEGY BACKTEST")
        logger.info("=" * 80)

        stocks_to_test = self.stocks[:max_stocks] if max_stocks else self.stocks

        for idx, symbol in enumerate(stocks_to_test, 1):
            logger.info(f"[{idx}/{len(stocks_to_test)}] Backtesting {symbol}...")

            trades = self.simulate_stock(symbol)

            if trades:
                self.all_trades.extend(trades)
                logger.info(f"  {symbol}: {len(trades)} trades generated")

            # Rate limiting
            if idx < len(stocks_to_test):
                import time
                time.sleep(config.REQUEST_DELAY_SECONDS)

        logger.info("=" * 80)
        logger.info(f"BACKTEST COMPLETE: {len(self.all_trades)} total trades")
        logger.info("=" * 80)

    def generate_excel_report(self, filename: str = "atr_backtest_results.xlsx"):
        """Generate comprehensive Excel report"""
        logger.info(f"Generating Excel report: {filename}")

        if not self.all_trades:
            logger.warning("No trades to report")
            return

        # Create DataFrame
        df_trades = pd.DataFrame(self.all_trades)

        # Calculate statistics
        stats = self._calculate_statistics(df_trades)

        # Create workbook
        wb = Workbook()

        # Sheet 1: Summary
        self._create_summary_sheet(wb, stats)

        # Sheet 2: All Trades
        self._create_trades_sheet(wb, df_trades)

        # Sheet 3: Monthly Performance
        self._create_monthly_sheet(wb, df_trades)

        # Sheet 4: Stock Performance
        self._create_stock_performance_sheet(wb, df_trades)

        # Sheet 5: Analysis & Recommendations
        self._create_analysis_sheet(wb, df_trades, stats)

        # Save workbook
        output_dir = "data/backtest_results"
        os.makedirs(output_dir, exist_ok=True)
        filepath = os.path.join(output_dir, filename)
        wb.save(filepath)

        logger.info(f"✓ Excel report saved: {filepath}")
        return filepath

    def _calculate_statistics(self, df: pd.DataFrame) -> Dict:
        """Calculate performance statistics"""
        stats = {}

        # Basic counts
        stats['total_trades'] = len(df)
        stats['winning_trades'] = len(df[df['result'] == 'WIN'])
        stats['losing_trades'] = len(df[df['result'] == 'LOSS'])

        # Win rate
        stats['win_rate'] = (stats['winning_trades'] / stats['total_trades'] * 100) if stats['total_trades'] > 0 else 0

        # P&L statistics
        stats['total_pnl'] = df['pnl_percent'].sum()
        stats['avg_pnl'] = df['pnl_percent'].mean()
        stats['avg_win'] = df[df['result'] == 'WIN']['pnl_percent'].mean() if stats['winning_trades'] > 0 else 0
        stats['avg_loss'] = df[df['result'] == 'LOSS']['pnl_percent'].mean() if stats['losing_trades'] > 0 else 0

        # Best/worst trades
        stats['best_trade'] = df.loc[df['pnl_percent'].idxmax()] if len(df) > 0 else None
        stats['worst_trade'] = df.loc[df['pnl_percent'].idxmin()] if len(df) > 0 else None

        # Holding period
        stats['avg_holding_days'] = df['holding_days'].mean()

        # Exit reasons
        stats['stop_loss_exits'] = len(df[df['exit_reason'] == 'STOP_LOSS'])
        stats['friday_exits'] = len(df[df['exit_reason'] == 'FRIDAY_EXIT'])

        # Risk/Reward
        if stats['avg_loss'] != 0:
            stats['risk_reward_ratio'] = abs(stats['avg_win'] / stats['avg_loss'])
        else:
            stats['risk_reward_ratio'] = 0

        # Unique stocks traded
        stats['unique_stocks'] = df['symbol'].nunique()

        return stats

    def _create_summary_sheet(self, wb: Workbook, stats: Dict):
        """Create summary sheet with key metrics"""
        ws = wb.active
        ws.title = "Summary"

        # Title
        ws['A1'] = "ATR BREAKOUT STRATEGY - BACKTEST RESULTS"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        ws['A2'] = f"Period: {self.start_date.date()} to {self.end_date.date()}"
        ws['A2'].font = Font(size=12)
        ws.merge_cells('A2:D2')

        # Key metrics
        row = 4
        metrics = [
            ("OVERALL PERFORMANCE", ""),
            ("Total Trades", stats['total_trades']),
            ("Winning Trades", stats['winning_trades']),
            ("Losing Trades", stats['losing_trades']),
            ("Win Rate", f"{stats['win_rate']:.2f}%"),
            ("", ""),
            ("PROFIT & LOSS", ""),
            ("Total P&L", f"{stats['total_pnl']:.2f}%"),
            ("Average P&L per Trade", f"{stats['avg_pnl']:.2f}%"),
            ("Average Win", f"{stats['avg_win']:.2f}%"),
            ("Average Loss", f"{stats['avg_loss']:.2f}%"),
            ("Risk/Reward Ratio", f"{stats['risk_reward_ratio']:.2f}"),
            ("", ""),
            ("HOLDING PERIOD", ""),
            ("Average Holding Days", f"{stats['avg_holding_days']:.1f}"),
            ("", ""),
            ("EXIT ANALYSIS", ""),
            ("Stop Loss Exits", stats['stop_loss_exits']),
            ("Friday Exits", stats['friday_exits']),
            ("", ""),
            ("COVERAGE", ""),
            ("Unique Stocks Traded", stats['unique_stocks']),
        ]

        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value

            if label and not value:  # Section headers
                ws[f'A{row}'].font = Font(bold=True, size=12)
                ws.merge_cells(f'A{row}:B{row}')

            row += 1

        # Best/worst trades
        row += 1
        ws[f'A{row}'] = "BEST TRADE"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        if stats['best_trade'] is not None:
            best = stats['best_trade']
            ws[f'A{row}'] = f"{best['symbol']}: {best['pnl_percent']:.2f}% ({best['entry_date']} → {best['exit_date']})"

        row += 2
        ws[f'A{row}'] = "WORST TRADE"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        if stats['worst_trade'] is not None:
            worst = stats['worst_trade']
            ws[f'A{row}'] = f"{worst['symbol']}: {worst['pnl_percent']:.2f}% ({worst['entry_date']} → {worst['exit_date']})"

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def _create_trades_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create sheet with all trades"""
        ws = wb.create_sheet("All Trades")

        # Headers
        headers = [
            'Symbol', 'Entry Date', 'Entry Day', 'Exit Date', 'Exit Day',
            'Holding Days', 'Entry Price', 'Exit Price', 'Stop Loss',
            'P&L ₹', 'P&L %', 'Result', 'Exit Reason',
            'ATR(20)', 'ATR(30)', 'Volatility Filter', 'Volume'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Data
        for row_idx, trade in enumerate(df.to_dict('records'), 2):
            ws.cell(row_idx, 1, trade['symbol'])
            ws.cell(row_idx, 2, trade['entry_date'])
            ws.cell(row_idx, 3, trade['entry_day'])
            ws.cell(row_idx, 4, trade['exit_date'])
            ws.cell(row_idx, 5, trade['exit_day'])
            ws.cell(row_idx, 6, trade['holding_days'])
            ws.cell(row_idx, 7, f"₹{trade['entry_price']:.2f}")
            ws.cell(row_idx, 8, f"₹{trade['exit_price']:.2f}")
            ws.cell(row_idx, 9, f"₹{trade['stop_loss']:.2f}")
            ws.cell(row_idx, 10, f"₹{trade['pnl']:.2f}")
            ws.cell(row_idx, 11, f"{trade['pnl_percent']:.2f}%")
            ws.cell(row_idx, 12, trade['result'])
            ws.cell(row_idx, 13, trade['exit_reason'])
            ws.cell(row_idx, 14, f"₹{trade['atr_20']:.2f}")
            ws.cell(row_idx, 15, f"₹{trade['atr_30']:.2f}")
            ws.cell(row_idx, 16, trade['volatility_filter'])
            ws.cell(row_idx, 17, int(trade['volume']))

            # Color code result
            result_cell = ws.cell(row_idx, 12)
            if trade['result'] == 'WIN':
                result_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                result_cell.font = Font(color="006100")
            else:
                result_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                result_cell.font = Font(color="9C0006")

        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_monthly_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create monthly performance sheet"""
        ws = wb.create_sheet("Monthly Performance")

        # Parse dates and create month column
        df['entry_month'] = pd.to_datetime(df['entry_date']).dt.to_period('M')

        # Group by month
        monthly = df.groupby('entry_month').agg({
            'pnl_percent': ['count', 'sum', 'mean'],
            'result': lambda x: (x == 'WIN').sum()
        })

        monthly.columns = ['trades', 'total_pnl', 'avg_pnl', 'wins']
        monthly['losses'] = monthly['trades'] - monthly['wins']
        monthly['win_rate'] = (monthly['wins'] / monthly['trades'] * 100)
        monthly = monthly.reset_index()

        # Headers
        headers = ['Month', 'Trades', 'Wins', 'Losses', 'Win Rate %', 'Total P&L %', 'Avg P&L %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Data
        for row_idx, row in enumerate(monthly.to_dict('records'), 2):
            ws.cell(row_idx, 1, str(row['entry_month']))
            ws.cell(row_idx, 2, int(row['trades']))
            ws.cell(row_idx, 3, int(row['wins']))
            ws.cell(row_idx, 4, int(row['losses']))
            ws.cell(row_idx, 5, f"{row['win_rate']:.2f}%")
            ws.cell(row_idx, 6, f"{row['total_pnl']:.2f}%")
            ws.cell(row_idx, 7, f"{row['avg_pnl']:.2f}%")

        # Auto-size
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_stock_performance_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create per-stock performance sheet"""
        ws = wb.create_sheet("Stock Performance")

        # Group by stock
        stock_perf = df.groupby('symbol').agg({
            'pnl_percent': ['count', 'sum', 'mean'],
            'result': lambda x: (x == 'WIN').sum()
        })

        stock_perf.columns = ['trades', 'total_pnl', 'avg_pnl', 'wins']
        stock_perf['losses'] = stock_perf['trades'] - stock_perf['wins']
        stock_perf['win_rate'] = (stock_perf['wins'] / stock_perf['trades'] * 100)
        stock_perf = stock_perf.sort_values('total_pnl', ascending=False).reset_index()

        # Headers
        headers = ['Symbol', 'Trades', 'Wins', 'Losses', 'Win Rate %', 'Total P&L %', 'Avg P&L %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Data
        for row_idx, row in enumerate(stock_perf.to_dict('records'), 2):
            ws.cell(row_idx, 1, row['symbol'])
            ws.cell(row_idx, 2, int(row['trades']))
            ws.cell(row_idx, 3, int(row['wins']))
            ws.cell(row_idx, 4, int(row['losses']))
            ws.cell(row_idx, 5, f"{row['win_rate']:.2f}%")
            ws.cell(row_idx, 6, f"{row['total_pnl']:.2f}%")
            ws.cell(row_idx, 7, f"{row['avg_pnl']:.2f}%")

        # Auto-size
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_analysis_sheet(self, wb: Workbook, df: pd.DataFrame, stats: Dict):
        """Create analysis and recommendations sheet"""
        ws = wb.create_sheet("Analysis")

        # Title
        ws['A1'] = "STRATEGY ANALYSIS & RECOMMENDATIONS"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')

        row = 3

        # Win rate analysis
        ws[f'A{row}'] = "1. WIN RATE ANALYSIS"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1

        ws[f'A{row}'] = f"Overall Win Rate: {stats['win_rate']:.2f}%"
        row += 1

        if stats['win_rate'] >= 60:
            ws[f'A{row}'] = "✓ EXCELLENT: Win rate above 60% indicates strong strategy"
            ws[f'A{row}'].font = Font(color="006100")
        elif stats['win_rate'] >= 50:
            ws[f'A{row}'] = "✓ GOOD: Win rate above 50% is profitable"
            ws[f'A{row}'].font = Font(color="FF6600")
        else:
            ws[f'A{row}'] = "⚠ WARNING: Win rate below 50% - strategy needs improvement"
            ws[f'A{row}'].font = Font(color="9C0006")

        row += 2

        # Risk/Reward analysis
        ws[f'A{row}'] = "2. RISK/REWARD ANALYSIS"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1

        ws[f'A{row}'] = f"Average Win: {stats['avg_win']:.2f}%"
        row += 1
        ws[f'A{row}'] = f"Average Loss: {stats['avg_loss']:.2f}%"
        row += 1
        ws[f'A{row}'] = f"Risk/Reward Ratio: {stats['risk_reward_ratio']:.2f}"
        row += 1

        if stats['risk_reward_ratio'] >= 2:
            ws[f'A{row}'] = "✓ EXCELLENT: R:R ratio above 2:1"
            ws[f'A{row}'].font = Font(color="006100")
        elif stats['risk_reward_ratio'] >= 1.5:
            ws[f'A{row}'] = "✓ GOOD: R:R ratio acceptable"
            ws[f'A{row}'].font = Font(color="FF6600")
        else:
            ws[f'A{row}'] = "⚠ WARNING: R:R ratio too low"
            ws[f'A{row}'].font = Font(color="9C0006")

        row += 2

        # Exit analysis
        ws[f'A{row}'] = "3. EXIT STRATEGY ANALYSIS"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1

        stop_loss_pct = (stats['stop_loss_exits'] / stats['total_trades'] * 100) if stats['total_trades'] > 0 else 0
        ws[f'A{row}'] = f"Stop Loss Exits: {stats['stop_loss_exits']} ({stop_loss_pct:.1f}%)"
        row += 1

        friday_pct = (stats['friday_exits'] / stats['total_trades'] * 100) if stats['total_trades'] > 0 else 0
        ws[f'A{row}'] = f"Friday Exits: {stats['friday_exits']} ({friday_pct:.1f}%)"
        row += 1

        # Analyze exit effectiveness
        stop_loss_df = df[df['exit_reason'] == 'STOP_LOSS']
        friday_df = df[df['exit_reason'] == 'FRIDAY_EXIT']

        if len(stop_loss_df) > 0:
            stop_loss_protected = len(stop_loss_df[stop_loss_df['pnl_percent'] > -2])
            row += 1
            ws[f'A{row}'] = f"Stop loss protected {stop_loss_protected} trades from larger losses"

        row += 2

        # Recommendations
        ws[f'A{row}'] = "4. RECOMMENDATIONS"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1

        recommendations = []

        if stats['win_rate'] < 50:
            recommendations.append("• Increase ATR_ENTRY_MULTIPLIER to reduce false breakouts")
            recommendations.append("• Consider adding volume confirmation filter")

        if stats['risk_reward_ratio'] < 1.5:
            recommendations.append("• Tighten stop loss (reduce ATR_STOP_MULTIPLIER)")
            recommendations.append("• Or use trailing stop to capture more profits")

        if stats['avg_holding_days'] > 3:
            recommendations.append("• Consider adding time-based exit (e.g., 3 days max)")

        if stop_loss_pct > 70:
            recommendations.append("• Stop loss too tight - increase ATR_STOP_MULTIPLIER")

        if not recommendations:
            recommendations.append("✓ Strategy parameters are well-optimized")

        for rec in recommendations:
            ws[f'A{row}'] = rec
            row += 1

        # Column width
        ws.column_dimensions['A'].width = 80


def main():
    """Main entry point"""
    import argparse

    parser = argparse.ArgumentParser(description='Backtest ATR Breakout Strategy')
    parser.add_argument('--months', type=int, default=12, help='Number of months to backtest (default: 12)')
    parser.add_argument('--stocks', type=int, help='Limit number of stocks to test')
    args = parser.parse_args()

    # Calculate date range
    end_date = datetime.now()
    start_date = end_date - timedelta(days=args.months * 30)

    logger.info("=" * 80)
    logger.info("ATR BREAKOUT STRATEGY - BACKTESTING")
    logger.info("=" * 80)
    logger.info(f"Period: {start_date.date()} to {end_date.date()}")
    logger.info(f"Duration: {args.months} months")
    if args.stocks:
        logger.info(f"Testing: First {args.stocks} stocks")
    logger.info("=" * 80)

    # Create backtester
    backtester = ATRBacktester(start_date, end_date)

    # Run backtest
    backtester.run_backtest(max_stocks=args.stocks)

    # Generate report
    if backtester.all_trades:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"atr_backtest_{args.months}months_{timestamp}.xlsx"
        filepath = backtester.generate_excel_report(filename)

        logger.info("")
        logger.info("=" * 80)
        logger.info("BACKTEST COMPLETE")
        logger.info("=" * 80)
        logger.info(f"Total Trades: {len(backtester.all_trades)}")
        logger.info(f"Excel Report: {filepath}")
        logger.info("=" * 80)
    else:
        logger.warning("No trades generated. Check date range and stock data availability.")


if __name__ == "__main__":
    main()
