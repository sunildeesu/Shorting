#!/usr/bin/env python3
"""
Verify that all price fields are populated correctly in EOD report
"""

import openpyxl
import sys

def verify_report_prices(report_path: str):
    """Check all price columns in the Excel report"""

    print(f"Opening report: {report_path}")
    wb = openpyxl.load_workbook(report_path)
    ws = wb.active

    print(f"\n{'='*100}")
    print(f"{'Stock':<15} {'Current Price':<15} {'Buy Price':<15} {'Target Price':<15} {'Stop Loss':<15} {'Pattern':<30}")
    print(f"{'='*100}")

    issues = []
    total_rows = 0

    # Start from row 4 (data rows)
    for row_num in range(4, ws.max_row + 1):
        stock = ws.cell(row=row_num, column=1).value
        if not stock:
            break

        total_rows += 1

        # Column mappings:
        # 9 = Current Price
        # 11 = Buy/Entry Price
        # 12 = Target Price
        # 13 = Stop Loss
        # 8 = Chart Patterns

        current_price = ws.cell(row=row_num, column=9).value
        buy_price = ws.cell(row=row_num, column=11).value
        target_price = ws.cell(row=row_num, column=12).value
        stop_loss = ws.cell(row=row_num, column=13).value
        pattern = ws.cell(row=row_num, column=8).value

        # Extract numeric values
        def extract_price(val):
            if not val or val == '-':
                return None
            if isinstance(val, str):
                # Remove ₹ and commas
                return val.replace('₹', '').replace(',', '').strip()
            return str(val)

        current_price_str = extract_price(current_price)
        buy_price_str = extract_price(buy_price)
        target_price_str = extract_price(target_price)
        stop_loss_str = extract_price(stop_loss)

        print(f"{stock:<15} {current_price or '-':<15} {buy_price or '-':<15} {target_price or '-':<15} {stop_loss or '-':<15} {pattern or '-':<30}")

        # Check for issues
        has_pattern = pattern and pattern.strip() != ''

        if has_pattern:
            # If there's a pattern, these prices should be populated
            if not buy_price or buy_price == '-':
                issues.append(f"{stock}: Has pattern '{pattern}' but missing Buy Price")
            if not target_price or target_price == '-':
                issues.append(f"{stock}: Has pattern '{pattern}' but missing Target Price")
            if not stop_loss or stop_loss == '-':
                issues.append(f"{stock}: Has pattern '{pattern}' but missing Stop Loss")

        if not current_price or current_price == '-':
            issues.append(f"{stock}: Missing Current Price")

    print(f"{'='*100}")
    print(f"\nTotal stocks in report: {total_rows}")

    if issues:
        print(f"\n❌ Issues Found ({len(issues)}):")
        print(f"{'='*100}")
        for issue in issues:
            print(f"  - {issue}")
        return False
    else:
        print(f"\n✅ All price fields populated correctly!")
        return True

if __name__ == "__main__":
    if len(sys.argv) > 1:
        report_path = sys.argv[1]
    else:
        report_path = "data/eod_reports/2025/11/eod_analysis_2025-11-04.xlsx"

    try:
        success = verify_report_prices(report_path)
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
