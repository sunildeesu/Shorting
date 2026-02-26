#!/usr/bin/env python3
"""
Pre-Market Report Generator - Creates Excel reports for pre-market pattern alerts

Generates reports with two sheets:
1. Top 3 Patterns - Highlighted green for immediate action
2. All Detected Patterns - Reference for lower-priority setups

Folder structure: data/premarket_reports/YYYY/MM/premarket_analysis_YYYY-MM-DD.xlsx

Author: Sunil Kumar Durganaik
"""

import os
from datetime import datetime
from typing import List, Dict
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pattern_utils as pu

logger = logging.getLogger(__name__)


class PreMarketReportGenerator:
    """Generates Excel reports for pre-market pattern analysis"""

    def __init__(self, base_dir: str = "data/premarket_reports"):
        """
        Initialize report generator

        Args:
            base_dir: Base directory for reports (default: data/premarket_reports)
        """
        self.base_dir = base_dir

    def get_report_path(self, analysis_date: datetime = None) -> str:
        """
        Get report file path with YYYY/MM folder structure

        Args:
            analysis_date: Date of analysis (default: today)

        Returns:
            Full path to report file
        """
        if analysis_date is None:
            analysis_date = datetime.now()

        year = analysis_date.strftime('%Y')
        month = analysis_date.strftime('%m')
        filename = f"premarket_analysis_{analysis_date.strftime('%Y-%m-%d')}.xlsx"

        # Create directory structure: data/premarket_reports/YYYY/MM/
        report_dir = os.path.join(self.base_dir, year, month)
        os.makedirs(report_dir, exist_ok=True)

        return os.path.join(report_dir, filename)

    def generate_report(
        self,
        top_patterns: List[Dict],
        all_daily_patterns: List[Dict],
        all_hourly_patterns: List[Dict],
        market_regime: str = "NEUTRAL",
        analysis_date: datetime = None
    ) -> str:
        """
        Generate Excel report with top patterns and all patterns

        Args:
            top_patterns: Top 1-3 ranked patterns for alerts
            all_daily_patterns: All detected daily patterns
            all_hourly_patterns: All detected hourly patterns
            market_regime: Current market regime
            analysis_date: Date of analysis (default: today)

        Returns:
            Path to generated report file
        """
        if analysis_date is None:
            analysis_date = datetime.now()

        # Create workbook with two sheets
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Sheet 1: Top 3 Patterns (green highlight)
        ws_top = wb.create_sheet("Top 3 Patterns", 0)
        self._write_top_patterns_sheet(ws_top, top_patterns, market_regime, analysis_date)

        # Sheet 2: All Detected Patterns (reference)
        ws_all = wb.create_sheet("All Patterns", 1)
        all_patterns = all_daily_patterns + all_hourly_patterns
        self._write_all_patterns_sheet(ws_all, all_patterns, market_regime, analysis_date)

        # Save report
        report_path = self.get_report_path(analysis_date)
        wb.save(report_path)
        from google_drive_sync import sync_to_drive
        sync_to_drive(report_path, "PreMarket")

        logger.info(f"Report generated: {report_path}")
        logger.info(f"  Top patterns: {len(top_patterns)}")
        logger.info(f"  All patterns: {len(all_patterns)} ({len(all_daily_patterns)} daily + {len(all_hourly_patterns)} hourly)")

        return report_path

    def _write_top_patterns_sheet(
        self,
        ws,
        top_patterns: List[Dict],
        market_regime: str,
        analysis_date: datetime
    ):
        """Write Top 3 Patterns sheet with green highlighting"""

        # Header info
        ws['A1'] = "PRE-MARKET PATTERN ALERTS - TOP 3 SETUPS"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Analysis Date: {analysis_date.strftime('%Y-%m-%d %A')}"
        ws['A3'] = f"Market Regime: {market_regime}"
        ws['A4'] = f"Market Opens: 09:15 AM"

        # Column headers (row 6)
        headers = [
            "Rank", "Stock", "Pattern", "Timeframe", "Confidence", "Priority",
            "Entry", "Target", "Target%", "Stop Loss", "Stop%",
            "R:R", "Volume", "Freshness", "Notes"
        ]

        row = 6
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data (green background for top patterns)
        row = 7
        for rank, pattern in enumerate(top_patterns, 1):
            details = pattern['details']

            # Calculate target and stop percentages
            entry = details.get('buy_price', 0)
            target = details.get('target_price', 0)
            stop = details.get('stop_loss', 0)

            target_pct = ((target - entry) / entry * 100) if entry > 0 else 0
            stop_pct = ((entry - stop) / entry * 100) if entry > 0 else 0

            # Risk-reward ratio
            rr_ratio = pu.calculate_risk_reward_ratio(entry, target, stop)

            # Freshness
            candles_ago = pattern.get('candles_ago', 0)
            timeframe = pattern['timeframe']
            freshness_text = self._format_freshness(candles_ago, timeframe)

            # Notes (pattern-specific)
            notes = self._format_notes(pattern)

            # Write row
            row_data = [
                rank,
                pattern['symbol'],
                pu.format_pattern_name(pattern['pattern_name']),
                timeframe.upper(),
                f"{details.get('confidence_score', 0):.1f}/10",
                f"{pattern.get('priority_score', 0):.2f}/10",
                f"₹{entry:.2f}",
                f"₹{target:.2f}",
                f"+{target_pct:.1f}%",
                f"₹{stop:.2f}",
                f"-{stop_pct:.1f}%",
                f"1:{rr_ratio:.1f}",
                f"{details.get('volume_ratio', 0):.1f}x",
                freshness_text,
                notes
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                # Green background for top patterns
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                cell.alignment = Alignment(horizontal="left" if col in [2, 3, 15] else "center", vertical="center")

            row += 1

        # Format columns
        ws.column_dimensions['A'].width = 6   # Rank
        ws.column_dimensions['B'].width = 12  # Stock
        ws.column_dimensions['C'].width = 20  # Pattern
        ws.column_dimensions['D'].width = 10  # Timeframe
        ws.column_dimensions['E'].width = 11  # Confidence
        ws.column_dimensions['F'].width = 11  # Priority
        ws.column_dimensions['G'].width = 11  # Entry
        ws.column_dimensions['H'].width = 11  # Target
        ws.column_dimensions['I'].width = 10  # Target%
        ws.column_dimensions['J'].width = 11  # Stop
        ws.column_dimensions['K'].width = 10  # Stop%
        ws.column_dimensions['L'].width = 8   # R:R
        ws.column_dimensions['M'].width = 10  # Volume
        ws.column_dimensions['N'].width = 15  # Freshness
        ws.column_dimensions['O'].width = 40  # Notes

        # Freeze header rows
        ws.freeze_panes = 'A7'

    def _write_all_patterns_sheet(
        self,
        ws,
        all_patterns: List[Dict],
        market_regime: str,
        analysis_date: datetime
    ):
        """Write All Patterns sheet (reference)"""

        # Header info
        ws['A1'] = "ALL DETECTED PATTERNS - REFERENCE"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Analysis Date: {analysis_date.strftime('%Y-%m-%d %A')}"
        ws['A3'] = f"Market Regime: {market_regime}"
        ws['A4'] = f"Total Patterns: {len(all_patterns)}"

        # Column headers (row 6)
        headers = [
            "Stock", "Pattern", "Timeframe", "Confidence", "Entry",
            "Target", "Target%", "Stop Loss", "Stop%", "R:R", "Volume"
        ]

        row = 6
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Sort patterns by confidence score
        sorted_patterns = sorted(all_patterns, key=lambda x: x['details'].get('confidence_score', 0), reverse=True)

        # Write data
        row = 7
        for pattern in sorted_patterns:
            details = pattern['details']

            # Calculate percentages
            entry = details.get('buy_price', 0)
            target = details.get('target_price', 0)
            stop = details.get('stop_loss', 0)

            target_pct = ((target - entry) / entry * 100) if entry > 0 else 0
            stop_pct = ((entry - stop) / entry * 100) if entry > 0 else 0
            rr_ratio = pu.calculate_risk_reward_ratio(entry, target, stop)

            row_data = [
                pattern['symbol'],
                pu.format_pattern_name(pattern['pattern_name']),
                pattern['timeframe'].upper(),
                f"{details.get('confidence_score', 0):.1f}/10",
                f"₹{entry:.2f}",
                f"₹{target:.2f}",
                f"+{target_pct:.1f}%",
                f"₹{stop:.2f}",
                f"-{stop_pct:.1f}%",
                f"1:{rr_ratio:.1f}",
                f"{details.get('volume_ratio', 0):.1f}x"
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.alignment = Alignment(horizontal="left" if col in [1, 2] else "center", vertical="center")

            row += 1

        # Format columns
        ws.column_dimensions['A'].width = 12  # Stock
        ws.column_dimensions['B'].width = 20  # Pattern
        ws.column_dimensions['C'].width = 10  # Timeframe
        ws.column_dimensions['D'].width = 11  # Confidence
        ws.column_dimensions['E'].width = 11  # Entry
        ws.column_dimensions['F'].width = 11  # Target
        ws.column_dimensions['G'].width = 10  # Target%
        ws.column_dimensions['H'].width = 11  # Stop
        ws.column_dimensions['I'].width = 10  # Stop%
        ws.column_dimensions['J'].width = 8   # R:R
        ws.column_dimensions['K'].width = 10  # Volume

        # Freeze header rows
        ws.freeze_panes = 'A7'

    def _format_freshness(self, candles_ago: int, timeframe: str) -> str:
        """Format freshness text based on timeframe"""
        if candles_ago == 0:
            return "Just now"
        elif timeframe == 'daily':
            return f"{candles_ago} day(s) ago"
        else:  # hourly
            return f"{candles_ago} hour(s) ago"

    def _format_notes(self, pattern: Dict) -> str:
        """Format pattern-specific notes"""
        details = pattern['details']
        pattern_name = pattern['pattern_name'].upper()
        notes = []

        # Pattern strength
        strength = details.get('pattern_strength', '')
        if strength:
            notes.append(f"Strength: {strength}")

        # Pattern-specific details
        if 'DOUBLE_BOTTOM' in pattern_name:
            first_low = details.get('first_low', 0)
            second_low = details.get('second_low', 0)
            peak = details.get('peak_between', 0)
            notes.append(f"Lows: ₹{first_low:.0f}, ₹{second_low:.0f} | Peak: ₹{peak:.0f}")

        elif 'RESISTANCE_BREAKOUT' in pattern_name:
            resistance = details.get('resistance_level', 0)
            support = details.get('support_level', 0)
            notes.append(f"Resistance: ₹{resistance:.0f} | Support: ₹{support:.0f}")

        return " | ".join(notes) if notes else "Pattern detected"


def main():
    """Test/demonstration of PreMarketReportGenerator"""
    print("=" * 60)
    print("PRE-MARKET REPORT GENERATOR - TEST")
    print("=" * 60)

    # Mock data
    top_patterns = [
        {
            'symbol': 'RELIANCE',
            'pattern_name': 'DOUBLE_BOTTOM',
            'timeframe': 'daily',
            'priority_score': 8.5,
            'candles_ago': 0,
            'details': {
                'confidence_score': 8.7,
                'buy_price': 2450.0,
                'target_price': 2550.0,
                'stop_loss': 2420.0,
                'volume_ratio': 2.3,
                'pattern_strength': 'Strong',
                'first_low': 2400,
                'second_low': 2405,
                'peak_between': 2480
            }
        }
    ]

    all_daily_patterns = top_patterns
    all_hourly_patterns = []

    # Generate report
    generator = PreMarketReportGenerator()
    report_path = generator.generate_report(
        top_patterns=top_patterns,
        all_daily_patterns=all_daily_patterns,
        all_hourly_patterns=all_hourly_patterns,
        market_regime="BULLISH"
    )

    print(f"\nReport generated: {report_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()
