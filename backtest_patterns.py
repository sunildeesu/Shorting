#!/usr/bin/env python3
"""
Backtest Pattern Detection Logic
Tests buy/target price recommendations against historical data
"""

import sys
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Tuple
from kiteconnect import KiteConnect
import config
from eod_pattern_detector import EODPatternDetector
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import time

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/backtest.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class PatternBacktester:
    """Backtests pattern detection and price targets"""

    def __init__(self, lookback_days: int = 60, forward_days: int = 30):
        """
        Initialize backtester

        Args:
            lookback_days: How many days back to test (default: 60)
            forward_days: How many days forward to check if target hit (default: 30)
        """
        self.kite = KiteConnect(api_key=config.KITE_API_KEY)
        self.kite.set_access_token(config.KITE_ACCESS_TOKEN)
        self.pattern_detector = EODPatternDetector(pattern_tolerance=2.0)
        self.lookback_days = lookback_days
        self.forward_days = forward_days

        # Build instrument token map
        self.instrument_tokens = self._build_instrument_token_map()

        logger.info(f"Backtester initialized: {lookback_days} days back, {forward_days} days forward")

    def _build_instrument_token_map(self) -> Dict[str, int]:
        """Build mapping of stock symbol to instrument token"""
        logger.info("Building instrument token map...")
        try:
            instruments = self.kite.instruments("NSE")
            token_map = {}

            # Top liquid F&O stocks for testing
            test_stocks = [
                'RELIANCE', 'TCS', 'INFY', 'HDFCBANK', 'ICICIBANK',
                'SBIN', 'BHARTIARTL', 'ITC', 'KOTAKBANK', 'LT',
                'AXISBANK', 'HINDUNILVR', 'BAJFINANCE', 'MARUTI', 'ASIANPAINT',
                'WIPRO', 'TITAN', 'NTPC', 'ONGC', 'POWERGRID'
            ]

            for instrument in instruments:
                if instrument['segment'] == 'NSE' and instrument['tradingsymbol'] in test_stocks:
                    token_map[instrument['tradingsymbol']] = instrument['instrument_token']

            logger.info(f"Built token map for {len(token_map)} stocks")
            return token_map

        except Exception as e:
            logger.error(f"Error building instrument token map: {e}")
            return {}

    def fetch_historical_data(self, symbol: str, from_date: datetime, to_date: datetime) -> List[Dict]:
        """
        Fetch historical daily data

        Args:
            symbol: Stock symbol
            from_date: Start date
            to_date: End date

        Returns:
            List of daily OHLCV candles
        """
        try:
            instrument_token = self.instrument_tokens.get(symbol)
            if not instrument_token:
                logger.warning(f"{symbol}: No instrument token found")
                return []

            data = self.kite.historical_data(
                instrument_token=instrument_token,
                from_date=from_date,
                to_date=to_date,
                interval="day"
            )

            return data

        except Exception as e:
            logger.error(f"{symbol}: Error fetching historical data - {e}")
            return []

    def test_pattern(self, symbol: str, test_date: datetime) -> List[Dict]:
        """
        Test pattern detection on a specific date

        Args:
            symbol: Stock symbol
            test_date: Date to test pattern detection

        Returns:
            List of trade results
        """
        # Fetch 30 days before test_date for pattern detection
        pattern_start = test_date - timedelta(days=30)
        pattern_data = self.fetch_historical_data(symbol, pattern_start, test_date)

        if len(pattern_data) < 10:
            return []

        # Detect patterns
        result = self.pattern_detector.detect_patterns(symbol, pattern_data)

        if not result['has_patterns']:
            return []

        # For each pattern, test forward performance
        trades = []
        pattern_details = result.get('pattern_details', {})

        for pattern_name, details in pattern_details.items():
            if not details or 'buy_price' not in details:
                continue

            # Fetch forward data to see if target was hit
            forward_start = test_date + timedelta(days=1)
            forward_end = test_date + timedelta(days=self.forward_days)
            forward_data = self.fetch_historical_data(symbol, forward_start, forward_end)

            # Analyze forward performance
            performance = self._analyze_forward_performance(
                details, forward_data, test_date
            )

            trades.append({
                'symbol': symbol,
                'pattern': pattern_name,
                'test_date': test_date,
                'buy_price': details.get('buy_price'),
                'target_price': details.get('target_price'),
                'pattern_type': details.get('pattern_type'),
                'current_price': details.get('current_price'),
                **performance
            })

        return trades

    def _analyze_forward_performance(
        self,
        pattern_details: Dict,
        forward_data: List[Dict],
        entry_date: datetime
    ) -> Dict:
        """
        Analyze if target was hit and calculate metrics

        Args:
            pattern_details: Pattern details with buy/target prices
            forward_data: Forward looking price data
            entry_date: Entry date

        Returns:
            Dict with performance metrics
        """
        buy_price = pattern_details.get('buy_price')
        target_price = pattern_details.get('target_price')
        pattern_type = pattern_details.get('pattern_type')

        if not forward_data:
            return {
                'target_hit': False,
                'days_to_target': None,
                'max_gain_pct': 0,
                'max_loss_pct': 0,
                'final_price': None,
                'final_pnl_pct': 0
            }

        target_hit = False
        days_to_target = None
        max_gain_pct = 0
        max_loss_pct = 0

        for day_num, candle in enumerate(forward_data, 1):
            high = candle['high']
            low = candle['low']
            close = candle['close']

            # Check if target was hit
            if pattern_type == 'BULLISH':
                # For bullish, check if high reached target
                if high >= target_price and not target_hit:
                    target_hit = True
                    days_to_target = day_num

                # Track max gain (price went up from buy)
                gain_pct = ((high - buy_price) / buy_price) * 100
                max_gain_pct = max(max_gain_pct, gain_pct)

                # Track max loss (price went down from buy)
                loss_pct = ((low - buy_price) / buy_price) * 100
                max_loss_pct = min(max_loss_pct, loss_pct)

            else:  # BEARISH
                # For bearish, check if low reached target
                if low <= target_price and not target_hit:
                    target_hit = True
                    days_to_target = day_num

                # Track max gain (price went down from short entry)
                gain_pct = ((buy_price - low) / buy_price) * 100
                max_gain_pct = max(max_gain_pct, gain_pct)

                # Track max loss (price went up from short entry)
                loss_pct = ((buy_price - high) / buy_price) * 100
                max_loss_pct = min(max_loss_pct, loss_pct)

        # Final P&L at end of forward period
        final_price = forward_data[-1]['close']
        if pattern_type == 'BULLISH':
            final_pnl_pct = ((final_price - buy_price) / buy_price) * 100
        else:
            final_pnl_pct = ((buy_price - final_price) / buy_price) * 100

        return {
            'target_hit': target_hit,
            'days_to_target': days_to_target,
            'max_gain_pct': max_gain_pct,
            'max_loss_pct': max_loss_pct,
            'final_price': final_price,
            'final_pnl_pct': final_pnl_pct
        }

    def run_backtest(self) -> List[Dict]:
        """
        Run complete backtest across all stocks and dates

        Returns:
            List of all trades with performance metrics
        """
        logger.info("="*80)
        logger.info("Starting Pattern Backtest")
        logger.info("="*80)

        all_trades = []
        test_stocks = list(self.instrument_tokens.keys())

        # Calculate test date range
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=self.lookback_days)

        logger.info(f"Testing {len(test_stocks)} stocks from {start_date} to {end_date}")
        logger.info(f"Forward test window: {self.forward_days} days")

        # Test each stock
        for stock_num, symbol in enumerate(test_stocks, 1):
            logger.info(f"[{stock_num}/{len(test_stocks)}] Testing {symbol}...")

            # Test pattern detection weekly (every 7 days) to reduce API calls
            test_dates = []
            current_date = start_date
            while current_date <= end_date:
                test_dates.append(datetime.combine(current_date, datetime.min.time()))
                current_date += timedelta(days=7)  # Weekly tests

            for test_date in test_dates:
                trades = self.test_pattern(symbol, test_date)
                all_trades.extend(trades)

                # Rate limiting
                time.sleep(config.REQUEST_DELAY_SECONDS)

            logger.info(f"{symbol}: {len([t for t in all_trades if t['symbol'] == symbol])} patterns detected")

        logger.info(f"Backtest complete: {len(all_trades)} total trades")
        return all_trades

    def generate_excel_report(self, trades: List[Dict], output_file: str):
        """
        Generate Excel report with backtest results

        Args:
            trades: List of trade results
            output_file: Output Excel file path
        """
        wb = openpyxl.Workbook()

        # Sheet 1: Trade-by-Trade Results
        ws_trades = wb.active
        ws_trades.title = "Trade Results"
        self._write_trades_sheet(ws_trades, trades)

        # Sheet 2: Summary Statistics
        ws_summary = wb.create_sheet("Summary Statistics")
        self._write_summary_sheet(ws_summary, trades)

        # Sheet 3: Pattern Analysis
        ws_patterns = wb.create_sheet("Pattern Analysis")
        self._write_pattern_analysis(ws_patterns, trades)

        # Save workbook
        wb.save(output_file)
        logger.info(f"Backtest report saved: {output_file}")

    def _write_trades_sheet(self, ws, trades: List[Dict]):
        """Write trade-by-trade results"""
        # Header
        ws.merge_cells('A1:O1')
        title = ws['A1']
        title.value = f"Pattern Backtest Results - {len(trades)} Trades"
        title.font = Font(size=14, bold=True, color="FFFFFF")
        title.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        title.alignment = Alignment(horizontal="center", vertical="center")

        # Column headers
        headers = [
            'Date', 'Stock', 'Pattern', 'Type', 'Buy Price', 'Target Price',
            'Target Hit?', 'Days to Target', 'Max Gain %', 'Max Loss %',
            'Final Price', 'Final P&L %', 'Expected Gain %', 'Actual vs Expected', 'Result'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Data rows
        for row, trade in enumerate(trades, 4):
            buy_price = trade.get('buy_price', 0)
            target_price = trade.get('target_price', 0)
            expected_gain = ((target_price - buy_price) / buy_price * 100) if buy_price > 0 else 0
            if trade.get('pattern_type') == 'BEARISH':
                expected_gain = ((buy_price - target_price) / buy_price * 100) if buy_price > 0 else 0

            final_pnl = trade.get('final_pnl_pct', 0)
            actual_vs_expected = (final_pnl / expected_gain * 100) if expected_gain != 0 else 0

            result = 'WIN' if trade.get('target_hit') else 'LOSS'

            ws.cell(row=row, column=1, value=trade.get('test_date').strftime('%Y-%m-%d') if trade.get('test_date') else '')
            ws.cell(row=row, column=2, value=trade.get('symbol'))
            ws.cell(row=row, column=3, value=trade.get('pattern'))
            ws.cell(row=row, column=4, value=trade.get('pattern_type'))
            ws.cell(row=row, column=5, value=f"₹{buy_price:.2f}" if buy_price else '')
            ws.cell(row=row, column=6, value=f"₹{target_price:.2f}" if target_price else '')
            ws.cell(row=row, column=7, value='YES' if trade.get('target_hit') else 'NO')
            ws.cell(row=row, column=8, value=trade.get('days_to_target') or '-')
            ws.cell(row=row, column=9, value=f"{trade.get('max_gain_pct', 0):.2f}%")
            ws.cell(row=row, column=10, value=f"{trade.get('max_loss_pct', 0):.2f}%")
            ws.cell(row=row, column=11, value=f"₹{trade.get('final_price', 0):.2f}")
            ws.cell(row=row, column=12, value=f"{final_pnl:.2f}%")
            ws.cell(row=row, column=13, value=f"{expected_gain:.2f}%")
            ws.cell(row=row, column=14, value=f"{actual_vs_expected:.0f}%")
            ws.cell(row=row, column=15, value=result)

            # Color code result
            result_cell = ws.cell(row=row, column=15)
            if result == 'WIN':
                result_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                result_cell.font = Font(color="006100", bold=True)
            else:
                result_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                result_cell.font = Font(color="9C0006", bold=True)

        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 14
        ws.column_dimensions['N'].width = 16
        ws.column_dimensions['O'].width = 10

    def _write_summary_sheet(self, ws, trades: List[Dict]):
        """Write summary statistics"""
        ws['A1'] = "BACKTEST SUMMARY STATISTICS"
        ws['A1'].font = Font(size=14, bold=True)

        total_trades = len(trades)
        winning_trades = sum(1 for t in trades if t.get('target_hit'))
        losing_trades = total_trades - winning_trades
        win_rate = (winning_trades / total_trades * 100) if total_trades > 0 else 0

        avg_days_to_target = sum(t.get('days_to_target', 0) for t in trades if t.get('days_to_target')) / winning_trades if winning_trades > 0 else 0
        avg_gain = sum(t.get('final_pnl_pct', 0) for t in trades if t.get('target_hit')) / winning_trades if winning_trades > 0 else 0
        avg_loss = sum(t.get('final_pnl_pct', 0) for t in trades if not t.get('target_hit')) / losing_trades if losing_trades > 0 else 0

        summary = [
            ['', ''],
            ['Total Trades', total_trades],
            ['Winning Trades', winning_trades],
            ['Losing Trades', losing_trades],
            ['Win Rate', f"{win_rate:.1f}%"],
            ['', ''],
            ['Average Days to Target', f"{avg_days_to_target:.1f} days"],
            ['Average Gain (Winners)', f"{avg_gain:.2f}%"],
            ['Average Loss (Losers)', f"{avg_loss:.2f}%"],
            ['', ''],
            ['Best Trade', f"{max(t.get('final_pnl_pct', 0) for t in trades):.2f}%"],
            ['Worst Trade', f"{min(t.get('final_pnl_pct', 0) for t in trades):.2f}%"],
        ]

        for row, (label, value) in enumerate(summary, 3):
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _write_pattern_analysis(self, ws, trades: List[Dict]):
        """Write pattern-by-pattern analysis"""
        ws['A1'] = "PATTERN-WISE PERFORMANCE"
        ws['A1'].font = Font(size=14, bold=True)

        # Group by pattern
        patterns = {}
        for trade in trades:
            pattern = trade.get('pattern')
            if pattern not in patterns:
                patterns[pattern] = []
            patterns[pattern].append(trade)

        headers = ['Pattern', 'Total', 'Wins', 'Losses', 'Win Rate', 'Avg Gain', 'Avg Days']
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header).font = Font(bold=True)

        row = 4
        for pattern, pattern_trades in patterns.items():
            total = len(pattern_trades)
            wins = sum(1 for t in pattern_trades if t.get('target_hit'))
            losses = total - wins
            win_rate = (wins / total * 100) if total > 0 else 0
            avg_gain = sum(t.get('final_pnl_pct', 0) for t in pattern_trades) / total if total > 0 else 0
            avg_days = sum(t.get('days_to_target', 0) for t in pattern_trades if t.get('days_to_target')) / wins if wins > 0 else 0

            ws.cell(row=row, column=1, value=pattern)
            ws.cell(row=row, column=2, value=total)
            ws.cell(row=row, column=3, value=wins)
            ws.cell(row=row, column=4, value=losses)
            ws.cell(row=row, column=5, value=f"{win_rate:.1f}%")
            ws.cell(row=row, column=6, value=f"{avg_gain:.2f}%")
            ws.cell(row=row, column=7, value=f"{avg_days:.1f} days")

            row += 1

        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20


def main():
    """Main entry point"""
    logger.info("Starting Pattern Backtest...")

    # Create backtester (60 days back, 30 days forward)
    backtester = PatternBacktester(lookback_days=60, forward_days=30)

    # Run backtest
    trades = backtester.run_backtest()

    if not trades:
        logger.warning("No trades found in backtest period")
        return

    # Generate report
    report_file = "data/eod_reports/pattern_backtest_results.xlsx"
    backtester.generate_excel_report(trades, report_file)

    print(f"\n✅ Backtest complete!")
    print(f"📊 Total trades: {len(trades)}")
    print(f"📁 Report: {report_file}\n")


if __name__ == "__main__":
    main()
