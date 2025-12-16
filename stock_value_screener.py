#!/Users/sunildeesu/myProjects/ShortIndicator/venv/bin/python3
"""
Stock Value Screener - Find Undervalued Stocks with Strong Fundamentals

Screens NSE stocks (market cap > ₹10,000 Cr) for:
1. Price is 20-30% below 3-year peak
2. Both Revenue and PAT are growing (YoY preferred, QoQ fallback)

Generates detailed Excel report with 15 columns.

Author: Sunil Kumar Durganaik
"""

import json
import os
import time
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple
import pandas as pd
from kiteconnect import KiteConnect
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import requests
from bs4 import BeautifulSoup

import config
from unified_quote_cache import UnifiedQuoteCache
from unified_data_cache import UnifiedDataCache
from sector_manager import get_sector_manager
from trend_analyzer import TrendAnalyzer

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/value_screener.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class StockValueScreener:
    """Screen stocks for value opportunities with improving fundamentals"""

    def __init__(self):
        """Initialize the stock value screener"""
        # Initialize Kite Connect
        if not config.KITE_API_KEY or not config.KITE_ACCESS_TOKEN:
            raise ValueError("Kite Connect requires KITE_API_KEY and KITE_ACCESS_TOKEN")

        self.kite = KiteConnect(api_key=config.KITE_API_KEY)
        self.kite.set_access_token(config.KITE_ACCESS_TOKEN)
        logger.info("Kite Connect initialized")

        # Load stock lists
        self.stocks = self._load_stock_list()
        self.instrument_tokens = self._load_instrument_tokens()
        self.shares_outstanding = self._load_shares_outstanding()

        # Initialize caching
        self.quote_cache = UnifiedQuoteCache(ttl_seconds=60)
        self.data_cache = UnifiedDataCache(cache_dir="data/unified_cache")
        logger.info("Unified caching enabled")

        # Initialize sector manager
        try:
            self.sector_manager = get_sector_manager()
            logger.info(f"Sector manager loaded ({len(self.sector_manager.get_all_sectors())} sectors)")
        except Exception as e:
            logger.warning(f"Sector manager not available: {e}")
            self.sector_manager = None

        # Quotes storage
        self.quotes = {}

        logger.info(f"Stock value screener initialized for {len(self.stocks)} stocks")

    def _load_stock_list(self) -> List[str]:
        """Load stock list from all_nse_stocks.json"""
        try:
            with open('data/all_nse_stocks.json', 'r') as f:
                data = json.load(f)
                stocks = data.get('stocks', [])
                logger.info(f"Loaded {len(stocks)} stocks from all_nse_stocks.json")
                return stocks
        except FileNotFoundError:
            logger.error("data/all_nse_stocks.json not found. Run fetch_all_nse_stocks.py first!")
            raise
        except Exception as e:
            logger.error(f"Error loading stock list: {e}")
            raise

    def _load_instrument_tokens(self) -> Dict[str, int]:
        """Load instrument tokens from all_instrument_tokens.json"""
        try:
            with open('data/all_instrument_tokens.json', 'r') as f:
                tokens = json.load(f)
                logger.info(f"Loaded {len(tokens)} instrument tokens")
                return tokens
        except FileNotFoundError:
            logger.error("data/all_instrument_tokens.json not found. Run fetch_all_nse_stocks.py first!")
            raise
        except Exception as e:
            logger.error(f"Error loading instrument tokens: {e}")
            raise

    def _load_shares_outstanding(self) -> Dict[str, float]:
        """Load shares outstanding from all_shares_outstanding.json"""
        try:
            with open('data/all_shares_outstanding.json', 'r') as f:
                shares = json.load(f)
                logger.info(f"Loaded {len(shares)} shares outstanding values")
                return shares
        except FileNotFoundError:
            logger.warning("data/all_shares_outstanding.json not found. Market cap calculation may fail.")
            return {}
        except Exception as e:
            logger.error(f"Error loading shares outstanding: {e}")
            return {}

    def fetch_3year_historical_data(self, symbol: str) -> Optional[pd.DataFrame]:
        """
        Fetch 3 years of historical data with caching

        Args:
            symbol: Stock symbol (e.g., "RELIANCE")

        Returns:
            DataFrame with OHLCV data or None if failed
        """
        # Check cache first
        cached = self.data_cache.get_data(symbol, 'historical_3year')
        if cached:
            logger.debug(f"{symbol}: 3-year data from cache")
            return pd.DataFrame(cached)

        # Fetch from Kite API
        try:
            token = self.instrument_tokens.get(symbol)
            if not token:
                logger.warning(f"{symbol}: No instrument token found")
                return None

            to_date = datetime.now().date()
            from_date = to_date - timedelta(days=1095)  # 3 years

            data = self.kite.historical_data(
                instrument_token=token,
                from_date=from_date,
                to_date=to_date,
                interval="day"
            )

            if not data:
                logger.warning(f"{symbol}: No historical data returned")
                return None

            # Cache it
            self.data_cache.set_data(symbol, data, 'historical_3year')

            df = pd.DataFrame(data)
            logger.debug(f"{symbol}: Fetched {len(df)} candles (3 years)")
            return df

        except Exception as e:
            logger.error(f"{symbol}: Failed to fetch historical data: {e}")
            return None

    def calculate_market_cap(self, symbol: str, price: float) -> Optional[float]:
        """Calculate market cap in crores"""
        shares = self.shares_outstanding.get(symbol)
        if not shares:
            return None

        market_cap_cr = (price * shares) / 10_000_000
        return market_cap_cr

    def check_price_criteria(self, symbol: str) -> Tuple[bool, Optional[Dict]]:
        """
        Check if stock price is 30% or more below 3-year peak

        Args:
            symbol: Stock symbol

        Returns:
            Tuple of (passes_criteria, price_data_dict) or (False, None)
        """
        # Fetch historical data
        df = self.fetch_3year_historical_data(symbol)
        if df is None or len(df) < 365:
            return False, None

        # Get current price
        current_price = self.quotes.get(f"NSE:{symbol}", {}).get("last_price")
        if not current_price:
            return False, None

        # Calculate 3-year peak
        three_year_high = df['high'].max()

        # Calculate drawdown percentage
        drawdown_pct = ((three_year_high - current_price) / three_year_high) * 100

        # Check if 30% or more below peak
        if drawdown_pct >= 30:
            return True, {
                'current_price': current_price,
                '3year_peak': three_year_high,
                'drawdown_pct': drawdown_pct,
                '52w_high': df['high'].tail(252).max() if len(df) >= 252 else df['high'].max(),
                '52w_low': df['low'].tail(252).min() if len(df) >= 252 else df['low'].min()
            }

        return False, None

    def check_short_price_criteria(self, symbol: str) -> Tuple[bool, Optional[Dict]]:
        """
        Check if stock price is within 5% of 52-week high OR 3-year peak (potential short candidates)

        Args:
            symbol: Stock symbol

        Returns:
            Tuple of (passes_criteria, price_data_dict) or (False, None)
        """
        # Fetch historical data
        df = self.fetch_3year_historical_data(symbol)
        if df is None or len(df) < 365:
            return False, None

        # Get current price
        current_price = self.quotes.get(f"NSE:{symbol}", {}).get("last_price")
        if not current_price:
            return False, None

        # Calculate 52-week high and 3-year peak
        fifty_two_week_high = df['high'].tail(252).max() if len(df) >= 252 else df['high'].max()
        three_year_peak = df['high'].max()

        # Calculate distance from highs
        pct_from_52w_high = ((fifty_two_week_high - current_price) / fifty_two_week_high) * 100
        pct_from_3y_peak = ((three_year_peak - current_price) / three_year_peak) * 100

        # Check if within 5% of either high (overvalued zone - potential resistance)
        if pct_from_52w_high <= 5 or pct_from_3y_peak <= 5:
            return True, {
                'current_price': current_price,
                '52w_high': fifty_two_week_high,
                '3year_peak': three_year_peak,
                'pct_from_52w_high': pct_from_52w_high,
                'pct_from_3y_peak': pct_from_3y_peak,
                '52w_low': df['low'].tail(252).min() if len(df) >= 252 else df['low'].min(),
                'near_52w_high': pct_from_52w_high <= 5,
                'near_3y_peak': pct_from_3y_peak <= 5
            }

        return False, None

    def scrape_fundamentals_from_screener(self, symbol: str) -> Optional[List[Dict]]:
        """
        Scrape Revenue and PAT from screener.in

        Args:
            symbol: Stock symbol

        Returns:
            List of quarterly data dicts with 'quarter', 'revenue', 'pat' or None if failed
        """
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'
            }

            # Try original symbol first, then variations if it fails
            symbols_to_try = [symbol]

            # If symbol has suffix (e.g., "SANWARIA-BZ"), try without suffix
            if '-' in symbol:
                base_symbol = symbol.split('-')[0]
                symbols_to_try.append(base_symbol)

            response = None
            for try_symbol in symbols_to_try:
                url = f"https://www.screener.in/company/{try_symbol}/"
                response = requests.get(url, headers=headers, timeout=15)

                if response.status_code == 200:
                    break

            if not response or response.status_code != 200:
                logger.warning(f"{symbol}: Screener.in returned status {response.status_code}")
                return None

            soup = BeautifulSoup(response.text, 'html.parser')

            # Find quarterly results table
            quarterly_table = None
            for section in soup.find_all('section', id='quarters'):
                table = section.find('table', class_='data-table')
                if table:
                    quarterly_table = table
                    break

            if not quarterly_table:
                logger.warning(f"{symbol}: Quarterly results table not found")
                return None

            # Extract data from table
            # Table structure: rows are metrics (Sales, Net Profit), columns are quarters
            rows = quarterly_table.find_all('tr')
            if len(rows) < 2:
                logger.warning(f"{symbol}: Quarterly table has insufficient rows")
                return None

            # Extract quarter names from header row (row 0)
            header_row = rows[0]
            header_cells = header_row.find_all(['th', 'td'])
            quarters = [cell.text.strip() for cell in header_cells[1:]]  # Skip first empty cell

            if len(quarters) < 3:
                logger.warning(f"{symbol}: Insufficient quarters in header")
                return None

            # Find the "Sales+" row (Revenue) and "Net Profit+" row (PAT)
            revenue_row = None
            pat_row = None

            for row in rows[1:]:
                cells = row.find_all('td')
                if not cells:
                    continue

                row_label = cells[0].text.strip()
                if 'Sales' in row_label:
                    revenue_row = cells
                elif 'Net Profit' in row_label and 'NP' not in row_label:  # Avoid "NP %"
                    pat_row = cells

            if not revenue_row or not pat_row:
                logger.warning(f"{symbol}: Could not find Sales or Net Profit row")
                return None

            # Extract data quarter by quarter
            quarters_data = []
            for i in range(min(len(quarters), len(revenue_row) - 1, len(pat_row) - 1, 12)):
                try:
                    quarter_name = quarters[i]
                    revenue_text = revenue_row[i + 1].text.strip().replace(',', '')
                    pat_text = pat_row[i + 1].text.strip().replace(',', '')

                    # Convert to float (handle negative values)
                    revenue = float(revenue_text)
                    pat = float(pat_text)

                    quarters_data.append({
                        'quarter': quarter_name,
                        'revenue': revenue,
                        'pat': pat
                    })

                except (ValueError, IndexError) as e:
                    logger.debug(f"{symbol}: Error parsing quarter {i}: {e}")
                    continue

            if len(quarters_data) < 3:
                logger.warning(f"{symbol}: Insufficient quarterly data ({len(quarters_data)} quarters)")
                return None

            logger.debug(f"{symbol}: Extracted {len(quarters_data)} quarters of data")
            return quarters_data

        except Exception as e:
            logger.error(f"{symbol}: Error scraping fundamentals: {e}")
            return None

    def calculate_growth_metrics(self, quarters_data: List[Dict]) -> Dict:
        """
        Calculate YoY and QoQ growth metrics

        Args:
            quarters_data: List of quarterly data dicts

        Returns:
            Dict with growth metrics
        """
        # YoY Growth (compare Q1'25 vs Q1'24 vs Q1'23, etc.)
        yoy_revenue_growth = []
        yoy_pat_growth = []

        # Need at least 8 quarters for 2 years of YoY comparison
        if len(quarters_data) >= 8:
            for i in range(min(8, len(quarters_data) - 4)):
                current_q = quarters_data[i]
                year_ago_q = quarters_data[i + 4]

                if year_ago_q['revenue'] != 0:
                    rev_growth = ((current_q['revenue'] - year_ago_q['revenue']) / abs(year_ago_q['revenue'])) * 100
                    yoy_revenue_growth.append(rev_growth)

                if year_ago_q['pat'] != 0:
                    pat_growth = ((current_q['pat'] - year_ago_q['pat']) / abs(year_ago_q['pat'])) * 100
                    yoy_pat_growth.append(pat_growth)

        # QoQ Growth (Q1 vs Q2 vs Q3)
        qoq_revenue_growth = []
        qoq_pat_growth = []

        # Need at least 3 quarters for QoQ comparison
        if len(quarters_data) >= 3:
            for i in range(min(3, len(quarters_data) - 1)):
                current_q = quarters_data[i]
                prev_q = quarters_data[i + 1]

                if prev_q['revenue'] != 0:
                    rev_growth = ((current_q['revenue'] - prev_q['revenue']) / abs(prev_q['revenue'])) * 100
                    qoq_revenue_growth.append(rev_growth)

                if prev_q['pat'] != 0:
                    pat_growth = ((current_q['pat'] - prev_q['pat']) / abs(prev_q['pat'])) * 100
                    qoq_pat_growth.append(pat_growth)

        # For YoY: Check if majority of recent quarters show growth
        # (At least 2 out of 3 comparisons should be positive)
        yoy_revenue_growing = False
        yoy_pat_growing = False

        if len(yoy_revenue_growth) >= 3:
            recent_rev = yoy_revenue_growth[:3]  # Last 3 YoY comparisons
            recent_pat = yoy_pat_growth[:3] if len(yoy_pat_growth) >= 3 else []

            yoy_revenue_growing = sum(1 for g in recent_rev if g > 0) >= 2  # At least 2 out of 3
            yoy_pat_growing = sum(1 for g in recent_pat if g > 0) >= 2 if recent_pat else False

        # For QoQ: Check if average growth is positive (allows for seasonal variations)
        qoq_revenue_growing = False
        qoq_pat_growing = False

        if len(qoq_revenue_growth) >= 2:
            avg_qoq_rev = sum(qoq_revenue_growth) / len(qoq_revenue_growth)
            avg_qoq_pat = sum(qoq_pat_growth) / len(qoq_pat_growth) if qoq_pat_growth else 0

            qoq_revenue_growing = avg_qoq_rev > 0
            qoq_pat_growing = avg_qoq_pat > 0

        return {
            'yoy_revenue_avg': sum(yoy_revenue_growth) / len(yoy_revenue_growth) if yoy_revenue_growth else 0,
            'yoy_pat_avg': sum(yoy_pat_growth) / len(yoy_pat_growth) if yoy_pat_growth else 0,
            'qoq_revenue_avg': sum(qoq_revenue_growth) / len(qoq_revenue_growth) if qoq_revenue_growth else 0,
            'qoq_pat_avg': sum(qoq_pat_growth) / len(qoq_pat_growth) if qoq_pat_growth else 0,
            'yoy_revenue_growing': yoy_revenue_growing,
            'yoy_pat_growing': yoy_pat_growing,
            'qoq_revenue_growing': qoq_revenue_growing,
            'qoq_pat_growing': qoq_pat_growing,
            # Additional debug info
            'yoy_rev_detail': yoy_revenue_growth[:3] if len(yoy_revenue_growth) >= 3 else yoy_revenue_growth,
            'yoy_pat_detail': yoy_pat_growth[:3] if len(yoy_pat_growth) >= 3 else yoy_pat_growth,
            'qoq_rev_detail': qoq_revenue_growth,
            'qoq_pat_detail': qoq_pat_growth
        }

    def check_fundamental_criteria(self, growth_metrics: Dict) -> Tuple[bool, Optional[str]]:
        """
        Check if BOTH Revenue AND PAT are growing (YoY preferred, QoQ fallback)

        Args:
            growth_metrics: Dict with growth metrics

        Returns:
            Tuple of (passes_criteria, growth_type) where growth_type is "YoY", "QoQ", or None
        """
        # Priority 1: Check YoY growth (all years must be positive)
        if growth_metrics['yoy_revenue_growing'] and growth_metrics['yoy_pat_growing']:
            return True, "YoY"

        # Priority 2: Fallback to QoQ if YoY not growing (all quarters must be positive)
        if growth_metrics['qoq_revenue_growing'] and growth_metrics['qoq_pat_growing']:
            return True, "QoQ"

        return False, None

    def check_short_fundamental_criteria(self, growth_metrics: Dict) -> Tuple[bool, Optional[str]]:
        """
        Check if BOTH Revenue AND PAT are DECLINING (for short candidates)

        Args:
            growth_metrics: Dict with growth metrics

        Returns:
            Tuple of (passes_criteria, decline_type) where decline_type is "YoY Decline", "QoQ Decline", or None
        """
        # Calculate if majority of recent quarters show DECLINE
        yoy_revenue_declining = False
        yoy_pat_declining = False

        yoy_rev_detail = growth_metrics.get('yoy_rev_detail', [])
        yoy_pat_detail = growth_metrics.get('yoy_pat_detail', [])

        if len(yoy_rev_detail) >= 3:
            # At least 2 out of 3 quarters should show negative growth
            yoy_revenue_declining = sum(1 for g in yoy_rev_detail[:3] if g < 0) >= 2
            yoy_pat_declining = sum(1 for g in yoy_pat_detail[:3] if g < 0) >= 2 if len(yoy_pat_detail) >= 3 else False

        # Check QoQ decline
        qoq_revenue_declining = growth_metrics['qoq_revenue_avg'] < 0
        qoq_pat_declining = growth_metrics['qoq_pat_avg'] < 0

        # Priority 1: YoY decline (stronger signal)
        if yoy_revenue_declining and yoy_pat_declining:
            return True, "YoY Decline"

        # Priority 2: QoQ decline (weaker signal but still valid)
        if qoq_revenue_declining and qoq_pat_declining:
            return True, "QoQ Decline"

        return False, None

    def screen_all_stocks(self) -> List[Dict]:
        """
        Main screening workflow

        Returns:
            List of stocks that pass all criteria
        """
        results = []

        # Step 1: Batch fetch all quotes
        logger.info(f"Fetching quotes for {len(self.stocks)} stocks...")
        try:
            self.quotes = self.quote_cache.get_or_fetch_quotes(self.stocks, self.kite)
            logger.info(f"✓ Fetched {len(self.quotes)} quotes")
        except Exception as e:
            logger.error(f"Failed to fetch quotes: {e}")
            return results

        # Step 2: Screen each stock
        logger.info("Starting stock screening...")
        logger.info("")

        price_qualified = 0
        fundamental_qualified = 0

        for idx, symbol in enumerate(self.stocks, 1):
            # Progress update every 50 stocks
            if idx % 50 == 0:
                logger.info(f"Progress: {idx}/{len(self.stocks)} ({idx/len(self.stocks)*100:.1f}%) | Price qualified: {price_qualified} | Results: {fundamental_qualified}")

            logger.debug(f"[{idx}/{len(self.stocks)}] Screening {symbol}...")

            # Price criteria check
            passes_price, price_data = self.check_price_criteria(symbol)

            if not passes_price:
                logger.debug(f"  ✗ {symbol}: Price criteria not met")
                continue

            price_qualified += 1
            logger.info(f"  ✓ [{idx}/{len(self.stocks)}] {symbol}: Price qualified ({price_data['drawdown_pct']:.1f}% below peak)")

            # Calculate additional metrics
            market_cap_cr = self.calculate_market_cap(symbol, price_data['current_price'])
            sector = self.sector_manager.get_sector(symbol) if self.sector_manager else "UNKNOWN"
            volume_lakhs = self.quotes.get(f"NSE:{symbol}", {}).get("volume", 0) / 100000

            # === TREND ANALYSIS (NEW) ===
            # Analyze trend using 3-year historical data (already cached!)
            trend_analysis = None
            try:
                df_hist = self.fetch_3year_historical_data(symbol)
                if df_hist is not None and len(df_hist) >= 200:
                    analyzer = TrendAnalyzer(df_hist, symbol=symbol)
                    trend_analysis = analyzer.get_comprehensive_analysis(
                        account_size=1000000,  # 10L account for position sizing
                        risk_pct=2.0
                    )
                    logger.info(f"    → Trend: {trend_analysis['trend_status']} (Score: {trend_analysis['trend_score']}/10)")
                    logger.info(f"    → Entry Signal: {trend_analysis['entry_signal']} ({trend_analysis['entry_type']})")
                else:
                    logger.warning(f"    ✗ {symbol}: Insufficient data for trend analysis")
            except Exception as e:
                logger.error(f"    ✗ {symbol}: Error in trend analysis: {e}")

            # Try to fetch fundamentals (optional - for green highlighting)
            growth_metrics = None
            growth_type = "No Data"
            growth_status = "Data Not Available"

            try:
                quarters_data = self.scrape_fundamentals_from_screener(symbol)

                if quarters_data:
                    growth_metrics = self.calculate_growth_metrics(quarters_data)
                    passes_fundamentals, growth_type = self.check_fundamental_criteria(growth_metrics)

                    if passes_fundamentals:
                        fundamental_qualified += 1
                        growth_status = f"{growth_type} Growth"
                        logger.info(f"    ✓ {symbol}: Has {growth_type} growth!")
                    else:
                        growth_status = "No Growth"
                        logger.info(f"    ○ {symbol}: No consistent growth")
                else:
                    logger.info(f"    ○ {symbol}: Fundamental data unavailable")

            except Exception as e:
                logger.warning(f"    ○ {symbol}: Error fetching fundamentals: {e}")

            # Add to results (ALL price-qualified stocks)
            result_data = {
                'symbol': symbol,
                'current_price': price_data['current_price'],
                '3year_peak': price_data['3year_peak'],
                'drawdown_pct': price_data['drawdown_pct'],
                'market_cap_cr': market_cap_cr if market_cap_cr else 0,
                'sector': sector,
                'volume_lakhs': volume_lakhs,
                '52w_high': price_data['52w_high'],
                '52w_low': price_data['52w_low'],
                'revenue_growth_yoy': growth_metrics['yoy_revenue_avg'] if growth_metrics else 0,
                'pat_growth_yoy': growth_metrics['yoy_pat_avg'] if growth_metrics else 0,
                'revenue_growth_qoq': growth_metrics['qoq_revenue_avg'] if growth_metrics else 0,
                'pat_growth_qoq': growth_metrics['qoq_pat_avg'] if growth_metrics else 0,
                'growth_type': growth_type,
                'growth_status': growth_status,
                'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }

            # Add trend analysis data if available
            if trend_analysis:
                result_data.update({
                    # Trend Analysis
                    'trend_status': trend_analysis['trend_status'],
                    'trend_score': trend_analysis['trend_score'],
                    'trend_confidence': trend_analysis['trend_confidence'],
                    'pct_from_ema20': trend_analysis['pct_from_ema20'],
                    'pct_from_sma50': trend_analysis['pct_from_sma50'],
                    'pct_from_sma200': trend_analysis['pct_from_sma200'],
                    'adx': trend_analysis['adx'],
                    'adx_signal': trend_analysis['adx_signal'],

                    # Entry Signal
                    'entry_signal': trend_analysis['entry_signal'],
                    'entry_type': trend_analysis['entry_type'],
                    'entry_price': trend_analysis['entry_price'],
                    'entry_score': trend_analysis['entry_score'],

                    # Risk Management
                    'stop_loss': trend_analysis['stop_loss'],
                    'stop_loss_pct': trend_analysis['stop_loss_pct'],
                    'target_1': trend_analysis['target_1'],
                    'target_1_pct': trend_analysis['target_1_pct'],
                    'target_2': trend_analysis['target_2'],
                    'target_2_pct': trend_analysis['target_2_pct'],
                    'risk_reward': trend_analysis['risk_reward'],

                    # Support/Resistance
                    'nearest_support': trend_analysis['nearest_support'],
                    'nearest_resistance': trend_analysis['nearest_resistance'],
                    's1_pivot': trend_analysis['s1_pivot'],
                    'r1_pivot': trend_analysis['r1_pivot'],

                    # Position Sizing
                    'position_size': trend_analysis['position_size'],
                    'position_value': trend_analysis['position_value']
                })
            else:
                # Default values if trend analysis failed
                result_data.update({
                    'trend_status': 'N/A',
                    'trend_score': 0,
                    'trend_confidence': 'N/A',
                    'pct_from_ema20': 0,
                    'pct_from_sma50': 0,
                    'pct_from_sma200': 0,
                    'adx': 0,
                    'adx_signal': 'N/A',
                    'entry_signal': 'N/A',
                    'entry_type': 'N/A',
                    'entry_price': 0,
                    'entry_score': 0,
                    'stop_loss': 0,
                    'stop_loss_pct': 0,
                    'target_1': 0,
                    'target_1_pct': 0,
                    'target_2': 0,
                    'target_2_pct': 0,
                    'risk_reward': 'N/A',
                    'nearest_support': 0,
                    'nearest_resistance': 0,
                    's1_pivot': 0,
                    'r1_pivot': 0,
                    'position_size': 0,
                    'position_value': 0
                })

            results.append(result_data)

            logger.info(f"    → Added to results (Total: {len(results)}, With Growth: {fundamental_qualified})")

            # Rate limiting for screener.in
            time.sleep(3)

        logger.info("")
        logger.info(f"Screening complete:")
        logger.info(f"  Total stocks scanned: {len(self.stocks)}")
        logger.info(f"  Price filter passed: {price_qualified} ({price_qualified/len(self.stocks)*100:.1f}%)")
        logger.info(f"  Final results: {len(results)} ({len(results)/len(self.stocks)*100:.1f}%)")

        return results

    def screen_short_candidates(self) -> List[Dict]:
        """
        Screen for SHORT selling opportunities

        Criteria:
        1. Price within 5% of 52W high OR 3Y peak (resistance zone)
        2. Strong bearish trend (trend score ≤ -4)
        3. Optional: Declining fundamentals (YoY/QoQ)

        Returns:
            List of short candidate stocks
        """
        results = []

        # Step 1: Batch fetch all quotes (reuse if already fetched)
        if not self.quotes:
            logger.info(f"Fetching quotes for {len(self.stocks)} stocks...")
            try:
                self.quotes = self.quote_cache.get_or_fetch_quotes(self.stocks, self.kite)
                logger.info(f"✓ Fetched {len(self.quotes)} quotes")
            except Exception as e:
                logger.error(f"Failed to fetch quotes: {e}")
                return results

        # Step 2: Screen each stock for SHORT opportunities
        logger.info("")
        logger.info("="*80)
        logger.info("SCREENING FOR SHORT CANDIDATES")
        logger.info("="*80)
        logger.info("")

        price_qualified = 0
        trend_qualified = 0
        short_signals = 0

        for idx, symbol in enumerate(self.stocks, 1):
            # Progress update every 50 stocks
            if idx % 50 == 0:
                logger.info(f"Short Screening Progress: {idx}/{len(self.stocks)} ({idx/len(self.stocks)*100:.1f}%) | Near High: {price_qualified} | Bearish: {trend_qualified} | SHORT Signals: {short_signals}")

            logger.debug(f"[SHORT] [{idx}/{len(self.stocks)}] Screening {symbol}...")

            # Price criteria check (near highs)
            passes_price, price_data = self.check_short_price_criteria(symbol)

            if not passes_price:
                logger.debug(f"  ✗ {symbol}: Not near highs")
                continue

            price_qualified += 1
            near_what = []
            if price_data['near_52w_high']:
                near_what.append(f"52W high ({price_data['pct_from_52w_high']:.1f}% away)")
            if price_data['near_3y_peak']:
                near_what.append(f"3Y peak ({price_data['pct_from_3y_peak']:.1f}% away)")

            logger.info(f"  ✓ [{idx}/{len(self.stocks)}] {symbol}: Near {' and '.join(near_what)}")

            # Calculate additional metrics
            market_cap_cr = self.calculate_market_cap(symbol, price_data['current_price'])
            sector = self.sector_manager.get_sector(symbol) if self.sector_manager else "UNKNOWN"
            volume_lakhs = self.quotes.get(f"NSE:{symbol}", {}).get("volume", 0) / 100000

            # === SHORT TREND ANALYSIS ===
            short_analysis = None
            try:
                df_hist = self.fetch_3year_historical_data(symbol)
                if df_hist is not None and len(df_hist) >= 200:
                    analyzer = TrendAnalyzer(df_hist, symbol=symbol)
                    short_analysis = analyzer.get_short_comprehensive_analysis(
                        account_size=1000000,  # 10L account
                        risk_pct=2.0  # Conservative for shorts
                    )

                    # Filter by trend score (need bearish trend)
                    if short_analysis['trend_score'] > -4:
                        logger.debug(f"    ✗ {symbol}: Not bearish enough (Score: {short_analysis['trend_score']})")
                        continue

                    trend_qualified += 1
                    logger.info(f"    ✓ {symbol}: {short_analysis['trend_status']} (Score: {short_analysis['trend_score']}/10)")
                    logger.info(f"    → Entry Signal: {short_analysis['entry_signal']} ({short_analysis['entry_type']})")

                    if short_analysis['entry_signal'] == 'SHORT':
                        short_signals += 1
                else:
                    logger.warning(f"    ✗ {symbol}: Insufficient data for trend analysis")
                    continue
            except Exception as e:
                logger.error(f"    ✗ {symbol}: Error in short trend analysis: {e}")
                continue

            # Try to fetch fundamentals (optional - for better confidence)
            growth_metrics = None
            decline_type = "No Data"
            fundamental_status = "Data Not Available"

            try:
                quarters_data = self.scrape_fundamentals_from_screener(symbol)

                if quarters_data:
                    growth_metrics = self.calculate_growth_metrics(quarters_data)
                    has_decline, decline_type = self.check_short_fundamental_criteria(growth_metrics)

                    if has_decline:
                        fundamental_status = f"{decline_type}"
                        logger.info(f"    ✓ {symbol}: Has {decline_type} - Strong short candidate!")
                    else:
                        fundamental_status = "Stable/Growing"
                        logger.info(f"    ○ {symbol}: No fundamental decline")
                else:
                    logger.info(f"    ○ {symbol}: Fundamental data unavailable")

            except Exception as e:
                logger.warning(f"    ○ {symbol}: Error fetching fundamentals: {e}")

            # Add to results (stocks that are near highs AND bearish)
            result_data = {
                'symbol': symbol,
                'current_price': price_data['current_price'],
                '52w_high': price_data['52w_high'],
                '3year_peak': price_data['3year_peak'],
                'pct_from_52w_high': price_data['pct_from_52w_high'],
                'pct_from_3y_peak': price_data['pct_from_3y_peak'],
                'market_cap_cr': market_cap_cr if market_cap_cr else 0,
                'sector': sector,
                'volume_lakhs': volume_lakhs,
                '52w_low': price_data['52w_low'],
                'revenue_growth_yoy': growth_metrics['yoy_revenue_avg'] if growth_metrics else 0,
                'pat_growth_yoy': growth_metrics['yoy_pat_avg'] if growth_metrics else 0,
                'revenue_growth_qoq': growth_metrics['qoq_revenue_avg'] if growth_metrics else 0,
                'pat_growth_qoq': growth_metrics['qoq_pat_avg'] if growth_metrics else 0,
                'decline_type': decline_type,
                'fundamental_status': fundamental_status,
                'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }

            # Add SHORT trend analysis data
            if short_analysis:
                result_data.update({
                    # Trend Analysis (bearish)
                    'trend_status': short_analysis['trend_status'],
                    'trend_score': short_analysis['trend_score'],
                    'trend_confidence': short_analysis['trend_confidence'],
                    'pct_from_ema20': short_analysis['pct_from_ema20'],
                    'pct_from_sma50': short_analysis['pct_from_sma50'],
                    'pct_from_sma200': short_analysis['pct_from_sma200'],
                    'adx': short_analysis['adx'],
                    'adx_signal': short_analysis['adx_signal'],

                    # SHORT Entry Signal
                    'entry_signal': short_analysis['entry_signal'],
                    'entry_type': short_analysis['entry_type'],
                    'entry_price': short_analysis['entry_price'],
                    'entry_score': short_analysis['entry_score'],

                    # Risk Management (SHORT - stops ABOVE, targets BELOW)
                    'stop_loss': short_analysis['stop_loss'],
                    'stop_loss_pct': short_analysis['stop_loss_pct'],
                    'target_1': short_analysis['target_1'],
                    'target_1_pct': short_analysis['target_1_pct'],
                    'target_2': short_analysis['target_2'],
                    'target_2_pct': short_analysis['target_2_pct'],
                    'risk_reward': short_analysis['risk_reward'],

                    # Support/Resistance (inverted for shorts)
                    'nearest_support': short_analysis['nearest_support'],  # Target
                    'nearest_resistance': short_analysis['nearest_resistance'],  # Stop
                    's1_pivot': short_analysis['s1_pivot'],
                    'r1_pivot': short_analysis['r1_pivot'],

                    # Position Sizing
                    'position_size': short_analysis['position_size'],
                    'position_value': short_analysis['position_value']
                })

            results.append(result_data)

        # Summary
        logger.info("")
        logger.info("="*80)
        logger.info("SHORT SCREENING COMPLETE")
        logger.info("="*80)
        logger.info(f"Total stocks scanned: {len(self.stocks)}")
        logger.info(f"Near highs (resistance zone): {price_qualified}")
        logger.info(f"Bearish trend confirmed: {trend_qualified}")
        logger.info(f"SHORT entry signals: {short_signals}")
        logger.info(f"Final short candidates: {len(results)}")
        logger.info("="*80)
        logger.info("")

        return results

    def generate_excel_report(self, long_results: List[Dict], short_results: List[Dict], filename: str) -> str:
        """
        Generate 2-sheet Excel report with long and short opportunities

        Args:
            long_results: List of long opportunity dicts
            short_results: List of short opportunity dicts
            filename: Output filename (without path)

        Returns:
            Full path to generated Excel file
        """
        # Create output directory
        output_dir = "data/screener_results"
        os.makedirs(output_dir, exist_ok=True)
        filepath = os.path.join(output_dir, filename)

        # Create workbook
        wb = openpyxl.Workbook()

        # ===================================
        # SHEET 1: LONG OPPORTUNITIES
        # ===================================
        ws_long = wb.active
        ws_long.title = "Long Opportunities"

        # Define headers (Extended with Trend Analysis)
        headers = [
            # Basic Info
            'Symbol', 'Current Price (₹)', '3-Year Peak (₹)', 'Drawdown (%)',
            'Market Cap (Cr)', 'Sector', 'Volume (Lakhs)',
            '52W High (₹)', '52W Low (₹)',

            # Fundamental Growth
            'Revenue Growth YoY (%)', 'PAT Growth YoY (%)',
            'Revenue Growth QoQ (%)', 'PAT Growth QoQ (%)',
            'Growth Status',

            # Trend Analysis
            'Trend Status', 'Trend Score (0-10)', 'Trend Confidence',
            '% from 20 EMA', '% from 50 SMA', '% from 200 SMA',
            'ADX', 'ADX Signal',

            # Entry Signal
            'Entry Signal', 'Entry Type', 'Entry Price (₹)', 'Entry Score',

            # Risk Management
            'Stop Loss (₹)', 'Stop Loss (%)',
            'Target 1 (₹)', 'Target 1 (%)',
            'Target 2 (₹)', 'Target 2 (%)',
            'Risk:Reward',

            # Support/Resistance
            'Nearest Support (₹)', 'Nearest Resistance (₹)',
            'S1 Pivot (₹)', 'R1 Pivot (₹)',

            # Position Sizing
            'Position Size (Qty)', 'Position Value (₹)',

            'Last Updated'
        ]

        # Write headers
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, header in enumerate(headers, 1):
            cell = ws_long.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data rows with 4-tier color coding
        for row_idx, stock in enumerate(long_results, 2):
            # === 4-TIER COLOR CODING LOGIC ===
            # Tier 1 (Priority 1): Dark Green - Strong Uptrend + BUY signal
            # Tier 2 (Priority 2): Light Green - Growth stocks
            # Tier 3 (Priority 3): Yellow - Potential reversal (oversold downtrend)
            # Tier 4 (Priority 4): Red - Strong Downtrend (avoid)

            has_growth = 'Growth' in stock['growth_status'] and stock['growth_status'] != "No Growth"
            trend_score = stock.get('trend_score', 0)
            entry_signal = stock.get('entry_signal', 'N/A')

            # Determine color tier
            if trend_score >= 7 and entry_signal == 'BUY':
                # TIER 1: Dark Green - BEST (Strong Uptrend + BUY)
                row_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                row_font = Font(bold=True, color="006100")
                priority = "⭐ TIER 1"
            elif has_growth:
                # TIER 2: Light Green - GOOD (Growth stocks)
                row_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                row_font = Font(color="000000")
                priority = "TIER 2"
            elif trend_score <= -4 and trend_score > -7:
                # TIER 3: Yellow - WATCH (Potential reversal)
                row_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                row_font = Font(color="7F6000")
                priority = "TIER 3"
            elif trend_score <= -7:
                # TIER 4: Red - AVOID (Strong Downtrend)
                row_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                row_font = Font(color="9C0006")
                priority = "TIER 4"
            else:
                # DEFAULT: No highlighting
                row_fill = None
                row_font = Font(color="000000")
                priority = "---"

            col_num = 1

            # Basic Info
            def write_cell(value, apply_fill=True):
                nonlocal col_num
                cell = ws_long.cell(row_idx, col_num, value)
                if apply_fill and row_fill:
                    cell.fill = row_fill
                col_num += 1
                return cell

            # Symbol (bold for Tier 1)
            symbol_cell = write_cell(stock['symbol'])
            if priority == "⭐ TIER 1":
                symbol_cell.font = Font(bold=True, color="006100", size=11)

            # Basic price data
            write_cell(round(stock['current_price'], 2))
            write_cell(round(stock['3year_peak'], 2))
            write_cell(round(stock['drawdown_pct'], 2))
            write_cell(round(stock['market_cap_cr'], 0))
            write_cell(stock['sector'])
            write_cell(round(stock['volume_lakhs'], 2))
            write_cell(round(stock['52w_high'], 2))
            write_cell(round(stock['52w_low'], 2))

            # Growth metrics with conditional formatting
            for key in ['revenue_growth_yoy', 'pat_growth_yoy', 'revenue_growth_qoq', 'pat_growth_qoq']:
                growth_cell = ws_long.cell(row_idx, col_num, round(stock[key], 2))
                if row_fill:
                    growth_cell.fill = row_fill
                # Add green/red overlay for positive/negative
                if stock[key] > 5:
                    growth_cell.font = Font(color="006100", bold=True)
                elif stock[key] < -5:
                    growth_cell.font = Font(color="9C0006")
                col_num += 1

            # Growth Status
            status_cell = write_cell(stock['growth_status'])
            if has_growth:
                status_cell.font = Font(bold=True, color="006100")

            # === TREND ANALYSIS COLUMNS ===

            # Trend Status (with color coding)
            trend_cell = ws_long.cell(row_idx, col_num, stock.get('trend_status', 'N/A'))
            trend_status = stock.get('trend_status', 'N/A')
            if 'Strong Uptrend' in trend_status:
                trend_cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                trend_cell.font = Font(bold=True, color="006100")
            elif 'Uptrend' in trend_status:
                trend_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                trend_cell.font = Font(color="006100")
            elif 'Strong Downtrend' in trend_status:
                trend_cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                trend_cell.font = Font(color="9C0006")
            elif 'Downtrend' in trend_status:
                trend_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                trend_cell.font = Font(color="7F6000")
            col_num += 1

            # Trend Score (0-10)
            score_cell = ws_long.cell(row_idx, col_num, stock.get('trend_score', 0))
            if trend_score >= 7:
                score_cell.font = Font(bold=True, color="006100")
            elif trend_score <= -7:
                score_cell.font = Font(bold=True, color="9C0006")
            col_num += 1

            # Trend Confidence
            write_cell(stock.get('trend_confidence', 'N/A'))

            # % from EMAs/SMAs
            write_cell(round(stock.get('pct_from_ema20', 0), 2))
            write_cell(round(stock.get('pct_from_sma50', 0), 2))
            write_cell(round(stock.get('pct_from_sma200', 0), 2))

            # ADX
            write_cell(round(stock.get('adx', 0), 2))
            write_cell(stock.get('adx_signal', 'N/A'))

            # Entry Signal (BUY highlighted)
            entry_cell = ws_long.cell(row_idx, col_num, stock.get('entry_signal', 'N/A'))
            if entry_signal == 'BUY':
                entry_cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
                entry_cell.font = Font(bold=True, color="006100")
            elif entry_signal == 'WAIT':
                entry_cell.font = Font(color="7F6000")
            elif entry_signal == 'AVOID':
                entry_cell.font = Font(color="9C0006")
            col_num += 1

            # Entry Type
            write_cell(stock.get('entry_type', 'N/A'))

            # Entry Price
            write_cell(round(stock.get('entry_price', 0), 2))

            # Entry Score
            write_cell(stock.get('entry_score', 0))

            # Risk Management
            write_cell(round(stock.get('stop_loss', 0), 2))
            write_cell(round(stock.get('stop_loss_pct', 0), 2))
            write_cell(round(stock.get('target_1', 0), 2))
            write_cell(round(stock.get('target_1_pct', 0), 2))
            write_cell(round(stock.get('target_2', 0), 2))
            write_cell(round(stock.get('target_2_pct', 0), 2))
            write_cell(stock.get('risk_reward', 'N/A'))

            # Support/Resistance
            write_cell(round(stock.get('nearest_support', 0), 2))
            write_cell(round(stock.get('nearest_resistance', 0), 2))
            write_cell(round(stock.get('s1_pivot', 0), 2))
            write_cell(round(stock.get('r1_pivot', 0), 2))

            # Position Sizing
            write_cell(stock.get('position_size', 0))
            write_cell(round(stock.get('position_value', 0), 0))

            # Last Updated
            write_cell(stock['last_updated'])

        # Auto-size columns
        for column in ws_long.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_long.column_dimensions[column_letter].width = adjusted_width

        # Freeze top row
        ws_long.freeze_panes = "A2"

        # Enable auto-filter
        ws_long.auto_filter.ref = ws_long.dimensions

        # ===================================
        # SHEET 2: SHORT OPPORTUNITIES
        # ===================================
        ws_short = wb.create_sheet("Short Opportunities")

        # Define headers for short sheet
        short_headers = [
            # Basic Info
            'Symbol', 'Current Price (₹)', '52W High (₹)', '3Y Peak (₹)',
            '% from 52W High', '% from 3Y Peak',
            'Market Cap (Cr)', 'Sector', 'Volume (Lakhs)', '52W Low (₹)',

            # Fundamental Decline
            'Revenue Growth YoY (%)', 'PAT Growth YoY (%)',
            'Revenue Growth QoQ (%)', 'PAT Growth QoQ (%)',
            'Fundamental Status',

            # Trend Analysis
            'Trend Status', 'Trend Score (-10 to 0)', 'Trend Confidence',
            '% from 20 EMA', '% from 50 SMA', '% from 200 SMA',
            'ADX', 'ADX Signal',

            # SHORT Entry Signal
            'Entry Signal', 'Entry Type', 'Entry Price (₹)', 'Entry Score',

            # Risk Management (SHORT - stops ABOVE, targets BELOW)
            'Stop Loss (₹)', 'Stop Loss (%)',
            'Target 1 (₹)', 'Target 1 (%)',
            'Target 2 (₹)', 'Target 2 (%)',
            'Risk:Reward',

            # Support/Resistance
            'Nearest Support (₹)', 'Nearest Resistance (₹)',
            'S1 Pivot (₹)', 'R1 Pivot (₹)',

            # Position Sizing
            'Position Size (Qty)', 'Position Value (₹)',

            'Last Updated'
        ]

        # Write SHORT sheet headers
        for col_num, header in enumerate(short_headers, 1):
            cell = ws_short.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")  # Dark red
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write SHORT data rows with inverted 4-tier color coding
        for row_idx, stock in enumerate(short_results, 2):
            # === 4-TIER COLOR CODING FOR SHORTS (INVERTED) ===
            trend_score = stock.get('trend_score', 0)
            entry_signal = stock.get('entry_signal', 'N/A')
            has_decline = 'Decline' in stock.get('fundamental_status', '')

            # Determine color tier (inverted - red for strong bearish)
            if trend_score <= -7 and entry_signal == 'SHORT':
                # TIER 1: Dark Red - BEST (Strong Downtrend + SHORT)
                row_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                row_font = Font(bold=True, color="9C0006")
                priority = "⭐ TIER 1"
            elif has_decline:
                # TIER 2: Light Red - GOOD (Fundamental decline)
                row_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                row_font = Font(color="9C0006")
                priority = "TIER 2"
            elif -7 < trend_score <= -4:
                # TIER 3: Yellow - WATCH (Moderate downtrend)
                row_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
                row_font = Font(color="7F6000")
                priority = "TIER 3"
            else:
                # TIER 4: Light Yellow - WEAK
                row_fill = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")
                row_font = Font(color="000000")
                priority = "TIER 4"

            col_num = 1

            def write_short_cell(value, apply_fill=True):
                nonlocal col_num
                cell = ws_short.cell(row_idx, col_num, value)
                if apply_fill and row_fill:
                    cell.fill = row_fill
                col_num += 1
                return cell

            # Symbol (bold for Tier 1)
            symbol_cell = write_short_cell(stock['symbol'])
            if priority == "⭐ TIER 1":
                symbol_cell.font = Font(bold=True, color="9C0006", size=11)

            # Basic price data
            write_short_cell(round(stock['current_price'], 2))
            write_short_cell(round(stock['52w_high'], 2))
            write_short_cell(round(stock['3year_peak'], 2))
            write_short_cell(round(stock['pct_from_52w_high'], 2))
            write_short_cell(round(stock['pct_from_3y_peak'], 2))
            write_short_cell(round(stock['market_cap_cr'], 0))
            write_short_cell(stock['sector'])
            write_short_cell(round(stock['volume_lakhs'], 2))
            write_short_cell(round(stock['52w_low'], 2))

            # Fundamental metrics (declining is bad for fundamentals but good for shorts)
            for key in ['revenue_growth_yoy', 'pat_growth_yoy', 'revenue_growth_qoq', 'pat_growth_qoq']:
                decline_cell = ws_short.cell(row_idx, col_num, round(stock[key], 2))
                if row_fill:
                    decline_cell.fill = row_fill
                # Negative growth = red (bad fundamentals, good for shorts)
                if stock[key] < -5:
                    decline_cell.font = Font(color="9C0006", bold=True)
                elif stock[key] > 5:
                    decline_cell.font = Font(color="006100")
                col_num += 1

            # Fundamental Status
            status_cell = write_short_cell(stock['fundamental_status'])
            if has_decline:
                status_cell.font = Font(bold=True, color="9C0006")

            # Trend Status (bearish)
            trend_cell = ws_short.cell(row_idx, col_num, stock.get('trend_status', 'N/A'))
            trend_status_val = stock.get('trend_status', 'N/A')
            if 'Strong Downtrend' in trend_status_val:
                trend_cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                trend_cell.font = Font(bold=True, color="9C0006")
            elif 'Downtrend' in trend_status_val:
                trend_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                trend_cell.font = Font(color="9C0006")
            col_num += 1

            # Trend Score
            score_cell = ws_short.cell(row_idx, col_num, stock.get('trend_score', 0))
            if trend_score <= -7:
                score_cell.font = Font(bold=True, color="9C0006")
            col_num += 1

            # Rest of trend columns
            write_short_cell(stock.get('trend_confidence', 'N/A'))
            write_short_cell(round(stock.get('pct_from_ema20', 0), 2))
            write_short_cell(round(stock.get('pct_from_sma50', 0), 2))
            write_short_cell(round(stock.get('pct_from_sma200', 0), 2))
            write_short_cell(round(stock.get('adx', 0), 2))
            write_short_cell(stock.get('adx_signal', 'N/A'))

            # SHORT Entry Signal (highlighted)
            entry_cell = ws_short.cell(row_idx, col_num, stock.get('entry_signal', 'N/A'))
            if entry_signal == 'SHORT':
                entry_cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                entry_cell.font = Font(bold=True, color="9C0006")
            col_num += 1

            # Entry details
            write_short_cell(stock.get('entry_type', 'N/A'))
            write_short_cell(round(stock.get('entry_price', 0), 2))
            write_short_cell(stock.get('entry_score', 0))

            # Risk management (shorts)
            write_short_cell(round(stock.get('stop_loss', 0), 2))
            write_short_cell(round(stock.get('stop_loss_pct', 0), 2))
            write_short_cell(round(stock.get('target_1', 0), 2))
            write_short_cell(round(stock.get('target_1_pct', 0), 2))
            write_short_cell(round(stock.get('target_2', 0), 2))
            write_short_cell(round(stock.get('target_2_pct', 0), 2))
            write_short_cell(stock.get('risk_reward', 'N/A'))

            # Support/Resistance
            write_short_cell(round(stock.get('nearest_support', 0), 2))
            write_short_cell(round(stock.get('nearest_resistance', 0), 2))
            write_short_cell(round(stock.get('s1_pivot', 0), 2))
            write_short_cell(round(stock.get('r1_pivot', 0), 2))

            # Position Sizing
            write_short_cell(stock.get('position_size', 0))
            write_short_cell(round(stock.get('position_value', 0), 0))

            # Last Updated
            write_short_cell(stock['last_updated'])

        # Auto-size columns for SHORT sheet
        for column in ws_short.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_short.column_dimensions[column_letter].width = adjusted_width

        # Freeze top row
        ws_short.freeze_panes = "A2"

        # Enable auto-filter
        ws_short.auto_filter.ref = ws_short.dimensions

        # Save workbook
        wb.save(filepath)
        logger.info(f"✓ Excel report saved: {filepath}")

        return filepath

    def run(self):
        """Main execution method - screens for LONG and SHORT opportunities"""
        start_time = time.time()

        logger.info("=" * 80)
        logger.info("STOCK VALUE SCREENER - NSE (Market Cap > ₹10,000 Cr)")
        logger.info("=" * 80)
        logger.info(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info(f"Stocks to scan: {len(self.stocks)}")
        logger.info("")
        logger.info("Dual Screening Strategy:")
        logger.info("  LONG OPPORTUNITIES:")
        logger.info("    • Price is 30% or more below 3-year peak")
        logger.info("    • Strong uptrend with BUY signals")
        logger.info("    • GREEN rows = Has fundamental growth")
        logger.info("")
        logger.info("  SHORT OPPORTUNITIES:")
        logger.info("    • Price within 5% of 52W/3Y high (resistance)")
        logger.info("    • Strong downtrend (bearish)")
        logger.info("    • RED rows = Strong short candidates")
        logger.info("=" * 80)
        logger.info("")

        # Screen LONG opportunities
        long_results = self.screen_all_stocks()

        # Screen SHORT opportunities
        short_results = self.screen_short_candidates()

        # Generate 2-sheet Excel report
        if long_results or short_results:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"trading_opportunities_{timestamp}.xlsx"
            filepath = self.generate_excel_report(long_results, short_results, filename)
        else:
            logger.warning("No stocks matched any criteria!")
            filepath = None

        # Summary
        elapsed_time = time.time() - start_time

        # Count stocks with growth/decline
        stocks_with_growth = sum(1 for r in long_results if 'Growth' in r['growth_status'] and r['growth_status'] != "No Growth")
        stocks_with_short_signal = sum(1 for r in short_results if r.get('entry_signal') == 'SHORT')

        logger.info("")
        logger.info("=" * 80)
        logger.info("SCREENING COMPLETE")
        logger.info("=" * 80)
        logger.info(f"Total stocks scanned:      {len(self.stocks)}")
        logger.info("")
        logger.info("LONG OPPORTUNITIES:")
        logger.info(f"  Price qualified:         {len(long_results)} stocks (30%+ below peak)")
        logger.info(f"  → With growth (GREEN):   {stocks_with_growth} stocks")
        logger.info(f"  → Without growth:        {len(long_results) - stocks_with_growth} stocks")
        logger.info("")
        logger.info("SHORT OPPORTUNITIES:")
        logger.info(f"  Total candidates:        {len(short_results)} stocks")
        logger.info(f"  → SHORT signals:         {stocks_with_short_signal} stocks")
        logger.info("")
        logger.info(f"Execution time:            {elapsed_time/60:.1f} minutes")
        if filepath:
            logger.info(f"Results saved to:          {filepath}")
        logger.info("=" * 80)

        if long_results or short_results:
            logger.info("")
            if long_results:
                logger.info(f"📈 LONGS: {len(long_results)} undervalued stocks found!")
                logger.info(f"   {stocks_with_growth} with growth (Sheet 1: Long Opportunities)")
            if short_results:
                logger.info(f"📉 SHORTS: {len(short_results)} overextended stocks at resistance!")
                logger.info(f"   {stocks_with_short_signal} with SHORT signals (Sheet 2: Short Opportunities)")
            logger.info("")
            logger.info("⚠️  SHORT SELLING WARNING: Shorts carry unlimited risk. Use strict stop losses!")
            logger.info("Check the Excel file for detailed analysis.")
        else:
            logger.info("")
            logger.info("No opportunities found. Try different screening criteria or run during market hours.")

        logger.info("")


def main():
    """Main entry point"""
    try:
        screener = StockValueScreener()
        screener.run()
    except KeyboardInterrupt:
        logger.info("\nInterrupted by user")
    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)


if __name__ == "__main__":
    main()
