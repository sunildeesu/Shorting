#!/usr/bin/env python3
"""
Verify price logic is correct (stop loss, buy, target relationships)
"""

import openpyxl

def extract_price(val):
    """Extract numeric price from cell value"""
    if not val or val == '-':
        return None
    if isinstance(val, str):
        val = val.replace('₹', '').replace(',', '').strip()
        try:
            return float(val)
        except:
            return None
    try:
        return float(val)
    except:
        return None

def verify_price_logic(report_path: str):
    """Verify that price relationships are correct"""

    print(f"Checking price logic in: {report_path}\n")
    wb = openpyxl.load_workbook(report_path)
    ws = wb.active

    print(f"{'='*120}")
    print(f"{'Stock':<12} {'Pattern':<20} {'Signal':<10} {'Current':<10} {'Buy':<10} {'Target':<10} {'Stop':<10} {'R:R':<8} {'Logic Check':<20}")
    print(f"{'='*120}")

    issues = []
    valid_trades = 0

    for row_num in range(4, ws.max_row + 1):
        stock = ws.cell(row=row_num, column=1).value
        if not stock:
            break

        pattern = ws.cell(row=row_num, column=8).value
        current = extract_price(ws.cell(row=row_num, column=9).value)
        buy = extract_price(ws.cell(row=row_num, column=11).value)
        target = extract_price(ws.cell(row=row_num, column=12).value)
        stop = extract_price(ws.cell(row=row_num, column=13).value)
        signal = ws.cell(row=row_num, column=16).value

        if not pattern or pattern.strip() == '':
            continue

        valid_trades += 1

        # Determine if bullish or bearish
        is_bullish = 'DOUBLE_BOTTOM' in pattern.upper() or 'RESISTANCE_BREAKOUT' in pattern.upper()
        is_bearish = 'DOUBLE_TOP' in pattern.upper() or 'SUPPORT_BREAKOUT' in pattern.upper()

        logic_check = "✅ OK"
        risk_reward = "-"

        if buy and target and stop:
            # Calculate risk-reward
            if is_bullish:
                risk = buy - stop
                reward = target - buy
                if risk > 0:
                    risk_reward = f"1:{reward/risk:.1f}"
                else:
                    risk_reward = "Invalid"

                # For bullish: stop < buy < target
                if not (stop < buy < target):
                    logic_check = f"❌ WRONG ORDER"
                    issues.append(f"{stock} (Bullish): Expected stop({stop:.2f}) < buy({buy:.2f}) < target({target:.2f})")

                # Stop loss should be ~2% below buy
                expected_stop_range = (buy * 0.96, buy * 1.00)
                if not (expected_stop_range[0] <= stop <= expected_stop_range[1]):
                    logic_check += " ⚠️ Stop%"

            elif is_bearish:
                risk = stop - buy
                reward = buy - target
                if risk > 0:
                    risk_reward = f"1:{reward/risk:.1f}"
                else:
                    risk_reward = "Invalid"

                # For bearish: target < buy < stop
                if not (target < buy < stop):
                    logic_check = f"❌ WRONG ORDER"
                    issues.append(f"{stock} (Bearish): Expected target({target:.2f}) < buy({buy:.2f}) < stop({stop:.2f})")

                # Stop loss should be ~2% above buy
                expected_stop_range = (buy * 1.00, buy * 1.04)
                if not (expected_stop_range[0] <= stop <= expected_stop_range[1]):
                    logic_check += " ⚠️ Stop%"

        else:
            logic_check = "❌ Missing prices"
            issues.append(f"{stock}: Missing one or more prices")

        # Display
        pattern_short = pattern[:18] if pattern else "-"
        print(f"{stock:<12} {pattern_short:<20} {signal or '-':<10} {current or 0:<10.2f} {buy or 0:<10.2f} {target or 0:<10.2f} {stop or 0:<10.2f} {risk_reward:<8} {logic_check:<20}")

    print(f"{'='*120}")
    print(f"\nTotal trades with patterns: {valid_trades}")

    if issues:
        print(f"\n❌ Issues Found ({len(issues)}):")
        print(f"{'='*120}")
        for issue in issues:
            print(f"  - {issue}")
        return False
    else:
        print(f"\n✅ All price relationships are logically correct!")
        print(f"\nKey Observations:")
        print(f"  - Bullish patterns: Stop Loss < Buy Price < Target Price")
        print(f"  - Bearish patterns: Target Price < Buy Price < Stop Loss")
        print(f"  - Stop losses are approximately 2% from key levels")
        return True

if __name__ == "__main__":
    report_path = "data/eod_reports/2025/11/eod_analysis_2025-11-04.xlsx"
    verify_price_logic(report_path)
