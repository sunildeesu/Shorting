#!/usr/bin/env python3
"""
Backtest: 5-Min Alerts - First Alert After 12 PM (2-Min Delayed Entry)

Strategy:
- Consider only the FIRST 5-min alert per stock per day
- That first alert must have come AFTER 12:00 PM IST
- Position taken 2 minutes after the alert (simulating decision + execution lag)
- For DROP: short position (profit if price falls further)
- For RISE: long position (profit if price rises further)

Runs multiple scenarios for comprehensive analysis:
1. STRICT: First alert post 12 PM, excluding results days
2. WITH RESULTS: First alert post 12 PM, including results days
3. ALL POST-NOON: All alerts post 12 PM (not just first)
4. BENCHMARK: All alerts (any time) for comparison

Analyzes P&L at different exit windows: 5, 10, 15, 20, 30 minutes after entry.
"""

import sqlite3
from datetime import datetime, timedelta
from collections import defaultdict
import statistics
import openpyxl
import config
import requests
import json
import os

NOON_CUTOFF_HOUR = 12
ENTRY_DELAY_MINUTES = 2
EXIT_TIMES = [5, 10, 15, 20, 30]


def load_5min_alerts(days: int = 60):
    """Load 5-min alerts from Excel for last N days."""
    excel_path = config.ALERT_EXCEL_PATH

    try:
        workbook = openpyxl.load_workbook(excel_path, read_only=True)
    except Exception as e:
        print(f"Failed to load Excel: {e}")
        return []

    if "5min_alerts" not in workbook.sheetnames:
        workbook.close()
        return []

    sheet = workbook["5min_alerts"]
    cutoff_date = datetime.now() - timedelta(days=days)

    alerts = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        try:
            date_val = row[0]
            time_val = row[1]

            if isinstance(date_val, datetime):
                date_str = date_val.strftime("%Y-%m-%d")
            else:
                date_str = str(date_val)

            if isinstance(time_val, datetime):
                time_str = time_val.strftime("%H:%M:%S")
            else:
                time_str = str(time_val)

            try:
                alert_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
            except ValueError:
                alert_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")

            if alert_dt < cutoff_date:
                continue

            direction = str(row[3]).strip() if row[3] else 'Drop'
            change_pct = float(row[6]) if row[6] else 0
            current_price = float(row[4]) if row[4] else 0

            alerts.append({
                'datetime': alert_dt,
                'date': date_str,
                'symbol': str(row[2]).strip() if row[2] else '',
                'direction': 'DROP' if 'drop' in direction.lower() else 'RISE',
                'change_pct': change_pct,
                'alert_price': current_price,
            })

        except Exception:
            continue

    workbook.close()
    return alerts


def load_results_schedule():
    """Load historical quarterly results announcements."""
    historical_file = "data/results_cache/historical_results.json"
    if not os.path.exists(historical_file):
        return {}
    try:
        with open(historical_file, 'r') as f:
            data = json.load(f)
            return data.get('results_dates', {})
    except Exception:
        return {}


def was_results_day(symbol: str, alert_date: str, schedule: dict) -> bool:
    symbol = symbol.upper()
    if symbol not in schedule:
        return False
    return alert_date in schedule[symbol]


def get_price_at_time(conn, symbol: str, target_time: datetime, tolerance_minutes: int = 2):
    """Get closest price to target time within tolerance."""
    cursor = conn.cursor()
    start = (target_time - timedelta(minutes=tolerance_minutes)).strftime('%Y-%m-%d %H:%M:%S')
    end = (target_time + timedelta(minutes=tolerance_minutes)).strftime('%Y-%m-%d %H:%M:%S')

    cursor.execute("""
        SELECT price FROM stock_quotes
        WHERE symbol = ?
        AND timestamp >= ? AND timestamp <= ?
        ORDER BY ABS(strftime('%s', timestamp) - strftime('%s', ?))
        LIMIT 1
    """, (symbol, start, end, target_time.strftime('%Y-%m-%d %H:%M:%S')))

    row = cursor.fetchone()
    return row[0] if row else None


def filter_first_alert_post_noon(alerts):
    """Keep only the FIRST alert per stock per day, and only if after noon."""
    by_stock_day = defaultdict(list)
    for alert in alerts:
        key = (alert['symbol'], alert['date'])
        by_stock_day[key].append(alert)

    filtered = []
    for key, day_alerts in by_stock_day.items():
        day_alerts.sort(key=lambda x: x['datetime'])
        first_alert = day_alerts[0]
        if first_alert['datetime'].hour >= NOON_CUTOFF_HOUR:
            filtered.append(first_alert)

    return filtered


def filter_all_post_noon(alerts):
    """Keep ALL alerts that come after noon."""
    return [a for a in alerts if a['datetime'].hour >= NOON_CUTOFF_HOUR]


def analyze_trades(alerts, conn, label=""):
    """Run trade analysis on a set of alerts. Returns results dict."""
    results = {t: {'all': [], 'DROP': [], 'RISE': []} for t in EXIT_TIMES}
    day_results = defaultdict(lambda: {t: [] for t in EXIT_TIMES})
    no_entry = 0

    for alert in alerts:
        symbol = alert['symbol']
        alert_time = alert['datetime']
        direction = alert['direction']

        entry_time = alert_time + timedelta(minutes=ENTRY_DELAY_MINUTES)
        entry_price = get_price_at_time(conn, symbol, entry_time)

        if not entry_price or entry_price <= 0:
            no_entry += 1
            continue

        for exit_min in EXIT_TIMES:
            exit_time = entry_time + timedelta(minutes=exit_min)
            market_close = alert_time.replace(hour=15, minute=25, second=0)
            if exit_time > market_close:
                continue

            exit_price = get_price_at_time(conn, symbol, exit_time)
            if not exit_price or exit_price <= 0:
                continue

            if direction == 'DROP':
                pnl_pct = ((entry_price - exit_price) / entry_price) * 100
            else:
                pnl_pct = ((exit_price - entry_price) / entry_price) * 100

            trade = {
                'symbol': symbol,
                'date': alert['date'],
                'direction': direction,
                'alert_time': alert_time.strftime('%H:%M'),
                'entry_time': entry_time.strftime('%H:%M'),
                'alert_price': alert['alert_price'],
                'entry_price': entry_price,
                'exit_price': exit_price,
                'pnl_pct': pnl_pct,
                'profitable': pnl_pct > 0,
                'change_pct': alert['change_pct'],
            }

            results[exit_min]['all'].append(trade)
            results[exit_min][direction].append(trade)
            day_results[alert['date']][exit_min].append(trade)

    return results, day_results, no_entry


def print_scenario_results(results, day_results, label, no_entry=0, total_alerts=0):
    """Print results for a scenario."""
    print(f"\n{'='*80}")
    print(f"  {label}")
    print(f"{'='*80}")

    if total_alerts:
        has_data = sum(1 for t in results[10]['all']) if results[10]['all'] else 0
        print(f"\n  Qualifying alerts: {total_alerts} | With price data: {has_data + no_entry} | No entry price: {no_entry}")

    # Check if we have any trades
    has_trades = any(results[t]['all'] for t in EXIT_TIMES)
    if not has_trades:
        print(f"\n  No trades with price data available for this scenario.")
        return

    print(f"\n  {'Exit':<7} {'Trades':<8} {'Win%':<8} {'Avg P&L':<10} {'Total P&L':<11} {'Avg Win':<10} {'Avg Loss':<10} {'MaxWin':<9} {'MaxLoss':<9}")
    print(f"  {'-'*85}")

    for exit_min in EXIT_TIMES:
        trades = results[exit_min]['all']
        if not trades:
            continue

        winners = [t for t in trades if t['profitable']]
        losers = [t for t in trades if not t['profitable']]
        all_pnl = [t['pnl_pct'] for t in trades]
        win_rate = len(winners) / len(trades) * 100
        avg_pnl = statistics.mean(all_pnl)
        total_pnl = sum(all_pnl)
        avg_win = statistics.mean([t['pnl_pct'] for t in winners]) if winners else 0
        avg_loss = statistics.mean([t['pnl_pct'] for t in losers]) if losers else 0
        max_win = max(all_pnl)
        max_loss = min(all_pnl)

        print(f"  {exit_min}min{'':<3} {len(trades):<8} {win_rate:.1f}%{'':<3} {avg_pnl:+.3f}%{'':<3} {total_pnl:+.2f}%{'':<4} {avg_win:+.3f}%{'':<3} {avg_loss:+.3f}%{'':<3} {max_win:+.2f}%{'':<2} {max_loss:+.2f}%")

    # By direction
    for direction in ['DROP', 'RISE']:
        dir_trades = results[10][direction]
        if not dir_trades:
            continue

        winners = [t for t in dir_trades if t['profitable']]
        all_pnl = [t['pnl_pct'] for t in dir_trades]
        win_rate = len(winners) / len(dir_trades) * 100
        avg_pnl = statistics.mean(all_pnl)
        total_pnl = sum(all_pnl)

        emoji = "SHORT" if direction == "DROP" else "LONG"
        print(f"\n  {direction} ({emoji}) @ 10min: {len(dir_trades)} trades | {win_rate:.1f}% win | {avg_pnl:+.3f}% avg | {total_pnl:+.2f}% total")

    # Trade details
    if results[10]['all']:
        print(f"\n  Trade details (10min exit):")
        for t in sorted(results[10]['all'], key=lambda x: x['date']):
            marker = "+" if t['profitable'] else "-"
            print(f"    {marker} {t['date']} {t['symbol']:<15} {t['direction']:<5} Alert:{t['alert_time']} Entry:Rs{t['entry_price']:>8.2f} Exit:Rs{t['exit_price']:>8.2f} P&L:{t['pnl_pct']:+.2f}%")


def print_detailed_analysis(results, day_results):
    """Print detailed analysis for the best scenario."""
    if not results[10]['all']:
        return

    # P&L Distribution
    print(f"\n{'='*80}")
    print("P&L DISTRIBUTION (10 min exit)")
    print(f"{'='*80}\n")

    trades_10 = results[10]['all']
    ranges = [
        ('Big Loss (< -2%)', lambda x: x < -2),
        ('Loss (-2% to -1%)', lambda x: -2 <= x < -1),
        ('Small Loss (-1% to 0%)', lambda x: -1 <= x < 0),
        ('Small Win (0% to 0.5%)', lambda x: 0 <= x < 0.5),
        ('Win (0.5% to 1%)', lambda x: 0.5 <= x < 1),
        ('Good Win (1% to 2%)', lambda x: 1 <= x < 2),
        ('Big Win (> 2%)', lambda x: x >= 2),
    ]

    for label, condition in ranges:
        count = sum(1 for t in trades_10 if condition(t['pnl_pct']))
        pct = count / len(trades_10) * 100
        bar = '#' * int(pct / 2)
        print(f"  {label:<25} {count:>4} ({pct:>5.1f}%) {bar}")

    # Time-of-day analysis
    print(f"\n{'='*80}")
    print("TIME-OF-DAY ANALYSIS (10 min exit)")
    print(f"{'='*80}\n")

    hour_buckets = defaultdict(list)
    for t in trades_10:
        hour = int(t['alert_time'].split(':')[0])
        hour_buckets[hour].append(t['pnl_pct'])

    print(f"  {'Hour':<10} {'Trades':<8} {'Win%':<8} {'Avg P&L':<10}")
    print(f"  {'-'*40}")
    for hour in sorted(hour_buckets.keys()):
        pnls = hour_buckets[hour]
        winners = sum(1 for p in pnls if p > 0)
        win_rate = winners / len(pnls) * 100
        avg_p = statistics.mean(pnls)
        print(f"  {hour}:00{'':<5} {len(pnls):<8} {win_rate:.1f}%{'':<3} {avg_p:+.3f}%")

    # Day-by-day performance
    print(f"\n{'='*80}")
    print("DAY-BY-DAY PERFORMANCE (10 min exit)")
    print(f"{'='*80}\n")

    print(f"  {'Date':<12} {'Trades':<8} {'Win%':<8} {'Total P&L':<12} {'Avg P&L':<10}")
    print(f"  {'-'*55}")

    total_day_pnl = 0
    winning_days = 0
    total_days = 0

    for date in sorted(day_results.keys()):
        trades = day_results[date][10]
        if not trades:
            continue

        total_days += 1
        pnls = [t['pnl_pct'] for t in trades]
        day_total = sum(pnls)
        total_day_pnl += day_total
        winners = sum(1 for p in pnls if p > 0)
        win_rate = winners / len(pnls) * 100
        avg_p = statistics.mean(pnls)

        if day_total > 0:
            winning_days += 1

        marker = "+" if day_total > 0 else "-"
        print(f"  {date:<12} {len(trades):<8} {win_rate:.0f}%{'':<4} {day_total:+.2f}%{'':<5} {avg_p:+.3f}%  {marker}")

    if total_days > 0:
        print(f"\n  Winning Days: {winning_days}/{total_days} ({winning_days/total_days*100:.0f}%)")
        print(f"  Cumulative P&L: {total_day_pnl:+.2f}%")

    # Top winners and losers
    if len(trades_10) >= 5:
        sorted_trades = sorted(trades_10, key=lambda x: -x['pnl_pct'])

        print(f"\n{'='*80}")
        print("TOP 5 WINNERS (10 min exit)")
        print(f"{'='*80}")
        for t in sorted_trades[:5]:
            print(f"  {t['date']} {t['symbol']:<15} {t['direction']:<5} Entry:Rs{t['entry_price']:>8.2f} Exit:Rs{t['exit_price']:>8.2f}  P&L: {t['pnl_pct']:+.2f}%")

        print(f"\nTOP 5 LOSERS (10 min exit)")
        print(f"{'='*80}")
        for t in sorted_trades[-5:]:
            print(f"  {t['date']} {t['symbol']:<15} {t['direction']:<5} Entry:Rs{t['entry_price']:>8.2f} Exit:Rs{t['exit_price']:>8.2f}  P&L: {t['pnl_pct']:+.2f}%")


def run_backtest(days: int = 60):
    """Run multi-scenario backtest."""
    print(f"\n{'='*80}")
    print("  5-MIN ALERT BACKTEST: POST 12 PM ANALYSIS (2-MIN DELAYED ENTRY)")
    print(f"  Period: Last {days} days | Entry: 2 min after alert")
    print(f"{'='*80}")

    # Load alerts
    all_alerts = load_5min_alerts(days)
    print(f"\nLoaded {len(all_alerts)} total 5-min alerts from last {days} days")

    if not all_alerts:
        print("No alerts found!")
        return None

    # Load results schedule
    schedule = load_results_schedule()
    non_results = [a for a in all_alerts if not was_results_day(a['symbol'], a['date'], schedule)]
    print(f"After excluding results days: {len(non_results)} alerts")

    # Note DB date range limitation
    conn = sqlite3.connect("data/central_quotes.db", timeout=30)
    cursor = conn.cursor()
    cursor.execute("SELECT MIN(timestamp), MAX(timestamp) FROM stock_quotes")
    db_min, db_max = cursor.fetchone()
    print(f"Price DB range: {db_min} to {db_max}")

    # ===== SCENARIO 1: STRICT (First alert post noon, excl results) =====
    s1_alerts = filter_first_alert_post_noon(non_results)
    s1_results, s1_days, s1_no_entry = analyze_trades(s1_alerts, conn)
    print_scenario_results(s1_results, s1_days,
                          "SCENARIO 1: First Alert Post 12PM (Excl Results)",
                          s1_no_entry, len(s1_alerts))

    # ===== SCENARIO 2: First alert post noon, INCLUDING results =====
    s2_alerts = filter_first_alert_post_noon(all_alerts)
    s2_results, s2_days, s2_no_entry = analyze_trades(s2_alerts, conn)
    print_scenario_results(s2_results, s2_days,
                          "SCENARIO 2: First Alert Post 12PM (Incl Results)",
                          s2_no_entry, len(s2_alerts))

    # ===== SCENARIO 3: ALL alerts post noon (excl results) =====
    s3_alerts = filter_all_post_noon(non_results)
    s3_results, s3_days, s3_no_entry = analyze_trades(s3_alerts, conn)
    print_scenario_results(s3_results, s3_days,
                          "SCENARIO 3: ALL Alerts Post 12PM (Excl Results)",
                          s3_no_entry, len(s3_alerts))

    # ===== SCENARIO 4: ALL alerts any time (excl results) - BENCHMARK =====
    s4_results, s4_days, s4_no_entry = analyze_trades(non_results, conn)
    print_scenario_results(s4_results, s4_days,
                          "SCENARIO 4: ALL Alerts Any Time (Benchmark)",
                          s4_no_entry, len(non_results))

    # Find best scenario with most data for detailed analysis
    best_scenario = None
    best_results = None
    best_days = None
    best_label = ""

    for label, res, days_res in [
        ("Scenario 1 (Strict)", s1_results, s1_days),
        ("Scenario 2 (With Results)", s2_results, s2_days),
        ("Scenario 3 (All Post-Noon)", s3_results, s3_days),
        ("Scenario 4 (Benchmark)", s4_results, s4_days),
    ]:
        trade_count = len(res[10]['all'])
        if trade_count > 0 and (best_scenario is None or trade_count > len(best_results[10]['all'])):
            best_results = res
            best_days = days_res
            best_label = label

    # Print detailed analysis for scenario with most data
    if best_results and len(best_results[10]['all']) >= 3:
        print(f"\n\n{'#'*80}")
        print(f"  DETAILED ANALYSIS: {best_label}")
        print(f"{'#'*80}")
        print_detailed_analysis(best_results, best_days)

    # === RECOMMENDATION ===
    print(f"\n{'='*80}")
    print("RECOMMENDATION")
    print(f"{'='*80}\n")

    # Use scenario 3 (all post-noon) for recommendation as it has more data
    rec_results = s3_results if s3_results[10]['all'] else s4_results
    rec_label = "Post-Noon Alerts" if s3_results[10]['all'] else "All Alerts"

    if rec_results[10]['all']:
        trades = rec_results[10]['all']
        winners = [t for t in trades if t['profitable']]
        win_rate = len(winners) / len(trades) * 100
        avg_pnl = statistics.mean([t['pnl_pct'] for t in trades])
        total_pnl = sum([t['pnl_pct'] for t in trades])

        print(f"  Based on {rec_label} (10min exit):")
        print(f"  Trades: {len(trades)} | Win Rate: {win_rate:.1f}% | Avg P&L: {avg_pnl:+.3f}% | Total P&L: {total_pnl:+.2f}%")

        for direction in ['DROP', 'RISE']:
            dir_trades = rec_results[10][direction]
            if not dir_trades:
                continue

            dir_winners = [t for t in dir_trades if t['profitable']]
            dir_win_rate = len(dir_winners) / len(dir_trades) * 100
            dir_avg_pnl = statistics.mean([t['pnl_pct'] for t in dir_trades])

            if dir_win_rate >= 55 and dir_avg_pnl > 0:
                verdict = "PROFITABLE - Worth taking positions"
            elif dir_win_rate >= 50 and dir_avg_pnl > 0:
                verdict = "MARGINAL - Proceed with caution"
            elif dir_win_rate >= 45:
                verdict = "WEAK - Not recommended without additional filters"
            else:
                verdict = "UNPROFITABLE - Avoid taking positions"

            print(f"\n  {direction}: {verdict}")
            print(f"    Trades: {len(dir_trades)}, Win Rate: {dir_win_rate:.1f}%, Avg P&L: {dir_avg_pnl:+.3f}%")

        # Overall verdict
        print(f"\n  OVERALL VERDICT:")
        if win_rate >= 55 and avg_pnl > 0:
            print(f"  The strategy shows POSITIVE edge. Worth taking positions after 12 PM alerts.")
        elif win_rate >= 50 and avg_pnl > 0:
            print(f"  The strategy shows a MARGINAL edge. Can be taken with strict risk management.")
        elif win_rate >= 45:
            print(f"  The strategy shows WEAK results. Not recommended without additional confirmation.")
        else:
            print(f"  The strategy is UNPROFITABLE. Do NOT take positions based on post-noon 5-min alerts alone.")

        if len(trades) < 20:
            print(f"\n  WARNING: Sample size is small ({len(trades)} trades). Results may not be statistically significant.")
            print(f"  Need at least 30+ trades for reliable conclusions.")
    else:
        print(f"  Insufficient data for recommendation. No trades found with price data.")

    print(f"\n{'='*80}\n")

    conn.close()

    return {
        'scenario1': s1_results,
        'scenario2': s2_results,
        'scenario3': s3_results,
        'scenario4': s4_results,
    }


def generate_telegram_report(all_results):
    """Generate HTML-formatted Telegram report."""
    if not all_results:
        return None

    lines = []
    lines.append("<b>5-MIN ALERT BACKTEST: POST 12PM</b>")
    lines.append("<b>━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━</b>")
    lines.append("")
    lines.append("<b>Strategy:</b> First 5-min alert per stock/day")
    lines.append("<b>Filter:</b> Alert after 12:00 PM only")
    lines.append("<b>Entry:</b> 2 min after alert")
    lines.append(f"<b>Period:</b> Last 2 months")
    lines.append("")

    # Report each scenario briefly
    scenario_labels = [
        ("scenario1", "1. First Post-12PM (No Results)"),
        ("scenario2", "2. First Post-12PM (With Results)"),
        ("scenario3", "3. All Post-12PM"),
        ("scenario4", "4. All Alerts (Benchmark)"),
    ]

    lines.append("<b>SCENARIO COMPARISON (10min exit)</b>")
    lines.append("<pre>")
    lines.append(f"{'Scenario':<28}{'N':<5}{'Win%':<7}{'AvgP&L':<9}")
    lines.append("-" * 49)

    for key, label in scenario_labels:
        res = all_results[key]
        trades = res[10]['all']
        if trades:
            winners = [t for t in trades if t['profitable']]
            win_rate = len(winners) / len(trades) * 100
            avg_pnl = statistics.mean([t['pnl_pct'] for t in trades])
            lines.append(f"{label:<28}{len(trades):<5}{win_rate:.0f}%{'':<3}{avg_pnl:+.2f}%")
        else:
            lines.append(f"{label:<28}{'0':<5}{'N/A':<7}{'N/A'}")

    lines.append("</pre>")

    # Detailed results for best scenario with data
    for key in ['scenario3', 'scenario4', 'scenario2', 'scenario1']:
        res = all_results[key]
        if len(res[10]['all']) >= 3:
            lines.append("")
            lines.append(f"<b>DETAILED: {dict(scenario_labels)[key]}</b>")
            lines.append("<pre>")
            lines.append(f"{'Exit':<6}{'N':<5}{'Win%':<7}{'AvgP&L':<9}{'Total':<9}")
            lines.append("-" * 36)

            for exit_min in EXIT_TIMES:
                trades = res[exit_min]['all']
                if not trades:
                    continue
                winners = [t for t in trades if t['profitable']]
                all_pnl = [t['pnl_pct'] for t in trades]
                win_rate = len(winners) / len(trades) * 100
                avg_pnl = statistics.mean(all_pnl)
                total_pnl = sum(all_pnl)
                lines.append(f"{exit_min}min {len(trades):<5}{win_rate:.0f}%{'':<3}{avg_pnl:+.2f}%{'':<3}{total_pnl:+.1f}%")

            lines.append("</pre>")

            # Direction breakdown
            for direction in ['DROP', 'RISE']:
                dir_trades = res[10][direction]
                if not dir_trades:
                    continue
                dir_winners = [t for t in dir_trades if t['profitable']]
                dir_win_rate = len(dir_winners) / len(dir_trades) * 100
                dir_avg_pnl = statistics.mean([t['pnl_pct'] for t in dir_trades])
                emoji = "🔴" if direction == "DROP" else "🟢"
                lines.append(f"{emoji} {direction}: {len(dir_trades)} trades, {dir_win_rate:.0f}% win, {dir_avg_pnl:+.2f}% avg")

            break

    # Verdict
    lines.append("")
    lines.append("<b>VERDICT</b>")

    best_res = None
    for key in ['scenario3', 'scenario4']:
        if all_results[key][10]['all']:
            best_res = all_results[key]
            break

    if best_res and best_res[10]['all']:
        trades = best_res[10]['all']
        winners = [t for t in trades if t['profitable']]
        win_rate = len(winners) / len(trades) * 100
        avg_pnl = statistics.mean([t['pnl_pct'] for t in trades])

        if win_rate >= 55 and avg_pnl > 0:
            lines.append("✅ WORTH TAKING - Strategy has edge")
        elif win_rate >= 50 and avg_pnl > 0:
            lines.append("⚠️ MARGINAL - Use with risk mgmt")
        elif win_rate >= 45:
            lines.append("⚠️ WEAK - Need additional filters")
        else:
            lines.append("❌ AVOID - Strategy not profitable")

        lines.append(f"Win Rate: {win_rate:.0f}% | Avg P&L: {avg_pnl:+.2f}%")

        if len(trades) < 20:
            lines.append(f"⚠️ Small sample ({len(trades)} trades)")

        # Top trades
        sorted_trades = sorted(trades, key=lambda x: -x['pnl_pct'])
        if len(sorted_trades) >= 3:
            lines.append("")
            lines.append("<b>Best trades:</b>")
            for t in sorted_trades[:3]:
                lines.append(f"  {t['symbol']} {t['direction']} {t['date']} {t['pnl_pct']:+.2f}%")
            lines.append("<b>Worst trades:</b>")
            for t in sorted_trades[-3:]:
                lines.append(f"  {t['symbol']} {t['direction']} {t['date']} {t['pnl_pct']:+.2f}%")
    else:
        lines.append("❌ Insufficient data for conclusion")

    lines.append("")
    lines.append(f"<i>Generated: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}</i>")

    return "\n".join(lines)


def send_telegram_report(message: str) -> bool:
    """Send the report to Telegram channel."""
    bot_token = config.TELEGRAM_BOT_TOKEN
    channel_id = config.TELEGRAM_CHANNEL_ID

    if not bot_token or not channel_id:
        print("Telegram credentials not configured!")
        return False

    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"

    # Split if needed (4096 char limit)
    if len(message) <= 4096:
        chunks = [message]
    else:
        chunks = []
        current = ""
        for line in message.split("\n"):
            if len(current) + len(line) + 1 > 4000:
                chunks.append(current)
                current = line
            else:
                current += "\n" + line if current else line
        if current:
            chunks.append(current)

    success = True
    for chunk in chunks:
        payload = {
            "chat_id": channel_id,
            "text": chunk,
            "parse_mode": "HTML"
        }

        try:
            response = requests.post(url, json=payload, timeout=10)
            response.raise_for_status()
            print(f"Telegram message sent ({len(chunk)} chars)")
        except Exception as e:
            print(f"Failed to send Telegram message: {e}")
            success = False

    return success


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Backtest 5-min first alerts post 12 PM")
    parser.add_argument("--days", type=int, default=60, help="Days to analyze (default: 60)")
    parser.add_argument("--send-telegram", action="store_true", help="Send report to Telegram")
    parser.add_argument("--noon-hour", type=int, default=12, help="Cutoff hour (default: 12)")
    args = parser.parse_args()

    NOON_CUTOFF_HOUR = args.noon_hour

    all_results = run_backtest(args.days)

    if all_results and args.send_telegram:
        print("\nSending report to Telegram...")
        report = generate_telegram_report(all_results)
        if report:
            send_telegram_report(report)
    elif all_results:
        print("\nTo send to Telegram, run with --send-telegram flag")
