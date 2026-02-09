#!/usr/bin/env python3
"""
Backtest Early Warning Detector

Tests the early warning system against historical data to measure:
1. True Positives: Pre-alert fired AND 5-min alert followed
2. False Positives: Pre-alert would fire but NO 5-min alert followed
3. False Negatives: 5-min alert fired but NO pre-alert preceded it

Usage:
    python backtest_early_warning.py --days 30

Author: Claude Opus 4.5
Date: 2026-02-09
"""

import argparse
import logging
import sqlite3
from datetime import datetime, timedelta, time as dt_time
from typing import Dict, List, Set, Tuple
from collections import defaultdict

import openpyxl

import config

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class EarlyWarningBacktester:
    """
    Backtests the early warning system to measure accuracy and false positive rate.
    """

    def __init__(self, excel_path: str = None, db_path: str = None):
        self.excel_path = excel_path or config.ALERT_EXCEL_PATH
        self.db_path = db_path or "data/central_quotes.db"
        self.conn = None

        # Early warning thresholds (from config)
        self.ew_threshold = config.EARLY_WARNING_THRESHOLD  # 1.0%
        self.ew_lookback = config.EARLY_WARNING_LOOKBACK    # 3 min
        self.ew_volume_mult = config.EARLY_WARNING_VOLUME_MULT  # 1.2x

        # Filter flags (from config)
        self.require_obv = getattr(config, 'EARLY_WARNING_REQUIRE_OBV', True)
        self.require_oi = getattr(config, 'EARLY_WARNING_REQUIRE_OI', True)
        self.require_rsi = getattr(config, 'EARLY_WARNING_REQUIRE_RSI', True)
        self.require_vwap = getattr(config, 'EARLY_WARNING_REQUIRE_VWAP', True)

        # 5-min alert thresholds
        self.alert_threshold = config.DROP_THRESHOLD_5MIN  # 1.25%
        self.alert_volume_mult = 1.25  # From RapidAlertDetector

        # Cache for VWAP calculations
        self._vwap_cache = {}  # (symbol, date) -> vwap

    def _get_db_connection(self) -> sqlite3.Connection:
        if self.conn is None:
            self.conn = sqlite3.connect(self.db_path, timeout=30)
        return self.conn

    def load_5min_alerts(self, days: int = 30) -> List[Dict]:
        """Load 5-min alerts from Excel."""
        if not self.excel_path:
            return []

        try:
            workbook = openpyxl.load_workbook(self.excel_path, read_only=True)
        except Exception as e:
            logger.error(f"Failed to load Excel: {e}")
            return []

        if "5min_alerts" not in workbook.sheetnames:
            workbook.close()
            return []

        sheet = workbook["5min_alerts"]
        cutoff_date = datetime.now() - timedelta(days=days)

        alerts = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            try:
                date_val = row[0]
                time_val = row[1]

                if isinstance(date_val, datetime):
                    date_str = date_val.strftime("%Y-%m-%d")
                else:
                    date_str = str(date_val)

                if isinstance(time_val, datetime):
                    time_str = time_val.strftime("%H:%M:%S")
                else:
                    time_str = str(time_val)

                try:
                    alert_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    alert_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")

                if alert_dt < cutoff_date:
                    continue

                # Parse volume multiplier (may have 'x' suffix)
                vol_mult_raw = row[10]
                if vol_mult_raw:
                    vol_mult_str = str(vol_mult_raw).replace('x', '').strip()
                    try:
                        vol_mult = float(vol_mult_str)
                    except ValueError:
                        vol_mult = 0
                else:
                    vol_mult = 0

                alerts.append({
                    'datetime': alert_dt,
                    'symbol': str(row[2]) if row[2] else '',
                    'direction': str(row[3]) if row[3] else 'Drop',
                    'change_pct': float(row[6]) if row[6] else 0,
                    'volume_multiplier': vol_mult,
                })

            except Exception as e:
                continue

        workbook.close()
        logger.info(f"Loaded {len(alerts)} 5-min alerts from last {days} days")
        return alerts

    def get_unique_trading_periods(self, days: int = 30) -> List[Tuple[str, str]]:
        """Get unique (date, symbol) pairs from the database for scanning."""
        conn = self._get_db_connection()
        cursor = conn.cursor()

        cutoff = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')

        cursor.execute("""
            SELECT DISTINCT date(timestamp), symbol
            FROM stock_quotes
            WHERE timestamp >= ?
            ORDER BY timestamp
        """, (cutoff,))

        periods = [(row[0], row[1]) for row in cursor.fetchall()]
        logger.info(f"Found {len(periods)} unique (date, symbol) periods to scan")
        return periods

    def get_day_data(self, symbol: str, date_str: str) -> List[Dict]:
        """Get all minute data for a symbol on a specific date."""
        conn = self._get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT timestamp, price, volume
            FROM stock_quotes
            WHERE symbol = ?
            AND date(timestamp) = ?
            AND time(timestamp) >= '09:25:00'
            AND time(timestamp) <= '15:30:00'
            ORDER BY timestamp ASC
        """, (symbol, date_str))

        data = []
        for row in cursor.fetchall():
            data.append({
                'timestamp': row[0],
                'price': row[1],
                'volume': row[2] or 0
            })

        return data

    def calculate_obv(self, data: List[Dict]) -> List[float]:
        """Calculate OBV for price/volume data."""
        if len(data) < 2:
            return []
        obv = [0]
        for i in range(1, len(data)):
            prev_price = data[i-1]['price']
            curr_price = data[i]['price']
            volume = data[i]['volume'] or 0
            if curr_price > prev_price:
                obv.append(obv[-1] + volume)
            elif curr_price < prev_price:
                obv.append(obv[-1] - volume)
            else:
                obv.append(obv[-1])
        return obv

    def check_obv_confirmation(self, data: List[Dict], idx: int, direction: str) -> bool:
        """Check if OBV confirms price direction at given index."""
        if idx < 5:
            return True  # Not enough data

        # Get OBV up to this point
        slice_data = data[:idx+1]
        obv = self.calculate_obv(slice_data)

        if len(obv) < 5:
            return True

        # Calculate OBV slope over last 5 points
        recent = obv[-5:]
        n = len(recent)
        x_mean = (n - 1) / 2
        y_mean = sum(recent) / n

        num = sum((i - x_mean) * (y - y_mean) for i, y in enumerate(recent))
        den = sum((i - x_mean) ** 2 for i in range(n))
        slope = num / den if den != 0 else 0

        # Normalize
        avg_change = sum(abs(obv[i] - obv[i-1]) for i in range(1, len(obv))) / (len(obv) - 1) if len(obv) > 1 else 1
        normalized = slope / avg_change if avg_change > 0 else 0

        # Check confirmation
        obv_up = normalized > 0.1
        obv_down = normalized < -0.1

        if direction == 'drop':
            return obv_down  # OBV should be falling for drop confirmation
        else:
            return obv_up  # OBV should be rising for rise confirmation

    def check_oi_confirmation(self, symbol: str, date_str: str, direction: str) -> bool:
        """Check if OI pattern confirms the move direction."""
        try:
            conn = self._get_db_connection()
            cursor = conn.cursor()

            # Get day-start OI
            cursor.execute("""
                SELECT oi FROM stock_quotes
                WHERE symbol = ? AND date(timestamp) = ?
                AND time(timestamp) >= '09:15:00'
                ORDER BY timestamp ASC
                LIMIT 1
            """, (symbol, date_str))
            row = cursor.fetchone()
            if not row or not row[0]:
                return True  # Allow if no OI data

            day_start_oi = row[0]

            # Get latest OI
            cursor.execute("""
                SELECT oi FROM stock_quotes
                WHERE symbol = ? AND date(timestamp) = ?
                ORDER BY timestamp DESC
                LIMIT 1
            """, (symbol, date_str))
            row = cursor.fetchone()
            if not row or not row[0]:
                return True

            current_oi = row[0]
            oi_change_pct = ((current_oi - day_start_oi) / day_start_oi) * 100

            oi_increasing = oi_change_pct > 0.5
            oi_decreasing = oi_change_pct < -0.5

            if direction == 'drop':
                # For drops: Short Buildup (OI ↑) = strong
                # Long Unwinding (OI ↓) = weak, reject
                if oi_decreasing:
                    return False
            else:
                # For rises: Long Buildup (OI ↑) = strong
                # Short Covering (OI ↓) = weak, reject
                if oi_decreasing:
                    return False

            return True
        except:
            return True

    def check_rsi_momentum(self, data: List[Dict], idx: int, direction: str) -> bool:
        """Check if RSI has room to continue in the direction."""
        if idx < 10:
            return True  # Not enough data

        # Get prices for RSI calculation
        prices = [d['price'] for d in data[max(0, idx-14):idx+1] if d.get('price')]
        if len(prices) < 10:
            return True

        # Calculate RSI(9)
        gains, losses = [], []
        for i in range(1, len(prices)):
            change = prices[i] - prices[i-1]
            if change > 0:
                gains.append(change)
                losses.append(0)
            else:
                gains.append(0)
                losses.append(abs(change))

        period = min(9, len(gains))
        avg_gain = sum(gains[-period:]) / period
        avg_loss = sum(losses[-period:]) / period

        if avg_loss == 0:
            rsi = 100.0
        else:
            rs = avg_gain / avg_loss
            rsi = 100 - (100 / (1 + rs))

        if direction == 'drop':
            # For drops: RSI should have room to fall (not already oversold)
            return rsi > 40
        else:
            # For rises: RSI should have room to rise (not already overbought)
            return rsi < 60

    def calculate_vwap(self, symbol: str, date_str: str, data: List[Dict], idx: int) -> float:
        """Calculate VWAP up to current index."""
        cache_key = (symbol, date_str, idx)
        if cache_key in self._vwap_cache:
            return self._vwap_cache[cache_key]

        total_pv = 0
        total_volume = 0

        for i in range(idx + 1):
            price = data[i].get('price', 0)
            volume = data[i].get('volume', 0)
            if price and volume and price > 0 and volume > 0:
                total_pv += price * volume
                total_volume += volume

        if total_volume == 0:
            return 0

        vwap = total_pv / total_volume
        self._vwap_cache[cache_key] = vwap
        return vwap

    def check_vwap_position(self, symbol: str, date_str: str, data: List[Dict],
                            idx: int, direction: str) -> bool:
        """Check if price position relative to VWAP confirms the move."""
        vwap = self.calculate_vwap(symbol, date_str, data, idx)
        if vwap == 0:
            return True  # Allow if can't calculate

        current_price = data[idx].get('price', 0)
        if not current_price:
            return True

        if direction == 'drop':
            # For drops: Price should be below VWAP (weakness)
            return current_price <= vwap * 1.002
        else:
            # For rises: Price should be above VWAP (strength)
            return current_price >= vwap * 0.998

    def check_early_warning_signal(self, data: List[Dict], idx: int,
                                    symbol: str = None, date_str: str = None) -> Tuple[bool, str, float]:
        """
        Check if early warning signal would fire at given index.

        Matches the actual EarlyWarningDetector logic including volume check.

        Returns: (would_fire, direction, change_pct)
        """
        if idx < self.ew_lookback:
            return False, None, 0

        current = data[idx]
        prev = data[idx - self.ew_lookback]

        current_price = current['price']
        prev_price = prev['price']
        current_vol = current['volume']
        prev_vol = prev['volume']

        if not current_price or not prev_price or current_price <= 0 or prev_price <= 0:
            return False, None, 0

        # Volume check - REQUIRED (matches early_warning_detector.py)
        # Require volume to be at least ew_volume_mult (1.1x) of lookback period
        has_volume = False
        if prev_vol > 0 and current_vol > 0:
            vol_ratio = current_vol / prev_vol
            has_volume = vol_ratio >= self.ew_volume_mult
        else:
            # If we can't calculate volume ratio, skip
            return False, None, 0

        if not has_volume:
            return False, None, 0

        # Price change
        drop_pct = ((prev_price - current_price) / prev_price) * 100
        rise_pct = ((current_price - prev_price) / prev_price) * 100

        if drop_pct >= self.ew_threshold:
            # Filter 1: OBV confirmation
            if self.require_obv and not self.check_obv_confirmation(data, idx, 'drop'):
                return False, None, 0
            # Filter 2: OI pattern
            if self.require_oi and symbol and date_str:
                if not self.check_oi_confirmation(symbol, date_str, 'drop'):
                    return False, None, 0
            # Filter 3: RSI momentum
            if self.require_rsi and not self.check_rsi_momentum(data, idx, 'drop'):
                return False, None, 0
            # Filter 4: VWAP position
            if self.require_vwap and symbol and date_str:
                if not self.check_vwap_position(symbol, date_str, data, idx, 'drop'):
                    return False, None, 0
            return True, 'drop', drop_pct

        if rise_pct >= self.ew_threshold:
            # Filter 1: OBV confirmation
            if self.require_obv and not self.check_obv_confirmation(data, idx, 'rise'):
                return False, None, 0
            # Filter 2: OI pattern
            if self.require_oi and symbol and date_str:
                if not self.check_oi_confirmation(symbol, date_str, 'rise'):
                    return False, None, 0
            # Filter 3: RSI momentum
            if self.require_rsi and not self.check_rsi_momentum(data, idx, 'rise'):
                return False, None, 0
            # Filter 4: VWAP position
            if self.require_vwap and symbol and date_str:
                if not self.check_vwap_position(symbol, date_str, data, idx, 'rise'):
                    return False, None, 0
            return True, 'rise', rise_pct

        return False, None, 0

    def check_5min_alert_follows(self, data: List[Dict], idx: int, direction: str,
                                  max_minutes: int = 5) -> Tuple[bool, float, int]:
        """
        Check if a 5-min alert would fire within next N minutes.

        Returns: (alert_fired, change_pct, minutes_later)
        """
        if idx < 5:  # Need 5 minutes of history for 5-min alert
            return False, 0, 0

        # Check each minute in the window
        for offset in range(1, max_minutes + 1):
            future_idx = idx + offset
            if future_idx >= len(data):
                break

            # For the 5-min alert, we need to compare current to 5 minutes ago
            if future_idx < 5:
                continue

            current = data[future_idx]
            past = data[future_idx - 5]

            current_price = current['price']
            past_price = past['price']
            current_vol = current['volume']
            past_vol = past['volume']

            if not current_price or not past_price or current_price <= 0 or past_price <= 0:
                continue

            # Volume spike check
            has_volume_spike = False
            if past_vol > 0 and current_vol > 0:
                vol_ratio = current_vol / past_vol
                has_volume_spike = vol_ratio >= self.alert_volume_mult

            if not has_volume_spike:
                continue

            # Price change
            if direction == 'drop':
                change_pct = ((past_price - current_price) / past_price) * 100
                if change_pct >= self.alert_threshold:
                    return True, change_pct, offset
            else:
                change_pct = ((current_price - past_price) / past_price) * 100
                if change_pct >= self.alert_threshold:
                    return True, change_pct, offset

        return False, 0, 0

    def run_backtest(self, days: int = 30) -> Dict:
        """
        Run full backtest.

        Scans all data and counts:
        - True Positives: EW signal AND 5-min alert followed
        - False Positives: EW signal but NO 5-min alert followed
        - False Negatives: 5-min alert but NO EW signal preceded it
        """
        # Build filter list
        filters_enabled = []
        if self.require_obv: filters_enabled.append('OBV')
        if self.require_oi: filters_enabled.append('OI')
        if self.require_rsi: filters_enabled.append('RSI')
        if self.require_vwap: filters_enabled.append('VWAP')

        logger.info(f"Starting backtest for last {days} days...")
        logger.info(f"Early Warning: {self.ew_threshold}% in {self.ew_lookback} min")
        logger.info(f"Filters enabled: {'+'.join(filters_enabled) if filters_enabled else 'none'}")
        logger.info(f"5-Min Alert: {self.alert_threshold}% in 5 min with {self.alert_volume_mult}x volume")

        # Load actual 5-min alerts for false negative check
        actual_alerts = self.load_5min_alerts(days=days)
        actual_alert_set = set()
        for a in actual_alerts:
            # Key: symbol_date_hour_minute
            key = f"{a['symbol']}_{a['datetime'].strftime('%Y-%m-%d_%H:%M')}"
            actual_alert_set.add(key)

        # Get unique dates from database
        conn = self._get_db_connection()
        cursor = conn.cursor()
        cutoff = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')

        cursor.execute("""
            SELECT DISTINCT date(timestamp)
            FROM stock_quotes
            WHERE timestamp >= ?
            ORDER BY timestamp
        """, (cutoff,))

        dates = [row[0] for row in cursor.fetchall()]
        logger.info(f"Scanning {len(dates)} trading days...")

        # Results
        true_positives = []  # EW fired + alert followed
        false_positives = []  # EW fired but no alert
        ew_signals_by_symbol = defaultdict(list)

        # Track cooldowns to avoid counting same signal multiple times
        last_signal_time = {}  # symbol -> last signal timestamp

        total_minutes_scanned = 0
        symbols_scanned = set()

        for date_idx, date_str in enumerate(dates):
            if date_idx % 5 == 0:
                logger.info(f"Processing date {date_idx + 1}/{len(dates)}: {date_str}")

            # Get all symbols for this date
            cursor.execute("""
                SELECT DISTINCT symbol
                FROM stock_quotes
                WHERE date(timestamp) = ?
            """, (date_str,))

            symbols = [row[0] for row in cursor.fetchall()]

            for symbol in symbols:
                symbols_scanned.add(symbol)
                data = self.get_day_data(symbol, date_str)

                if len(data) < 10:
                    continue

                total_minutes_scanned += len(data)

                # Scan each minute
                for idx in range(self.ew_lookback, len(data)):
                    # Check cooldown (15 min between signals for same symbol)
                    current_ts = data[idx]['timestamp']
                    cooldown_key = symbol

                    if cooldown_key in last_signal_time:
                        last_ts = last_signal_time[cooldown_key]
                        try:
                            last_dt = datetime.strptime(last_ts, '%Y-%m-%d %H:%M:%S')
                            current_dt = datetime.strptime(current_ts, '%Y-%m-%d %H:%M:%S')
                            if (current_dt - last_dt).total_seconds() < 900:  # 15 min
                                continue
                        except:
                            pass

                    # Check if EW signal would fire
                    signal_fires, direction, change_pct = self.check_early_warning_signal(
                        data, idx, symbol=symbol, date_str=date_str
                    )

                    if not signal_fires:
                        continue

                    # Signal fired - check if 5-min alert follows
                    alert_follows, alert_pct, minutes_later = self.check_5min_alert_follows(
                        data, idx, direction, max_minutes=5
                    )

                    # Record signal
                    signal_record = {
                        'symbol': symbol,
                        'timestamp': current_ts,
                        'direction': direction,
                        'ew_change_pct': change_pct,
                        'alert_followed': alert_follows,
                        'alert_pct': alert_pct if alert_follows else 0,
                        'minutes_to_alert': minutes_later if alert_follows else 0
                    }

                    if alert_follows:
                        true_positives.append(signal_record)
                    else:
                        false_positives.append(signal_record)

                    ew_signals_by_symbol[symbol].append(signal_record)

                    # Update cooldown
                    last_signal_time[cooldown_key] = current_ts

        # Calculate metrics
        total_signals = len(true_positives) + len(false_positives)
        precision = len(true_positives) / total_signals if total_signals > 0 else 0
        false_positive_rate = len(false_positives) / total_signals if total_signals > 0 else 0

        # Find false negatives (alerts without preceding EW signal)
        # This is harder - need to check each actual alert
        false_negatives = 0
        for alert in actual_alerts:
            symbol = alert['symbol']
            alert_time = alert['datetime']

            # Check if there was an EW signal 1-5 minutes before
            found_preceding_signal = False
            for sig in ew_signals_by_symbol.get(symbol, []):
                try:
                    sig_time = datetime.strptime(sig['timestamp'], '%Y-%m-%d %H:%M:%S')
                    delta = (alert_time - sig_time).total_seconds() / 60
                    if 1 <= delta <= 5:
                        found_preceding_signal = True
                        break
                except:
                    continue

            if not found_preceding_signal:
                false_negatives += 1

        recall = (len(actual_alerts) - false_negatives) / len(actual_alerts) if actual_alerts else 0

        results = {
            'days_analyzed': days,
            'dates_scanned': len(dates),
            'symbols_scanned': len(symbols_scanned),
            'total_minutes_scanned': total_minutes_scanned,
            'actual_5min_alerts': len(actual_alerts),
            'total_ew_signals': total_signals,
            'true_positives': len(true_positives),
            'false_positives': len(false_positives),
            'false_negatives': false_negatives,
            'precision': precision,
            'recall': recall,
            'false_positive_rate': false_positive_rate,
            'true_positive_details': true_positives[:20],  # Sample
            'false_positive_details': false_positives[:20],  # Sample
        }

        return results

    def print_results(self, results: Dict):
        """Print backtest results."""
        print("\n" + "=" * 70)
        print("EARLY WARNING BACKTEST RESULTS")
        print("=" * 70)

        print(f"\nAnalysis Period: Last {results['days_analyzed']} days")
        print(f"Trading Days Scanned: {results['dates_scanned']}")
        print(f"Symbols Scanned: {results['symbols_scanned']}")
        print(f"Total Minutes Analyzed: {results['total_minutes_scanned']:,}")

        print("\n" + "-" * 70)
        print("SIGNAL COUNTS")
        print("-" * 70)
        print(f"Actual 5-Min Alerts (from Excel): {results['actual_5min_alerts']}")
        print(f"Early Warning Signals Generated:  {results['total_ew_signals']}")
        print()
        print(f"  True Positives (EW → Alert):    {results['true_positives']}")
        print(f"  False Positives (EW → No Alert): {results['false_positives']}")
        print(f"  False Negatives (Alert → No EW): {results['false_negatives']}")

        print("\n" + "-" * 70)
        print("ACCURACY METRICS")
        print("-" * 70)
        print(f"Precision (TP / All Signals):     {results['precision']:.1%}")
        print(f"  → When EW fires, {results['precision']:.1%} lead to actual alerts")
        print()
        print(f"False Positive Rate:              {results['false_positive_rate']:.1%}")
        print(f"  → {results['false_positive_rate']:.1%} of signals are false alarms")
        print()
        print(f"Recall (TP / All Alerts):         {results['recall']:.1%}")
        print(f"  → EW catches {results['recall']:.1%} of 5-min alerts")

        # Signal rate
        if results['total_minutes_scanned'] > 0:
            signals_per_day = results['total_ew_signals'] / results['dates_scanned'] if results['dates_scanned'] > 0 else 0
            print(f"\nSignals per Day:                  {signals_per_day:.1f}")

        print("\n" + "-" * 70)
        print("SAMPLE TRUE POSITIVES (EW → Alert worked)")
        print("-" * 70)
        for tp in results.get('true_positive_details', [])[:5]:
            print(f"  {tp['symbol']} @ {tp['timestamp']}: "
                  f"{tp['direction'].upper()} {tp['ew_change_pct']:.2f}% → "
                  f"Alert {tp['alert_pct']:.2f}% in {tp['minutes_to_alert']} min")

        print("\n" + "-" * 70)
        print("SAMPLE FALSE POSITIVES (EW → No Alert)")
        print("-" * 70)
        for fp in results.get('false_positive_details', [])[:5]:
            print(f"  {fp['symbol']} @ {fp['timestamp']}: "
                  f"{fp['direction'].upper()} {fp['ew_change_pct']:.2f}% → No alert followed")

        print("\n" + "=" * 70)

        # Recommendation
        print("\nRECOMMENDATION:")
        if results['precision'] >= 0.5:
            print("✅ Early warning is USEFUL - majority of signals lead to alerts")
        elif results['precision'] >= 0.3:
            print("⚠️ Early warning has MODERATE accuracy - consider tightening threshold")
        else:
            print("❌ Too many false positives - need to increase threshold or add filters")

        if results['false_positive_rate'] > 0.7:
            print(f"   Consider raising threshold from {config.EARLY_WARNING_THRESHOLD}% to 0.6-0.7%")

        print()

    def close(self):
        if self.conn:
            self.conn.close()
            self.conn = None


def main():
    parser = argparse.ArgumentParser(description="Backtest Early Warning Detector")
    parser.add_argument("--days", type=int, default=30, help="Days to analyze (default: 30)")

    args = parser.parse_args()

    backtester = EarlyWarningBacktester()

    try:
        results = backtester.run_backtest(days=args.days)
        backtester.print_results(results)
    finally:
        backtester.close()


if __name__ == "__main__":
    main()
