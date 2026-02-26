#!/usr/bin/env python3
"""
Alert P&L Tracker - Futures Simulation

Simulates futures trades on alerts to validate the alert system with real prices.
When a 5-min alert or pre-alert fires, simulates entering 1 futures lot 2 minutes later,
then tracks P&L at 15 min and 30 min exit windows.

Trade Logic:
- DROP/SELL alert -> SHORT 1 futures lot (profit if price falls)
- RISE/BUY alert -> LONG 1 futures lot (profit if price rises)
- P&L in both percentage AND absolute Rupees (lot_size * price_diff)

Filters:
- Only first alerts per stock (alert_count == 1)
- Only post-12 PM alerts (alert_time.hour >= 12)

Author: Claude Opus 4.6
Date: 2026-02-21
"""

import logging
import os
import fcntl
from datetime import datetime, date, timedelta
from typing import Dict, List, Optional
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import config

logger = logging.getLogger(__name__)


class AlertPnLTracker:
    """
    Tracks simulated futures P&L for alerts using real market prices
    from the central quote database.
    """

    # Excel headers for Daily_PnL sheet
    HEADERS = [
        "Date", "Time", "Symbol", "Type", "Direction", "Lot Size",
        "Alert Price", "Entry (T+2)", "Exit 15m", "Exit 30m",
        "P&L % 15m", "P&L % 30m", "P&L Rs 15m", "P&L Rs 30m"
    ]

    def __init__(self, central_db, telegram, kite_client):
        """
        Initialize Alert P&L Tracker.

        Args:
            central_db: CentralQuoteDB instance (reader mode)
            telegram: TelegramNotifier instance
            kite_client: Kite Connect client for loading lot sizes
        """
        self.db = central_db
        self.telegram = telegram
        self._lot_sizes: Dict[str, int] = {}
        self._pending_trades: List[Dict] = []
        self._completed_trades: List[Dict] = []
        self._current_date: Optional[date] = None
        self._report_sent_today: bool = False
        self.excel_path = config.ALERT_PNL_EXCEL_PATH

        # Ensure directory exists
        Path(self.excel_path).parent.mkdir(parents=True, exist_ok=True)

        # Load lot sizes from NFO instruments
        self._load_lot_sizes(kite_client)

        logger.info(f"AlertPnLTracker initialized (lot_sizes: {len(self._lot_sizes)} symbols, "
                   f"excel: {self.excel_path})")

    def _load_lot_sizes(self, kite_client):
        """
        Load futures lot sizes from Kite instruments.

        Filters for instrument_type == 'FUT', groups by symbol name,
        picks nearest expiry, and stores lot_size per symbol.
        """
        try:
            instruments = kite_client.instruments("NFO")
            if not instruments:
                logger.warning("PnLTracker: No NFO instruments returned")
                return

            # Group FUT instruments by name, pick nearest expiry
            fut_by_name: Dict[str, Dict] = {}
            today = date.today()

            for inst in instruments:
                if inst.get('instrument_type') != 'FUT':
                    continue

                name = inst.get('name', '')
                expiry = inst.get('expiry')
                lot_size = inst.get('lot_size', 0)

                if not name or not expiry or not lot_size:
                    continue

                # Convert expiry to date if needed
                if isinstance(expiry, str):
                    try:
                        expiry = datetime.strptime(expiry, '%Y-%m-%d').date()
                    except ValueError:
                        continue

                # Skip expired contracts
                if expiry < today:
                    continue

                # Keep nearest expiry per symbol
                if name not in fut_by_name or expiry < fut_by_name[name]['expiry']:
                    fut_by_name[name] = {
                        'expiry': expiry,
                        'lot_size': lot_size
                    }

            # Store lot sizes
            for name, data in fut_by_name.items():
                self._lot_sizes[name] = data['lot_size']

            logger.info(f"PnLTracker: Loaded {len(self._lot_sizes)} lot sizes from NFO instruments")

        except Exception as e:
            logger.error(f"PnLTracker: Failed to load lot sizes: {e}")

    def record_alerts(self, alerted_symbols: List[Dict]):
        """
        Record alerts for P&L tracking.

        Called after each detector's detect_all() returns.

        Args:
            alerted_symbols: List of dicts with keys:
                symbol, direction, price, time, alert_type, alert_count
        """
        # Daily reset check
        self._check_daily_reset()

        now = datetime.now()

        for alert in alerted_symbols:
            try:
                symbol = alert.get('symbol', '')
                alert_count = alert.get('alert_count', 0)
                alert_time_str = alert.get('time', '')

                # Filter: only first alerts
                if alert_count != 1:
                    continue

                # Filter: only post-12 PM
                try:
                    alert_time = datetime.fromisoformat(alert_time_str)
                except (ValueError, TypeError):
                    alert_time = now

                if alert_time.hour < 12:
                    continue

                # Get lot size for this symbol
                lot_size = self._lot_sizes.get(symbol, 0)
                if lot_size == 0:
                    logger.debug(f"PnLTracker: No lot size for {symbol}, skipping")
                    continue

                # Check for duplicate (same symbol + alert_type already pending today)
                alert_type = alert.get('alert_type', '5min')
                already_tracked = any(
                    t['symbol'] == symbol and t['date'] == now.strftime('%Y-%m-%d')
                    and t['alert_type'] == alert_type
                    for t in self._pending_trades + self._completed_trades
                )
                if already_tracked:
                    continue

                # Create pending trade
                trade = {
                    'date': now.strftime('%Y-%m-%d'),
                    'time': alert_time.strftime('%H:%M'),
                    'symbol': symbol,
                    'alert_type': alert.get('alert_type', '5min'),
                    'direction': alert.get('direction', 'drop'),
                    'alert_price': alert.get('price', 0),
                    'alert_timestamp': alert_time,
                    'lot_size': lot_size,
                    'entry_price': None,
                    'exit_price_15m': None,
                    'exit_price_30m': None,
                    'pnl_pct_15m': None,
                    'pnl_pct_30m': None,
                    'pnl_rs_15m': None,
                    'pnl_rs_30m': None,
                    'status': 'pending'
                }

                self._pending_trades.append(trade)
                logger.info(f"PnLTracker: Recorded alert for {symbol} "
                           f"({alert.get('direction', '?')} {alert.get('alert_type', '?')}, "
                           f"lot={lot_size})")

            except Exception as e:
                logger.error(f"PnLTracker: Error recording alert: {e}")

    def process_pending_prices(self):
        """
        Process pending trades to fill entry/exit prices.

        Called every cycle (1 min) from the main loop.
        """
        if not self._pending_trades:
            return

        now = datetime.now()
        completed_indices = []

        for i, trade in enumerate(self._pending_trades):
            try:
                alert_time = trade['alert_timestamp']
                elapsed = (now - alert_time).total_seconds() / 60  # minutes

                symbol = trade['symbol']

                # Fill entry price at T+2
                if elapsed >= 2 and trade['entry_price'] is None:
                    entry_time = alert_time + timedelta(minutes=2)
                    entry_ts = entry_time.strftime('%Y-%m-%d %H:%M:00')
                    price = self.db.get_stock_price_at_time(symbol, entry_ts)
                    if price:
                        trade['entry_price'] = price
                        trade['status'] = 'entry_filled'
                        logger.info(f"PnLTracker: {symbol} entry filled at {price:.2f} (T+2)")

                # Fill 15m exit price
                if elapsed >= 15 and trade['exit_price_15m'] is None and trade['entry_price'] is not None:
                    exit_time = alert_time + timedelta(minutes=15)
                    exit_ts = exit_time.strftime('%Y-%m-%d %H:%M:00')
                    price = self.db.get_stock_price_at_time(symbol, exit_ts)
                    if price:
                        trade['exit_price_15m'] = price
                        self._compute_pnl(trade, '15m')
                        logger.info(f"PnLTracker: {symbol} 15m exit at {price:.2f}, "
                                   f"P&L: Rs{trade['pnl_rs_15m']:+,.0f} ({trade['pnl_pct_15m']:+.2f}%)")

                # Fill 30m exit price
                if elapsed >= 30 and trade['exit_price_30m'] is None and trade['entry_price'] is not None:
                    exit_time = alert_time + timedelta(minutes=30)
                    exit_ts = exit_time.strftime('%Y-%m-%d %H:%M:00')
                    price = self.db.get_stock_price_at_time(symbol, exit_ts)
                    if price:
                        trade['exit_price_30m'] = price
                        self._compute_pnl(trade, '30m')
                        trade['status'] = 'completed'
                        logger.info(f"PnLTracker: {symbol} 30m exit at {price:.2f}, "
                                   f"P&L: Rs{trade['pnl_rs_30m']:+,.0f} ({trade['pnl_pct_30m']:+.2f}%)")

                        # Write to Excel and move to completed
                        self._write_to_excel(trade)
                        completed_indices.append(i)

            except Exception as e:
                logger.error(f"PnLTracker: Error processing {trade.get('symbol', '?')}: {e}")

        # Move completed trades
        for i in sorted(completed_indices, reverse=True):
            self._completed_trades.append(self._pending_trades.pop(i))

    def _compute_pnl(self, trade: Dict, exit_type: str):
        """
        Compute P&L for a trade at the given exit window.

        SHORT (drop): pnl = (entry - exit) * lot_size
        LONG (rise): pnl = (exit - entry) * lot_size
        """
        entry = trade['entry_price']
        exit_key = f'exit_price_{exit_type}'
        exit_price = trade[exit_key]

        if entry is None or exit_price is None or entry <= 0:
            return

        lot_size = trade['lot_size']

        if trade['direction'] == 'drop':
            # SHORT: profit when price falls
            pnl_rs = (entry - exit_price) * lot_size
            pnl_pct = ((entry - exit_price) / entry) * 100
        else:
            # LONG: profit when price rises
            pnl_rs = (exit_price - entry) * lot_size
            pnl_pct = ((exit_price - entry) / entry) * 100

        trade[f'pnl_pct_{exit_type}'] = round(pnl_pct, 2)
        trade[f'pnl_rs_{exit_type}'] = round(pnl_rs, 2)

    def send_eod_report(self):
        """
        Send end-of-day P&L report via Telegram.

        Called after the main loop exits (post 3:30 PM).
        Fills missing prices from latest DB data before reporting.
        """
        if self._report_sent_today:
            return

        # Fill any remaining pending trades with latest available prices
        self._fill_remaining_prices()

        all_trades = self._completed_trades + self._pending_trades

        if not all_trades:
            logger.info("PnLTracker: No trades to report today")
            self._report_sent_today = True
            return

        try:
            message = self._build_eod_message(all_trades)
            self._send_telegram_message(message)
            self._report_sent_today = True
            logger.info(f"PnLTracker: EOD report sent ({len(all_trades)} trades)")
        except Exception as e:
            logger.error(f"PnLTracker: Failed to send EOD report: {e}")

    def _fill_remaining_prices(self):
        """Fill missing prices for pending trades using latest available data."""
        for trade in self._pending_trades:
            try:
                symbol = trade['symbol']
                alert_time = trade['alert_timestamp']

                # Fill entry price if missing
                if trade['entry_price'] is None:
                    entry_time = alert_time + timedelta(minutes=2)
                    entry_ts = entry_time.strftime('%Y-%m-%d %H:%M:00')
                    price = self.db.get_stock_price_at_time(symbol, entry_ts)
                    if price:
                        trade['entry_price'] = price

                if trade['entry_price'] is None:
                    continue

                # Fill 15m exit if missing
                if trade['exit_price_15m'] is None:
                    exit_time = alert_time + timedelta(minutes=15)
                    exit_ts = exit_time.strftime('%Y-%m-%d %H:%M:00')
                    price = self.db.get_stock_price_at_time(symbol, exit_ts)
                    if price:
                        trade['exit_price_15m'] = price
                        self._compute_pnl(trade, '15m')

                # Fill 30m exit if missing
                if trade['exit_price_30m'] is None:
                    exit_time = alert_time + timedelta(minutes=30)
                    exit_ts = exit_time.strftime('%Y-%m-%d %H:%M:00')
                    price = self.db.get_stock_price_at_time(symbol, exit_ts)
                    if price:
                        trade['exit_price_30m'] = price
                        self._compute_pnl(trade, '30m')

                # Write to excel if not yet written (at EOD, write even with partial data)
                if trade['status'] != 'completed' and trade.get('entry_price') is not None:
                    trade['status'] = 'completed'
                    self._write_to_excel(trade)

            except Exception as e:
                logger.error(f"PnLTracker: Error filling remaining prices for {trade.get('symbol', '?')}: {e}")

    def _build_eod_message(self, trades: List[Dict]) -> str:
        """Build Telegram EOD report message."""
        today_str = datetime.now().strftime('%d-%b-%Y')
        total_trades = len(trades)

        # Separate by 15m P&L (primary metric)
        profitable_15m = []
        loss_15m = []

        for t in trades:
            if t.get('pnl_rs_15m') is not None:
                if t['pnl_rs_15m'] >= 0:
                    profitable_15m.append(t)
                else:
                    loss_15m.append(t)

        # Build message
        lines = [
            f"\U0001f4ca <b>DAILY ALERT P&L REPORT</b>",
            f"\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501",
            f"\U0001f4c5 {today_str} | Trades: {total_trades}",
            ""
        ]

        # Profitable trades
        if profitable_15m:
            lines.append(f"\U0001f7e2 <b>PROFITABLE ({len(profitable_15m)}):</b>")
            for t in sorted(profitable_15m, key=lambda x: x.get('pnl_rs_15m', 0), reverse=True):
                dir_arrow = "\u2193" if t['direction'] == 'drop' else "\u2191"
                type_short = "5m" if t['alert_type'] == '5min' else "pre"
                entry = t.get('entry_price', 0)
                exit_15 = t.get('exit_price_15m', 0)
                pnl_rs = t.get('pnl_rs_15m', 0)
                pnl_pct = t.get('pnl_pct_15m', 0)
                lot = t.get('lot_size', 0)
                lines.append(
                    f"  {t['symbol']} {dir_arrow}{type_short} | "
                    f"1 lot({lot}) | "
                    f"\u20b9{entry:.0f}\u2192\u20b9{exit_15:.0f} | "
                    f"+\u20b9{pnl_rs:,.0f} (+{pnl_pct:.2f}%)"
                )
            lines.append("")

        # Loss trades
        if loss_15m:
            lines.append(f"\U0001f534 <b>LOSS ({len(loss_15m)}):</b>")
            for t in sorted(loss_15m, key=lambda x: x.get('pnl_rs_15m', 0)):
                dir_arrow = "\u2193" if t['direction'] == 'drop' else "\u2191"
                type_short = "5m" if t['alert_type'] == '5min' else "pre"
                entry = t.get('entry_price', 0)
                exit_15 = t.get('exit_price_15m', 0)
                pnl_rs = t.get('pnl_rs_15m', 0)
                pnl_pct = t.get('pnl_pct_15m', 0)
                lot = t.get('lot_size', 0)
                lines.append(
                    f"  {t['symbol']} {dir_arrow}{type_short} | "
                    f"1 lot({lot}) | "
                    f"\u20b9{entry:.0f}\u2192\u20b9{exit_15:.0f} | "
                    f"\u20b9{pnl_rs:+,.0f} ({pnl_pct:+.2f}%)"
                )
            lines.append("")

        # 15m Summary
        trades_with_15m = [t for t in trades if t.get('pnl_rs_15m') is not None]
        if trades_with_15m:
            total_pnl_15m = sum(t['pnl_rs_15m'] for t in trades_with_15m)
            wins_15m = sum(1 for t in trades_with_15m if t['pnl_rs_15m'] >= 0)
            count_15m = len(trades_with_15m)
            avg_pnl_15m = total_pnl_15m / count_15m if count_15m > 0 else 0
            win_rate_15m = (wins_15m / count_15m * 100) if count_15m > 0 else 0

            best_15m = max(trades_with_15m, key=lambda t: t['pnl_rs_15m'])
            worst_15m = min(trades_with_15m, key=lambda t: t['pnl_rs_15m'])

            lines.append(f"\U0001f4c8 <b>SUMMARY (15m exits)</b>")
            lines.append(f"  Total P&L: \u20b9{total_pnl_15m:+,.0f}")
            lines.append(f"  Avg P&L: \u20b9{avg_pnl_15m:+,.0f}/trade")
            lines.append(f"  Win Rate: {win_rate_15m:.1f}% ({wins_15m}/{count_15m})")
            lines.append(f"  Best: {best_15m['symbol']} \u20b9{best_15m['pnl_rs_15m']:+,.0f}")
            lines.append(f"  Worst: {worst_15m['symbol']} \u20b9{worst_15m['pnl_rs_15m']:+,.0f}")
            lines.append("")

        # 30m Summary
        trades_with_30m = [t for t in trades if t.get('pnl_rs_30m') is not None]
        if trades_with_30m:
            total_pnl_30m = sum(t['pnl_rs_30m'] for t in trades_with_30m)
            wins_30m = sum(1 for t in trades_with_30m if t['pnl_rs_30m'] >= 0)
            count_30m = len(trades_with_30m)
            win_rate_30m = (wins_30m / count_30m * 100) if count_30m > 0 else 0

            lines.append(f"\U0001f4c8 <b>SUMMARY (30m exits)</b>")
            lines.append(f"  Total P&L: \u20b9{total_pnl_30m:+,.0f}")
            lines.append(f"  Win Rate: {win_rate_30m:.1f}% ({wins_30m}/{count_30m})")

        lines.append(f"\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501\u2501")

        return "\n".join(lines)

    def _write_to_excel(self, trade: Dict):
        """
        Append a completed trade row to the Excel file.

        Uses openpyxl with fcntl locking (same pattern as alert_excel_logger.py).
        """
        try:
            # Load or create workbook
            if os.path.exists(self.excel_path):
                wb = openpyxl.load_workbook(self.excel_path)
            else:
                wb = openpyxl.Workbook()
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]

            # Get or create Daily_PnL sheet
            sheet_name = "Daily_PnL"
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                self._write_excel_headers(ws)
            else:
                ws = wb[sheet_name]

            # Find next row
            next_row = ws.max_row + 1

            # Build row data
            row_data = [
                trade['date'],
                trade['time'],
                trade['symbol'],
                trade['alert_type'],
                trade['direction'],
                trade['lot_size'],
                trade.get('alert_price'),
                trade.get('entry_price'),
                trade.get('exit_price_15m'),
                trade.get('exit_price_30m'),
                trade.get('pnl_pct_15m'),
                trade.get('pnl_pct_30m'),
                trade.get('pnl_rs_15m'),
                trade.get('pnl_rs_30m'),
            ]

            # Write cells with formatting
            for col_num, value in enumerate(row_data, start=1):
                cell = ws.cell(row=next_row, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # Number formatting for price columns
                if col_num in (7, 8, 9, 10):  # Price columns
                    cell.number_format = '#,##0.00'
                elif col_num in (11, 12):  # P&L % columns
                    cell.number_format = '0.00'
                elif col_num in (13, 14):  # P&L Rs columns
                    cell.number_format = '#,##0'

            # Color-code P&L columns
            for col_num in (13, 14):  # P&L Rs 15m, 30m
                cell = ws.cell(row=next_row, column=col_num)
                val = cell.value
                if val is not None:
                    if val >= 0:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(color="006100")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006")

            # Save with file locking
            with open(self.excel_path, 'wb') as f:
                fcntl.flock(f.fileno(), fcntl.LOCK_EX)
                try:
                    wb.save(f)
                finally:
                    fcntl.flock(f.fileno(), fcntl.LOCK_UN)

            wb.close()
            from google_drive_sync import sync_to_drive
            sync_to_drive(self.excel_path, "AlertPnL")
            logger.debug(f"PnLTracker: Written {trade['symbol']} to Excel")

        except Exception as e:
            logger.error(f"PnLTracker: Excel write failed for {trade.get('symbol', '?')}: {e}")

    def _write_excel_headers(self, ws):
        """Write formatted headers to a new worksheet."""
        for col_num, header in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # Set column widths
        widths = [12, 8, 12, 8, 10, 10, 12, 12, 12, 12, 12, 12, 14, 14]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        ws.freeze_panes = "A2"

    def _send_telegram_message(self, message: str) -> bool:
        """Send message to Telegram using the notifier's connection."""
        try:
            import requests

            url = f"https://api.telegram.org/bot{self.telegram.bot_token}/sendMessage"
            payload = {
                "chat_id": self.telegram.channel_id,
                "text": message,
                "parse_mode": "HTML"
            }

            response = requests.post(url, json=payload, timeout=10)
            response.raise_for_status()
            return True

        except Exception as e:
            logger.error(f"PnLTracker: Telegram send failed: {e}")
            return False

    def _check_daily_reset(self):
        """Reset daily state if date has changed."""
        today = date.today()
        if self._current_date != today:
            if self._current_date is not None:
                logger.info(f"PnLTracker: Daily reset (was {self._current_date}, now {today})")
            self._pending_trades.clear()
            self._completed_trades.clear()
            self._report_sent_today = False
            self._current_date = today
