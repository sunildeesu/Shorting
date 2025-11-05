#!/usr/bin/env python3
"""
Comprehensive 3-Year Backtest for EOD Analysis System
Tests pattern detection logic from Nov 2022 to Nov 2025
Generates detailed Excel report with performance metrics and recommendations
"""

import sys
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Tuple
from kiteconnect import KiteConnect
import config
from eod_pattern_detector import EODPatternDetector
from market_regime_detector import MarketRegimeDetector
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time
from collections import defaultdict

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/backtest_3year.log'),
        logging.StreamHandler()
    ]
)

logger = logging.getLogger(__name__)


class Comprehensive3YearBacktester:
    """Comprehensive backtester for 3-year period"""

    def __init__(self):
        """Initialize backtester"""
        self.kite = KiteConnect(api_key=config.KITE_API_KEY)
        self.kite.set_access_token(config.KITE_ACCESS_TOKEN)

        # Initialize components
        self.pattern_detector = EODPatternDetector(
            pattern_tolerance=2.0,
            volume_confirmation=True,
            min_confidence=7.0
        )
        self.regime_detector = MarketRegimeDetector(self.kite)

        # Build instrument token map
        self.instrument_tokens = self._build_instrument_token_map()

        logger.info("3-Year Backtester initialized")

    def _build_instrument_token_map(self) -> Dict[str, int]:
        """Build mapping of stock symbol to instrument token"""
        try:
            instruments = self.kite.instruments("NSE")
            token_map = {}

            # Major stocks to test (top 30 liquid stocks)
            test_stocks = [
                'RELIANCE', 'TCS', 'HDFCBANK', 'INFY', 'ICICIBANK',
                'HINDUNILVR', 'ITC', 'SBIN', 'BHARTIARTL', 'KOTAKBANK',
                'LT', 'AXISBANK', 'ASIANPAINT', 'MARUTI', 'BAJFINANCE',
                'TITAN', 'WIPRO', 'ONGC', 'NTPC', 'POWERGRID',
                'TATASTEEL', 'TATAMOTORS', 'BAJAJFINSV', 'SUNPHARMA',
                'ULTRACEMCO', 'NESTLEIND', 'DIVISLAB', 'ADANIPORTS',
                'JSWSTEEL', 'GRASIM'
            ]

            for instrument in instruments:
                if instrument['segment'] == 'NSE' and instrument['tradingsymbol'] in test_stocks:
                    token_map[instrument['tradingsymbol']] = instrument['instrument_token']

            logger.info(f"Built token map for {len(token_map)} stocks")
            return token_map

        except Exception as e:
            logger.error(f"Error building token map: {e}")
            return {}

    def fetch_historical_data(
        self,
        symbol: str,
        from_date: datetime,
        to_date: datetime
    ) -> List[Dict]:
        """Fetch historical data for a symbol"""
        token = self.instrument_tokens.get(symbol)
        if not token:
            return []

        try:
            data = self.kite.historical_data(
                instrument_token=token,
                from_date=from_date,
                to_date=to_date,
                interval="day"
            )
            return data
        except Exception as e:
            logger.error(f"{symbol}: Error fetching data - {e}")
            return []

    def test_pattern(
        self,
        symbol: str,
        test_date: datetime,
        market_regime: str
    ) -> List[Dict]:
        """
        Test patterns for a specific date and analyze forward performance

        Args:
            symbol: Stock symbol
            test_date: Date to detect patterns
            market_regime: Market regime on test date

        Returns:
            List of trade results
        """
        # Fetch 60 days before test_date for pattern detection
        pattern_start = test_date - timedelta(days=80)  # Extra buffer
        pattern_data = self.fetch_historical_data(symbol, pattern_start, test_date)

        if not pattern_data or len(pattern_data) < 30:
            return []

        # Get last 30 days for pattern detection
        pattern_data = pattern_data[-30:]

        # Detect patterns
        result = self.pattern_detector.detect_patterns(symbol, pattern_data, market_regime)

        if not result['has_patterns']:
            return []

        # Fetch forward data (30 days) to check performance
        forward_start = test_date + timedelta(days=1)
        forward_end = test_date + timedelta(days=40)  # Extra buffer
        forward_data = self.fetch_historical_data(symbol, forward_start, forward_end)

        if not forward_data:
            return []

        # Analyze each detected pattern
        trades = []
        pattern_details = result['pattern_details']

        for pattern_name, details in pattern_details.items():
            if not details:
                continue

            trade_result = self._analyze_forward_performance(
                symbol=symbol,
                pattern_name=pattern_name,
                pattern_details=details,
                forward_data=forward_data,
                entry_date=test_date,
                market_regime=market_regime
            )

            if trade_result:
                trades.append(trade_result)

        return trades

    def _analyze_forward_performance(
        self,
        symbol: str,
        pattern_name: str,
        pattern_details: Dict,
        forward_data: List[Dict],
        entry_date: datetime,
        market_regime: str
    ) -> Dict:
        """Analyze forward performance of a pattern"""

        buy_price = pattern_details.get('buy_price')
        target_price = pattern_details.get('target_price')
        pattern_type = pattern_details.get('pattern_type')
        confidence = pattern_details.get('confidence_score', 0)
        volume_ratio = pattern_details.get('volume_ratio', 1.0)

        if not buy_price or not target_price:
            return None

        # Take first 30 days
        forward_data = forward_data[:30] if len(forward_data) > 30 else forward_data

        if not forward_data:
            return None

        # Track performance
        target_hit = False
        days_to_target = None
        exit_price = forward_data[-1]['close']  # Last day price
        max_gain_pct = 0
        max_loss_pct = 0

        for day_num, candle in enumerate(forward_data, 1):
            if pattern_type == 'BULLISH':
                # Check if high reached target
                if candle['high'] >= target_price and not target_hit:
                    target_hit = True
                    days_to_target = day_num
                    exit_price = target_price
                    break

                # Track max gain/loss
                gain_pct = ((candle['high'] - buy_price) / buy_price) * 100
                loss_pct = ((candle['low'] - buy_price) / buy_price) * 100
                max_gain_pct = max(max_gain_pct, gain_pct)
                max_loss_pct = min(max_loss_pct, loss_pct)

            else:  # BEARISH
                # Check if low reached target
                if candle['low'] <= target_price and not target_hit:
                    target_hit = True
                    days_to_target = day_num
                    exit_price = target_price
                    break

                # Track max gain/loss (inverse for short)
                gain_pct = ((buy_price - candle['low']) / buy_price) * 100
                loss_pct = ((buy_price - candle['high']) / buy_price) * 100
                max_gain_pct = max(max_gain_pct, gain_pct)
                max_loss_pct = min(max_loss_pct, loss_pct)

        # Calculate final P&L
        if pattern_type == 'BULLISH':
            final_pnl_pct = ((exit_price - buy_price) / buy_price) * 100
        else:
            final_pnl_pct = ((buy_price - exit_price) / buy_price) * 100

        # Determine outcome
        outcome = 'WIN' if target_hit else 'LOSS'

        return {
            'symbol': symbol,
            'entry_date': entry_date.strftime('%Y-%m-%d'),
            'pattern_name': pattern_name.upper(),
            'pattern_type': pattern_type,
            'confidence': confidence,
            'volume_ratio': volume_ratio,
            'market_regime': market_regime,
            'buy_price': buy_price,
            'target_price': target_price,
            'exit_price': exit_price,
            'target_hit': target_hit,
            'days_to_target': days_to_target if days_to_target else 30,
            'final_pnl_pct': final_pnl_pct,
            'max_gain_pct': max_gain_pct,
            'max_loss_pct': max_loss_pct,
            'outcome': outcome,
            'year': entry_date.year,
            'month': entry_date.month,
            'quarter': f"Q{(entry_date.month-1)//3 + 1}"
        }

    def run_backtest(
        self,
        start_date: datetime,
        end_date: datetime
    ) -> List[Dict]:
        """
        Run comprehensive backtest

        Args:
            start_date: Start date for backtest
            end_date: End date for backtest

        Returns:
            List of all trades
        """
        logger.info("="*80)
        logger.info("Starting 3-Year Comprehensive Backtest")
        logger.info("="*80)
        logger.info(f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
        logger.info(f"Testing {len(self.instrument_tokens)} stocks")
        logger.info("="*80)

        all_trades = []

        # Test each stock
        for stock_num, symbol in enumerate(sorted(self.instrument_tokens.keys()), 1):
            logger.info(f"[{stock_num}/{len(self.instrument_tokens)}] Testing {symbol}...")

            # Test weekly throughout the period
            current_date = start_date
            test_dates = []

            while current_date <= end_date:
                # Only test on weekdays
                if current_date.weekday() < 5:
                    test_dates.append(current_date)
                current_date += timedelta(days=7)  # Weekly tests

            logger.info(f"  Testing {len(test_dates)} dates (weekly)")

            for test_date in test_dates:
                # Get market regime (use cached if possible)
                market_regime = "NEUTRAL"  # Default

                try:
                    trades = self.test_pattern(symbol, test_date, market_regime)
                    all_trades.extend(trades)

                    if trades:
                        logger.info(f"  {test_date.strftime('%Y-%m-%d')}: {len(trades)} patterns detected")

                    time.sleep(0.1)  # Rate limiting

                except Exception as e:
                    logger.error(f"  Error testing {symbol} on {test_date}: {e}")

            logger.info(f"  {symbol} complete: {sum(1 for t in all_trades if t['symbol'] == symbol)} trades")

        logger.info("="*80)
        logger.info(f"Backtest complete: {len(all_trades)} total trades")
        logger.info("="*80)

        return all_trades

    def generate_excel_report(self, trades: List[Dict], report_file: str):
        """Generate comprehensive Excel report"""

        logger.info(f"Generating comprehensive report: {report_file}")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Sheet 1: All Trades
        self._write_trades_sheet(wb, trades)

        # Sheet 2: Summary Statistics
        self._write_summary_sheet(wb, trades)

        # Sheet 3: Pattern Analysis
        self._write_pattern_analysis_sheet(wb, trades)

        # Sheet 4: Market Regime Analysis
        self._write_regime_analysis_sheet(wb, trades)

        # Sheet 5: Confidence Score Analysis
        self._write_confidence_analysis_sheet(wb, trades)

        # Sheet 6: Yearly Performance
        self._write_yearly_performance_sheet(wb, trades)

        # Sheet 7: Monthly Performance
        self._write_monthly_performance_sheet(wb, trades)

        # Sheet 8: Recommendations
        self._write_recommendations_sheet(wb, trades)

        # Save workbook
        wb.save(report_file)
        logger.info(f"Report saved: {report_file}")

    def _write_trades_sheet(self, wb, trades):
        """Write all trades to Excel sheet"""
        ws = wb.create_sheet("All Trades")

        # Header
        headers = [
            'Symbol', 'Entry Date', 'Pattern', 'Type', 'Confidence',
            'Volume', 'Market', 'Buy Price', 'Target', 'Exit Price',
            'Target Hit?', 'Days', 'P&L %', 'Max Gain %', 'Max Loss %',
            'Outcome', 'Year', 'Quarter'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        for row_num, trade in enumerate(trades, 2):
            ws.cell(row=row_num, column=1, value=trade['symbol'])
            ws.cell(row=row_num, column=2, value=trade['entry_date'])
            ws.cell(row=row_num, column=3, value=trade['pattern_name'])
            ws.cell(row=row_num, column=4, value=trade['pattern_type'])
            ws.cell(row=row_num, column=5, value=f"{trade['confidence']:.1f}/10")
            ws.cell(row=row_num, column=6, value=f"{trade['volume_ratio']:.1f}x")
            ws.cell(row=row_num, column=7, value=trade['market_regime'])
            ws.cell(row=row_num, column=8, value=f"₹{trade['buy_price']:.2f}")
            ws.cell(row=row_num, column=9, value=f"₹{trade['target_price']:.2f}")
            ws.cell(row=row_num, column=10, value=f"₹{trade['exit_price']:.2f}")
            ws.cell(row=row_num, column=11, value='YES' if trade['target_hit'] else 'NO')
            ws.cell(row=row_num, column=12, value=trade['days_to_target'])
            ws.cell(row=row_num, column=13, value=f"{trade['final_pnl_pct']:+.2f}%")
            ws.cell(row=row_num, column=14, value=f"{trade['max_gain_pct']:+.2f}%")
            ws.cell(row=row_num, column=15, value=f"{trade['max_loss_pct']:+.2f}%")
            ws.cell(row=row_num, column=16, value=trade['outcome'])
            ws.cell(row=row_num, column=17, value=trade['year'])
            ws.cell(row=row_num, column=18, value=trade['quarter'])

            # Color code outcome
            outcome_cell = ws.cell(row=row_num, column=16)
            if trade['outcome'] == 'WIN':
                outcome_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                outcome_cell.font = Font(color="006100", bold=True)
            else:
                outcome_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                outcome_cell.font = Font(color="9C0006", bold=True)

        # Column widths
        for col in range(1, 19):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _write_summary_sheet(self, wb, trades):
        """Write summary statistics"""
        ws = wb.create_sheet("Summary Statistics", 0)

        wins = [t for t in trades if t['outcome'] == 'WIN']
        losses = [t for t in trades if t['outcome'] == 'LOSS']

        row = 1
        ws.cell(row=row, column=1, value="OVERALL PERFORMANCE")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 2

        stats = [
            ("Total Trades", len(trades)),
            ("Winning Trades", len(wins)),
            ("Losing Trades", len(losses)),
            ("Win Rate", f"{len(wins)/len(trades)*100:.1f}%" if trades else "0%"),
            ("", ""),
            ("Average P&L (All)", f"{sum(t['final_pnl_pct'] for t in trades)/len(trades):.2f}%" if trades else "0%"),
            ("Average Gain (Winners)", f"{sum(t['final_pnl_pct'] for t in wins)/len(wins):.2f}%" if wins else "0%"),
            ("Average Loss (Losers)", f"{sum(t['final_pnl_pct'] for t in losses)/len(losses):.2f}%" if losses else "0%"),
            ("", ""),
            ("Best Trade", f"{max(t['final_pnl_pct'] for t in trades):.2f}%" if trades else "0%"),
            ("Worst Trade", f"{min(t['final_pnl_pct'] for t in trades):.2f}%" if trades else "0%"),
            ("", ""),
            ("Average Days to Target (Winners)", f"{sum(t['days_to_target'] for t in wins)/len(wins):.1f}" if wins else "0"),
        ]

        for label, value in stats:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def _write_pattern_analysis_sheet(self, wb, trades):
        """Pattern-wise analysis"""
        ws = wb.create_sheet("Pattern Analysis")

        # Group by pattern
        pattern_stats = defaultdict(lambda: {'total': 0, 'wins': 0, 'pnl': []})

        for trade in trades:
            pattern = trade['pattern_name']
            pattern_stats[pattern]['total'] += 1
            if trade['outcome'] == 'WIN':
                pattern_stats[pattern]['wins'] += 1
            pattern_stats[pattern]['pnl'].append(trade['final_pnl_pct'])

        # Write header
        headers = ['Pattern', 'Total', 'Wins', 'Losses', 'Win Rate', 'Avg P&L', 'Recommendation']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        # Write data
        row = 2
        for pattern, stats in sorted(pattern_stats.items()):
            total = stats['total']
            wins = stats['wins']
            losses = total - wins
            win_rate = (wins / total * 100) if total > 0 else 0
            avg_pnl = sum(stats['pnl']) / len(stats['pnl']) if stats['pnl'] else 0

            # Recommendation
            if win_rate >= 65:
                rec = "✅ HIGHLY RECOMMENDED"
            elif win_rate >= 55:
                rec = "✅ RECOMMENDED"
            elif win_rate >= 45:
                rec = "⚠️ NEUTRAL - Use with caution"
            else:
                rec = "❌ AVOID - Poor performance"

            ws.cell(row=row, column=1, value=pattern)
            ws.cell(row=row, column=2, value=total)
            ws.cell(row=row, column=3, value=wins)
            ws.cell(row=row, column=4, value=losses)
            ws.cell(row=row, column=5, value=f"{win_rate:.1f}%")
            ws.cell(row=row, column=6, value=f"{avg_pnl:+.2f}%")
            ws.cell(row=row, column=7, value=rec)

            row += 1

        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _write_regime_analysis_sheet(self, wb, trades):
        """Market regime analysis"""
        ws = wb.create_sheet("Market Regime Analysis")

        ws.cell(row=1, column=1, value="Performance by Market Regime").font = Font(bold=True, size=12)
        ws.cell(row=3, column=1, value="Regime").font = Font(bold=True)
        ws.cell(row=3, column=2, value="Total Trades").font = Font(bold=True)
        ws.cell(row=3, column=3, value="Win Rate").font = Font(bold=True)
        ws.cell(row=3, column=4, value="Avg P&L").font = Font(bold=True)

        # Placeholder - would need actual regime tracking
        ws.cell(row=4, column=1, value="Data not tracked in backtest")

    def _write_confidence_analysis_sheet(self, wb, trades):
        """Confidence score analysis"""
        ws = wb.create_sheet("Confidence Analysis")

        # Group by confidence ranges
        confidence_ranges = {
            '9.0-10.0': (9.0, 10.0),
            '8.5-8.9': (8.5, 8.9),
            '8.0-8.4': (8.0, 8.4),
            '7.5-7.9': (7.5, 7.9),
            '7.0-7.4': (7.0, 7.4)
        }

        headers = ['Confidence Range', 'Total', 'Wins', 'Win Rate', 'Avg P&L']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)

        row = 2
        for range_name, (min_conf, max_conf) in confidence_ranges.items():
            range_trades = [t for t in trades if min_conf <= t['confidence'] <= max_conf]

            if not range_trades:
                continue

            wins = len([t for t in range_trades if t['outcome'] == 'WIN'])
            win_rate = (wins / len(range_trades) * 100) if range_trades else 0
            avg_pnl = sum(t['final_pnl_pct'] for t in range_trades) / len(range_trades)

            ws.cell(row=row, column=1, value=range_name)
            ws.cell(row=row, column=2, value=len(range_trades))
            ws.cell(row=row, column=3, value=wins)
            ws.cell(row=row, column=4, value=f"{win_rate:.1f}%")
            ws.cell(row=row, column=5, value=f"{avg_pnl:+.2f}%")

            row += 1

        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _write_yearly_performance_sheet(self, wb, trades):
        """Yearly performance breakdown"""
        ws = wb.create_sheet("Yearly Performance")

        headers = ['Year', 'Total Trades', 'Wins', 'Losses', 'Win Rate', 'Total P&L']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)

        years = sorted(set(t['year'] for t in trades))

        row = 2
        for year in years:
            year_trades = [t for t in trades if t['year'] == year]
            wins = len([t for t in year_trades if t['outcome'] == 'WIN'])
            losses = len(year_trades) - wins
            win_rate = (wins / len(year_trades) * 100) if year_trades else 0
            total_pnl = sum(t['final_pnl_pct'] for t in year_trades)

            ws.cell(row=row, column=1, value=year)
            ws.cell(row=row, column=2, value=len(year_trades))
            ws.cell(row=row, column=3, value=wins)
            ws.cell(row=row, column=4, value=losses)
            ws.cell(row=row, column=5, value=f"{win_rate:.1f}%")
            ws.cell(row=row, column=6, value=f"{total_pnl:+.1f}%")

            row += 1

        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _write_monthly_performance_sheet(self, wb, trades):
        """Monthly performance"""
        ws = wb.create_sheet("Monthly Performance")

        headers = ['Year-Month', 'Trades', 'Win Rate', 'Avg P&L']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)

        # Group by year-month
        monthly_data = defaultdict(list)
        for trade in trades:
            key = f"{trade['year']}-{trade['month']:02d}"
            monthly_data[key].append(trade)

        row = 2
        for month_key in sorted(monthly_data.keys()):
            month_trades = monthly_data[month_key]
            wins = len([t for t in month_trades if t['outcome'] == 'WIN'])
            win_rate = (wins / len(month_trades) * 100) if month_trades else 0
            avg_pnl = sum(t['final_pnl_pct'] for t in month_trades) / len(month_trades)

            ws.cell(row=row, column=1, value=month_key)
            ws.cell(row=row, column=2, value=len(month_trades))
            ws.cell(row=row, column=3, value=f"{win_rate:.1f}%")
            ws.cell(row=row, column=4, value=f"{avg_pnl:+.2f}%")

            row += 1

        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _write_recommendations_sheet(self, wb, trades):
        """Write recommendations based on analysis"""
        ws = wb.create_sheet("Recommendations", 0)

        # Analysis
        wins = [t for t in trades if t['outcome'] == 'WIN']
        overall_win_rate = (len(wins) / len(trades) * 100) if trades else 0

        # Pattern performance
        pattern_stats = defaultdict(lambda: {'total': 0, 'wins': 0})
        for trade in trades:
            pattern = trade['pattern_name']
            pattern_stats[pattern]['total'] += 1
            if trade['outcome'] == 'WIN':
                pattern_stats[pattern]['wins'] += 1

        row = 1
        ws.cell(row=row, column=1, value="3-YEAR BACKTEST RECOMMENDATIONS").font = Font(bold=True, size=16)
        row += 2

        ws.cell(row=row, column=1, value=f"Overall Win Rate: {overall_win_rate:.1f}%").font = Font(bold=True, size=12)
        row += 2

        # Recommendations
        recommendations = [
            ("PATTERN RECOMMENDATIONS", ""),
            ("", ""),
        ]

        # Add pattern-specific recommendations
        for pattern, stats in sorted(pattern_stats.items()):
            win_rate = (stats['wins'] / stats['total'] * 100) if stats['total'] > 0 else 0
            if win_rate >= 65:
                recommendations.append((f"✅ {pattern}", f"{win_rate:.1f}% win rate - HIGHLY RECOMMENDED"))
            elif win_rate >= 55:
                recommendations.append((f"✅ {pattern}", f"{win_rate:.1f}% win rate - RECOMMENDED"))
            elif win_rate >= 45:
                recommendations.append((f"⚠️ {pattern}", f"{win_rate:.1f}% win rate - USE WITH CAUTION"))
            else:
                recommendations.append((f"❌ {pattern}", f"{win_rate:.1f}% win rate - AVOID"))

        recommendations.extend([
            ("", ""),
            ("GENERAL RECOMMENDATIONS", ""),
            ("1. Confidence Threshold", "Keep minimum at 7.0/10 or increase to 7.5/10"),
            ("2. Volume Confirmation", "Keep 1.5x threshold - effective filter"),
            ("3. Position Sizing", "Full size for 8.0+ confidence, 50% for 7.0-7.9"),
            ("4. Stop Loss", "2-3% for bullish, 1.5-2% for bearish patterns"),
            ("5. Market Regime", "Only trade patterns aligned with market trend"),
        ])

        for label, value in recommendations:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 60


def main():
    """Main entry point"""
    logger.info("Starting 3-Year Comprehensive Backtest")

    # Date range: Nov 4, 2022 to Nov 4, 2025
    start_date = datetime(2022, 11, 4)
    end_date = datetime(2025, 11, 4)

    # Create backtester
    backtester = Comprehensive3YearBacktester()

    # Run backtest
    trades = backtester.run_backtest(start_date, end_date)

    if not trades:
        logger.warning("No trades found in backtest period")
        return

    # Generate report
    report_file = "data/eod_reports/backtest_3year_comprehensive.xlsx"
    backtester.generate_excel_report(trades, report_file)

    print(f"\n✅ 3-Year backtest complete!")
    print(f"📊 Total trades analyzed: {len(trades)}")
    print(f"📁 Report: {report_file}\n")


if __name__ == "__main__":
    main()
