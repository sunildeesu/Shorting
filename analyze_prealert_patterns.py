#!/usr/bin/env python3
"""
Analyze T-10 to T patterns for 5-min alerts.

Examines the 10 minutes preceding each 5-min alert to find:
1. When does the move typically start?
2. What % of the move happens at each minute?
3. Volume patterns before alerts
4. OBV patterns before alerts
5. Early indicators that could improve pre-alert detection

Author: Claude Opus 4.5
Date: 2026-02-09
"""

import sqlite3
from datetime import datetime, timedelta
from collections import defaultdict
import statistics

import openpyxl
import config


def load_5min_alerts(days: int = 30):
    """Load 5-min alerts from Excel."""
    excel_path = config.ALERT_EXCEL_PATH
    workbook = openpyxl.load_workbook(excel_path, read_only=True)

    if "5min_alerts" not in workbook.sheetnames:
        workbook.close()
        return []

    sheet = workbook["5min_alerts"]
    cutoff_date = datetime.now() - timedelta(days=days)

    alerts = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        try:
            date_val = row[0]
            time_val = row[1]

            if isinstance(date_val, datetime):
                date_str = date_val.strftime("%Y-%m-%d")
            else:
                date_str = str(date_val)

            if isinstance(time_val, datetime):
                time_str = time_val.strftime("%H:%M:%S")
            else:
                time_str = str(time_val)

            try:
                alert_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
            except ValueError:
                alert_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")

            if alert_dt < cutoff_date:
                continue

            direction = str(row[3]).lower() if row[3] else 'drop'
            change_pct = float(row[6]) if row[6] else 0

            alerts.append({
                'datetime': alert_dt,
                'symbol': str(row[2]) if row[2] else '',
                'direction': direction,
                'change_pct': change_pct,
            })
        except Exception as e:
            continue

    workbook.close()
    return alerts


def get_price_history(conn, symbol: str, alert_time: datetime, minutes_before: int = 10):
    """Get minute-by-minute price/volume data before alert."""
    cursor = conn.cursor()

    # Expand window slightly to catch data with minor timestamp differences
    start_time = (alert_time - timedelta(minutes=minutes_before + 1)).strftime('%Y-%m-%d %H:%M:%S')
    end_time = (alert_time + timedelta(minutes=1)).strftime('%Y-%m-%d %H:%M:%S')

    cursor.execute("""
        SELECT timestamp, price, volume, oi
        FROM stock_quotes
        WHERE symbol = ?
        AND timestamp >= ?
        AND timestamp <= ?
        ORDER BY timestamp ASC
    """, (symbol, start_time, end_time))

    return [{'timestamp': r[0], 'price': r[1], 'volume': r[2] or 0, 'oi': r[3] or 0}
            for r in cursor.fetchall()]


def get_price_at_offset(history, alert_time: datetime, minutes_before: int):
    """Get price at specific offset, with tolerance for missing data."""
    target_time = alert_time - timedelta(minutes=minutes_before)
    target_ts = target_time.strftime('%Y-%m-%d %H:%M')

    # Find closest match within 30 seconds
    for h in history:
        ts = h['timestamp'][:16]  # YYYY-MM-DD HH:MM
        if ts == target_ts:
            return h
    return None


def calculate_obv(history):
    """Calculate OBV series."""
    if len(history) < 2:
        return []

    obv = [0]
    for i in range(1, len(history)):
        prev_price = history[i-1]['price']
        curr_price = history[i]['price']
        volume = history[i]['volume'] or 0

        if curr_price > prev_price:
            obv.append(obv[-1] + volume)
        elif curr_price < prev_price:
            obv.append(obv[-1] - volume)
        else:
            obv.append(obv[-1])

    return obv


def analyze_single_alert(conn, alert, minutes_before=10):
    """Analyze a single alert's pre-alert pattern."""
    history = get_price_history(conn, alert['symbol'], alert['datetime'], minutes_before)

    if len(history) < 3:  # Need at least some data
        return None

    # Get price at alert time (T-0)
    alert_data = get_price_at_offset(history, alert['datetime'], 0)
    if not alert_data:
        # Use last data point as fallback
        alert_data = history[-1]

    alert_price = alert_data['price']
    if not alert_price or alert_price <= 0:
        return None

    # Build price/volume arrays at each minute offset
    prices = []
    volumes = []
    for offset in range(minutes_before, -1, -1):
        data = get_price_at_offset(history, alert['datetime'], offset)
        if data:
            prices.append(data['price'])
            volumes.append(data['volume'])
        else:
            prices.append(None)
            volumes.append(None)

    # Calculate price change from each minute to alert time
    # prices[0] = T-10, prices[10] = T-0 (alert)
    changes_from_t = []
    for i, price in enumerate(prices[:-1]):  # Exclude T-0
        if price and price > 0 and alert_price > 0:
            if 'drop' in alert['direction']:
                pct = ((price - alert_price) / price) * 100
            else:
                pct = ((alert_price - price) / price) * 100
            changes_from_t.append(pct)
        else:
            changes_from_t.append(None)

    # Calculate minute-to-minute changes
    minute_changes = []
    for i in range(1, len(prices)):
        if prices[i] and prices[i-1] and prices[i] > 0 and prices[i-1] > 0:
            if 'drop' in alert['direction']:
                pct = ((prices[i-1] - prices[i]) / prices[i-1]) * 100
            else:
                pct = ((prices[i] - prices[i-1]) / prices[i-1]) * 100
            minute_changes.append(pct)
        else:
            minute_changes.append(None)

    # Volume analysis - calculate average of non-None values
    valid_volumes = [v for v in volumes[:-1] if v is not None and v > 0]
    avg_volume = sum(valid_volumes) / len(valid_volumes) if valid_volumes else 0

    volume_ratios = []
    for v in volumes:
        if v is not None and avg_volume > 0:
            volume_ratios.append(v / avg_volume)
        else:
            volume_ratios.append(None)

    # OBV analysis
    obv = calculate_obv(history)
    obv_trend = 'neutral'
    if len(obv) >= 5:
        obv_start = sum(obv[:3]) / 3
        obv_end = sum(obv[-3:]) / 3
        if 'drop' in alert['direction']:
            if obv_end < obv_start * 0.95:
                obv_trend = 'confirming'
            elif obv_end > obv_start * 1.05:
                obv_trend = 'diverging'
        else:
            if obv_end > obv_start * 1.05:
                obv_trend = 'confirming'
            elif obv_end < obv_start * 0.95:
                obv_trend = 'diverging'

    # Count how many valid data points we have
    valid_prices = sum(1 for p in prices if p is not None)

    return {
        'symbol': alert['symbol'],
        'direction': alert['direction'],
        'alert_change': alert['change_pct'],
        'changes_from_t': changes_from_t,
        'minute_changes': minute_changes,
        'volume_ratios': volume_ratios,
        'obv_trend': obv_trend,
        'data_points': len(history),
        'valid_prices': valid_prices
    }


def run_analysis(days: int = 30):
    """Run full T-10 pattern analysis."""
    print(f"\n{'='*80}")
    print("T-10 MINUTE PATTERN ANALYSIS FOR 5-MIN ALERTS")
    print(f"{'='*80}\n")

    alerts = load_5min_alerts(days)
    print(f"Loaded {len(alerts)} 5-min alerts from last {days} days\n")

    conn = sqlite3.connect("data/central_quotes.db")

    # Analyze all alerts
    results = []
    for alert in alerts:
        analysis = analyze_single_alert(conn, alert, minutes_before=10)
        if analysis:
            results.append(analysis)

    conn.close()

    print(f"Successfully analyzed {len(results)} alerts (had sufficient data)\n")

    if not results:
        print("No results to analyze.")
        return

    # Aggregate analysis
    # 1. Average change at each minute before alert
    print(f"{'-'*80}")
    print("PRICE MOVEMENT TIMELINE (Average % move from T-X to Alert Time)")
    print(f"{'-'*80}")
    print(f"{'Minute':<10} {'Avg Change':<15} {'Median':<15} {'Samples':<10} {'>=0.5%':<10} {'>=1.0%':<10}")
    print(f"{'-'*80}")

    for t_minus in range(10, 0, -1):
        idx = 10 - t_minus  # Index in changes_from_t array

        changes = [r['changes_from_t'][idx] for r in results
                   if idx < len(r['changes_from_t']) and r['changes_from_t'][idx] is not None]

        if changes:
            avg = statistics.mean(changes)
            median = statistics.median(changes)
            count = len(changes)
            pct_05 = sum(1 for c in changes if c >= 0.5) / count * 100
            pct_10 = sum(1 for c in changes if c >= 1.0) / count * 100

            bar = "█" * int(avg * 10) if avg > 0 else ""
            print(f"T-{t_minus:<8} {avg:>+.2f}%{'':<8} {median:>+.2f}%{'':<8} {count:<10} {pct_05:.0f}%{'':<6} {pct_10:.0f}%")

    # 2. Minute-by-minute acceleration
    print(f"\n{'-'*80}")
    print("MINUTE-BY-MINUTE CHANGE (How much move happens each minute)")
    print(f"{'-'*80}")
    print(f"{'Transition':<15} {'Avg Change':<15} {'Contribution':<15}")
    print(f"{'-'*80}")

    total_change = []
    minute_contributions = defaultdict(list)

    for r in results:
        if len(r['minute_changes']) >= 10:
            valid_changes = [mc for mc in r['minute_changes'] if mc is not None]
            if not valid_changes:
                continue
            total = sum(valid_changes)
            if total > 0:
                for i, mc in enumerate(r['minute_changes']):
                    if mc is not None:
                        minute_contributions[i].append(mc / total * 100 if total > 0 else 0)

    for i in range(10):
        t_from = 10 - i
        t_to = 9 - i

        changes = [r['minute_changes'][i] for r in results
                   if i < len(r['minute_changes']) and r['minute_changes'][i] is not None]

        contribs = minute_contributions.get(i, [])

        if changes:
            avg = statistics.mean(changes)
            avg_contrib = statistics.mean(contribs) if contribs else 0

            bar = "█" * int(abs(avg) * 20) if avg != 0 else ""
            print(f"T-{t_from} → T-{t_to}{'':<5} {avg:>+.3f}%{'':<8} {avg_contrib:.1f}% {bar}")

    # 3. Volume patterns
    print(f"\n{'-'*80}")
    print("VOLUME PATTERN (Ratio vs Average)")
    print(f"{'-'*80}")
    print(f"{'Minute':<10} {'Avg Vol Ratio':<15} {'Spike (>1.5x)':<15}")
    print(f"{'-'*80}")

    for t_minus in range(10, 0, -1):
        idx = 10 - t_minus

        ratios = [r['volume_ratios'][idx] for r in results
                  if idx < len(r['volume_ratios']) and r['volume_ratios'][idx] is not None]

        if ratios:
            avg = statistics.mean(ratios)
            spike_pct = sum(1 for r in ratios if r >= 1.5) / len(ratios) * 100

            bar = "█" * int(avg * 5)
            print(f"T-{t_minus:<8} {avg:.2f}x{'':<10} {spike_pct:.0f}%{'':<10} {bar}")

    # Alert time volume
    alert_vols = [r['volume_ratios'][-1] for r in results
                  if r['volume_ratios'] and r['volume_ratios'][-1] is not None]
    if alert_vols:
        avg = statistics.mean(alert_vols)
        spike_pct = sum(1 for r in alert_vols if r >= 1.5) / len(alert_vols) * 100
        bar = "█" * int(avg * 5)
        print(f"T-0 (alert) {avg:.2f}x{'':<10} {spike_pct:.0f}%{'':<10} {bar}")

    # 4. OBV confirmation analysis
    print(f"\n{'-'*80}")
    print("OBV PATTERN ANALYSIS")
    print(f"{'-'*80}")

    obv_counts = defaultdict(int)
    for r in results:
        obv_counts[r['obv_trend']] += 1

    total = sum(obv_counts.values())
    for trend, count in sorted(obv_counts.items(), key=lambda x: -x[1]):
        pct = count / total * 100
        bar = "█" * int(pct / 2)
        print(f"  {trend:<12}: {count:>4} ({pct:.1f}%) {bar}")

    # 5. Key findings
    print(f"\n{'='*80}")
    print("KEY FINDINGS & RECOMMENDATIONS")
    print(f"{'='*80}\n")

    # Find when 0.5% threshold is typically hit
    threshold_times = defaultdict(int)
    for r in results:
        for idx, change in enumerate(r['changes_from_t']):
            if change and change >= 0.5:
                t_minus = 10 - idx
                threshold_times[t_minus] += 1
                break  # First time it hits threshold

    if threshold_times:
        most_common = max(threshold_times.items(), key=lambda x: x[1])
        print(f"1. 0.5% THRESHOLD TIMING:")
        print(f"   Most commonly first hit at: T-{most_common[0]} minutes")
        print(f"   Distribution:")
        for t in sorted(threshold_times.keys(), reverse=True):
            count = threshold_times[t]
            pct = count / len(results) * 100
            bar = "█" * int(pct / 2)
            print(f"     T-{t}: {count:>3} alerts ({pct:.1f}%) {bar}")

    # Find when 1.0% threshold is hit
    threshold_times_10 = defaultdict(int)
    for r in results:
        for idx, change in enumerate(r['changes_from_t']):
            if change and change >= 1.0:
                t_minus = 10 - idx
                threshold_times_10[t_minus] += 1
                break

    if threshold_times_10:
        most_common = max(threshold_times_10.items(), key=lambda x: x[1])
        print(f"\n2. 1.0% THRESHOLD TIMING:")
        print(f"   Most commonly first hit at: T-{most_common[0]} minutes")
        print(f"   Distribution:")
        for t in sorted(threshold_times_10.keys(), reverse=True):
            count = threshold_times_10[t]
            pct = count / len(results) * 100
            bar = "█" * int(pct / 2)
            print(f"     T-{t}: {count:>3} alerts ({pct:.1f}%) {bar}")

    # Volume spike timing
    print(f"\n3. VOLUME SPIKE TIMING (when >1.5x first appears):")
    vol_spike_times = defaultdict(int)
    for r in results:
        for idx, ratio in enumerate(r['volume_ratios']):
            if ratio is not None and ratio >= 1.5:
                t_minus = 10 - idx
                vol_spike_times[t_minus] += 1
                break

    if vol_spike_times:
        for t in sorted(vol_spike_times.keys(), reverse=True):
            count = vol_spike_times[t]
            pct = count / len(results) * 100
            if pct >= 5:  # Only show significant
                print(f"     T-{t}: {count:>3} alerts ({pct:.1f}%)")

    # Recommendations
    print(f"\n4. RECOMMENDATIONS:")

    # Calculate optimal lookback
    optimal_lookback = None
    best_detection = 0
    for lookback in [3, 4, 5, 6]:
        detection = sum(1 for r in results
                       if len(r['changes_from_t']) > 10 - lookback
                       and r['changes_from_t'][10 - lookback]
                       and r['changes_from_t'][10 - lookback] >= 0.5)
        if detection > best_detection:
            best_detection = detection
            optimal_lookback = lookback

    if optimal_lookback:
        detection_rate = best_detection / len(results) * 100
        print(f"   - Optimal lookback: {optimal_lookback} minutes (detects {detection_rate:.0f}% of alerts at 0.5% threshold)")

    # Calculate average lead time
    lead_times = []
    for r in results:
        for idx, change in enumerate(r['changes_from_t']):
            if change and change >= 1.0:
                lead_times.append(10 - idx)
                break

    if lead_times:
        avg_lead = statistics.mean(lead_times)
        print(f"   - Average lead time (1.0% threshold): {avg_lead:.1f} minutes before alert")

    print(f"\n{'='*80}\n")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--days", type=int, default=30)
    args = parser.parse_args()

    run_analysis(args.days)
