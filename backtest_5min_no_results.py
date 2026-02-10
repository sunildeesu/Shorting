#!/usr/bin/env python3
"""
Backtest 5-Min Alerts - Excluding Quarterly Results Days

Analyzes 5-min alerts from the past 30 days:
1. Excludes alerts on stocks with quarterly results that day
2. Checks profit at different exit times (5, 10, 15, 20, 30 min)
3. Compares DROP vs RISE performance

Author: Claude Opus 4.5
Date: 2026-02-10
"""

import sqlite3
from datetime import datetime, timedelta
from collections import defaultdict
import statistics
import openpyxl
import config


def load_5min_alerts(days: int = 30):
    """Load 5-min alerts from Excel."""
    excel_path = config.ALERT_EXCEL_PATH

    try:
        workbook = openpyxl.load_workbook(excel_path, read_only=True)
    except Exception as e:
        print(f"Failed to load Excel: {e}")
        return []

    if "5min_alerts" not in workbook.sheetnames:
        workbook.close()
        return []

    sheet = workbook["5min_alerts"]
    cutoff_date = datetime.now() - timedelta(days=days)

    alerts = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        try:
            date_val = row[0]
            time_val = row[1]

            if isinstance(date_val, datetime):
                date_str = date_val.strftime("%Y-%m-%d")
            else:
                date_str = str(date_val)

            if isinstance(time_val, datetime):
                time_str = time_val.strftime("%H:%M:%S")
            else:
                time_str = str(time_val)

            try:
                alert_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
            except ValueError:
                alert_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")

            if alert_dt < cutoff_date:
                continue

            direction = str(row[3]).lower() if row[3] else 'drop'
            change_pct = float(row[6]) if row[6] else 0

            # Get current price from row
            current_price = float(row[5]) if row[5] else 0

            alerts.append({
                'datetime': alert_dt,
                'date': date_str,
                'symbol': str(row[2]) if row[2] else '',
                'direction': 'DROP' if 'drop' in direction else 'RISE',
                'change_pct': change_pct,
                'entry_price': current_price,
            })

        except Exception as e:
            continue

    workbook.close()
    return alerts


def load_results_schedule():
    """Load historical quarterly results announcements from NSE data."""
    import json
    import os

    # Use historical results data (actual announcement dates from NSE)
    historical_file = "data/results_cache/historical_results.json"

    if not os.path.exists(historical_file):
        print("Warning: historical_results.json not found. Run with --fetch-history first.")
        return {}

    try:
        with open(historical_file, 'r') as f:
            data = json.load(f)
            # Returns dict: {symbol: [list of dates]}
            return data.get('results_dates', {})
    except:
        return {}


def was_results_day(symbol: str, alert_date: str, schedule: dict) -> bool:
    """Check if the stock had results/board meeting on the alert date."""
    symbol = symbol.upper()

    if symbol not in schedule:
        return False

    # schedule[symbol] is a list of dates in YYYY-MM-DD format
    result_dates = schedule[symbol]
    return alert_date in result_dates


def get_price_at_offset(conn, symbol: str, alert_time: datetime, minutes: int):
    """Get price at specified minutes after alert."""
    cursor = conn.cursor()

    target_time = alert_time + timedelta(minutes=minutes)

    # Find closest price within 1 minute of target
    start = (target_time - timedelta(minutes=1)).strftime('%Y-%m-%d %H:%M:%S')
    end = (target_time + timedelta(minutes=1)).strftime('%Y-%m-%d %H:%M:%S')

    cursor.execute("""
        SELECT price FROM stock_quotes
        WHERE symbol = ?
        AND timestamp >= ? AND timestamp <= ?
        ORDER BY ABS(strftime('%s', timestamp) - strftime('%s', ?))
        LIMIT 1
    """, (symbol, start, end, target_time.strftime('%Y-%m-%d %H:%M:%S')))

    row = cursor.fetchone()
    return row[0] if row else None


def run_backtest(days: int = 30):
    """Run the backtest analysis."""
    print(f"\n{'='*80}")
    print("5-MIN ALERT BACKTEST - EXCLUDING QUARTERLY RESULTS DAYS")
    print(f"{'='*80}\n")

    # Load alerts
    alerts = load_5min_alerts(days)
    print(f"Loaded {len(alerts)} 5-min alerts from last {days} days")

    # Load historical results announcements
    schedule = load_results_schedule()
    print(f"Loaded historical results data with {len(schedule)} stocks")

    # Filter out results days
    filtered_alerts = []
    results_day_alerts = []

    for alert in alerts:
        if was_results_day(alert['symbol'], alert['date'], schedule):
            results_day_alerts.append(alert)
        else:
            filtered_alerts.append(alert)

    print(f"\nFiltered: {len(filtered_alerts)} alerts (excluded {len(results_day_alerts)} on results days)")

    if not filtered_alerts:
        print("No alerts to analyze after filtering!")
        return

    # Connect to database
    conn = sqlite3.connect("data/central_quotes.db", timeout=30)

    # Exit times to test (minutes after alert)
    exit_times = [5, 10, 15, 20, 30]

    # Results storage
    results = {t: {'all': [], 'DROP': [], 'RISE': []} for t in exit_times}

    print(f"\nAnalyzing {len(filtered_alerts)} alerts at exit times: {exit_times} minutes...")

    for i, alert in enumerate(filtered_alerts):
        if i % 50 == 0:
            print(f"  Processing alert {i+1}/{len(filtered_alerts)}...")

        symbol = alert['symbol']
        alert_time = alert['datetime']
        direction = alert['direction']
        entry_price = alert['entry_price']

        if not entry_price or entry_price <= 0:
            continue

        for exit_min in exit_times:
            exit_price = get_price_at_offset(conn, symbol, alert_time, exit_min)

            if not exit_price or exit_price <= 0:
                continue

            # Calculate P&L based on direction
            if direction == 'DROP':
                # For DROP, we'd short: profit if price goes down
                pnl_pct = ((entry_price - exit_price) / entry_price) * 100
            else:
                # For RISE, we'd long: profit if price goes up
                pnl_pct = ((exit_price - entry_price) / entry_price) * 100

            results[exit_min]['all'].append({
                'symbol': symbol,
                'direction': direction,
                'entry_price': entry_price,
                'exit_price': exit_price,
                'pnl_pct': pnl_pct,
                'profitable': pnl_pct > 0
            })
            results[exit_min][direction].append({
                'pnl_pct': pnl_pct,
                'profitable': pnl_pct > 0
            })

    conn.close()

    # Print results
    print(f"\n{'='*80}")
    print("RESULTS SUMMARY")
    print(f"{'='*80}\n")

    print(f"{'Exit Time':<12} {'Trades':<10} {'Win Rate':<12} {'Avg P&L':<12} {'Total P&L':<12} {'Avg Win':<12} {'Avg Loss':<12}")
    print("-" * 80)

    for exit_min in exit_times:
        trades = results[exit_min]['all']
        if not trades:
            continue

        winners = [t for t in trades if t['profitable']]
        losers = [t for t in trades if not t['profitable']]

        win_rate = len(winners) / len(trades) * 100
        all_pnl = [t['pnl_pct'] for t in trades]
        avg_pnl = statistics.mean(all_pnl)
        total_pnl = sum(all_pnl)

        avg_win = statistics.mean([t['pnl_pct'] for t in winners]) if winners else 0
        avg_loss = statistics.mean([t['pnl_pct'] for t in losers]) if losers else 0

        print(f"{exit_min} min{'':<6} {len(trades):<10} {win_rate:.1f}%{'':<6} {avg_pnl:+.2f}%{'':<6} {total_pnl:+.1f}%{'':<6} {avg_win:+.2f}%{'':<6} {avg_loss:+.2f}%")

    # By direction
    print(f"\n{'='*80}")
    print("BY DIRECTION")
    print(f"{'='*80}\n")

    for direction in ['DROP', 'RISE']:
        print(f"\n{direction} ALERTS:")
        print(f"{'Exit Time':<12} {'Trades':<10} {'Win Rate':<12} {'Avg P&L':<12} {'Total P&L':<12}")
        print("-" * 60)

        for exit_min in exit_times:
            trades = results[exit_min][direction]
            if not trades:
                continue

            winners = [t for t in trades if t['profitable']]
            win_rate = len(winners) / len(trades) * 100
            all_pnl = [t['pnl_pct'] for t in trades]
            avg_pnl = statistics.mean(all_pnl)
            total_pnl = sum(all_pnl)

            print(f"{exit_min} min{'':<6} {len(trades):<10} {win_rate:.1f}%{'':<6} {avg_pnl:+.2f}%{'':<6} {total_pnl:+.1f}%")

    # P&L Distribution
    print(f"\n{'='*80}")
    print("P&L DISTRIBUTION (10 min exit)")
    print(f"{'='*80}\n")

    if results[10]['all']:
        trades_10 = results[10]['all']

        ranges = [
            ('Big Loss (< -2%)', lambda x: x < -2),
            ('Loss (-2% to -1%)', lambda x: -2 <= x < -1),
            ('Small Loss (-1% to 0%)', lambda x: -1 <= x < 0),
            ('Small Win (0% to 1%)', lambda x: 0 <= x < 1),
            ('Win (1% to 2%)', lambda x: 1 <= x < 2),
            ('Big Win (> 2%)', lambda x: x >= 2),
        ]

        for label, condition in ranges:
            count = sum(1 for t in trades_10 if condition(t['pnl_pct']))
            pct = count / len(trades_10) * 100
            bar = '█' * int(pct / 2)
            print(f"  {label:<25} {count:>4} ({pct:>5.1f}%) {bar}")

    # Top winners and losers
    print(f"\n{'='*80}")
    print("TOP 10 WINNERS (10 min exit)")
    print(f"{'='*80}")

    if results[10]['all']:
        sorted_trades = sorted(results[10]['all'], key=lambda x: -x['pnl_pct'])
        for t in sorted_trades[:10]:
            print(f"  {t['symbol']:<12} {t['direction']:<5} ₹{t['entry_price']:>8.2f} → ₹{t['exit_price']:>8.2f}  P&L: {t['pnl_pct']:+.2f}%")

    print(f"\n{'='*80}")
    print("TOP 10 LOSERS (10 min exit)")
    print(f"{'='*80}")

    if results[10]['all']:
        sorted_trades = sorted(results[10]['all'], key=lambda x: x['pnl_pct'])
        for t in sorted_trades[:10]:
            print(f"  {t['symbol']:<12} {t['direction']:<5} ₹{t['entry_price']:>8.2f} → ₹{t['exit_price']:>8.2f}  P&L: {t['pnl_pct']:+.2f}%")

    # Recommendation
    print(f"\n{'='*80}")
    print("RECOMMENDATION")
    print(f"{'='*80}\n")

    # Find best exit time
    best_exit = None
    best_pnl = -999

    for exit_min in exit_times:
        trades = results[exit_min]['all']
        if trades:
            avg_pnl = statistics.mean([t['pnl_pct'] for t in trades])
            if avg_pnl > best_pnl:
                best_pnl = avg_pnl
                best_exit = exit_min

    if best_exit:
        trades = results[best_exit]['all']
        winners = [t for t in trades if t['profitable']]
        win_rate = len(winners) / len(trades) * 100
        total_pnl = sum([t['pnl_pct'] for t in trades])

        print(f"Best Exit Time: {best_exit} minutes")
        print(f"  Win Rate: {win_rate:.1f}%")
        print(f"  Average P&L: {best_pnl:+.2f}%")
        print(f"  Total P&L: {total_pnl:+.1f}%")

        # Check direction performance
        for direction in ['DROP', 'RISE']:
            dir_trades = results[best_exit][direction]
            if dir_trades:
                dir_winners = [t for t in dir_trades if t['profitable']]
                dir_win_rate = len(dir_winners) / len(dir_trades) * 100
                dir_avg_pnl = statistics.mean([t['pnl_pct'] for t in dir_trades])

                if dir_win_rate >= 50 and dir_avg_pnl > 0:
                    status = "✅ PROFITABLE"
                elif dir_win_rate >= 45:
                    status = "⚠️ MARGINAL"
                else:
                    status = "❌ UNPROFITABLE"

                print(f"\n  {direction}: {status}")
                print(f"    Win Rate: {dir_win_rate:.1f}%, Avg P&L: {dir_avg_pnl:+.2f}%")

    print(f"\n{'='*80}\n")

    return results


def fetch_historical_results(days: int = 30):
    """Fetch historical quarterly results from NSE corporate announcements."""
    import requests
    import json
    from datetime import datetime, timedelta

    print("Fetching historical results from NSE...")

    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
        'Accept': 'application/json',
        'Referer': 'https://www.nseindia.com/',
    }

    session = requests.Session()
    session.headers.update(headers)

    # Get cookies
    try:
        session.get('https://www.nseindia.com', timeout=10)
    except:
        pass

    from_date = (datetime.now() - timedelta(days=days)).strftime("%d-%m-%Y")
    to_date = datetime.now().strftime("%d-%m-%Y")

    url = f"https://www.nseindia.com/api/corporate-announcements?index=equities&from_date={from_date}&to_date={to_date}"

    try:
        resp = session.get(url, timeout=30)
        if resp.status_code == 200:
            data = resp.json()

            # Filter for Financial Results and Board Meetings
            by_symbol = {}
            for item in data:
                desc = item.get('desc', '').lower()
                if 'financial result' in desc or 'quarterly result' in desc or 'board meeting' in desc:
                    symbol = item.get('symbol', '')
                    sort_date = item.get('sort_date', '')

                    try:
                        date_obj = datetime.strptime(sort_date.split()[0], '%Y-%m-%d')
                        date_str = date_obj.strftime('%Y-%m-%d')
                    except:
                        continue

                    if symbol not in by_symbol:
                        by_symbol[symbol] = []
                    if date_str not in by_symbol[symbol]:
                        by_symbol[symbol].append(date_str)

            # Sort dates for each symbol
            for symbol in by_symbol:
                by_symbol[symbol] = sorted(by_symbol[symbol])

            # Save to file
            output = {'results_dates': by_symbol, 'fetched_at': datetime.now().isoformat()}
            with open('data/results_cache/historical_results.json', 'w') as f:
                json.dump(output, f, indent=2)

            print(f"Saved {len(by_symbol)} stocks with results/board meetings to historical_results.json")
            return True
        else:
            print(f"Failed to fetch: HTTP {resp.status_code}")
            return False
    except Exception as e:
        print(f"Error: {e}")
        return False


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--days", type=int, default=30)
    parser.add_argument("--fetch-history", action="store_true", help="Fetch historical results from NSE")
    args = parser.parse_args()

    if args.fetch_history:
        fetch_historical_results(args.days)

    run_backtest(args.days)
