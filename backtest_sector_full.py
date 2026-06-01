#!/usr/bin/env python3
"""
Full sector-context backtest using recovered alert_tracking.xlsx data.

Hypothesis:
  DROP alerts: stock dropping while sector is flat/green (stock-specific) →
               better short than dropping with an already-weak sector.
  RISE alerts: stock rising while sector is flat/red (stock-specific) →
               better long than rising with an already-strong sector.

Data: ~1500 drop + ~1100 rise alerts from alert_tracking.xlsx (Oct 2025 – May 2026)
      with recovered Price 2min / Price 10min / Price EOD.
Sector state: Kite 5-min sector index candles, % change from day open to alert time.

P&L (drops): (alert_price - price_10min) / alert_price × 100  — positive = stock fell = WIN
P&L (rises): (price_10min - alert_price) / alert_price × 100  — positive = stock rose = WIN

Run: venv/bin/python3 backtest_sector_full.py
"""

import os, json, time, logging
from datetime import datetime, timedelta, date, time as dt_time
from collections import defaultdict

import openpyxl
from kiteconnect import KiteConnect

import config

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

TRACKING_FILE   = "data/alerts/alert_tracking.xlsx"
SECTOR_MAP_FILE = "data/stock_sectors.json"
CANDLE_CACHE    = "data/backtest_sector_full_cache.json"

SECTOR_INDEX_TOKENS = {
    "BANKING":            260105,
    "FINANCIAL_SERVICES": 260105,
    "IT":                 5459490,
    "PHARMA":             5577794,
    "AUTO":               5582082,
    "ENERGY":              14385,
    "INFRASTRUCTURE":      14385,
    "METAL":              234249,
    "FMCG":             6426369,
    "CONSUMER":         6426369,
}
NIFTY50_TOKEN = config.NIFTY_50_TOKEN

SHEETS = [
    ("10min_alerts",        14, 15, 16),
    ("5min_alerts",         14, 15, 16),
    ("30min_alerts",        14, 15, 16),
    ("Volume_Spike_alerts", 14, 15, 16),
]

RISING_THRESH  =  0.30
FALLING_THRESH = -0.30
MAX_PNL_PCT    =  15.0


def load_alerts(sector_map: dict, direction: str) -> list:
    """Load drop or rise alerts. direction = 'drop' | 'rise'"""
    wb = openpyxl.load_workbook(TRACKING_FILE)
    alerts = []
    for sheet_name, p2c, p10c, peodc in SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            row_dir = str(row[3]).lower() if row[3] else ""
            if direction not in row_dir:
                continue

            symbol      = str(row[2]) if row[2] else ""
            alert_price = row[4]
            price_10min = row[p10c - 1]
            price_eod   = row[peodc - 1]

            if not symbol or not alert_price or not price_10min:
                continue

            try:
                ap   = float(alert_price)
                p10  = float(price_10min)
                peod = float(price_eod) if price_eod else None
            except (TypeError, ValueError):
                continue

            if ap <= 0:
                continue

            # P&L: positive = WIN for the signal direction
            if direction == "drop":
                pnl_10m = (ap - p10) / ap * 100
                pnl_eod = (ap - peod) / ap * 100 if peod else None
            else:  # rise
                pnl_10m = (p10 - ap) / ap * 100
                pnl_eod = (peod - ap) / ap * 100 if peod else None

            if abs(pnl_10m) > MAX_PNL_PCT:
                continue

            date_val = row[0]
            time_val = row[1]
            try:
                if isinstance(date_val, str):
                    alert_date = date.fromisoformat(str(date_val)[:10])
                elif isinstance(date_val, datetime):
                    alert_date = date_val.date()
                else:
                    alert_date = date_val
            except Exception:
                continue

            try:
                if isinstance(time_val, str):
                    parts = time_val.replace(":", " ").split()
                    alert_time = dt_time(int(parts[0]), int(parts[1]))
                elif isinstance(time_val, dt_time):
                    alert_time = time_val
                elif isinstance(time_val, datetime):
                    alert_time = time_val.time()
                else:
                    alert_time = dt_time(9, 25)
            except Exception:
                alert_time = dt_time(9, 25)

            sector = sector_map.get(symbol)
            token  = SECTOR_INDEX_TOKENS.get(sector, NIFTY50_TOKEN)

            alerts.append({
                "symbol":      symbol,
                "sector":      sector,
                "token":       token,
                "date":        alert_date,
                "time":        alert_time,
                "alert_price": ap,
                "pnl_10m":     pnl_10m,
                "pnl_eod":     pnl_eod,
                "sheet":       sheet_name,
            })

    logger.info(f"Loaded {len(alerts)} {direction} alerts")
    return alerts


def load_cache():
    if os.path.exists(CANDLE_CACHE):
        with open(CANDLE_CACHE) as f:
            return json.load(f)
    return {}


def save_cache(cache):
    with open(CANDLE_CACHE, "w") as f:
        json.dump(cache, f)


def get_sector_pct(kite, token: int, alert_date: date, alert_time: dt_time, cache: dict) -> float | None:
    key = f"{token}_{alert_date}"
    if key not in cache:
        try:
            from_dt = datetime.combine(alert_date, dt_time(9, 0))
            to_dt   = datetime.combine(alert_date, dt_time(15, 35))
            raw = kite.historical_data(token, from_dt, to_dt, "5minute")
            cache[key] = [
                {
                    "date":  c["date"].isoformat() if hasattr(c["date"], "isoformat") else str(c["date"]),
                    "open":  c["open"],
                    "close": c["close"],
                }
                for c in raw
            ]
            time.sleep(0.30)
        except Exception as e:
            logger.debug(f"5min fetch failed token={token} date={alert_date}: {e}")
            cache[key] = []

    candles = cache[key]
    if not candles:
        return None

    day_open = candles[0]["open"]
    if day_open == 0:
        return None

    price_at = None
    for c in candles:
        try:
            ct = datetime.fromisoformat(c["date"]).time() if isinstance(c["date"], str) else c["date"].time()
        except Exception:
            continue
        if ct <= alert_time:
            price_at = c["close"]
        else:
            break

    if price_at is None:
        return None
    return (price_at - day_open) / day_open * 100


def categorize(sector_pct: float) -> str:
    if sector_pct > RISING_THRESH:
        return "RISING"
    if sector_pct < FALLING_THRESH:
        return "FALLING"
    return "FLAT"


def calc_stats(values: list) -> dict:
    vals = [v for v in values if v is not None]
    if not vals:
        return {"n": 0, "win_rate": None, "avg": None, "avg_win": 0, "avg_loss": 0, "n_wins": 0, "n_losses": 0}
    wins   = [v for v in vals if v > 0]
    losses = [v for v in vals if v <= 0]
    return {
        "n":        len(vals),
        "win_rate": len(wins) / len(vals) * 100,
        "avg":      sum(vals) / len(vals),
        "avg_win":  sum(wins) / len(wins) if wins else 0,
        "avg_loss": sum(losses) / len(losses) if losses else 0,
        "n_wins":   len(wins),
        "n_losses": len(losses),
    }


def run_direction(kite, alerts: list, cache: dict, direction: str):
    buckets = {
        "RISING":  {"10m": [], "eod": [], "sector_pcts": []},
        "FLAT":    {"10m": [], "eod": [], "sector_pcts": []},
        "FALLING": {"10m": [], "eod": [], "sector_pcts": []},
    }
    skipped = 0

    for i, a in enumerate(alerts):
        pct = get_sector_pct(kite, a["token"], a["date"], a["time"], cache)
        if pct is None:
            skipped += 1
            continue

        cat = categorize(pct)
        buckets[cat]["10m"].append(a["pnl_10m"])
        buckets[cat]["eod"].append(a["pnl_eod"])
        buckets[cat]["sector_pcts"].append(pct)

        if (i + 1) % 100 == 0:
            logger.info(f"  [{direction}] [{i+1}/{len(alerts)}] "
                        f"RISING={len(buckets['RISING']['10m'])} "
                        f"FLAT={len(buckets['FLAT']['10m'])} "
                        f"FALLING={len(buckets['FALLING']['10m'])} "
                        f"skipped={skipped}")
            save_cache(cache)

    return buckets, skipped


def print_results(alerts: list, buckets: dict, skipped: int, direction: str):
    analyzed = len(alerts) - skipped
    label = "DROP" if direction == "drop" else "RISE"
    win_desc = "stock fell (short WIN)" if direction == "drop" else "stock rose (long WIN)"

    print(f"\n{'='*76}")
    print(f"SECTOR CONTEXT BACKTEST — {label} Alerts vs Sector State at Alert Time")
    print(f"{'='*76}")
    print(f"Total: {len(alerts)}  |  Analyzed: {analyzed}  |  Skipped (no data): {skipped}")
    print(f"P&L: positive = {win_desc}")

    print(f"\n{'─'*76}")
    print(f"{'Sector':>8}  {'N':>5}  {'AvgIdx%':>8}  {'WR 10m':>7}  {'Avg 10m':>8}  {'WR EOD':>7}  {'Avg EOD':>8}")
    print(f"{'─'*76}")

    for cat in ["RISING", "FLAT", "FALLING"]:
        b    = buckets[cat]
        s10  = calc_stats(b["10m"])
        seod = calc_stats([v for v in b["eod"] if v is not None])
        sp   = b["sector_pcts"]
        avg_sp = sum(sp) / len(sp) if sp else 0

        wr10   = f"{s10['win_rate']:.0f}%"  if s10["win_rate"]  is not None else "N/A"
        avg10  = f"{s10['avg']:+.3f}%"      if s10["avg"]        is not None else "N/A"
        wreod  = f"{seod['win_rate']:.0f}%" if seod["win_rate"] is not None else "N/A"
        avgeod = f"{seod['avg']:+.3f}%"     if seod["avg"]       is not None else "N/A"

        print(f"{cat:>8}  {s10['n']:>5}  {avg_sp:>+7.2f}%  {wr10:>7}  {avg10:>8}  {wreod:>7}  {avgeod:>8}")

    rising  = calc_stats(buckets["RISING"]["10m"])
    falling = calc_stats(buckets["FALLING"]["10m"])
    flat    = calc_stats(buckets["FLAT"]["10m"])

    print(f"\n{'─'*76}")
    print("DETAIL (10-min P&L):")
    for cat, s in [("RISING", rising), ("FLAT", flat), ("FALLING", falling)]:
        if s["n"] > 0:
            print(f"  {cat:<8}: n={s['n']:>4}  WR={s['win_rate']:.0f}%  avg={s['avg']:+.3f}%  "
                  f"wins={s['n_wins']} ({s['avg_win']:+.3f}%)  losses={s['n_losses']} ({s['avg_loss']:+.3f}%)")

    print()
    if rising["n"] >= 10 and falling["n"] >= 10:
        wr_diff  = (rising["win_rate"] or 0) - (falling["win_rate"] or 0)
        avg_diff = (rising["avg"] or 0) - (falling["avg"] or 0)

        if direction == "drop":
            # For drops: RISING sector (stock-specific) is hypothesized to be better
            if wr_diff > 5:
                verdict = f"✅ CONFIRMED (+{wr_diff:.0f}pp): Stock-specific drops (vs rising sector) are better shorts"
            elif wr_diff > 2:
                verdict = f"⚠️  MARGINAL (+{wr_diff:.0f}pp): Slight edge for stock-specific drops"
            elif wr_diff < -5:
                verdict = f"❌ REVERSED ({wr_diff:.0f}pp): Sector weakness produced better shorts (momentum)"
            else:
                verdict = f"➖ NEUTRAL ({wr_diff:+.0f}pp): No meaningful difference from sector context"
        else:
            # For rises: FALLING sector (stock-specific) is hypothesized to be better
            wr_diff_rise = (falling["win_rate"] or 0) - (rising["win_rate"] or 0)
            avg_diff_rise = (falling["avg"] or 0) - (rising["avg"] or 0)
            if wr_diff_rise > 5:
                verdict = f"✅ CONFIRMED (+{wr_diff_rise:.0f}pp): Stock-specific rises (vs falling sector) are better longs"
            elif wr_diff_rise > 2:
                verdict = f"⚠️  MARGINAL (+{wr_diff_rise:.0f}pp): Slight edge for stock-specific rises"
            elif wr_diff_rise < -5:
                verdict = f"❌ REVERSED ({wr_diff_rise:.0f}pp): Sector momentum produced better longs"
            else:
                verdict = f"➖ NEUTRAL ({wr_diff_rise:+.0f}pp): No meaningful difference from sector context"
            wr_diff = wr_diff_rise

        print(f"  {verdict}")
        print(f"  Avg 10m: RISING={rising['avg']:+.3f}%  FLAT={flat['avg']:+.3f}%  FALLING={falling['avg']:+.3f}%")
        print(f"  Win rate: RISING={rising['win_rate']:.0f}%  FLAT={flat['win_rate']:.0f}%  FALLING={falling['win_rate']:.0f}%")
    else:
        print(f"  ⚠️  Insufficient data — RISING={rising['n']}, FLAT={flat['n']}, FALLING={falling['n']}")

    print(f"\n{'─'*76}")
    print("BY ALERT TYPE:")
    for sheet_name, _, _, _ in SHEETS:
        sheet_alerts = [a for a in alerts if a["sheet"] == sheet_name]
        pnls = [a["pnl_10m"] for a in sheet_alerts]
        if not pnls: continue
        wins = [p for p in pnls if p > 0]
        print(f"  {sheet_name:<22}: n={len(pnls):>4}  WR={len(wins)/len(pnls)*100:.0f}%  avg={sum(pnls)/len(pnls):+.3f}%")

    print(f"{'='*76}")


def run():
    if not config.KITE_API_KEY or not config.KITE_ACCESS_TOKEN:
        print("ERROR: KITE_API_KEY and KITE_ACCESS_TOKEN required.")
        return

    kite = KiteConnect(api_key=config.KITE_API_KEY)
    kite.set_access_token(config.KITE_ACCESS_TOKEN)
    logger.info("Kite connected")

    with open(SECTOR_MAP_FILE) as f:
        sector_map = json.load(f)

    cache = load_cache()

    # ── DROP alerts ────────────────────────────────────────────────────────────
    drop_alerts = load_alerts(sector_map, "drop")
    drop_buckets, drop_skipped = run_direction(kite, drop_alerts, cache, "drop")
    save_cache(cache)

    # ── RISE alerts ────────────────────────────────────────────────────────────
    rise_alerts = load_alerts(sector_map, "rise")
    rise_buckets, rise_skipped = run_direction(kite, rise_alerts, cache, "rise")
    save_cache(cache)

    # ── Print results ──────────────────────────────────────────────────────────
    print_results(drop_alerts, drop_buckets, drop_skipped, "drop")
    print_results(rise_alerts, rise_buckets, rise_skipped, "rise")

    # ── Cross-direction summary ────────────────────────────────────────────────
    print(f"\n{'='*76}")
    print("CROSS-DIRECTION SUMMARY — Best sector state per signal type")
    print(f"{'─'*76}")
    print(f"{'Signal':<14}  {'RISING sector':>16}  {'FLAT sector':>14}  {'FALLING sector':>16}")
    print(f"{'─'*76}")
    for direction, buckets in [("DROP (short)", drop_buckets), ("RISE (long)", rise_buckets)]:
        row = f"{direction:<14}"
        for cat in ["RISING", "FLAT", "FALLING"]:
            s = calc_stats(buckets[cat]["10m"])
            if s["n"] > 0:
                row += f"  {s['win_rate']:>5.0f}% WR / {s['avg']:>+6.3f}%"
            else:
                row += f"  {'N/A':>16}"
        print(row)
    print(f"{'='*76}\n")


if __name__ == "__main__":
    run()
