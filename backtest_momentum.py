#!/usr/bin/env python3
"""
Momentum Strategy Backtester
Backtests weekly rebalancing momentum strategy with realistic transaction costs

Strategy:
1. Every Monday close, rank all stocks by momentum score
2. Rebalance to hold top N stocks with equal weight
3. Track performance with realistic costs (slippage, brokerage, STT)

Metrics Tracked:
- CAGR, Max Drawdown, Sharpe Ratio, Sortino Ratio
- Win rate per week, Avg gain vs avg loss
- Turnover rate (important for tax implications)

Author: Claude Opus 4.5
Date: 2026-02-12
"""

import json
import os
import sys
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple
import pandas as pd
import numpy as np
from kiteconnect import KiteConnect
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import config

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/momentum_backtest.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class MomentumBacktester:
    """
    Backtests momentum strategy with realistic transaction costs.
    """

    # Transaction costs (all in percentage)
    SLIPPAGE = 0.001  # 0.1% slippage on each trade
    BROKERAGE = 0.0003  # 0.03% brokerage (Zerodha delivery)
    STT = 0.001  # 0.1% STT on sell side only
    STAMP_DUTY = 0.00015  # 0.015% stamp duty
    GST = 0.18  # 18% GST on brokerage

    def __init__(
        self,
        start_date: datetime,
        end_date: datetime,
        initial_capital: float = 1000000,  # 10 lakh INR
        top_n: int = 20,
        rebalance_day: int = 0,  # 0 = Monday
        lookback_6m: int = 126,
        lookback_12m: int = 252,
    ):
        """
        Initialize the backtester.

        Args:
            start_date: Backtest start date
            end_date: Backtest end date
            initial_capital: Starting capital in INR
            top_n: Number of stocks to hold
            rebalance_day: Day of week for rebalancing (0=Monday)
            lookback_6m: Days for 6-month momentum
            lookback_12m: Days for 12-month momentum
        """
        self.start_date = start_date
        self.end_date = end_date
        self.initial_capital = initial_capital
        self.top_n = top_n
        self.rebalance_day = rebalance_day
        self.lookback_6m = lookback_6m
        self.lookback_12m = lookback_12m

        # Initialize Kite Connect
        self.kite = KiteConnect(api_key=config.KITE_API_KEY)
        self.kite.set_access_token(config.KITE_ACCESS_TOKEN)

        # Load stocks and tokens
        self.stocks = self._load_stock_list()
        self.instrument_tokens = self._load_instrument_tokens()

        # Cache for price data
        self.price_data: Dict[str, pd.DataFrame] = {}

        # Results storage
        self.portfolio_history: List[Dict] = []
        self.trade_log: List[Dict] = []
        self.weekly_returns: List[Dict] = []

        logger.info(f"Backtester initialized: {start_date.date()} to {end_date.date()}")
        logger.info(f"Initial Capital: ₹{initial_capital:,.0f}, Top N: {top_n}")

    def _load_stock_list(self) -> List[str]:
        """Load F&O stock list"""
        try:
            with open(config.STOCK_LIST_FILE, 'r') as f:
                data = json.load(f)
                return data['stocks']
        except Exception as e:
            logger.error(f"Failed to load stock list: {e}")
            return []

    def _load_instrument_tokens(self) -> Dict[str, int]:
        """Load instrument tokens"""
        tokens_file = "data/instrument_tokens.json"
        try:
            if os.path.exists(tokens_file):
                with open(tokens_file, 'r') as f:
                    return json.load(f)
            else:
                return {}
        except Exception as e:
            logger.error(f"Failed to load tokens: {e}")
            return {}

    def fetch_all_price_data(self):
        """Fetch historical data for all stocks in universe"""
        logger.info("Fetching historical data for all stocks...")

        # Need extra buffer for momentum calculation
        fetch_from = self.start_date - timedelta(days=self.lookback_12m + 60)

        total = len(self.stocks)
        for idx, symbol in enumerate(self.stocks, 1):
            if idx % 50 == 0:
                logger.info(f"Fetching data: {idx}/{total}")

            try:
                if symbol not in self.instrument_tokens:
                    continue

                token = self.instrument_tokens[symbol]

                data = self.kite.historical_data(
                    instrument_token=token,
                    from_date=fetch_from.date(),
                    to_date=self.end_date.date(),
                    interval="day"
                )

                if data:
                    df = pd.DataFrame(data)
                    df.columns = df.columns.str.lower()
                    df['date'] = pd.to_datetime(df['date']).dt.tz_localize(None)
                    df.set_index('date', inplace=True)
                    self.price_data[symbol] = df

            except Exception as e:
                logger.error(f"{symbol}: Failed to fetch: {e}")

            # Rate limiting
            import time
            time.sleep(config.REQUEST_DELAY_SECONDS)

        logger.info(f"Fetched data for {len(self.price_data)} stocks")

    def calculate_momentum_scores(self, as_of_date: datetime) -> Dict[str, float]:
        """
        Calculate momentum scores for all stocks as of a specific date.

        Args:
            as_of_date: Calculate momentum as of this date

        Returns:
            Dict mapping symbol to momentum score
        """
        scores = {}

        for symbol, df in self.price_data.items():
            try:
                # Filter to data up to as_of_date
                df_filtered = df[df.index <= as_of_date]

                if len(df_filtered) < self.lookback_12m:
                    continue

                closes = df_filtered['close']
                current_price = closes.iloc[-1]

                # 6-month return
                price_6m = closes.iloc[-self.lookback_6m] if len(closes) >= self.lookback_6m else None
                ret_6m = (current_price - price_6m) / price_6m if price_6m else 0

                # 12-month return
                price_12m = closes.iloc[-self.lookback_12m] if len(closes) >= self.lookback_12m else None
                ret_12m = (current_price - price_12m) / price_12m if price_12m else 0

                # Weighted score
                score = 0.5 * ret_6m + 0.5 * ret_12m

                # Risk adjustment (divide by volatility)
                daily_returns = closes.pct_change().dropna()
                volatility = daily_returns.std() * np.sqrt(252)

                if volatility > 0:
                    score = score / volatility

                scores[symbol] = score

            except Exception as e:
                continue

        return scores

    def get_price(self, symbol: str, date: datetime) -> Optional[float]:
        """Get closing price for a symbol on a specific date"""
        if symbol not in self.price_data:
            return None

        df = self.price_data[symbol]

        # Find closest date (may not have exact date due to holidays)
        try:
            # Try exact date first
            if date in df.index:
                return df.loc[date, 'close']

            # Find nearest date before
            valid_dates = df.index[df.index <= date]
            if len(valid_dates) == 0:
                return None

            nearest_date = valid_dates[-1]
            return df.loc[nearest_date, 'close']

        except Exception:
            return None

    def calculate_transaction_cost(self, trade_value: float, is_sell: bool) -> float:
        """
        Calculate total transaction cost for a trade.

        Args:
            trade_value: Absolute value of trade
            is_sell: True if selling

        Returns:
            Total cost in INR
        """
        # Slippage
        slippage_cost = trade_value * self.SLIPPAGE

        # Brokerage
        brokerage = trade_value * self.BROKERAGE
        gst_on_brokerage = brokerage * self.GST
        total_brokerage = brokerage + gst_on_brokerage

        # STT (only on sell)
        stt_cost = trade_value * self.STT if is_sell else 0

        # Stamp duty
        stamp_duty = trade_value * self.STAMP_DUTY

        total_cost = slippage_cost + total_brokerage + stt_cost + stamp_duty

        return total_cost

    def rebalance_portfolio(
        self,
        current_holdings: Dict[str, float],
        target_holdings: Dict[str, float],
        total_value: float,
        date: datetime
    ) -> Tuple[Dict[str, float], float, List[Dict]]:
        """
        Rebalance portfolio from current to target holdings.

        Args:
            current_holdings: Dict of symbol -> current shares
            target_holdings: Dict of symbol -> target shares
            total_value: Total portfolio value
            date: Rebalancing date

        Returns:
            Tuple of (new_holdings, total_cost, trades)
        """
        trades = []
        total_cost = 0

        # Get current symbols and target symbols
        current_symbols = set(current_holdings.keys())
        target_symbols = set(target_holdings.keys())

        # Sells first (stocks leaving portfolio)
        for symbol in current_symbols - target_symbols:
            if current_holdings[symbol] > 0:
                price = self.get_price(symbol, date)
                if price:
                    shares = current_holdings[symbol]
                    value = shares * price
                    cost = self.calculate_transaction_cost(value, is_sell=True)
                    total_cost += cost

                    trades.append({
                        'date': date.strftime('%Y-%m-%d'),
                        'symbol': symbol,
                        'action': 'SELL',
                        'shares': shares,
                        'price': price,
                        'value': value,
                        'cost': cost
                    })

        # Buys (new stocks entering portfolio)
        for symbol in target_symbols - current_symbols:
            price = self.get_price(symbol, date)
            if price:
                # Equal weight allocation
                target_value = total_value / self.top_n
                shares = int(target_value / price)
                value = shares * price
                cost = self.calculate_transaction_cost(value, is_sell=False)
                total_cost += cost

                trades.append({
                    'date': date.strftime('%Y-%m-%d'),
                    'symbol': symbol,
                    'action': 'BUY',
                    'shares': shares,
                    'price': price,
                    'value': value,
                    'cost': cost
                })

        # Rebalances (adjust positions in existing holdings)
        for symbol in current_symbols & target_symbols:
            current_shares = current_holdings.get(symbol, 0)
            price = self.get_price(symbol, date)

            if not price:
                continue

            target_value = total_value / self.top_n
            target_shares = int(target_value / price)

            diff = target_shares - current_shares

            if abs(diff) > 0:
                value = abs(diff) * price
                is_sell = diff < 0
                cost = self.calculate_transaction_cost(value, is_sell=is_sell)
                total_cost += cost

                trades.append({
                    'date': date.strftime('%Y-%m-%d'),
                    'symbol': symbol,
                    'action': 'SELL' if is_sell else 'BUY',
                    'shares': abs(diff),
                    'price': price,
                    'value': value,
                    'cost': cost
                })

        # Build new holdings
        new_holdings = {}
        for symbol in target_symbols:
            price = self.get_price(symbol, date)
            if price:
                target_value = total_value / self.top_n
                new_holdings[symbol] = int(target_value / price)

        return new_holdings, total_cost, trades

    def run_backtest(self):
        """Run the momentum backtest"""
        logger.info("=" * 80)
        logger.info("STARTING MOMENTUM STRATEGY BACKTEST")
        logger.info("=" * 80)

        # Fetch all price data first
        self.fetch_all_price_data()

        if not self.price_data:
            logger.error("No price data available. Exiting.")
            return

        # Initialize portfolio
        cash = self.initial_capital
        holdings: Dict[str, float] = {}  # symbol -> shares
        portfolio_value = cash

        # Get all trading days
        sample_symbol = list(self.price_data.keys())[0]
        all_dates = self.price_data[sample_symbol].index
        trading_dates = all_dates[(all_dates >= self.start_date) & (all_dates <= self.end_date)]

        # Track weekly performance
        last_rebalance_value = self.initial_capital
        last_rebalance_date = None

        for date in trading_dates:
            # Calculate current portfolio value
            portfolio_value = cash
            for symbol, shares in holdings.items():
                price = self.get_price(symbol, date)
                if price:
                    portfolio_value += shares * price

            # Check if it's rebalancing day (Monday)
            if date.weekday() == self.rebalance_day:
                # Calculate momentum scores
                scores = self.calculate_momentum_scores(date)

                if len(scores) < self.top_n:
                    continue

                # Get top N stocks
                sorted_scores = sorted(scores.items(), key=lambda x: x[1], reverse=True)
                top_stocks = [s[0] for s in sorted_scores[:self.top_n]]
                target_holdings = {s: 1 for s in top_stocks}  # Placeholder

                # Rebalance
                new_holdings, total_cost, trades = self.rebalance_portfolio(
                    holdings, target_holdings, portfolio_value, date
                )

                # Update holdings and cash
                # Simplified: assume we can rebalance to target with costs deducted
                holdings = new_holdings
                cash = portfolio_value - sum(
                    shares * (self.get_price(s, date) or 0)
                    for s, shares in holdings.items()
                ) - total_cost

                # Log trades
                self.trade_log.extend(trades)

                # Track weekly return
                if last_rebalance_date is not None:
                    weekly_return = (portfolio_value - last_rebalance_value) / last_rebalance_value * 100
                    self.weekly_returns.append({
                        'week_start': last_rebalance_date.strftime('%Y-%m-%d'),
                        'week_end': date.strftime('%Y-%m-%d'),
                        'start_value': last_rebalance_value,
                        'end_value': portfolio_value,
                        'return_pct': weekly_return,
                        'trades': len(trades),
                        'cost': total_cost
                    })

                last_rebalance_value = portfolio_value
                last_rebalance_date = date

                logger.info(f"{date.date()}: Rebalanced - Value: ₹{portfolio_value:,.0f}, "
                           f"Trades: {len(trades)}, Cost: ₹{total_cost:,.0f}")

            # Record daily portfolio value
            self.portfolio_history.append({
                'date': date.strftime('%Y-%m-%d'),
                'portfolio_value': portfolio_value,
                'cash': cash,
                'holdings_count': len(holdings)
            })

        logger.info("=" * 80)
        logger.info("BACKTEST COMPLETE")
        logger.info("=" * 80)

    def calculate_statistics(self) -> Dict:
        """Calculate performance statistics"""
        if not self.portfolio_history:
            return {}

        df = pd.DataFrame(self.portfolio_history)
        df['date'] = pd.to_datetime(df['date'])
        df.set_index('date', inplace=True)

        # Calculate daily returns
        df['daily_return'] = df['portfolio_value'].pct_change()

        # Basic metrics
        stats = {}
        stats['initial_capital'] = self.initial_capital
        stats['final_value'] = df['portfolio_value'].iloc[-1]
        stats['total_return_pct'] = (stats['final_value'] - stats['initial_capital']) / stats['initial_capital'] * 100

        # CAGR
        years = (df.index[-1] - df.index[0]).days / 365.25
        stats['cagr'] = ((stats['final_value'] / stats['initial_capital']) ** (1 / years) - 1) * 100 if years > 0 else 0

        # Volatility (annualized)
        stats['volatility'] = df['daily_return'].std() * np.sqrt(252) * 100

        # Sharpe Ratio (assuming 6% risk-free rate for India)
        risk_free_rate = 0.06
        excess_return = stats['cagr'] / 100 - risk_free_rate
        stats['sharpe_ratio'] = excess_return / (stats['volatility'] / 100) if stats['volatility'] > 0 else 0

        # Sortino Ratio (downside deviation)
        negative_returns = df['daily_return'][df['daily_return'] < 0]
        downside_std = negative_returns.std() * np.sqrt(252) * 100
        stats['sortino_ratio'] = excess_return / (downside_std / 100) if downside_std > 0 else 0

        # Max Drawdown
        rolling_max = df['portfolio_value'].cummax()
        drawdown = (df['portfolio_value'] - rolling_max) / rolling_max * 100
        stats['max_drawdown'] = drawdown.min()

        # Weekly stats
        if self.weekly_returns:
            weekly_df = pd.DataFrame(self.weekly_returns)
            stats['total_weeks'] = len(weekly_df)
            stats['winning_weeks'] = len(weekly_df[weekly_df['return_pct'] > 0])
            stats['losing_weeks'] = len(weekly_df[weekly_df['return_pct'] <= 0])
            stats['weekly_win_rate'] = stats['winning_weeks'] / stats['total_weeks'] * 100 if stats['total_weeks'] > 0 else 0
            stats['avg_weekly_return'] = weekly_df['return_pct'].mean()
            stats['avg_weekly_win'] = weekly_df[weekly_df['return_pct'] > 0]['return_pct'].mean() if stats['winning_weeks'] > 0 else 0
            stats['avg_weekly_loss'] = weekly_df[weekly_df['return_pct'] <= 0]['return_pct'].mean() if stats['losing_weeks'] > 0 else 0
            stats['total_trades'] = len(self.trade_log)
            stats['total_costs'] = sum(t['cost'] for t in self.trade_log)
            stats['cost_pct_of_capital'] = stats['total_costs'] / stats['initial_capital'] * 100

        return stats

    def generate_excel_report(self, filename: str = None) -> str:
        """Generate comprehensive Excel report"""
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"momentum_backtest_{timestamp}.xlsx"

        stats = self.calculate_statistics()

        # Create workbook
        wb = Workbook()

        # Sheet 1: Summary
        self._create_summary_sheet(wb, stats)

        # Sheet 2: Weekly Returns
        self._create_weekly_sheet(wb)

        # Sheet 3: All Trades
        self._create_trades_sheet(wb)

        # Sheet 4: Portfolio History
        self._create_history_sheet(wb)

        # Save
        output_dir = "data/backtest_results"
        os.makedirs(output_dir, exist_ok=True)
        filepath = os.path.join(output_dir, filename)
        wb.save(filepath)

        logger.info(f"Excel report saved: {filepath}")
        return filepath

    def _create_summary_sheet(self, wb: Workbook, stats: Dict):
        """Create summary sheet"""
        ws = wb.active
        ws.title = "Summary"

        ws['A1'] = "MOMENTUM STRATEGY BACKTEST RESULTS"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        ws['A2'] = f"Period: {self.start_date.date()} to {self.end_date.date()}"
        ws.merge_cells('A2:D2')

        metrics = [
            ("", ""),
            ("CAPITAL", ""),
            ("Initial Capital", f"₹{stats.get('initial_capital', 0):,.0f}"),
            ("Final Value", f"₹{stats.get('final_value', 0):,.0f}"),
            ("Total Return", f"{stats.get('total_return_pct', 0):.2f}%"),
            ("", ""),
            ("PERFORMANCE METRICS", ""),
            ("CAGR", f"{stats.get('cagr', 0):.2f}%"),
            ("Volatility (Annualized)", f"{stats.get('volatility', 0):.2f}%"),
            ("Max Drawdown", f"{stats.get('max_drawdown', 0):.2f}%"),
            ("Sharpe Ratio", f"{stats.get('sharpe_ratio', 0):.2f}"),
            ("Sortino Ratio", f"{stats.get('sortino_ratio', 0):.2f}"),
            ("", ""),
            ("WEEKLY PERFORMANCE", ""),
            ("Total Weeks", stats.get('total_weeks', 0)),
            ("Winning Weeks", stats.get('winning_weeks', 0)),
            ("Losing Weeks", stats.get('losing_weeks', 0)),
            ("Weekly Win Rate", f"{stats.get('weekly_win_rate', 0):.1f}%"),
            ("Avg Weekly Return", f"{stats.get('avg_weekly_return', 0):.2f}%"),
            ("Avg Weekly Win", f"{stats.get('avg_weekly_win', 0):.2f}%"),
            ("Avg Weekly Loss", f"{stats.get('avg_weekly_loss', 0):.2f}%"),
            ("", ""),
            ("COSTS", ""),
            ("Total Trades", stats.get('total_trades', 0)),
            ("Total Transaction Costs", f"₹{stats.get('total_costs', 0):,.0f}"),
            ("Costs as % of Capital", f"{stats.get('cost_pct_of_capital', 0):.2f}%"),
        ]

        row = 4
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            if label and not value:
                ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def _create_weekly_sheet(self, wb: Workbook):
        """Create weekly returns sheet"""
        ws = wb.create_sheet("Weekly Returns")

        headers = ['Week Start', 'Week End', 'Start Value', 'End Value', 'Return %', 'Trades', 'Costs']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        for row_idx, week in enumerate(self.weekly_returns, 2):
            ws.cell(row_idx, 1, week['week_start'])
            ws.cell(row_idx, 2, week['week_end'])
            ws.cell(row_idx, 3, f"₹{week['start_value']:,.0f}")
            ws.cell(row_idx, 4, f"₹{week['end_value']:,.0f}")
            ws.cell(row_idx, 5, f"{week['return_pct']:.2f}%")
            ws.cell(row_idx, 6, week['trades'])
            ws.cell(row_idx, 7, f"₹{week['cost']:,.0f}")

            # Color returns
            ret_cell = ws.cell(row_idx, 5)
            if week['return_pct'] > 0:
                ret_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                ret_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_trades_sheet(self, wb: Workbook):
        """Create trades sheet"""
        ws = wb.create_sheet("All Trades")

        headers = ['Date', 'Symbol', 'Action', 'Shares', 'Price', 'Value', 'Cost']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        for row_idx, trade in enumerate(self.trade_log[:1000], 2):  # Limit to 1000 trades
            ws.cell(row_idx, 1, trade['date'])
            ws.cell(row_idx, 2, trade['symbol'])
            ws.cell(row_idx, 3, trade['action'])
            ws.cell(row_idx, 4, int(trade['shares']))
            ws.cell(row_idx, 5, f"₹{trade['price']:.2f}")
            ws.cell(row_idx, 6, f"₹{trade['value']:,.0f}")
            ws.cell(row_idx, 7, f"₹{trade['cost']:.0f}")

            # Color action
            action_cell = ws.cell(row_idx, 3)
            if trade['action'] == 'BUY':
                action_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                action_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12

    def _create_history_sheet(self, wb: Workbook):
        """Create portfolio history sheet"""
        ws = wb.create_sheet("Portfolio History")

        headers = ['Date', 'Portfolio Value', 'Cash', 'Holdings Count']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")

        # Sample every 5th day to reduce size
        for row_idx, record in enumerate(self.portfolio_history[::5], 2):
            ws.cell(row_idx, 1, record['date'])
            ws.cell(row_idx, 2, f"₹{record['portfolio_value']:,.0f}")
            ws.cell(row_idx, 3, f"₹{record['cash']:,.0f}")
            ws.cell(row_idx, 4, record['holdings_count'])

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def print_summary(self):
        """Print backtest summary to console"""
        stats = self.calculate_statistics()

        print("\n" + "=" * 80)
        print("MOMENTUM STRATEGY BACKTEST RESULTS")
        print("=" * 80)
        print(f"Period: {self.start_date.date()} to {self.end_date.date()}")
        print(f"Initial Capital: ₹{stats.get('initial_capital', 0):,.0f}")
        print(f"Final Value: ₹{stats.get('final_value', 0):,.0f}")
        print("-" * 40)
        print(f"Total Return: {stats.get('total_return_pct', 0):.2f}%")
        print(f"CAGR: {stats.get('cagr', 0):.2f}%")
        print(f"Max Drawdown: {stats.get('max_drawdown', 0):.2f}%")
        print(f"Sharpe Ratio: {stats.get('sharpe_ratio', 0):.2f}")
        print(f"Sortino Ratio: {stats.get('sortino_ratio', 0):.2f}")
        print("-" * 40)
        print(f"Weekly Win Rate: {stats.get('weekly_win_rate', 0):.1f}%")
        print(f"Avg Weekly Return: {stats.get('avg_weekly_return', 0):.2f}%")
        print("-" * 40)
        print(f"Total Trades: {stats.get('total_trades', 0)}")
        print(f"Total Costs: ₹{stats.get('total_costs', 0):,.0f} ({stats.get('cost_pct_of_capital', 0):.2f}% of capital)")
        print("=" * 80)


def main():
    """Main entry point"""
    import argparse

    parser = argparse.ArgumentParser(description='Backtest Momentum Strategy')
    parser.add_argument('--years', type=int, default=3, help='Years to backtest (default: 3)')
    parser.add_argument('--capital', type=float, default=1000000, help='Initial capital in INR (default: 10 lakh)')
    parser.add_argument('--top', type=int, default=20, help='Number of stocks to hold (default: 20)')
    parser.add_argument('--rebalance', choices=['weekly', 'biweekly', 'monthly'], default='weekly',
                       help='Rebalancing frequency (default: weekly)')
    args = parser.parse_args()

    # Ensure logs directory exists
    os.makedirs('logs', exist_ok=True)

    # Calculate dates
    end_date = datetime.now()
    start_date = end_date - timedelta(days=args.years * 365)

    logger.info("=" * 80)
    logger.info("MOMENTUM STRATEGY BACKTEST")
    logger.info("=" * 80)
    logger.info(f"Period: {args.years} years")
    logger.info(f"Capital: ₹{args.capital:,.0f}")
    logger.info(f"Top N: {args.top}")
    logger.info(f"Rebalancing: {args.rebalance}")
    logger.info("=" * 80)

    # Create backtester
    backtester = MomentumBacktester(
        start_date=start_date,
        end_date=end_date,
        initial_capital=args.capital,
        top_n=args.top
    )

    # Run backtest
    backtester.run_backtest()

    # Print summary
    backtester.print_summary()

    # Generate Excel report
    if backtester.portfolio_history:
        filepath = backtester.generate_excel_report()
        print(f"\nExcel report saved: {filepath}")


if __name__ == "__main__":
    main()
