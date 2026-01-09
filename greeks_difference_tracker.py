#!/usr/bin/env python3
"""
Greeks Difference Tracker - Intraday Greeks Change Analysis

Tracks changes in Delta, Theta, and Vega throughout the trading day by comparing
live Greeks against a 9:15 AM baseline for ATM and OTM strikes.

Output:
- Excel report with time-series differences (updated every 15 min)
- Single Telegram notification at 9:30 AM with cloud link
- Cloud storage (Google Drive/Dropbox) for multi-device access
"""

import os
import sys
import logging
from datetime import datetime, time, timedelta
from typing import Dict, List, Tuple, Optional
import json
import pickle

# External libraries
from kiteconnect import KiteConnect
import schedule

# Project imports
import config
from unified_data_cache import UnifiedDataCache
from telegram_notifier import TelegramNotifier
from black_scholes_greeks import BlackScholesGreeks

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class GreeksDifferenceTracker:
    """
    Tracks intraday changes in option Greeks (Delta, Theta, Vega) by comparing
    live values against a 9:15 AM baseline.

    Workflow:
    1. 9:15 AM: Capture baseline Greeks for all strikes
    2. Every 15 min: Fetch current Greeks, calculate differences, update Excel
    3. 9:30 AM: Upload to cloud storage, send Telegram notification with link
    4. 9:45 AM - 3:30 PM: Silent updates to cloud file
    """

    def __init__(self, kite: KiteConnect):
        """
        Initialize the Greeks Difference Tracker.

        Args:
            kite: Authenticated KiteConnect instance
        """
        self.kite = kite
        self.cache = UnifiedDataCache()
        self.telegram = TelegramNotifier()
        self.bs_calculator = BlackScholesGreeks()  # Black-Scholes calculator

        # Runtime state
        self.baseline_greeks: Dict = {}
        self.current_greeks: Dict = {}
        self.history: List[Dict] = []
        self.telegram_sent: bool = False
        self.cloud_link: Optional[str] = None
        self.current_vix: float = 0.10  # Default 10%
        self.current_threshold: float = 0.100  # Will be updated based on VIX

        # Configuration
        self.strike_offsets = config.GREEKS_DIFF_STRIKE_OFFSETS
        self.update_interval = config.GREEKS_UPDATE_INTERVAL_MINUTES
        self.market_start = config.GREEKS_MARKET_START
        self.market_end = config.GREEKS_MARKET_END
        self.report_dir = config.GREEKS_DIFF_REPORT_DIR

        logger.info("GreeksDifferenceTracker initialized with VIX-adaptive threshold")

    # ==================== PUBLIC METHODS ====================

    def capture_baseline_greeks(self) -> bool:
        """
        Capture baseline Greeks at 9:15 AM for all strikes.
        This becomes the static reference for the entire trading day.

        Returns:
            True if baseline captured successfully, False otherwise
        """
        logger.info("=" * 60)
        logger.info("CAPTURING BASELINE GREEKS AT 9:15 AM")
        logger.info("=" * 60)

        try:
            # Fetch India VIX and set adaptive threshold
            self.current_vix = self._get_india_vix()
            self.current_threshold = self._get_vix_adaptive_threshold(self.current_vix)

            # Get NIFTY spot price and determine ATM strike
            nifty_spot = self._get_nifty_spot_price()
            atm_strike = self._get_atm_strike(nifty_spot)

            # Get next week expiry
            expiries = self._get_next_expiries(count=1)
            if not expiries:
                raise ValueError("No valid expiries found")
            expiry = expiries[0]

            # Define all strikes (CE and PE)
            ce_strikes = [atm_strike + offset for offset in self.strike_offsets]
            pe_strikes = [atm_strike - offset for offset in self.strike_offsets]
            all_strikes = sorted(set(ce_strikes + pe_strikes))

            logger.info(f"India VIX: {self.current_vix*100:.2f}%")
            logger.info(f"Adaptive Delta Threshold: ±{self.current_threshold:.3f}")
            logger.info(f"NIFTY Spot: {nifty_spot}")
            logger.info(f"ATM Strike: {atm_strike}")
            logger.info(f"Expiry: {expiry.strftime('%Y-%m-%d')}")
            logger.info(f"CE Strikes: {ce_strikes}")
            logger.info(f"PE Strikes: {pe_strikes}")

            # Fetch Greeks for all strikes
            baseline = self._fetch_greeks_for_strikes(expiry, all_strikes)

            if not baseline or len(baseline) < 6:  # Need at least 3 CE + 3 PE
                raise ValueError(f"Insufficient baseline data: {len(baseline)} strikes")

            # Store baseline
            self.baseline_greeks = {
                'timestamp': datetime.now().isoformat(),
                'nifty_spot': nifty_spot,
                'atm_strike': atm_strike,
                'expiry': expiry.strftime('%Y-%m-%d'),
                'strikes': baseline
            }

            # Cache baseline for the day
            cache_key = config.GREEKS_BASELINE_CACHE_KEY.format(
                date=datetime.now().strftime('%Y%m%d')
            )
            self.cache.set_data(cache_key, self.baseline_greeks, 'greeks_diff')

            # Initialize history with baseline (all diffs = 0.00)
            self.history = [{
                'time': '09:15',
                'nifty': nifty_spot,
                'CE_delta': 0.00,
                'CE_theta': 0.00,
                'CE_vega': 0.00,
                'PE_delta': 0.00,
                'PE_theta': 0.00,
                'PE_vega': 0.00
            }]

            logger.info(f"✓ Baseline captured successfully: {len(baseline)} strikes")
            return True

        except Exception as e:
            logger.error(f"Failed to capture baseline: {e}", exc_info=True)
            return False

    def fetch_live_and_calculate_diff(self) -> Optional[Dict]:
        """
        Fetch live Greeks and calculate differences from baseline.
        Called every 15 minutes.

        Returns:
            Aggregated differences dict, or None if error
        """
        logger.info("Fetching live Greeks and calculating differences...")

        try:
            # Check if baseline exists
            if not self.baseline_greeks:
                self._load_baseline_from_cache()
                if not self.baseline_greeks:
                    logger.error("No baseline found. Run capture_baseline_greeks() first.")
                    return None

            # Get current NIFTY price
            nifty_spot = self._get_nifty_spot_price()

            # Fetch current Greeks for same strikes as baseline
            expiry_str = self.baseline_greeks['expiry']
            expiry = datetime.strptime(expiry_str, '%Y-%m-%d')
            strikes = list(self.baseline_greeks['strikes'].keys())

            current_greeks = self._fetch_greeks_for_strikes(expiry, strikes)

            if not current_greeks:
                logger.warning("Failed to fetch current Greeks")
                return None

            # Calculate differences
            differences = self._calculate_differences(current_greeks, self.baseline_greeks['strikes'])

            # Aggregate by option type
            aggregated = self._aggregate_by_type(differences)
            aggregated['timestamp'] = datetime.now().isoformat()
            aggregated['nifty_spot'] = nifty_spot

            # Generate prediction using VIX-adaptive threshold
            prediction, confidence = self.predict_daily_outcome(aggregated)
            aggregated['prediction'] = prediction
            aggregated['confidence'] = confidence
            aggregated['vix'] = self.current_vix
            aggregated['threshold'] = self.current_threshold

            # Append to history
            self._append_to_history(aggregated)

            logger.info(f"✓ Differences calculated: CE Δ = {aggregated['CE']['delta_diff_sum']:+.3f}, "
                       f"PE Δ = {aggregated['PE']['delta_diff_sum']:+.3f}")
            logger.info(f"✓ Prediction: {prediction} ({confidence*100:.1f}% confidence)")

            return aggregated

        except Exception as e:
            logger.error(f"Error in fetch_live_and_calculate_diff: {e}", exc_info=True)
            return None

    def export_to_excel(self) -> Optional[str]:
        """
        Generate formatted Excel report from history.

        Returns:
            Path to Excel file, or None if error
        """
        logger.info("Exporting to Excel...")

        try:
            # Import here to avoid dependency issues
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            if not self.history:
                logger.warning("No history data to export")
                return None

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Greeks Differences"

            # Headers
            headers = ['Time', 'NIFTY', 'CE Δ Diff', 'CE Θ Diff', 'CE V Diff',
                      'PE Δ Diff', 'PE Θ Diff', 'PE V Diff', 'Prediction', 'Confidence', 'VIX', 'Threshold']
            ws.append(headers)

            # Header formatting
            header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Data rows
            for row_data in self.history:
                ws.append([
                    row_data['time'],
                    row_data['nifty'],
                    row_data['CE_delta'],
                    row_data['CE_theta'],
                    row_data['CE_vega'],
                    row_data['PE_delta'],
                    row_data['PE_theta'],
                    row_data['PE_vega'],
                    row_data.get('prediction', ''),
                    row_data.get('confidence', 0),
                    row_data.get('vix', 0),
                    row_data.get('threshold', 0)
                ])

            # Format numbers and colors
            for row_idx in range(2, ws.max_row + 1):
                for col_idx in range(3, 9):  # Difference columns
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value

                    if isinstance(value, (int, float)):
                        cell.number_format = '0.00'

                        # Color based on positive/negative
                        if value > 0:
                            cell.font = Font(color="008000")  # Green
                        elif value < 0:
                            cell.font = Font(color="FF0000")  # Red

            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=8):
                for cell in row:
                    cell.border = thin_border

            # Freeze top row
            ws.freeze_panes = "A2"

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 15)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Create directory structure
            today = datetime.now()
            year_month_dir = os.path.join(
                self.report_dir,
                today.strftime('%Y'),
                today.strftime('%m')
            )
            os.makedirs(year_month_dir, exist_ok=True)

            # Save file
            filename = f"greeks_diff_{today.strftime('%Y%m%d')}.xlsx"
            filepath = os.path.join(year_month_dir, filename)
            wb.save(filepath)

            logger.info(f"✓ Excel exported: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}", exc_info=True)
            return None

    def send_telegram_notification(self, cloud_link: str) -> bool:
        """
        Send single Telegram notification with cloud link (9:30 AM only).

        Args:
            cloud_link: Shareable cloud storage link

        Returns:
            True if sent successfully, False otherwise
        """
        if self.telegram_sent:
            logger.info("Telegram already sent today. Skipping.")
            return False

        try:
            message = f"""📊 GREEKS DIFFERENCE TRACKER - LIVE REPORT

🎯 Tracking Started: 9:15 AM
📅 Date: {datetime.now().strftime('%Y-%m-%d')}

📄 Live Excel File (Google Drive):
{cloud_link}

⏰ Updates: Every 15 minutes (9:15 AM - 3:30 PM)
📊 Total Updates: 25 rows by end of day

💡 This file updates automatically in the cloud throughout the day.
   Click the link from ANY device (mobile/desktop) to see the latest Greeks differences!

🌐 Accessible from anywhere - no downloads needed!"""

            self.telegram._send_message(message)
            self.telegram_sent = True
            self.cloud_link = cloud_link

            logger.info(f"✓ Telegram notification sent with cloud link")
            return True

        except Exception as e:
            logger.error(f"Error sending Telegram: {e}", exc_info=True)
            return False

    def start_monitoring(self):
        """
        Main scheduler loop. Runs baseline capture at 9:15 AM and
        updates every 15 minutes thereafter.
        """
        logger.info("=" * 60)
        logger.info("STARTING GREEKS DIFFERENCE TRACKER")
        logger.info("=" * 60)

        # Schedule baseline capture at 9:15 AM
        schedule.every().day.at(self.market_start).do(self.capture_baseline_greeks)

        # Schedule updates every 15 minutes
        schedule.every(self.update_interval).minutes.do(self._scheduled_update)

        logger.info(f"Scheduler configured:")
        logger.info(f"  - Baseline: Daily at {self.market_start}")
        logger.info(f"  - Updates: Every {self.update_interval} minutes")

        # Main loop
        while True:
            try:
                if self._is_market_day() and self._is_market_hours():
                    schedule.run_pending()
                else:
                    logger.debug("Outside market hours. Sleeping...")

                # Check every minute
                import time
                time.sleep(60)

            except KeyboardInterrupt:
                logger.info("Shutting down gracefully...")
                break
            except Exception as e:
                logger.error(f"Error in main loop: {e}", exc_info=True)

    # ==================== PRIVATE METHODS ====================

    def _fetch_greeks_for_strikes(self, expiry: datetime, strikes: List[int]) -> Dict:
        """
        Fetch Delta, Theta, Vega for all strikes (both CE and PE).

        Args:
            expiry: Option expiry date
            strikes: List of strike prices

        Returns:
            Dict mapping strike to greeks: {strike: {'CE': {...}, 'PE': {...}}}
        """
        greeks_map = {}

        for strike in strikes:
            try:
                # Fetch CE option data
                ce_data = self._get_option_data('CE', expiry, strike)
                ce_greeks = ce_data.get('greeks', {})

                # Fetch PE option data
                pe_data = self._get_option_data('PE', expiry, strike)
                pe_greeks = pe_data.get('greeks', {})

                if ce_greeks and pe_greeks:
                    greeks_map[strike] = {
                        'CE': {
                            'delta': ce_greeks.get('delta', 0),
                            'theta': ce_greeks.get('theta', 0),
                            'vega': ce_greeks.get('vega', 0)
                        },
                        'PE': {
                            'delta': pe_greeks.get('delta', 0),
                            'theta': pe_greeks.get('theta', 0),
                            'vega': pe_greeks.get('vega', 0)
                        }
                    }
                else:
                    logger.warning(f"Missing Greeks for strike {strike}")

            except Exception as e:
                logger.error(f"Error fetching Greeks for strike {strike}: {e}")

        return greeks_map

    def _get_option_data(self, option_type: str, expiry: datetime, strike: int) -> Dict:
        """
        Fetch option data for a specific strike.
        Adapted from nifty_option_analyzer.py

        Args:
            option_type: 'CE' or 'PE'
            expiry: Expiry date
            strike: Strike price

        Returns:
            Option data dict with greeks
        """
        # Format symbol (e.g., NIFTY2601097550CE)
        symbol = f"NIFTY{expiry.strftime('%y%m%d')}{strike}{option_type}"

        try:
            # Fetch quote from Kite
            quote = self.kite.quote([f"NFO:{symbol}"])

            if not quote or f"NFO:{symbol}" not in quote:
                logger.warning(f"No quote data for {symbol}")
                return {}

            option_data = quote[f"NFO:{symbol}"]

            # Check if Greeks are available in API response
            if 'greeks' in option_data and option_data['greeks']:
                return option_data
            else:
                # Calculate Greeks using Black-Scholes
                logger.info(f"Greeks not in API response for {symbol}. Calculating using Black-Scholes...")

                # Get spot price
                spot_price = self._get_nifty_spot_price()
                if not spot_price:
                    logger.error("Cannot calculate Greeks: Failed to get spot price")
                    return {}

                # Get option price
                option_price = option_data.get('last_price', 0)
                if option_price <= 0:
                    logger.warning(f"Invalid option price for {symbol}: {option_price}")
                    return {}

                # Calculate time to expiry
                time_to_expiry = self.bs_calculator.calculate_time_to_expiry(expiry)

                # Calculate Greeks using Black-Scholes
                greeks = self.bs_calculator.calculate_greeks_from_price(
                    spot_price=spot_price,
                    strike_price=strike,
                    time_to_expiry=time_to_expiry,
                    option_price=option_price,
                    option_type=option_type
                )

                # Add Greeks to option_data
                option_data['greeks'] = greeks
                logger.info(f"✓ Calculated Greeks for {symbol}: Δ={greeks['delta']:.4f}, Θ={greeks['theta']:.4f}, V={greeks['vega']:.4f}")

                return option_data

        except Exception as e:
            logger.error(f"Error fetching option data for {symbol}: {e}")
            return {}

    def _calculate_differences(self, current: Dict, baseline: Dict) -> Dict:
        """
        Calculate differences between current and baseline Greeks.

        Args:
            current: Current Greeks {strike: {'CE': {...}, 'PE': {...}}}
            baseline: Baseline Greeks (same structure)

        Returns:
            Differences dict (same structure)
        """
        differences = {}

        for strike in current.keys():
            if strike not in baseline:
                logger.warning(f"Strike {strike} not in baseline. Skipping.")
                continue

            differences[strike] = {}

            for opt_type in ['CE', 'PE']:
                differences[strike][opt_type] = {
                    'delta': current[strike][opt_type]['delta'] - baseline[strike][opt_type]['delta'],
                    'theta': current[strike][opt_type]['theta'] - baseline[strike][opt_type]['theta'],
                    'vega': current[strike][opt_type]['vega'] - baseline[strike][opt_type]['vega']
                }

        return differences

    def _aggregate_by_type(self, differences: Dict) -> Dict:
        """
        Aggregate differences by option type (CE vs PE).

        Args:
            differences: Strike-level differences

        Returns:
            Aggregated sums: {'CE': {...}, 'PE': {...}}
        """
        ce_sums = {'delta_diff_sum': 0, 'theta_diff_sum': 0, 'vega_diff_sum': 0}
        pe_sums = {'delta_diff_sum': 0, 'theta_diff_sum': 0, 'vega_diff_sum': 0}

        for strike, data in differences.items():
            ce_sums['delta_diff_sum'] += data['CE']['delta']
            ce_sums['theta_diff_sum'] += data['CE']['theta']
            ce_sums['vega_diff_sum'] += data['CE']['vega']

            pe_sums['delta_diff_sum'] += data['PE']['delta']
            pe_sums['theta_diff_sum'] += data['PE']['theta']
            pe_sums['vega_diff_sum'] += data['PE']['vega']

        return {'CE': ce_sums, 'PE': pe_sums}

    def _append_to_history(self, aggregated: Dict):
        """
        Append current snapshot to history.

        Args:
            aggregated: Aggregated differences with timestamp and nifty_spot
        """
        current_time = datetime.now().strftime('%H:%M')

        self.history.append({
            'time': current_time,
            'nifty': aggregated['nifty_spot'],
            'CE_delta': round(aggregated['CE']['delta_diff_sum'], 2),
            'CE_theta': round(aggregated['CE']['theta_diff_sum'], 2),
            'CE_vega': round(aggregated['CE']['vega_diff_sum'], 2),
            'PE_delta': round(aggregated['PE']['delta_diff_sum'], 2),
            'PE_theta': round(aggregated['PE']['theta_diff_sum'], 2),
            'PE_vega': round(aggregated['PE']['vega_diff_sum'], 2)
        })

    def _upload_to_cloud(self, excel_path: str) -> Optional[str]:
        """
        Upload Excel file to cloud storage (Google Drive or Dropbox).

        Args:
            excel_path: Local path to Excel file

        Returns:
            Shareable cloud link, or None if error
        """
        provider = config.GREEKS_DIFF_CLOUD_PROVIDER

        try:
            if provider == 'google_drive':
                return self._upload_to_google_drive(excel_path)
            elif provider == 'dropbox':
                return self._upload_to_dropbox(excel_path)
            else:
                logger.error(f"Unknown cloud provider: {provider}")
                return None
        except Exception as e:
            logger.error(f"Error uploading to cloud: {e}", exc_info=True)
            return None

    def _upload_to_google_drive(self, excel_path: str) -> Optional[str]:
        """
        Upload Excel file to Google Drive and return shareable link.

        Args:
            excel_path: Local path to Excel file

        Returns:
            Shareable Google Drive link, or None if error
        """
        try:
            # Import Google Drive dependencies
            from googleapiclient.discovery import build
            from googleapiclient.http import MediaFileUpload
            from google.oauth2.service_account import Credentials

            logger.info("Uploading to Google Drive...")

            # Authenticate
            credentials_path = config.GREEKS_DIFF_GOOGLE_CREDENTIALS_PATH
            if not os.path.exists(credentials_path):
                logger.error(f"Google Drive credentials not found: {credentials_path}")
                return None

            creds = Credentials.from_service_account_file(
                credentials_path,
                scopes=['https://www.googleapis.com/auth/drive.file']
            )
            service = build('drive', 'v3', credentials=creds)

            # File name for Google Drive
            file_name = f"greeks_diff_{datetime.now().strftime('%Y%m%d')}.xlsx"

            # Check if file already exists today (to update instead of creating new)
            folder_id = config.GREEKS_DIFF_GOOGLE_DRIVE_FOLDER_ID
            query = f"name='{file_name}'"
            if folder_id:
                query += f" and '{folder_id}' in parents"

            # Include support for shared folders
            list_params = {
                'q': query,
                'fields': 'files(id, name)',
                'supportsAllDrives': True,
                'includeItemsFromAllDrives': True
            }

            existing_files = service.files().list(**list_params).execute()
            existing_file = existing_files.get('files', [])

            # Prepare file media
            media = MediaFileUpload(
                excel_path,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

            if existing_file:
                # Update existing file
                file_id = existing_file[0]['id']
                service.files().update(
                    fileId=file_id,
                    media_body=media,
                    supportsAllDrives=True
                ).execute()
                logger.info(f"✓ Updated existing file in Google Drive: {file_name}")
            else:
                # Create new file
                file_metadata = {'name': file_name}
                if folder_id:
                    file_metadata['parents'] = [folder_id]

                file = service.files().create(
                    body=file_metadata,
                    media_body=media,
                    fields='id',
                    supportsAllDrives=True
                ).execute()
                file_id = file['id']

                # Make publicly accessible
                service.permissions().create(
                    fileId=file_id,
                    body={'type': 'anyone', 'role': 'reader'},
                    supportsAllDrives=True
                ).execute()
                logger.info(f"✓ Created new file in Google Drive: {file_name}")

            # Return shareable link
            shareable_link = f"https://drive.google.com/file/d/{file_id}/view?usp=sharing"
            logger.info(f"✓ Google Drive link: {shareable_link}")
            return shareable_link

        except ImportError as e:
            logger.error(f"Google Drive libraries not installed: {e}")
            logger.error("Install: pip install google-auth google-auth-oauthlib google-api-python-client")
            return None
        except Exception as e:
            logger.error(f"Error uploading to Google Drive: {e}", exc_info=True)
            return None

    def _upload_to_dropbox(self, excel_path: str) -> Optional[str]:
        """
        Upload Excel file to Dropbox and return shareable link.

        Args:
            excel_path: Local path to Excel file

        Returns:
            Shareable Dropbox link, or None if error
        """
        try:
            # Import Dropbox dependencies
            import dropbox
            from dropbox.files import WriteMode

            logger.info("Uploading to Dropbox...")

            # Authenticate
            token = config.GREEKS_DIFF_DROPBOX_TOKEN
            if not token:
                logger.error("Dropbox token not configured")
                return None

            dbx = dropbox.Dropbox(token)

            # File path in Dropbox
            file_name = f"/greeks_diff_{datetime.now().strftime('%Y%m%d')}.xlsx"

            # Upload file (overwrite if exists)
            with open(excel_path, 'rb') as f:
                dbx.files_upload(
                    f.read(),
                    file_name,
                    mode=WriteMode.overwrite
                )

            logger.info(f"✓ Uploaded to Dropbox: {file_name}")

            # Create or get shareable link
            try:
                # Try to get existing link
                links = dbx.sharing_list_shared_links(path=file_name)
                if links.links:
                    link_url = links.links[0].url
                else:
                    # Create new link
                    link = dbx.sharing_create_shared_link_with_settings(file_name)
                    link_url = link.url

                # Convert to direct download link
                shareable_link = link_url.replace('?dl=0', '?dl=1')
                logger.info(f"✓ Dropbox link: {shareable_link}")
                return shareable_link

            except dropbox.exceptions.ApiError as e:
                logger.error(f"Error creating Dropbox link: {e}")
                return None

        except ImportError as e:
            logger.error(f"Dropbox library not installed: {e}")
            logger.error("Install: pip install dropbox")
            return None
        except Exception as e:
            logger.error(f"Error uploading to Dropbox: {e}", exc_info=True)
            return None

    def _scheduled_update(self):
        """
        Scheduled update function called every 15 minutes.
        Handles Excel export, cloud upload, and Telegram notification.
        """
        # Calculate differences
        aggregated = self.fetch_live_and_calculate_diff()
        if not aggregated:
            return

        # Export to Excel
        excel_path = self.export_to_excel()
        if not excel_path:
            return

        # Upload to cloud
        cloud_link = self._upload_to_cloud(excel_path)
        if not cloud_link:
            logger.warning("Cloud upload failed. Using local path.")
            cloud_link = excel_path

        # Send Telegram notification (only first time at 9:30 AM)
        if not self.telegram_sent:
            self.send_telegram_notification(cloud_link)
        else:
            logger.info(f"Excel updated and uploaded to cloud at {datetime.now()}")

    def _load_baseline_from_cache(self) -> bool:
        """Load baseline from cache if available"""
        cache_key = config.GREEKS_BASELINE_CACHE_KEY.format(
            date=datetime.now().strftime('%Y%m%d')
        )

        cached_baseline = self.cache.get_data(cache_key, 'greeks_diff')
        if cached_baseline:
            self.baseline_greeks = cached_baseline
            logger.info("Baseline loaded from cache")
            return True
        return False

    def _is_market_day(self) -> bool:
        """Check if today is a market day (Monday-Friday)"""
        return datetime.now().weekday() < 5

    def _is_market_hours(self) -> bool:
        """Check if current time is within market hours"""
        now = datetime.now().time()
        start = datetime.strptime(self.market_start, '%H:%M').time()
        end = datetime.strptime(self.market_end, '%H:%M').time()
        return start <= now <= end

    # Helper methods from nifty_option_analyzer.py (simplified versions)

    def _get_nifty_spot_price(self) -> float:
        """Get current NIFTY 50 spot price"""
        try:
            quote = self.kite.quote(["NSE:NIFTY 50"])
            return quote["NSE:NIFTY 50"]["last_price"]
        except Exception as e:
            logger.error(f"Error fetching NIFTY price: {e}")
            return 0.0

    def _get_india_vix(self) -> float:
        """
        Get current India VIX value.

        Returns:
            VIX value as decimal (e.g., 0.15 for 15%)
        """
        try:
            # India VIX instrument token: 264969
            quote = self.kite.quote(["NSE:INDIA VIX"])
            vix_value = quote["NSE:INDIA VIX"]["last_price"]
            logger.info(f"India VIX: {vix_value:.2f}%")
            return vix_value / 100.0  # Convert to decimal
        except Exception as e:
            logger.warning(f"Error fetching India VIX: {e}, using default 10%")
            return 0.10

    def _get_vix_adaptive_threshold(self, vix: float) -> float:
        """
        Get optimal Delta threshold based on current VIX level.

        VIX-Adaptive Thresholds (based on backtest analysis):
        - VIX <10%: ±0.100 (Low volatility, sensitive detection)
        - VIX 10-12%: ±0.100 (Normal conditions)
        - VIX 12-15%: ±0.125 (Elevated volatility)
        - VIX 15-20%: ±0.150 (High volatility, filter noise)
        - VIX >20%: ±0.200 (Extreme volatility, strict filter)

        Args:
            vix: Current VIX value as decimal (e.g., 0.15 for 15%)

        Returns:
            Optimal Delta threshold
        """
        vix_percent = vix * 100  # Convert to percentage

        if vix_percent < 10:
            threshold = 0.100
            regime = "Low"
        elif vix_percent < 12:
            threshold = 0.100
            regime = "Normal"
        elif vix_percent < 15:
            threshold = 0.125
            regime = "Elevated"
        elif vix_percent < 20:
            threshold = 0.150
            regime = "High"
        else:
            threshold = 0.200
            regime = "Very High"

        logger.info(f"VIX-Adaptive Threshold: {threshold:.3f} (VIX {vix_percent:.2f}% = {regime} volatility)")
        return threshold

    def predict_daily_outcome(self, aggregated: Dict) -> Tuple[str, float]:
        """
        Predict daily market outcome based on Delta differences.

        Uses VIX-adaptive threshold for robust predictions across volatility regimes.

        Args:
            aggregated: Dictionary with CE and PE Delta/Theta/Vega sums

        Returns:
            Tuple of (prediction, confidence)
            - prediction: 'Bullish', 'Bearish', or 'Neutral'
            - confidence: Expected accuracy (0-1)
        """
        ce_delta = aggregated['CE']['delta_diff_sum']
        pe_delta = aggregated['PE']['delta_diff_sum']

        threshold = self.current_threshold

        # Predict based on Delta with adaptive threshold
        if ce_delta > threshold and pe_delta > threshold:
            prediction = 'Bullish'
            confidence = 0.825  # 82.5% from backtest
        elif ce_delta < -threshold and pe_delta < -threshold:
            prediction = 'Bearish'
            confidence = 0.714  # 71.4% from backtest
        else:
            prediction = 'Neutral'
            confidence = 0.625  # 62.5% from backtest

        logger.info(f"Prediction: {prediction} (Confidence: {confidence*100:.1f}%, "
                   f"CE Δ: {ce_delta:+.3f}, PE Δ: {pe_delta:+.3f}, Threshold: ±{threshold:.3f})")

        return prediction, confidence

    def _get_atm_strike(self, spot_price: float) -> int:
        """Get ATM strike (rounded to nearest 50)"""
        return round(spot_price / 50) * 50

    def _get_next_expiries(self, count: int = 1) -> List[datetime]:
        """
        Get next N weekly expiries for NIFTY options.
        Returns expiries with > 7 days remaining.
        """
        try:
            # Get all NFO instruments
            instruments = self.kite.instruments('NFO')

            # Filter NIFTY options
            nifty_options = [
                i for i in instruments
                if i['name'] == 'NIFTY' and i['instrument_type'] in ['CE', 'PE']
            ]

            # Get unique expiries
            expiries = sorted(set([i['expiry'] for i in nifty_options]))

            # Filter expiries > 7 days away
            today = datetime.now().date()
            valid_expiries = [
                exp for exp in expiries
                if (exp - today).days > 7
            ]

            return valid_expiries[:count]

        except Exception as e:
            logger.error(f"Error getting expiries: {e}")
            return []


# ==================== CLI INTERFACE ====================

def main():
    """Main entry point for standalone execution"""
    import argparse

    parser = argparse.ArgumentParser(description='Greeks Difference Tracker')
    parser.add_argument('--capture-baseline', action='store_true',
                       help='Capture baseline Greeks (run at 9:15 AM)')
    parser.add_argument('--update', action='store_true',
                       help='Fetch live Greeks and update differences')
    parser.add_argument('--monitor', action='store_true',
                       help='Start continuous monitoring (scheduler)')

    args = parser.parse_args()

    # Initialize Kite
    kite = KiteConnect(api_key=config.KITE_API_KEY)
    kite.set_access_token(config.KITE_ACCESS_TOKEN)

    # Create tracker
    tracker = GreeksDifferenceTracker(kite)

    if args.capture_baseline:
        tracker.capture_baseline_greeks()
    elif args.update:
        tracker.fetch_live_and_calculate_diff()
        tracker.export_to_excel()
    elif args.monitor:
        tracker.start_monitoring()
    else:
        print("Usage: python greeks_difference_tracker.py [--capture-baseline|--update|--monitor]")
        sys.exit(1)


if __name__ == '__main__':
    main()
