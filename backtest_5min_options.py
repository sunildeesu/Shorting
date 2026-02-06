#!/usr/bin/env python3
"""
5-Minute Alert Options Backtesting System

Simulates buying OTM options when 5-minute alerts trigger (1.25% drops/rises),
and evaluates P&L across three exit strategies:
- Same-day (3:25 PM)
- 3 trading days
- 5 trading days (1 week)

Uses REAL option data from Kite Connect (limited to ~10-15 days history).

Author: Sunil Kumar Durganaik
Date: February 2026
"""

import os
import sys
import argparse
import logging
from datetime import datetime, timedelta, time, date
from typing import Dict, List, Optional, Tuple
from pathlib import Path
import time as time_module

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))

import config
from kiteconnect import KiteConnect
from token_manager import TokenManager
from market_utils import is_nse_holiday, get_current_ist_time
from historical_data_cache import get_historical_cache

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ============================================
# CONFIGURATION
# ============================================

BACKTEST_MIN_PREMIUM_COST = 3000   # Minimum total cost for option position
BACKTEST_MAX_PREMIUM_COST = 15000  # Maximum total cost for option position
BACKTEST_ENTRY_DELAY_MINUTES = 2   # Wait 2 minutes after alert before buying
BACKTEST_EXIT_TIME = "15:25"       # Same-day exit time
BACKTEST_3DAY_HOLDING = 3          # 3 trading days holding
BACKTEST_1WEEK_HOLDING = 5         # 5 trading days (1 week) holding
BACKTEST_OPTION_EXPIRY_MIN_DAYS = 21  # Minimum 3 weeks to expiry
BACKTEST_SEND_TELEGRAM = True      # Send summary to Telegram on completion
KITE_API_RATE_LIMIT_DELAY = 0.35   # Seconds between Kite API calls (3 req/sec limit)


class KiteRateLimiter:
    """Simple rate limiter for Kite API calls."""

    _last_call_time = 0.0

    @classmethod
    def wait(cls):
        """Wait if needed to respect rate limit."""
        elapsed = time_module.time() - cls._last_call_time
        if elapsed < KITE_API_RATE_LIMIT_DELAY:
            time_module.sleep(KITE_API_RATE_LIMIT_DELAY - elapsed)
        cls._last_call_time = time_module.time()


class AlertLoader:
    """Loads and filters alerts from alert_tracking.xlsx"""

    def __init__(self, excel_path: str = None):
        """
        Initialize AlertLoader.

        Args:
            excel_path: Path to alert_tracking.xlsx
        """
        self.excel_path = excel_path or config.ALERT_EXCEL_PATH
        self.workbook = None

    def load_alerts(self, start_date: date, end_date: date) -> List[Dict]:
        """
        Load 5-minute alerts from Excel file within date range.

        Args:
            start_date: Start date for filtering
            end_date: End date for filtering

        Returns:
            List of alert dictionaries sorted by datetime
        """
        if not os.path.exists(self.excel_path):
            logger.error(f"Alert file not found: {self.excel_path}")
            return []

        try:
            wb = openpyxl.load_workbook(self.excel_path, read_only=True)
            ws = wb['5min_alerts']

            alerts = []
            headers = [cell.value for cell in ws[1]]

            # Map column indices
            col_map = {h: i for i, h in enumerate(headers)}

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Skip empty rows
                    continue

                try:
                    # Parse date and time
                    date_str = str(row[col_map['Date']])
                    time_str = str(row[col_map['Time']])

                    # Handle datetime objects
                    if isinstance(row[col_map['Date']], datetime):
                        alert_date = row[col_map['Date']].date()
                    else:
                        alert_date = datetime.strptime(date_str, "%Y-%m-%d").date()

                    # Filter by date range
                    if alert_date < start_date or alert_date > end_date:
                        continue

                    # Parse time
                    if isinstance(row[col_map['Time']], time):
                        alert_time = row[col_map['Time']]
                    else:
                        alert_time = datetime.strptime(time_str, "%H:%M:%S").time()

                    alert_datetime = datetime.combine(alert_date, alert_time)

                    # Get direction
                    direction = row[col_map['Direction']]
                    if not direction:
                        continue

                    alert = {
                        'datetime': alert_datetime,
                        'date': alert_date,
                        'time': alert_time,
                        'symbol': row[col_map['Symbol']],
                        'direction': direction,
                        'alert_price': float(row[col_map['Alert Price']]) if row[col_map['Alert Price']] else 0,
                        'change_percent': float(row[col_map['Change %']]) if row[col_map['Change %']] else 0,
                    }
                    alerts.append(alert)

                except Exception as e:
                    logger.debug(f"Error parsing row: {e}")
                    continue

            wb.close()

            # Sort by datetime
            alerts.sort(key=lambda x: x['datetime'])
            logger.info(f"Loaded {len(alerts)} alerts from {start_date} to {end_date}")
            return alerts

        except Exception as e:
            logger.error(f"Error loading alerts: {e}")
            return []

    def get_first_alert_per_stock_day(self, alerts: List[Dict]) -> List[Dict]:
        """
        Filter to get only the first alert per stock per day.

        Args:
            alerts: List of all alerts (should be sorted by datetime)

        Returns:
            Filtered list with first alert per stock per day
        """
        seen = set()
        filtered = []

        for alert in alerts:
            key = (alert['symbol'], alert['date'])
            if key not in seen:
                seen.add(key)
                filtered.append(alert)

        logger.info(f"Filtered to {len(filtered)} alerts (first per stock per day)")
        return filtered


class StockOptionChainBuilder:
    """Builds option chain and finds suitable expiry/lot sizes"""

    def __init__(self, kite: KiteConnect):
        """
        Initialize with Kite instance.

        Args:
            kite: KiteConnect instance
        """
        self.kite = kite
        self.instruments_cache = None
        self.lot_size_cache = {}

    def get_nfo_instruments(self) -> List[Dict]:
        """Get NFO instruments (cached)."""
        if not self.instruments_cache:
            logger.info("Fetching NFO instruments...")
            KiteRateLimiter.wait()
            self.instruments_cache = self.kite.instruments("NFO")
            logger.info(f"Cached {len(self.instruments_cache)} NFO instruments")
        return self.instruments_cache

    def get_stock_options(
        self,
        symbol: str,
        expiry: date,
        option_type: str
    ) -> List[Dict]:
        """
        Get all options for a stock with specific expiry and type.

        Args:
            symbol: Stock symbol (e.g., "RELIANCE")
            expiry: Expiry date
            option_type: 'CE' or 'PE'

        Returns:
            List of matching option instruments
        """
        instruments = self.get_nfo_instruments()

        matches = [
            i for i in instruments
            if i['name'] == symbol
            and i['instrument_type'] == option_type
            and i['expiry'] == expiry
        ]

        return sorted(matches, key=lambda x: x['strike'])

    def find_suitable_expiry(
        self,
        symbol: str,
        alert_date: date,
        min_days: int = BACKTEST_OPTION_EXPIRY_MIN_DAYS
    ) -> Optional[date]:
        """
        Find expiry at least min_days away.

        Args:
            symbol: Stock symbol
            alert_date: Date of the alert
            min_days: Minimum days to expiry (default 21 = 3 weeks)

        Returns:
            Expiry date or None if not found
        """
        instruments = self.get_nfo_instruments()

        # Get all expiries for this stock's options
        stock_options = [
            i for i in instruments
            if i['name'] == symbol
            and i['instrument_type'] in ('CE', 'PE')
        ]

        if not stock_options:
            return None

        available_expiries = sorted(set(i['expiry'] for i in stock_options))

        for expiry in available_expiries:
            days_to_expiry = (expiry - alert_date).days
            if days_to_expiry >= min_days:
                return expiry

        # Fallback: next month expiry (find furthest available)
        if available_expiries:
            return available_expiries[-1]

        return None

    def get_lot_size(self, symbol: str) -> int:
        """
        Get lot size for a stock's options.

        Args:
            symbol: Stock symbol

        Returns:
            Lot size or 0 if not found
        """
        if symbol in self.lot_size_cache:
            return self.lot_size_cache[symbol]

        instruments = self.get_nfo_instruments()

        for inst in instruments:
            if inst['name'] == symbol and inst['instrument_type'] in ('CE', 'PE'):
                lot_size = inst.get('lot_size', 0)
                self.lot_size_cache[symbol] = lot_size
                return lot_size

        return 0


class OTMOptionFinder:
    """Finds suitable OTM options with premium in target range"""

    def __init__(self, kite: KiteConnect, option_chain_builder: StockOptionChainBuilder):
        """
        Initialize with Kite and option chain builder.

        Args:
            kite: KiteConnect instance
            option_chain_builder: StockOptionChainBuilder instance
        """
        self.kite = kite
        self.option_chain_builder = option_chain_builder
        self.historical_cache = get_historical_cache()
        self.intraday_cache = {}  # Cache for intraday 5min data: {(token, date): data}

    def _get_intraday_data(self, token: int, target_date: date) -> Optional[List]:
        """Get intraday 5-minute data with caching."""
        cache_key = (token, target_date)
        if cache_key in self.intraday_cache:
            return self.intraday_cache[cache_key]

        from_dt = datetime.combine(target_date, time(9, 15))
        to_dt = datetime.combine(target_date, time(15, 30))

        KiteRateLimiter.wait()
        try:
            data = self.kite.historical_data(
                instrument_token=token,
                from_date=from_dt,
                to_date=to_dt,
                interval="5minute"
            )
            self.intraday_cache[cache_key] = data
            return data
        except Exception as e:
            logger.debug(f"Error fetching intraday data for {token}: {e}")
            return None

    def find_suitable_option(
        self,
        symbol: str,
        spot_price: float,
        direction: str,
        expiry: date,
        alert_datetime: datetime,
        min_cost: int = BACKTEST_MIN_PREMIUM_COST,
        max_cost: int = BACKTEST_MAX_PREMIUM_COST,
        strike_type: str = 'otm'
    ) -> Optional[Dict]:
        """
        Find option with total cost in target range.

        For DROP alerts: Buy PUT (OTM: strike < spot, ATM: nearest to spot)
        For RISE alerts: Buy CALL (OTM: strike > spot, ATM: nearest to spot)

        Args:
            symbol: Stock symbol
            spot_price: Current spot price
            direction: 'Drop' or 'Rise'
            expiry: Option expiry date
            alert_datetime: Datetime of the alert
            min_cost: Minimum total cost (default 5000)
            max_cost: Maximum total cost (default 10000)
            strike_type: 'otm' (out-of-the-money) or 'atm' (at-the-money)

        Returns:
            Dict with option details or None
        """
        lot_size = self.option_chain_builder.get_lot_size(symbol)
        if lot_size <= 0:
            logger.debug(f"No lot size found for {symbol}")
            return None

        # Determine option type based on direction
        if direction == 'Drop':
            option_type = 'PE'
        else:  # Rise
            option_type = 'CE'

        options = self.option_chain_builder.get_stock_options(symbol, expiry, option_type)
        if not options:
            logger.debug(f"No {option_type} options found for {symbol} expiry {expiry}")
            return None

        # Filter and sort strikes based on strike_type
        if strike_type == 'atm':
            # ATM: Sort ALL options by distance from spot (nearest first)
            options.sort(key=lambda x: abs(x['strike'] - spot_price))
            candidates = options
        else:
            # OTM: Filter and sort
            if direction == 'Drop':
                # PE: OTM means strike < spot, sort from nearest to furthest
                candidates = [o for o in options if o['strike'] < spot_price]
                candidates.sort(key=lambda x: spot_price - x['strike'])  # Nearest first
            else:
                # CE: OTM means strike > spot, sort from nearest to furthest
                candidates = [o for o in options if o['strike'] > spot_price]
                candidates.sort(key=lambda x: x['strike'] - spot_price)  # Nearest first

        if not candidates:
            logger.debug(f"No {strike_type.upper()} {option_type} options found for {symbol}")
            return None

        # Iterate through candidate strikes to find one with suitable premium
        for option in candidates:
            premium = self._get_premium_at_time(option['instrument_token'], alert_datetime)

            if premium is None or premium <= 0:
                continue

            total_cost = premium * lot_size

            if min_cost <= total_cost <= max_cost:
                return {
                    'instrument_token': option['instrument_token'],
                    'tradingsymbol': option['tradingsymbol'],
                    'strike': option['strike'],
                    'option_type': option_type,
                    'expiry': expiry,
                    'lot_size': lot_size,
                    'entry_premium': premium,
                    'entry_cost': total_cost,
                }

        # If no exact match, try to find closest one
        logger.debug(f"No option with cost {min_cost}-{max_cost} found for {symbol}")
        return None

    def _get_premium_at_time(
        self,
        instrument_token: int,
        target_datetime: datetime
    ) -> Optional[float]:
        """
        Get option premium at specific datetime.

        Args:
            instrument_token: Option instrument token
            target_datetime: Target datetime

        Returns:
            Premium price or None
        """
        data = self._get_intraday_data(instrument_token, target_datetime.date())
        if not data:
            return None

        target_time = target_datetime.time()

        # Find candle closest to target time
        for candle in data:
            candle_time = candle['date'].time()
            # Use candle if within 10 minutes of target
            time_diff = abs(
                (candle['date'].hour * 60 + candle['date'].minute) -
                (target_time.hour * 60 + target_time.minute)
            )
            if time_diff <= 10:
                return candle['close']

        return None


class HistoricalOptionFetcher:
    """Fetches historical option prices for entry/exit calculations"""

    def __init__(self, kite: KiteConnect):
        """
        Initialize with Kite instance.

        Args:
            kite: KiteConnect instance
        """
        self.kite = kite
        self.historical_cache = get_historical_cache()
        self.nse_instruments_cache = None
        self.intraday_cache = {}  # Cache for intraday 5min data: {(token, date): data}

    def _get_intraday_data(self, token: int, target_date: date) -> Optional[List]:
        """Get intraday 5-minute data with caching."""
        cache_key = (token, target_date)
        if cache_key in self.intraday_cache:
            return self.intraday_cache[cache_key]

        from_dt = datetime.combine(target_date, time(9, 15))
        to_dt = datetime.combine(target_date, time(15, 30))

        KiteRateLimiter.wait()
        try:
            data = self.kite.historical_data(
                instrument_token=token,
                from_date=from_dt,
                to_date=to_dt,
                interval="5minute"
            )
            self.intraday_cache[cache_key] = data
            return data
        except Exception as e:
            logger.debug(f"Error fetching intraday data for {token}: {e}")
            return None

    def _get_nse_instruments(self) -> List[Dict]:
        """Get NSE instruments (cached)."""
        if not self.nse_instruments_cache:
            KiteRateLimiter.wait()
            self.nse_instruments_cache = self.kite.instruments("NSE")
        return self.nse_instruments_cache

    def get_option_price_at_time(
        self,
        token: int,
        target_datetime: datetime
    ) -> Optional[float]:
        """
        Get option price at specific datetime.

        Args:
            token: Instrument token
            target_datetime: Target datetime

        Returns:
            Price or None
        """
        data = self._get_intraday_data(token, target_datetime.date())
        if not data:
            return None

        target_time = target_datetime.time()

        # Find candle closest to target time
        best_candle = None
        best_diff = float('inf')

        for candle in data:
            candle_time = candle['date'].time()
            time_diff = abs(
                (candle['date'].hour * 60 + candle['date'].minute) -
                (target_time.hour * 60 + target_time.minute)
            )
            if time_diff < best_diff:
                best_diff = time_diff
                best_candle = candle

        if best_candle and best_diff <= 15:  # Within 15 minutes
            return best_candle['close']

        return None

    def get_stock_price_at_time(
        self,
        symbol: str,
        target_datetime: datetime
    ) -> Optional[float]:
        """
        Get stock price at specific datetime.

        Args:
            symbol: Stock symbol
            target_datetime: Target datetime

        Returns:
            Price or None
        """
        try:
            # Get instrument token for stock (cached)
            instruments = self._get_nse_instruments()
            stock_inst = next(
                (i for i in instruments if i['tradingsymbol'] == symbol),
                None
            )
            if not stock_inst:
                return None

            data = self._get_intraday_data(stock_inst['instrument_token'], target_datetime.date())
            if not data:
                return None

            target_time = target_datetime.time()

            best_candle = None
            best_diff = float('inf')

            for candle in data:
                time_diff = abs(
                    (candle['date'].hour * 60 + candle['date'].minute) -
                    (target_time.hour * 60 + target_time.minute)
                )
                if time_diff < best_diff:
                    best_diff = time_diff
                    best_candle = candle

            if best_candle and best_diff <= 15:
                return best_candle['close']

            return None

        except Exception as e:
            logger.debug(f"Error fetching stock price for {symbol}: {e}")
            return None

    def get_exit_price(
        self,
        token: int,
        exit_datetime: datetime,
        exit_type: str
    ) -> Optional[float]:
        """
        Get option exit price.

        Args:
            token: Instrument token
            exit_datetime: Exit datetime
            exit_type: 'same_day', '3day', or '1week'

        Returns:
            Exit price or None
        """
        return self.get_option_price_at_time(token, exit_datetime)

    def get_trading_day_offset(
        self,
        start_date: date,
        offset_days: int
    ) -> date:
        """
        Calculate date that is offset_days trading days from start_date.

        Args:
            start_date: Starting date
            offset_days: Number of trading days to offset

        Returns:
            Target date
        """
        current = start_date
        days_counted = 0

        while days_counted < offset_days:
            current += timedelta(days=1)
            # Skip weekends
            if current.weekday() >= 5:
                continue
            # Skip holidays
            if is_nse_holiday(current):
                continue
            days_counted += 1

        return current


class TradeSimulator:
    """Simulates trades and calculates P&L for different exit strategies"""

    def __init__(
        self,
        kite: KiteConnect,
        alert_loader: AlertLoader,
        option_chain_builder: StockOptionChainBuilder,
        otm_finder: OTMOptionFinder,
        historical_fetcher: HistoricalOptionFetcher,
        strike_type: str = 'otm'
    ):
        """
        Initialize TradeSimulator.

        Args:
            kite: KiteConnect instance
            alert_loader: AlertLoader instance
            option_chain_builder: StockOptionChainBuilder instance
            otm_finder: OTMOptionFinder instance
            historical_fetcher: HistoricalOptionFetcher instance
            strike_type: 'otm' (out-of-the-money) or 'atm' (at-the-money)
        """
        self.kite = kite
        self.alert_loader = alert_loader
        self.option_chain_builder = option_chain_builder
        self.otm_finder = otm_finder
        self.historical_fetcher = historical_fetcher
        self.strike_type = strike_type
        self.nse_instruments_cache = None

    def _get_nse_instruments(self) -> List[Dict]:
        """Get NSE instruments (cached)."""
        if not self.nse_instruments_cache:
            KiteRateLimiter.wait()
            self.nse_instruments_cache = self.kite.instruments("NSE")
        return self.nse_instruments_cache

    def run_backtest(
        self,
        start_date: date,
        end_date: date
    ) -> Dict[str, List]:
        """
        Run backtest for all exit strategies.

        Args:
            start_date: Start date for backtest
            end_date: End date for backtest

        Returns:
            Dict with keys 'same_day', '3day', '1week' containing trade lists
        """
        # Load and filter alerts
        all_alerts = self.alert_loader.load_alerts(start_date, end_date)
        alerts = self.alert_loader.get_first_alert_per_stock_day(all_alerts)

        if not alerts:
            logger.warning("No alerts found for backtest period")
            return {'same_day': [], '3day': [], '1week': []}

        logger.info(f"\n{'='*60}")
        logger.info(f"Starting backtest: {len(alerts)} alerts")
        logger.info(f"Period: {start_date} to {end_date}")
        logger.info(f"{'='*60}\n")

        # Results by strategy
        results = {
            'same_day': [],
            '3day': [],
            '1week': []
        }

        # Track active positions for multi-day strategies
        active_positions = {
            '3day': {},   # {symbol: exit_date}
            '1week': {}   # {symbol: exit_date}
        }

        for i, alert in enumerate(alerts, 1):
            logger.info(f"\n[{i}/{len(alerts)}] Processing {alert['symbol']} "
                       f"{alert['direction']} at {alert['datetime']}")

            # Same-day: no filtering needed
            same_day_trade = self.simulate_single_trade(alert, 'same_day')
            if same_day_trade:
                results['same_day'].append(same_day_trade)

            # 3-day: check if position already active
            if not self._should_ignore_alert(alert, active_positions['3day']):
                trade_3day = self.simulate_single_trade(alert, '3day')
                if trade_3day:
                    results['3day'].append(trade_3day)
                    exit_date = trade_3day.get('exit_date')
                    if exit_date:
                        active_positions['3day'][alert['symbol']] = exit_date

            # 1-week: check if position already active
            if not self._should_ignore_alert(alert, active_positions['1week']):
                trade_1week = self.simulate_single_trade(alert, '1week')
                if trade_1week:
                    results['1week'].append(trade_1week)
                    exit_date = trade_1week.get('exit_date')
                    if exit_date:
                        active_positions['1week'][alert['symbol']] = exit_date

        # Log summary
        for strategy, trades in results.items():
            valid_trades = [t for t in trades if t.get('status') != 'data_missing']
            logger.info(f"\n{strategy}: {len(valid_trades)} valid trades, "
                       f"{len(trades) - len(valid_trades)} data missing")

        return results

    def simulate_single_trade(
        self,
        alert: Dict,
        exit_strategy: str
    ) -> Optional[Dict]:
        """
        Simulate a single trade for given alert and exit strategy.

        Args:
            alert: Alert dictionary
            exit_strategy: 'same_day', '3day', or '1week'

        Returns:
            Trade result dictionary or None
        """
        symbol = alert['symbol']
        alert_datetime = alert['datetime']
        direction = alert['direction']
        spot_price = alert['alert_price']

        # Calculate entry time (2 minutes after alert)
        entry_datetime = alert_datetime + timedelta(minutes=BACKTEST_ENTRY_DELAY_MINUTES)
        entry_time = entry_datetime.time()

        # Initialize trade result
        trade = {
            'date': alert['date'].strftime('%Y-%m-%d'),
            'alert_time': alert['time'].strftime('%H:%M:%S'),
            'entry_time': entry_time.strftime('%H:%M:%S'),
            'symbol': symbol,
            'direction': direction,
            'stock_entry_price': spot_price,
            'exit_strategy': exit_strategy,
            'status': 'pending'
        }

        try:
            # Find suitable expiry
            expiry = self.option_chain_builder.find_suitable_expiry(
                symbol, alert['date']
            )
            if not expiry:
                trade['status'] = 'data_missing'
                trade['error'] = 'No suitable expiry found'
                return trade

            trade['expiry'] = expiry.strftime('%Y-%m-%d')
            trade['days_to_expiry'] = (expiry - alert['date']).days

            # Find suitable option (entry 2 mins after alert)
            option = self.otm_finder.find_suitable_option(
                symbol=symbol,
                spot_price=spot_price,
                direction=direction,
                expiry=expiry,
                alert_datetime=entry_datetime,  # Use entry time, not alert time
                strike_type=self.strike_type
            )

            if not option:
                trade['status'] = 'data_missing'
                trade['error'] = 'No suitable option found in premium range'
                return trade

            # Fill option details
            trade['option_symbol'] = option['tradingsymbol']
            trade['strike'] = option['strike']
            trade['option_type'] = option['option_type']
            trade['lot_size'] = option['lot_size']
            trade['entry_premium'] = round(option['entry_premium'], 2)
            trade['entry_cost'] = round(option['entry_cost'], 2)

            # Calculate exit datetime based on strategy
            if exit_strategy == 'same_day':
                exit_time = datetime.strptime(BACKTEST_EXIT_TIME, "%H:%M").time()
                exit_datetime = datetime.combine(alert['date'], exit_time)
                exit_date = alert['date']
            elif exit_strategy == '3day':
                exit_date = self.historical_fetcher.get_trading_day_offset(
                    alert['date'], BACKTEST_3DAY_HOLDING
                )
                exit_time = datetime.strptime(BACKTEST_EXIT_TIME, "%H:%M").time()
                exit_datetime = datetime.combine(exit_date, exit_time)
            else:  # 1week
                exit_date = self.historical_fetcher.get_trading_day_offset(
                    alert['date'], BACKTEST_1WEEK_HOLDING
                )
                exit_time = datetime.strptime(BACKTEST_EXIT_TIME, "%H:%M").time()
                exit_datetime = datetime.combine(exit_date, exit_time)

            trade['exit_date'] = exit_date

            # Get stock exit price
            stock_exit_price = self.historical_fetcher.get_stock_price_at_time(
                symbol, exit_datetime
            )
            if stock_exit_price:
                trade['stock_exit_price'] = round(stock_exit_price, 2)
                stock_move = ((stock_exit_price - spot_price) / spot_price) * 100
                trade['stock_move_pct'] = round(stock_move, 2)
            else:
                trade['stock_exit_price'] = None
                trade['stock_move_pct'] = None

            # Get option exit price
            exit_premium = self.historical_fetcher.get_exit_price(
                option['instrument_token'],
                exit_datetime,
                exit_strategy
            )

            if exit_premium is None:
                trade['status'] = 'data_missing'
                trade['error'] = f'No exit price data for {exit_datetime}'
                return trade

            trade['exit_premium'] = round(exit_premium, 2)
            trade['exit_value'] = round(exit_premium * option['lot_size'], 2)

            # Calculate P&L
            pnl = (exit_premium - option['entry_premium']) * option['lot_size']
            pnl_pct = ((exit_premium - option['entry_premium']) / option['entry_premium']) * 100

            trade['pnl'] = round(pnl, 2)
            trade['pnl_pct'] = round(pnl_pct, 2)
            trade['status'] = 'Win' if pnl > 0 else 'Loss'

            logger.info(f"  {exit_strategy}: {trade['option_symbol']} "
                       f"P&L: ₹{pnl:,.2f} ({pnl_pct:+.1f}%)")

            return trade

        except Exception as e:
            logger.error(f"Error simulating trade for {symbol}: {e}")
            trade['status'] = 'error'
            trade['error'] = str(e)
            return trade

    def _should_ignore_alert(
        self,
        alert: Dict,
        active_positions: Dict[str, date]
    ) -> bool:
        """
        Check if alert should be ignored due to active position.

        Args:
            alert: Alert dictionary
            active_positions: Dict of {symbol: exit_date}

        Returns:
            True if should ignore, False otherwise
        """
        symbol = alert['symbol']
        alert_date = alert['date']

        if symbol in active_positions:
            exit_date = active_positions[symbol]
            if alert_date < exit_date:
                logger.debug(f"  Skipping {symbol} - position active until {exit_date}")
                return True
            else:
                # Position expired, remove from active
                del active_positions[symbol]

        return False


class ReportGenerator:
    """Generates Excel reports and sends Telegram notifications"""

    def __init__(self):
        """Initialize ReportGenerator."""
        self.output_dir = Path("data/backtest_results")
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate_report(
        self,
        results: Dict[str, List],
        start_date: date,
        end_date: date,
        strike_type: str = 'otm'
    ) -> str:
        """
        Generate Excel report with summary and trade details.

        Args:
            results: Dict with 'same_day', '3day', '1week' trade lists
            start_date: Backtest start date
            end_date: Backtest end date
            strike_type: 'otm' or 'atm' strike selection mode

        Returns:
            Path to generated Excel file
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = self.output_dir / f"5min_options_backtest_{strike_type}_{timestamp}.xlsx"

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Create Summary sheet
        self._create_summary_sheet(wb, results, start_date, end_date)

        # Create detail sheets for each strategy
        for strategy in ['same_day', '3day', '1week']:
            self._create_detail_sheet(wb, strategy, results[strategy])

        wb.save(output_path)
        logger.info(f"Report saved to: {output_path}")
        return str(output_path)

    def _create_summary_sheet(
        self,
        wb: openpyxl.Workbook,
        results: Dict[str, List],
        start_date: date,
        end_date: date
    ):
        """Create summary sheet with statistics."""
        ws = wb.create_sheet("Summary")

        # Header styling
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Title
        ws['A1'] = "5-Minute Alert Options Backtest Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        ws['A2'] = f"Period: {start_date} to {end_date}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # Statistics table header
        row = 5
        headers = ['Metric', 'Same Day', '3-Day', '1-Week']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Calculate statistics for each strategy
        strategies = ['same_day', '3day', '1week']
        stats = {}

        for strategy in strategies:
            trades = results[strategy]
            valid_trades = [t for t in trades if t.get('status') in ('Win', 'Loss')]

            if valid_trades:
                wins = len([t for t in valid_trades if t['status'] == 'Win'])
                total_pnl = sum(t['pnl'] for t in valid_trades)
                avg_pnl = total_pnl / len(valid_trades)
                pnls = [t['pnl'] for t in valid_trades]
                best_trade = max(pnls) if pnls else 0
                worst_trade = min(pnls) if pnls else 0
                win_rate = (wins / len(valid_trades)) * 100 if valid_trades else 0
            else:
                wins = 0
                total_pnl = 0
                avg_pnl = 0
                best_trade = 0
                worst_trade = 0
                win_rate = 0

            stats[strategy] = {
                'total_trades': len(valid_trades),
                'wins': wins,
                'losses': len(valid_trades) - wins,
                'win_rate': win_rate,
                'total_pnl': total_pnl,
                'avg_pnl': avg_pnl,
                'best_trade': best_trade,
                'worst_trade': worst_trade
            }

        # Metrics to display
        metrics = [
            ('Total Trades', 'total_trades'),
            ('Wins', 'wins'),
            ('Losses', 'losses'),
            ('Win Rate %', 'win_rate'),
            ('Total P&L (₹)', 'total_pnl'),
            ('Avg P&L (₹)', 'avg_pnl'),
            ('Best Trade (₹)', 'best_trade'),
            ('Worst Trade (₹)', 'worst_trade'),
        ]

        for i, (metric_name, metric_key) in enumerate(metrics):
            row = 6 + i
            ws.cell(row=row, column=1, value=metric_name)

            for j, strategy in enumerate(strategies, 2):
                value = stats[strategy][metric_key]
                if metric_key == 'win_rate':
                    ws.cell(row=row, column=j, value=f"{value:.1f}%")
                elif 'pnl' in metric_key or 'trade' in metric_key.lower():
                    cell = ws.cell(row=row, column=j, value=f"₹{value:,.2f}")
                    if value > 0:
                        cell.font = Font(color="006400")  # Green
                    elif value < 0:
                        cell.font = Font(color="8B0000")  # Red
                else:
                    ws.cell(row=row, column=j, value=value)

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

    def _create_detail_sheet(
        self,
        wb: openpyxl.Workbook,
        strategy: str,
        trades: List[Dict]
    ):
        """Create detail sheet for a strategy."""
        sheet_names = {
            'same_day': 'Same Day Trades',
            '3day': '3-Day Trades',
            '1week': '1-Week Trades'
        }
        ws = wb.create_sheet(sheet_names[strategy])

        # Headers (includes Alert Time and Entry Time - 2 min delay)
        headers = [
            'Date', 'Alert Time', 'Entry Time', 'Symbol', 'Direction',
            'Stock Entry', 'Stock Exit', 'Stock Move %',
            'Option Symbol', 'Strike', 'Type', 'Expiry',
            'Entry Premium', 'Exit Premium', 'Lot Size',
            'Entry Cost', 'Exit Value', 'P&L', 'P&L %', 'Status'
        ]

        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # Data rows
        for row_idx, trade in enumerate(trades, 2):
            data = [
                trade.get('date', ''),
                trade.get('alert_time', trade.get('time', '')),  # Fallback to 'time' for compatibility
                trade.get('entry_time', ''),
                trade.get('symbol', ''),
                trade.get('direction', ''),
                trade.get('stock_entry_price', ''),
                trade.get('stock_exit_price', ''),
                trade.get('stock_move_pct', ''),
                trade.get('option_symbol', ''),
                trade.get('strike', ''),
                trade.get('option_type', ''),
                trade.get('expiry', ''),
                trade.get('entry_premium', ''),
                trade.get('exit_premium', ''),
                trade.get('lot_size', ''),
                trade.get('entry_cost', ''),
                trade.get('exit_value', ''),
                trade.get('pnl', ''),
                trade.get('pnl_pct', ''),
                trade.get('status', ''),
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = Alignment(horizontal='center')

                # Color P&L cells
                if col == 18:  # P&L column (shifted by 1 due to new Entry Time column)
                    if isinstance(value, (int, float)):
                        if value > 0:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif value < 0:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                # Color status
                if col == 20:  # Status column (shifted by 1)
                    if value == 'Win':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value == 'Loss':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif value == 'data_missing':
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12

        # Freeze header row
        ws.freeze_panes = 'A2'

    def _calculate_statistics(self, trades: List[Dict]) -> Dict:
        """Calculate statistics for a list of trades."""
        valid_trades = [t for t in trades if t.get('status') in ('Win', 'Loss')]

        if not valid_trades:
            return {
                'total_trades': 0,
                'wins': 0,
                'losses': 0,
                'win_rate': 0,
                'total_pnl': 0,
                'avg_pnl': 0,
                'best_trade': 0,
                'worst_trade': 0
            }

        wins = len([t for t in valid_trades if t['status'] == 'Win'])
        pnls = [t['pnl'] for t in valid_trades]

        return {
            'total_trades': len(valid_trades),
            'wins': wins,
            'losses': len(valid_trades) - wins,
            'win_rate': (wins / len(valid_trades)) * 100,
            'total_pnl': sum(pnls),
            'avg_pnl': sum(pnls) / len(pnls),
            'best_trade': max(pnls),
            'worst_trade': min(pnls)
        }

    def send_telegram_report(
        self,
        excel_path: str,
        start_date: date,
        end_date: date,
        results: Dict[str, List]
    ) -> bool:
        """
        Send Excel report as Telegram document.

        Args:
            excel_path: Path to Excel file
            start_date: Backtest start date
            end_date: Backtest end date
            results: Results dictionary for summary

        Returns:
            True if sent successfully
        """
        import requests

        bot_token = config.TELEGRAM_BOT_TOKEN
        channel_id = config.TELEGRAM_CHANNEL_ID

        if not bot_token or not channel_id:
            logger.warning("Telegram credentials not configured")
            return False

        # Calculate summary stats
        total_trades = sum(
            len([t for t in trades if t.get('status') in ('Win', 'Loss')])
            for trades in results.values()
        )

        # Format caption
        caption = (
            f"📊 <b>5-Min Options Backtest Report</b>\n\n"
            f"📅 Period: {start_date} to {end_date}\n"
            f"📈 Total Trades: {total_trades}\n\n"
            f"<i>See attached Excel for detailed analysis</i>"
        )

        url = f"https://api.telegram.org/bot{bot_token}/sendDocument"

        try:
            with open(excel_path, 'rb') as doc:
                files = {'document': doc}
                data = {
                    'chat_id': channel_id,
                    'caption': caption,
                    'parse_mode': 'HTML'
                }
                response = requests.post(url, data=data, files=files, timeout=30)
                response.raise_for_status()

            logger.info("Telegram report sent successfully")
            return True

        except Exception as e:
            logger.error(f"Failed to send Telegram report: {e}")
            return False


def initialize_kite() -> Optional[KiteConnect]:
    """Initialize Kite Connect client."""
    try:
        kite = KiteConnect(api_key=config.KITE_API_KEY)
        kite.set_access_token(config.KITE_ACCESS_TOKEN)

        profile = kite.profile()
        logger.info(f"Connected to Kite as: {profile.get('user_name', 'Unknown')}")
        return kite

    except Exception as e:
        logger.error(f"Failed to initialize Kite: {e}")
        return None


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="5-Minute Alert Options Backtesting System"
    )
    parser.add_argument(
        '--start',
        type=str,
        help='Start date (YYYY-MM-DD)',
        default=(datetime.now() - timedelta(days=10)).strftime('%Y-%m-%d')
    )
    parser.add_argument(
        '--end',
        type=str,
        help='End date (YYYY-MM-DD)',
        default=datetime.now().strftime('%Y-%m-%d')
    )
    parser.add_argument(
        '--no-telegram',
        action='store_true',
        help='Skip sending Telegram notification'
    )
    parser.add_argument(
        '--strike',
        type=str,
        choices=['otm', 'atm'],
        default='otm',
        help='Strike selection: otm (out-of-the-money) or atm (at-the-money)'
    )

    args = parser.parse_args()

    start_date = datetime.strptime(args.start, '%Y-%m-%d').date()
    end_date = datetime.strptime(args.end, '%Y-%m-%d').date()

    logger.info("=" * 70)
    logger.info("5-MINUTE ALERT OPTIONS BACKTEST")
    logger.info("=" * 70)
    logger.info(f"Period: {start_date} to {end_date}")
    logger.info(f"Strike selection: {args.strike.upper()}")
    logger.info(f"Exit strategies: Same Day, 3-Day, 1-Week")
    logger.info(f"Entry delay: {BACKTEST_ENTRY_DELAY_MINUTES} minutes after alert")
    logger.info(f"Premium range: ₹{BACKTEST_MIN_PREMIUM_COST:,} - ₹{BACKTEST_MAX_PREMIUM_COST:,}")
    logger.info(f"Min expiry: {BACKTEST_OPTION_EXPIRY_MIN_DAYS} days")
    logger.info("=" * 70)

    # Initialize Kite
    kite = initialize_kite()
    if not kite:
        logger.error("Cannot proceed without Kite connection")
        return 1

    # Initialize components
    alert_loader = AlertLoader()
    option_chain_builder = StockOptionChainBuilder(kite)
    otm_finder = OTMOptionFinder(kite, option_chain_builder)
    historical_fetcher = HistoricalOptionFetcher(kite)
    trade_simulator = TradeSimulator(
        kite, alert_loader, option_chain_builder, otm_finder, historical_fetcher,
        strike_type=args.strike
    )
    report_generator = ReportGenerator()

    # Run backtest
    results = trade_simulator.run_backtest(start_date, end_date)

    # Generate report
    excel_path = report_generator.generate_report(results, start_date, end_date, args.strike)

    # Print summary
    logger.info("\n" + "=" * 70)
    logger.info("BACKTEST SUMMARY")
    logger.info("=" * 70)

    for strategy in ['same_day', '3day', '1week']:
        trades = results[strategy]
        valid_trades = [t for t in trades if t.get('status') in ('Win', 'Loss')]

        if valid_trades:
            wins = len([t for t in valid_trades if t['status'] == 'Win'])
            total_pnl = sum(t['pnl'] for t in valid_trades)
            win_rate = (wins / len(valid_trades)) * 100

            logger.info(f"\n{strategy.upper()}:")
            logger.info(f"  Trades: {len(valid_trades)}")
            logger.info(f"  Win Rate: {win_rate:.1f}%")
            logger.info(f"  Total P&L: ₹{total_pnl:,.2f}")
        else:
            logger.info(f"\n{strategy.upper()}: No valid trades")

    logger.info("\n" + "=" * 70)

    # Send Telegram notification
    if BACKTEST_SEND_TELEGRAM and not args.no_telegram:
        report_generator.send_telegram_report(excel_path, start_date, end_date, results)

    logger.info(f"\n✅ Backtest complete! Report: {excel_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
