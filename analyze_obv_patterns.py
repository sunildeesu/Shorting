#!/usr/bin/env python3
"""
Analyze OBV (On-Balance Volume) patterns before 5-minute alerts.

OBV = cumulative sum of:
  +volume when price closes higher than previous
  -volume when price closes lower than previous

Key patterns to check:
1. OBV Confirmation: Price rising + OBV rising = strong move (likely to continue)
2. OBV Divergence: Price rising + OBV falling = weak move (likely to reverse)
3. OBV Breakout: OBV breaks trend before price does

Usage:
    python analyze_obv_patterns.py --days 30
"""

import argparse
import sqlite3
from datetime import datetime, timedelta
from typing import Dict, List, Tuple
from collections import defaultdict

import openpyxl
import config

import logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class OBVAnalyzer:
    """Analyzes OBV patterns before 5-minute alerts."""

    def __init__(self, db_path: str = "data/central_quotes.db"):
        self.db_path = db_path
        self.conn = None

    def _get_conn(self):
        if self.conn is None:
            self.conn = sqlite3.connect(self.db_path)
        return self.conn

    def calculate_obv(self, data: List[Dict]) -> List[float]:
        """
        Calculate OBV for a series of price/volume data.

        Args:
            data: List of {price, volume} dicts in chronological order

        Returns:
            List of cumulative OBV values
        """
        if len(data) < 2:
            return []

        obv = [0]  # Start at 0

        for i in range(1, len(data)):
            prev_price = data[i-1]['price']
            curr_price = data[i]['price']
            volume = data[i]['volume'] or 0

            if curr_price > prev_price:
                obv.append(obv[-1] + volume)
            elif curr_price < prev_price:
                obv.append(obv[-1] - volume)
            else:
                obv.append(obv[-1])  # No change

        return obv

    def calculate_obv_slope(self, obv: List[float], window: int = 5) -> float:
        """Calculate the slope of OBV over the last N points."""
        if len(obv) < window:
            return 0

        recent = obv[-window:]
        # Simple linear regression slope
        n = len(recent)
        x_mean = (n - 1) / 2
        y_mean = sum(recent) / n

        numerator = sum((i - x_mean) * (y - y_mean) for i, y in enumerate(recent))
        denominator = sum((i - x_mean) ** 2 for i in range(n))

        if denominator == 0:
            return 0
        return numerator / denominator

    def check_obv_divergence(self, prices: List[float], obv: List[float],
                              window: int = 5) -> Tuple[str, float]:
        """
        Check for OBV divergence.

        Returns:
            (divergence_type, strength)
            - 'bullish_div': Price down, OBV up (potential reversal up)
            - 'bearish_div': Price up, OBV down (potential reversal down)
            - 'confirmation': Price and OBV moving same direction
            - 'neutral': No clear pattern
        """
        if len(prices) < window or len(obv) < window:
            return 'neutral', 0

        # Price change over window
        price_change = (prices[-1] - prices[-window]) / prices[-window] * 100

        # OBV trend (using slope instead of raw change)
        obv_slope = self.calculate_obv_slope(obv, window)

        # Normalize OBV slope relative to average volume
        avg_vol = sum(abs(obv[i] - obv[i-1]) for i in range(1, len(obv))) / (len(obv) - 1) if len(obv) > 1 else 1
        if avg_vol > 0:
            normalized_slope = obv_slope / avg_vol
        else:
            normalized_slope = 0

        # More sensitive thresholds
        price_threshold = 0.2  # 0.2% price move
        obv_threshold = 0.1    # Normalized slope threshold

        # Classify based on direction match
        price_up = price_change > price_threshold
        price_down = price_change < -price_threshold
        obv_up = normalized_slope > obv_threshold
        obv_down = normalized_slope < -obv_threshold

        if price_up and obv_down:
            return 'bearish_div', abs(normalized_slope)
        elif price_down and obv_up:
            return 'bullish_div', normalized_slope
        elif (price_up and obv_up) or (price_down and obv_down):
            return 'confirmation', abs(normalized_slope)
        else:
            return 'neutral', 0

    def get_pre_alert_data(self, symbol: str, alert_time: datetime,
                           minutes_back: int = 15) -> List[Dict]:
        """Get minute-by-minute data before an alert."""
        conn = self._get_conn()
        cursor = conn.cursor()

        start_time = (alert_time - timedelta(minutes=minutes_back)).strftime('%Y-%m-%d %H:%M:%S')
        end_time = alert_time.strftime('%Y-%m-%d %H:%M:%S')

        cursor.execute("""
            SELECT timestamp, price, volume
            FROM stock_quotes
            WHERE symbol = ?
            AND timestamp >= ?
            AND timestamp <= ?
            ORDER BY timestamp ASC
        """, (symbol, start_time, end_time))

        return [{'timestamp': r[0], 'price': r[1], 'volume': r[2] or 0}
                for r in cursor.fetchall()]

    def load_alerts(self, days: int = 30) -> List[Dict]:
        """Load 5-min alerts from Excel."""
        excel_path = config.ALERT_EXCEL_PATH
        workbook = openpyxl.load_workbook(excel_path, read_only=True)

        if "5min_alerts" not in workbook.sheetnames:
            workbook.close()
            return []

        sheet = workbook["5min_alerts"]
        cutoff = datetime.now() - timedelta(days=days)

        alerts = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            try:
                date_val, time_val = row[0], row[1]
                if isinstance(date_val, datetime):
                    date_str = date_val.strftime("%Y-%m-%d")
                else:
                    date_str = str(date_val)
                if isinstance(time_val, datetime):
                    time_str = time_val.strftime("%H:%M:%S")
                else:
                    time_str = str(time_val)

                try:
                    alert_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")
                except:
                    alert_dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")

                if alert_dt >= cutoff:
                    alerts.append({
                        'datetime': alert_dt,
                        'symbol': str(row[2]),
                        'direction': str(row[3]).lower() if row[3] else 'drop',
                        'change_pct': float(row[6]) if row[6] else 0,
                    })
            except:
                continue

        workbook.close()
        return alerts

    def analyze_alerts(self, days: int = 30) -> Dict:
        """
        Analyze OBV patterns for all alerts.

        Checks at different points before alert:
        - T-10, T-5, T-3, T-1 minutes
        """
        alerts = self.load_alerts(days)
        logger.info(f"Analyzing OBV for {len(alerts)} alerts...")

        results = {
            'total_alerts': len(alerts),
            'obv_confirmation': 0,  # OBV confirms price direction
            'obv_divergence': 0,    # OBV diverges from price
            'obv_neutral': 0,
            'details': []
        }

        # Track patterns at different timeframes
        patterns_at_t5 = defaultdict(int)
        patterns_at_t3 = defaultdict(int)

        # Track if confirmation leads to stronger moves
        confirmation_changes = []
        divergence_changes = []

        for alert in alerts:
            data = self.get_pre_alert_data(alert['symbol'], alert['datetime'], 15)

            if len(data) < 10:
                continue

            prices = [d['price'] for d in data]
            obv = self.calculate_obv(data)

            if len(obv) < 10:
                continue

            # Check at T-5 (5 minutes before alert)
            if len(obv) >= 10:
                t5_idx = len(obv) - 5
                obv_at_t5 = obv[:t5_idx]
                prices_at_t5 = prices[:t5_idx]

                if len(obv_at_t5) >= 5:
                    pattern, strength = self.check_obv_divergence(prices_at_t5, obv_at_t5, 5)
                    patterns_at_t5[pattern] += 1

            # Check at T-3 (3 minutes before alert)
            if len(obv) >= 8:
                t3_idx = len(obv) - 3
                obv_at_t3 = obv[:t3_idx]
                prices_at_t3 = prices[:t3_idx]

                if len(obv_at_t3) >= 5:
                    pattern, strength = self.check_obv_divergence(prices_at_t3, obv_at_t3, 5)
                    patterns_at_t3[pattern] += 1

                    # Track if pattern correlates with move strength
                    if pattern == 'confirmation':
                        confirmation_changes.append(abs(alert['change_pct']))
                        results['obv_confirmation'] += 1
                    elif pattern in ['bullish_div', 'bearish_div']:
                        divergence_changes.append(abs(alert['change_pct']))
                        results['obv_divergence'] += 1
                    else:
                        results['obv_neutral'] += 1

            # OBV slope before alert
            obv_slope = self.calculate_obv_slope(obv[:-1], 5)

            results['details'].append({
                'symbol': alert['symbol'],
                'time': alert['datetime'].strftime('%Y-%m-%d %H:%M'),
                'direction': alert['direction'],
                'change_pct': alert['change_pct'],
                'obv_slope': obv_slope,
                'pattern_at_t3': pattern if len(obv) >= 8 else 'unknown'
            })

        # Calculate statistics
        results['patterns_at_t5'] = dict(patterns_at_t5)
        results['patterns_at_t3'] = dict(patterns_at_t3)

        if confirmation_changes:
            results['avg_move_with_confirmation'] = sum(confirmation_changes) / len(confirmation_changes)
        if divergence_changes:
            results['avg_move_with_divergence'] = sum(divergence_changes) / len(divergence_changes)

        return results

    def analyze_false_positives(self, days: int = 30) -> Dict:
        """
        Analyze OBV patterns for early warning signals that did NOT lead to alerts.

        This helps determine if OBV can filter out false positives.
        """
        logger.info("Analyzing OBV for potential false positives...")

        conn = self._get_conn()
        cursor = conn.cursor()

        cutoff = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')

        # Get sample of dates
        cursor.execute("""
            SELECT DISTINCT date(timestamp)
            FROM stock_quotes
            WHERE timestamp >= ?
            ORDER BY timestamp
            LIMIT 5
        """, (cutoff,))

        dates = [row[0] for row in cursor.fetchall()]

        # Load actual alerts for comparison
        alerts = self.load_alerts(days)
        alert_keys = set()
        for a in alerts:
            # Key: symbol + time rounded to minute
            key = f"{a['symbol']}_{a['datetime'].strftime('%Y-%m-%d_%H:%M')}"
            alert_keys.add(key)

        fp_with_confirmation = 0
        fp_with_divergence = 0
        fp_with_neutral = 0

        tp_with_confirmation = 0
        tp_with_divergence = 0
        tp_with_neutral = 0

        signals_checked = 0

        for date_str in dates[:5]:  # Check 5 days
            cursor.execute("""
                SELECT DISTINCT symbol
                FROM stock_quotes
                WHERE date(timestamp) = ?
            """, (date_str,))

            symbols = [row[0] for row in cursor.fetchall()]

            for symbol in symbols[:100]:  # Check 100 symbols per day
                cursor.execute("""
                    SELECT timestamp, price, volume
                    FROM stock_quotes
                    WHERE symbol = ?
                    AND date(timestamp) = ?
                    AND time(timestamp) >= '09:25:00'
                    ORDER BY timestamp ASC
                """, (symbol, date_str))

                data = [{'timestamp': r[0], 'price': r[1], 'volume': r[2] or 0}
                        for r in cursor.fetchall()]

                if len(data) < 15:
                    continue

                prices = [d['price'] for d in data]
                obv = self.calculate_obv(data)

                last_signal_idx = -15

                # Scan for early warning signals (1% move in 3 min)
                for i in range(8, len(data) - 5):
                    if i - last_signal_idx < 15:
                        continue

                    # Check for 1% move
                    if i < 3:
                        continue

                    price_now = prices[i]
                    price_3min_ago = prices[i-3]

                    if price_3min_ago <= 0:
                        continue

                    change = abs(price_now - price_3min_ago) / price_3min_ago * 100

                    if change < 1.0:
                        continue

                    signals_checked += 1
                    last_signal_idx = i

                    # Check OBV pattern at this point
                    obv_slice = obv[:i+1]
                    price_slice = prices[:i+1]

                    if len(obv_slice) >= 5:
                        pattern, _ = self.check_obv_divergence(price_slice, obv_slice, 5)
                    else:
                        pattern = 'neutral'

                    # Check if 5-min alert followed within 5 minutes
                    alert_followed = False
                    for j in range(1, 6):
                        if i + j >= len(data):
                            break
                        ts = data[i + j]['timestamp']
                        key = f"{symbol}_{ts[:16]}"  # YYYY-MM-DD_HH:MM
                        if key in alert_keys:
                            alert_followed = True
                            break

                    # Also check by price move
                    if not alert_followed:
                        for j in range(1, 6):
                            if i + j >= len(data) or i + j < 5:
                                continue
                            fut_price = prices[i + j]
                            past_price = prices[i + j - 5]
                            if past_price > 0:
                                move = abs(fut_price - past_price) / past_price * 100
                                if move >= 1.25:
                                    alert_followed = True
                                    break

                    if alert_followed:
                        if pattern == 'confirmation':
                            tp_with_confirmation += 1
                        elif pattern in ['bullish_div', 'bearish_div']:
                            tp_with_divergence += 1
                        else:
                            tp_with_neutral += 1
                    else:
                        if pattern == 'confirmation':
                            fp_with_confirmation += 1
                        elif pattern in ['bullish_div', 'bearish_div']:
                            fp_with_divergence += 1
                        else:
                            fp_with_neutral += 1

        return {
            'signals_checked': signals_checked,
            'true_positives': {
                'confirmation': tp_with_confirmation,
                'divergence': tp_with_divergence,
                'neutral': tp_with_neutral,
                'total': tp_with_confirmation + tp_with_divergence + tp_with_neutral
            },
            'false_positives': {
                'confirmation': fp_with_confirmation,
                'divergence': fp_with_divergence,
                'neutral': fp_with_neutral,
                'total': fp_with_confirmation + fp_with_divergence + fp_with_neutral
            }
        }

    def close(self):
        if self.conn:
            self.conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--days", type=int, default=30)
    args = parser.parse_args()

    analyzer = OBVAnalyzer()

    try:
        # Analyze actual alerts
        print("\n" + "=" * 70)
        print("OBV PATTERN ANALYSIS FOR 5-MIN ALERTS")
        print("=" * 70)

        results = analyzer.analyze_alerts(args.days)

        print(f"\nAlerts Analyzed: {results['total_alerts']}")
        print(f"\nOBV Patterns at T-3 (3 min before alert):")
        for pattern, count in sorted(results.get('patterns_at_t3', {}).items()):
            pct = count / results['total_alerts'] * 100 if results['total_alerts'] > 0 else 0
            print(f"  {pattern}: {count} ({pct:.1f}%)")

        print(f"\nOBV Pattern Summary:")
        print(f"  Confirmation (OBV confirms move): {results['obv_confirmation']}")
        print(f"  Divergence (OBV opposes move):    {results['obv_divergence']}")
        print(f"  Neutral:                          {results['obv_neutral']}")

        if 'avg_move_with_confirmation' in results:
            print(f"\nAvg move strength with OBV confirmation: {results['avg_move_with_confirmation']:.2f}%")
        if 'avg_move_with_divergence' in results:
            print(f"Avg move strength with OBV divergence:   {results['avg_move_with_divergence']:.2f}%")

        # Analyze false positives
        print("\n" + "=" * 70)
        print("OBV PATTERN FOR TRUE VS FALSE POSITIVES")
        print("=" * 70)

        fp_results = analyzer.analyze_false_positives(args.days)

        print(f"\nSignals Checked: {fp_results['signals_checked']}")

        tp = fp_results['true_positives']
        fp = fp_results['false_positives']

        print(f"\nTrue Positives (led to alert):")
        print(f"  With OBV Confirmation: {tp['confirmation']}")
        print(f"  With OBV Divergence:   {tp['divergence']}")
        print(f"  Neutral:               {tp['neutral']}")

        print(f"\nFalse Positives (no alert followed):")
        print(f"  With OBV Confirmation: {fp['confirmation']}")
        print(f"  With OBV Divergence:   {fp['divergence']}")
        print(f"  Neutral:               {fp['neutral']}")

        # Calculate precision by OBV pattern
        print("\n" + "-" * 70)
        print("PRECISION BY OBV PATTERN")
        print("-" * 70)

        for pattern in ['confirmation', 'divergence', 'neutral']:
            tp_count = tp[pattern]
            fp_count = fp[pattern]
            total = tp_count + fp_count
            if total > 0:
                precision = tp_count / total * 100
                print(f"  {pattern.capitalize()}: {precision:.1f}% precision ({tp_count}/{total})")

        # Recommendation
        print("\n" + "=" * 70)
        print("RECOMMENDATION")
        print("=" * 70)

        conf_precision = tp['confirmation'] / (tp['confirmation'] + fp['confirmation']) * 100 if (tp['confirmation'] + fp['confirmation']) > 0 else 0
        div_precision = tp['divergence'] / (tp['divergence'] + fp['divergence']) * 100 if (tp['divergence'] + fp['divergence']) > 0 else 0

        if conf_precision > div_precision + 10:
            print(f"✅ OBV Confirmation is a USEFUL filter")
            print(f"   Signals with OBV confirmation have {conf_precision:.0f}% precision")
            print(f"   Consider ONLY sending pre-alerts when OBV confirms price direction")
        elif div_precision > conf_precision + 10:
            print(f"⚠️ OBV Divergence signals are actually BETTER")
            print(f"   This is unexpected - may indicate reversal patterns")
        else:
            print(f"❌ OBV pattern doesn't significantly improve prediction")
            print(f"   Confirmation: {conf_precision:.0f}%, Divergence: {div_precision:.0f}%")

        print()

    finally:
        analyzer.close()


if __name__ == "__main__":
    main()
